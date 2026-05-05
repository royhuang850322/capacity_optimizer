"""Generate a customer-facing workbook that explains five example products."""
from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.data_loader import (
    load_direct_mode_a_with_capacity_bases,
    load_direct_mode_b_with_capacity_bases,
)
from app.main import _merge_capacity_records
from app.models import AllocationResult, LoadRecord
from app.optimizer import run_optimization_mode_a, run_optimization_mode_b
from app.runtime_paths import resolve_runtime_paths


RUNTIME_PATHS = resolve_runtime_paths()
DEFAULT_INPUT_DIR = RUNTIME_PATHS.sample_data_dir
DEFAULT_OUTPUT_DIR = RUNTIME_PATHS.outputs_dir
DEFAULT_SCENARIO = "Expansion"
DEFAULT_START_MONTH = "2027-01"
DEFAULT_HORIZON = 60
DEFAULT_MODEB_BASIS = "Planned"

CASE_PRODUCTS: list[tuple[str, str]] = [
    ("TRS-02", "ModeA 已经完全消化需求，不需要 routing。"),
    ("LMB-03", "ModeA 先由 baseline capacity 消化一部分，ModeB 再用 routing 补上大部分剩余需求。"),
    ("AUF-01", "ModeA 先由内部 baseline 消化一部分，ModeB 再把剩余需求的一部分送去 Toller。"),
    ("VCG-03", "ModeB 的 routing 有帮助，但仍不足以消化全部剩余需求。"),
    ("AUP-04", "没有 routing，baseline capacity 只消化一部分，剩余继续记为 unmet。"),
]

THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=15, bold=True, color="1F1F1F")
SECTION_FONT = Font(size=11, bold=True, color="1F1F1F")

HEADER_LABELS = {
    "Product": "产品",
    "Example": "案例说明",
    "Planner_File": "Planner 来源文件",
    "Planner_Load_Resource": "Planner 填写资源",
    "Master_Capacity_WC": "master_capacity 工作中心",
    "Master_Routing_WC": "master_routing 内部路径工作中心",
    "Toller_WC": "Toller 工作中心",
    "ModeA_Internal_Tons": "ModeA 内部分配吨位",
    "ModeA_Unmet_Tons": "ModeA 未满足吨位",
    "ModeB_Capacity_Base_Tons": "ModeB Capacity_Base 吨位",
    "ModeB_Routing_Reroute_Tons": "ModeB Routing_Reroute 吨位",
    "ModeB_Outsourced_Tons": "ModeB 外协吨位",
    "ModeB_Unmet_Tons": "ModeB 未满足吨位",
    "Source_File": "来源文件",
    "Planner_Name": "Planner 名称",
    "Product_Family": "产品族",
    "Plant": "工厂",
    "Scenario": "场景",
    "Month_Span": "月份范围",
    "Demand_Tons": "需求吨位",
    "Record_Count": "记录数",
    "WorkCenter": "工作中心",
    "Annual_Capacity_Tons": "年产能吨位",
    "Monthly_Capacity_Tons": "月产能吨位",
    "Utilization_Target": "Utilization_Target",
    "Route_Type": "路径类型",
    "Max_Capacity_Tons": "Max 产能吨位",
    "Planned_Capacity_Tons": "Planner 产能吨位",
    "Eligible_Flag": "Eligible 标记",
    "Allocation_Source": "分配来源",
    "Internal_Tons": "内部承接吨位",
    "Share_of_Product_Demand_Pct": "占该产品需求比例(%)",
    "WC_Role": "工作中心角色",
    "Focus_Product_Tons": "当前产品吨位",
    "Total_WC_Internal_Tons": "该 WC 总内部吨位",
    "Focus_Product_Share_of_WC_Pct": "当前产品占该 WC 比例(%)",
    "Other_Product": "同 WC 其他产品",
    "Other_Product_Internal_Tons": "其他产品内部吨位",
    "Focus_Product_Outsourced_Tons": "当前产品外协吨位",
    "Other_Product_Outsourced_Tons": "其他产品外协吨位",
    "What_This_Case_Shows": "该案例说明",
    "Planner_Source_File": "Planner 来源文件",
    "Routing_Definition": "Routing 定义",
    "ModeA_Demand_Tons": "ModeA 需求吨位",
}

ROUTE_TYPE_LABELS = {
    "Primary": "主路径",
    "Alternative": "替代路径",
    "Toller": "外协路径",
}

ALLOCATION_SOURCE_LABELS = {
    "ModeA_Internal": "ModeA 内部分配",
    "Capacity_Base": "Capacity_Base",
    "Routing_Reroute": "Routing_Reroute",
    "ModeB_Internal": "ModeB 内部分配",
}

WC_ROLE_LABELS = {
    "ModeA_Internal_WC": "ModeA 内部分配 WC",
    "Capacity_Base": "ModeB Capacity_Base WC",
    "Routing_Reroute": "ModeB Routing_Reroute WC",
    "ModeB_WC": "ModeB 工作中心",
    "ModeB_Toller_Route": "ModeB Toller 路径",
}


def _label(text: str) -> str:
    return HEADER_LABELS.get(text, text)


def _route_type_label(text: str) -> str:
    return ROUTE_TYPE_LABELS.get(text, text)


def _allocation_source_label(text: str) -> str:
    return ALLOCATION_SOURCE_LABELS.get(text, text)


def _wc_role_label(text: str) -> str:
    return WC_ROLE_LABELS.get(text, text)


@dataclass
class ProductMetrics:
    demand: float = 0.0
    internal: float = 0.0
    outsourced: float = 0.0
    unmet: float = 0.0
    capacity_base: float = 0.0
    routing_reroute: float = 0.0


def _month_list(start_month: str, horizon: int) -> list[str]:
    year = int(start_month[:4])
    month = int(start_month[5:])
    months: list[str] = []
    for _ in range(horizon):
        months.append(f"{year}-{month:02d}")
        month += 1
        if month > 12:
            month = 1
            year += 1
    return months


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _safe_pct(numerator: float, denominator: float) -> float:
    if abs(denominator) <= 1e-9:
        return 0.0
    return 100.0 * numerator / denominator


def _product_metrics(results: Sequence[AllocationResult]) -> dict[str, ProductMetrics]:
    metrics = defaultdict(ProductMetrics)
    seen_demand: set[tuple[str, str]] = set()
    for result in results:
        product_metrics = metrics[result.product]
        demand_key = (result.month, result.product)
        if demand_key not in seen_demand:
            product_metrics.demand += float(result.demand_tons or 0.0)
            seen_demand.add(demand_key)
        if result.allocation_type == "Internal":
            product_metrics.internal += float(result.allocated_tons or 0.0)
            if result.allocation_source == "Capacity_Base":
                product_metrics.capacity_base += float(result.allocated_tons or 0.0)
            elif result.allocation_source == "Routing_Reroute":
                product_metrics.routing_reroute += float(result.allocated_tons or 0.0)
        elif result.allocation_type == "Outsourced":
            product_metrics.outsourced += float(result.outsourced_tons or 0.0)
        elif result.allocation_type == "Unmet":
            product_metrics.unmet += float(result.unmet_tons or 0.0)
    return metrics


def _internal_by_product_wc(results: Sequence[AllocationResult]) -> dict[tuple[str, str], float]:
    totals: dict[tuple[str, str], float] = defaultdict(float)
    for result in results:
        if result.allocation_type != "Internal":
            continue
        totals[(result.product, result.work_center)] += float(result.allocated_tons or 0.0)
    return totals


def _internal_by_product_wc_and_source(results: Sequence[AllocationResult]) -> dict[tuple[str, str, str], float]:
    totals: dict[tuple[str, str, str], float] = defaultdict(float)
    for result in results:
        if result.allocation_type != "Internal":
            continue
        source = result.allocation_source or "ModeA_Internal"
        totals[(result.product, source, result.work_center)] += float(result.allocated_tons or 0.0)
    return totals


def _wc_internal_totals(results: Sequence[AllocationResult]) -> dict[str, dict[str, float]]:
    totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for result in results:
        if result.allocation_type != "Internal":
            continue
        totals[result.work_center][result.product] += float(result.allocated_tons or 0.0)
    return totals


def _toller_wc_by_product(routing_rows: Sequence[dict[str, str]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for row in routing_rows:
        route_type = str(row.get("Router Type", "")).strip().lower()
        product = str(row.get("Product", "")).strip()
        wc = str(row.get("Resource", "")).strip()
        if product and wc and route_type == "toller":
            mapping[product] = wc
    return mapping


def _outsourced_by_toller_wc(
    results: Sequence[AllocationResult],
    toller_wc_lookup: dict[str, str],
) -> dict[str, dict[str, float]]:
    totals: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for result in results:
        if result.allocation_type != "Outsourced":
            continue
        wc = toller_wc_lookup.get(result.product)
        if not wc:
            continue
        totals[wc][result.product] += float(result.outsourced_tons or 0.0)
    return totals


def _planner_input_rows(loads: Sequence[LoadRecord], product: str) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str, str, str], dict[str, Any]] = {}
    product_loads = [load for load in loads if load.product == product]
    if not product_loads:
        return []

    months = sorted({load.month for load in product_loads})
    month_span = f"{months[0]} -> {months[-1]}" if months else ""
    for load in product_loads:
        key = (
            str(load.source_file or ""),
            str(load.planner_name or ""),
            str(load.product_family or ""),
            str(load.plant or ""),
            str(load.resource_group_owner or ""),
            str(load.scenario_version or load.scenario or ""),
        )
        row = grouped.setdefault(
            key,
            {
                "Source_File": key[0],
                "Planner_Name": key[1],
                "Product_Family": key[2],
                "Plant": key[3],
                "Planner_Load_Resource": key[4],
                "Scenario": key[5],
                "Month_Span": month_span,
                "Demand_Tons": 0.0,
                "Record_Count": 0,
            },
        )
        row["Demand_Tons"] += float(load.forecast_tons or 0.0)
        row["Record_Count"] += 1

    rows = list(grouped.values())
    rows.sort(key=lambda row: (row["Source_File"], row["Planner_Name"], row["Plant"]))
    return rows


def _master_capacity_rows(raw_capacity_rows: Sequence[dict[str, str]], product: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in raw_capacity_rows:
        if str(row.get("Product", "")).strip() != product:
            continue
        annual_capacity = float(row.get("Annual Capacity Tons", 0) or 0)
        rows.append(
            {
                "Source_File": "master_capacity.csv",
                "WorkCenter": str(row.get("Resource", "")).strip(),
                "Annual_Capacity_Tons": annual_capacity,
                "Monthly_Capacity_Tons": annual_capacity / 12.0,
                "Utilization_Target": str(row.get("Utilization Target", "")).strip(),
            }
        )
    return rows


def _master_routing_rows(raw_routing_rows: Sequence[dict[str, str]], product: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in raw_routing_rows:
        if str(row.get("Product", "")).strip() != product:
            continue
        rows.append(
            {
                "Source_File": "master_routing.csv",
                "WorkCenter": str(row.get("Resource", "")).strip(),
                "Route_Type": _route_type_label(str(row.get("Router Type", "")).strip()),
                "Max_Capacity_Tons": float(row.get("Max Capacity Ton", 0) or 0),
                "Planned_Capacity_Tons": float(row.get("Planned Capacity Ton", row.get("Planner Capacity Ton", 0)) or 0),
                "Eligible_Flag": str(row.get("EligibleFalg", "")).strip(),
            }
        )
    return rows


def _mode_allocation_rows(
    product: str,
    internal_totals: dict[tuple[str, str, str], float],
    demand_tons: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for (prod, source, wc), tons in sorted(
        internal_totals.items(),
        key=lambda item: (item[0][1], -item[1], item[0][2]),
    ):
        if prod != product or tons <= 0:
            continue
        rows.append(
            {
                "Allocation_Source": _allocation_source_label(source),
                "WorkCenter": wc,
                "Internal_Tons": tons,
                "Share_of_Product_Demand_Pct": _safe_pct(tons, demand_tons),
            }
        )
    return rows


def _modea_allocation_rows(
    product: str,
    internal_totals: dict[tuple[str, str], float],
    demand_tons: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for (prod, wc), tons in sorted(internal_totals.items(), key=lambda item: (-item[1], item[0][1])):
        if prod != product or tons <= 0:
            continue
        rows.append(
            {
                "Allocation_Source": _allocation_source_label("ModeA_Internal"),
                "WorkCenter": wc,
                "Internal_Tons": tons,
                "Share_of_Product_Demand_Pct": _safe_pct(tons, demand_tons),
            }
        )
    return rows


def _other_products_on_wc_rows(
    focus_product: str,
    wc_roles: Sequence[tuple[str, str]],
    wc_internal_totals: dict[str, dict[str, float]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for wc_role, wc in wc_roles:
        wc_products = wc_internal_totals.get(wc, {})
        total_wc_tons = sum(wc_products.values())
        other_rows = [
            (product, tons)
            for product, tons in sorted(wc_products.items(), key=lambda item: (-item[1], item[0]))
            if product != focus_product and tons > 0
        ]
        focus_tons = wc_products.get(focus_product, 0.0)
        if not other_rows:
            rows.append(
                {
                    "WC_Role": _wc_role_label(wc_role),
                    "WorkCenter": wc,
                    "Focus_Product_Tons": focus_tons,
                    "Total_WC_Internal_Tons": total_wc_tons,
                    "Focus_Product_Share_of_WC_Pct": _safe_pct(focus_tons, total_wc_tons),
                    "Other_Product": "（无）",
                    "Other_Product_Internal_Tons": 0.0,
                }
            )
            continue
        for other_product, other_tons in other_rows:
            rows.append(
                {
                    "WC_Role": _wc_role_label(wc_role),
                    "WorkCenter": wc,
                    "Focus_Product_Tons": focus_tons,
                    "Total_WC_Internal_Tons": total_wc_tons,
                    "Focus_Product_Share_of_WC_Pct": _safe_pct(focus_tons, total_wc_tons),
                    "Other_Product": other_product,
                    "Other_Product_Internal_Tons": other_tons,
                }
            )
    return rows


def _other_products_on_toller_rows(
    focus_product: str,
    toller_wc: str | None,
    focus_outsourced_tons: float,
    outsourced_by_wc: dict[str, dict[str, float]],
) -> list[dict[str, Any]]:
    if not toller_wc:
        return []
    wc_products = outsourced_by_wc.get(toller_wc, {})
    rows: list[dict[str, Any]] = []
    other_rows = [
        (product, tons)
        for product, tons in sorted(wc_products.items(), key=lambda item: (-item[1], item[0]))
        if product != focus_product and tons > 0
    ]
    if not other_rows:
        return [
            {
                "WC_Role": _wc_role_label("ModeB_Toller_Route"),
                "WorkCenter": toller_wc,
                "Focus_Product_Outsourced_Tons": focus_outsourced_tons,
                "Other_Product": "（无）",
                "Other_Product_Outsourced_Tons": 0.0,
            }
        ]
    for other_product, other_tons in other_rows:
        rows.append(
            {
                "WC_Role": _wc_role_label("ModeB_Toller_Route"),
                "WorkCenter": toller_wc,
                "Focus_Product_Outsourced_Tons": focus_outsourced_tons,
                "Other_Product": other_product,
                "Other_Product_Outsourced_Tons": other_tons,
            }
        )
    return rows


def _case_overview_rows(
    products: Sequence[tuple[str, str]],
    loads: Sequence[LoadRecord],
    raw_capacity_rows: Sequence[dict[str, str]],
    raw_routing_rows: Sequence[dict[str, str]],
    modea_metrics: dict[str, ProductMetrics],
    modeb_metrics: dict[str, ProductMetrics],
) -> list[dict[str, Any]]:
    planner_source_lookup: dict[str, str] = {}
    planner_resource_lookup: dict[str, str] = {}
    for product, _description in products:
        product_rows = [load for load in loads if load.product == product]
        source_files = sorted({str(load.source_file or "") for load in product_rows if load.source_file})
        resources = sorted({str(load.resource_group_owner or "") for load in product_rows if load.resource_group_owner})
        planner_source_lookup[product] = " | ".join(source_files)
        planner_resource_lookup[product] = " | ".join(resources)

    rows: list[dict[str, Any]] = []
    for product, description in products:
        cap_wcs = sorted(
            {
                str(row.get("Resource", "")).strip()
                for row in raw_capacity_rows
                if str(row.get("Product", "")).strip() == product
            }
        )
        routing_wcs = sorted(
            {
                str(row.get("Resource", "")).strip()
                for row in raw_routing_rows
                if str(row.get("Product", "")).strip() == product
                and str(row.get("Router Type", "")).strip().lower() != "toller"
            }
        )
        toller_wcs = sorted(
            {
                str(row.get("Resource", "")).strip()
                for row in raw_routing_rows
                if str(row.get("Product", "")).strip() == product
                and str(row.get("Router Type", "")).strip().lower() == "toller"
            }
        )
        modea = modea_metrics[product]
        modeb = modeb_metrics[product]
        rows.append(
            {
                "Product": product,
                "Example": description,
                "Planner_File": planner_source_lookup.get(product, ""),
                "Planner_Load_Resource": planner_resource_lookup.get(product, ""),
                "Master_Capacity_WC": " | ".join(cap_wcs),
                "Master_Routing_WC": " | ".join(routing_wcs) or "（无）",
                "Toller_WC": " | ".join(toller_wcs) or "（无）",
                "ModeA_Internal_Tons": modea.internal,
                "ModeA_Unmet_Tons": modea.unmet,
                "ModeB_Capacity_Base_Tons": modeb.capacity_base,
                "ModeB_Routing_Reroute_Tons": modeb.routing_reroute,
                "ModeB_Outsourced_Tons": modeb.outsourced,
                "ModeB_Unmet_Tons": modeb.unmet,
            }
        )
    return rows


def _style_title(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_note(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = NOTE_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_section(ws, row: int, text: str, end_col: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_table(
    ws,
    start_row: int,
    rows: Sequence[dict[str, Any]],
    table_name: str,
    pct_columns: Iterable[str] | None = None,
    ton_columns: Iterable[str] | None = None,
) -> int:
    if not rows:
        ws.cell(start_row, 1).value = "（无）"
        return start_row

    pct_columns = set(pct_columns or [])
    ton_columns = set(ton_columns or [])
    headers = list(rows[0].keys())

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_index)
        cell.value = _label(header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for row_offset, row_data in enumerate(rows, start=1):
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(start_row + row_offset, col_index)
            value = row_data.get(header)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if header in ton_columns and isinstance(value, (int, float)):
                cell.number_format = '#,##0.0'
            elif header in pct_columns and isinstance(value, (int, float)):
                cell.number_format = '0.0'

    end_row = start_row + len(rows)
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    return end_row


def _autofit(ws) -> None:
    max_widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            max_widths[cell.column] = max(max_widths[cell.column], len(value))
    for col_index, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 12), 36)


def _case_summary_row(
    product: str,
    description: str,
    planner_rows: Sequence[dict[str, Any]],
    modea: ProductMetrics,
    modeb: ProductMetrics,
    routing_rows: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    planner_files = " | ".join(sorted({str(row["Source_File"]) for row in planner_rows})) or "（无）"
    routing_summary = " | ".join(
        f'{row["Route_Type"]}:{row["WorkCenter"]}'
        for row in routing_rows
    ) or "（无）"
    return [
        {
            "Product": product,
            "What_This_Case_Shows": description,
            "Planner_Source_File": planner_files,
            "Routing_Definition": routing_summary,
            "ModeA_Demand_Tons": modea.demand,
            "ModeA_Internal_Tons": modea.internal,
            "ModeA_Unmet_Tons": modea.unmet,
            "ModeB_Capacity_Base_Tons": modeb.capacity_base,
            "ModeB_Routing_Reroute_Tons": modeb.routing_reroute,
            "ModeB_Outsourced_Tons": modeb.outsourced,
            "ModeB_Unmet_Tons": modeb.unmet,
        }
    ]


def generate_customer_case_report(
    *,
    input_dir: Path,
    output_dir: Path,
    scenario: str = DEFAULT_SCENARIO,
    start_month: str = DEFAULT_START_MONTH,
    horizon: int = DEFAULT_HORIZON,
    case_products: Sequence[tuple[str, str]] = tuple(CASE_PRODUCTS),
    modeb_basis: str = DEFAULT_MODEB_BASIS,
) -> Path:
    months = _month_list(start_month, horizon)
    loads_a, modea_capacities_by_basis, _ = load_direct_mode_a_with_capacity_bases(
        str(input_dir),
        str(input_dir),
        selected_scenario=scenario,
    )
    modea_results = run_optimization_mode_a(
        months,
        loads_a,
        modea_capacities_by_basis["Planned"],
        verbose=False,
    )

    loads_b, modeb_baseline_capacities_by_basis, capacities_by_basis, routings = load_direct_mode_b_with_capacity_bases(
        str(input_dir),
        str(input_dir),
        selected_scenario=scenario,
    )
    modeb_results, _toller_products = run_optimization_mode_b(
        months,
        loads_b,
        modeb_baseline_capacities_by_basis[modeb_basis],
        capacities_by_basis[modeb_basis],
        routings,
        verbose=False,
    )

    raw_capacity_rows = _read_csv_rows(input_dir / "master_capacity.csv")
    raw_routing_rows = _read_csv_rows(input_dir / "master_routing.csv")

    modea_metrics = _product_metrics(modea_results)
    modeb_metrics = _product_metrics(modeb_results)
    modea_internal = _internal_by_product_wc(modea_results)
    modeb_internal = _internal_by_product_wc_and_source(modeb_results)
    modea_wc_totals = _wc_internal_totals(modea_results)
    modeb_wc_totals = _wc_internal_totals(modeb_results)
    toller_lookup = _toller_wc_by_product(raw_routing_rows)
    outsourced_by_toller_wc = _outsourced_by_toller_wc(modeb_results, toller_lookup)

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_ws = workbook.create_sheet("总览")
    _style_title(summary_ws, 1, "客户案例讲解 - 5 个示例产品", 14)
    _style_note(
        summary_ws,
        2,
        (
            f"场景 = {scenario} | 期间 = 从 {start_month} 开始，共 {horizon} 个月 | "
            f"ModeB 口径 = {modeb_basis}。这份 workbook 把原本分散在 planner 输入、master_capacity、"
            "master_routing、Allocation_Detail、WC_Heatmap 和 Bottleneck 里的逻辑，集中解释给客户看。"
        ),
        14,
    )
    next_row = _write_section(summary_ws, 4, "案例总览", 14)
    overview_rows = _case_overview_rows(
        case_products,
        loads_b,
        raw_capacity_rows,
        raw_routing_rows,
        modea_metrics,
        modeb_metrics,
    )
    next_row = _write_table(
        summary_ws,
        next_row,
        overview_rows,
        table_name="CaseOverview",
        ton_columns={
            "ModeA_Internal_Tons",
            "ModeA_Unmet_Tons",
            "ModeB_Capacity_Base_Tons",
            "ModeB_Routing_Reroute_Tons",
            "ModeB_Outsourced_Tons",
            "ModeB_Unmet_Tons",
        },
    ) + 2
    _write_section(summary_ws, next_row, "如何阅读每个产品页", 14)
    summary_ws.cell(next_row + 1, 1).value = "1. Planner 输入汇总：说明需求来自哪个 planner 文件，以及 planner 原始填写了哪个资源。"
    summary_ws.cell(next_row + 2, 1).value = "2. Master Capacity / Master Routing：说明这个产品在主数据里到底有哪些工作中心和路径。"
    summary_ws.cell(next_row + 3, 1).value = "3. ModeA / ModeB 分配结果：说明 baseline capacity、routing reroute、toller、unmet 各吃掉了多少。"
    summary_ws.cell(next_row + 4, 1).value = "4. 工作中心上下文：说明同一个 WC 里还排进了哪些其他产品，帮助客户理解资源竞争关系。"

    for sheet_index, (product, description) in enumerate(case_products, start=1):
        ws = workbook.create_sheet(f"{sheet_index}_{product}")
        _style_title(ws, 1, f"{product} - 客户示例", 12)
        _style_note(ws, 2, description, 12)
        planner_rows = _planner_input_rows(loads_b, product)
        capacity_rows = _master_capacity_rows(raw_capacity_rows, product)
        routing_rows = _master_routing_rows(raw_routing_rows, product)
        modea = modea_metrics[product]
        modeb = modeb_metrics[product]

        current_row = _write_section(ws, 4, "案例摘要", 12)
        current_row = _write_table(
            ws,
            current_row,
            _case_summary_row(product, description, planner_rows, modea, modeb, routing_rows),
            table_name=f"CaseSummary{sheet_index}",
            ton_columns={
                "ModeA_Demand_Tons",
                "ModeA_Internal_Tons",
                "ModeA_Unmet_Tons",
                "ModeB_Capacity_Base_Tons",
                "ModeB_Routing_Reroute_Tons",
                "ModeB_Outsourced_Tons",
                "ModeB_Unmet_Tons",
            },
        ) + 2

        current_row = _write_section(ws, current_row, "Planner 输入汇总", 12)
        current_row = _write_table(
            ws,
            current_row,
            planner_rows,
            table_name=f"PlannerInput{sheet_index}",
            ton_columns={"Demand_Tons"},
        ) + 2

        current_row = _write_section(ws, current_row, "这个产品用到的 master_capacity 行", 12)
        current_row = _write_table(
            ws,
            current_row,
            capacity_rows,
            table_name=f"CapacityInput{sheet_index}",
            ton_columns={"Annual_Capacity_Tons", "Monthly_Capacity_Tons"},
        ) + 2

        current_row = _write_section(ws, current_row, "这个产品用到的 master_routing 行", 12)
        current_row = _write_table(
            ws,
            current_row,
            routing_rows
            if routing_rows
            else [
                {
                    "Source_File": "master_routing.csv",
                    "WorkCenter": "（无）",
                    "Route_Type": "没有 product-level routing 行",
                    "Max_Capacity_Tons": 0.0,
                    "Planned_Capacity_Tons": 0.0,
                    "Eligible_Flag": "",
                }
            ],
            table_name=f"RoutingInput{sheet_index}",
            ton_columns={"Max_Capacity_Tons", "Planned_Capacity_Tons"},
        ) + 2

        current_row = _write_section(ws, current_row, "ModeA 分配路径", 12)
        modea_rows = _modea_allocation_rows(product, modea_internal, modea.demand)
        current_row = _write_table(
            ws,
            current_row,
            modea_rows
            if modea_rows
            else [
                {
                    "Allocation_Source": _allocation_source_label("ModeA_Internal"),
                    "WorkCenter": "（无）",
                    "Internal_Tons": 0.0,
                    "Share_of_Product_Demand_Pct": 0.0,
                }
            ],
            table_name=f"ModeAAlloc{sheet_index}",
            ton_columns={"Internal_Tons"},
            pct_columns={"Share_of_Product_Demand_Pct"},
        ) + 2

        modea_wcs = [("ModeA_Internal_WC", row["WorkCenter"]) for row in modea_rows]
        current_row = _write_section(ws, current_row, "同一个 ModeA 工作中心里还有哪些其他产品", 12)
        current_row = _write_table(
            ws,
            current_row,
            _other_products_on_wc_rows(product, modea_wcs, modea_wc_totals)
            if modea_wcs
            else [
                {
                    "WC_Role": _wc_role_label("ModeA_Internal_WC"),
                    "WorkCenter": "（无）",
                    "Focus_Product_Tons": 0.0,
                    "Total_WC_Internal_Tons": 0.0,
                    "Focus_Product_Share_of_WC_Pct": 0.0,
                    "Other_Product": "（无）",
                    "Other_Product_Internal_Tons": 0.0,
                }
            ],
            table_name=f"ModeAWC{sheet_index}",
            ton_columns={"Focus_Product_Tons", "Total_WC_Internal_Tons", "Other_Product_Internal_Tons"},
            pct_columns={"Focus_Product_Share_of_WC_Pct"},
        ) + 2

        current_row = _write_section(ws, current_row, f"ModeB 分配路径（{modeb_basis} 口径）", 12)
        modeb_rows = _mode_allocation_rows(product, modeb_internal, modeb.demand)
        current_row = _write_table(
            ws,
            current_row,
            modeb_rows
            if modeb_rows
            else [
                {
                    "Allocation_Source": _allocation_source_label("ModeB_Internal"),
                    "WorkCenter": "（无）",
                    "Internal_Tons": 0.0,
                    "Share_of_Product_Demand_Pct": 0.0,
                }
            ],
            table_name=f"ModeBAlloc{sheet_index}",
            ton_columns={"Internal_Tons"},
            pct_columns={"Share_of_Product_Demand_Pct"},
        ) + 2

        modeb_wcs = [(row["Allocation_Source"], row["WorkCenter"]) for row in modeb_rows]
        current_row = _write_section(ws, current_row, "同一个 ModeB 工作中心里还有哪些其他产品", 12)
        current_row = _write_table(
            ws,
            current_row,
            _other_products_on_wc_rows(product, modeb_wcs, modeb_wc_totals)
            if modeb_wcs
            else [
                {
                    "WC_Role": _wc_role_label("ModeB_WC"),
                    "WorkCenter": "（无）",
                    "Focus_Product_Tons": 0.0,
                    "Total_WC_Internal_Tons": 0.0,
                    "Focus_Product_Share_of_WC_Pct": 0.0,
                    "Other_Product": "（无）",
                    "Other_Product_Internal_Tons": 0.0,
                }
            ],
            table_name=f"ModeBWC{sheet_index}",
            ton_columns={"Focus_Product_Tons", "Total_WC_Internal_Tons", "Other_Product_Internal_Tons"},
            pct_columns={"Focus_Product_Share_of_WC_Pct"},
        ) + 2

        if modeb.outsourced > 0:
            current_row = _write_section(ws, current_row, "同一条 Toller 路径上还有哪些其他产品", 12)
            current_row = _write_table(
                ws,
                current_row,
                _other_products_on_toller_rows(product, toller_lookup.get(product), modeb.outsourced, outsourced_by_toller_wc),
                table_name=f"TollerWC{sheet_index}",
                ton_columns={"Focus_Product_Outsourced_Tons", "Other_Product_Outsourced_Tons"},
            ) + 2

        if not routing_rows:
            _style_note(
                ws,
                current_row,
                "这个产品在 master_routing.csv 里没有 product-level routing 行，所以 ModeB 只能复用 Stage 1 baseline capacity 的结果，不存在 reroute 或 toller 逻辑。",
                12,
            )
            current_row += 2
        elif modeb.routing_reroute <= 0 and modeb.outsourced <= 0 and modeb.unmet > 0:
            _style_note(
                ws,
                current_row,
                "输入里虽然定义了 routing，但相关 routing 工作中心在排完其他产品后没有足够余量，因此剩余需求仍然保留为 unmet。",
                12,
            )
            current_row += 2

        _autofit(ws)

    _autofit(summary_ws)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"customer_case_walkthrough_cn_{scenario}_{timestamp}.xlsx"
    workbook.save(output_path)
    workbook.close()
    return output_path


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate the customer-facing 5-product example workbook.")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--scenario", default=DEFAULT_SCENARIO)
    parser.add_argument("--start-month", default=DEFAULT_START_MONTH)
    parser.add_argument("--horizon", type=int, default=DEFAULT_HORIZON)
    parser.add_argument("--modeb-basis", choices=("Max", "Planned"), default=DEFAULT_MODEB_BASIS)
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    output_path = generate_customer_case_report(
        input_dir=Path(args.input_dir),
        output_dir=Path(args.output_dir),
        scenario=args.scenario,
        start_month=args.start_month,
        horizon=args.horizon,
        modeb_basis=args.modeb_basis,
    )
    print(f"客户案例 workbook 已生成: {output_path}")


if __name__ == "__main__":
    main()
