import glob
import logging
import os
import tempfile
import unittest
from contextlib import contextmanager
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
from click.testing import CliRunner
from openpyxl import load_workbook

from app.create_template import write_control_workbook
from app.data_loader import load_config
from app.main import main as main_cli
from app.runtime_paths import ensure_workspace_dirs, resolve_runtime_paths
from app.workspace_init import initialize_user_workspace


TEST_TMP_ROOT = os.path.join(os.path.dirname(__file__), "_tmp")
os.makedirs(TEST_TMP_ROOT, exist_ok=True)


@contextmanager
def workspace_tempdir():
    with tempfile.TemporaryDirectory(dir=TEST_TMP_ROOT) as tmpdir:
        yield tmpdir


class SmokeM8Tests(unittest.TestCase):
    def tearDown(self) -> None:
        self._close_logger_handlers()

    @staticmethod
    def _close_logger_handlers() -> None:
        logger = logging.getLogger("capacity_optimizer")
        for handler in list(logger.handlers):
            logger.removeHandler(handler)
            try:
                handler.close()
            except Exception:
                pass

    def test_smoke_paths_workspace_initialization(self):
        with workspace_tempdir() as tmpdir:
            env = {"CAPACITY_OPTIMIZER_WORKSPACE": tmpdir}
            with patch.dict(os.environ, env, clear=False):
                paths = ensure_workspace_dirs(resolve_runtime_paths())
                result = initialize_user_workspace(paths)

            self.assertTrue(paths.user_workspace_dir.exists())
            self.assertTrue(paths.templates_dir.exists())
            self.assertTrue(paths.outputs_dir.exists())
            self.assertTrue(paths.logs_dir.exists())
            self.assertTrue(paths.license_active_dir.exists())
            self.assertTrue(result.paths.control_workbook_path.exists())
            self.assertTrue(result.paths.workspace_manifest_path.exists())

    def test_smoke_config_load_from_control_workbook(self):
        with workspace_tempdir() as tmpdir:
            project_root = os.path.join(tmpdir, "project")
            workbook_path = self._prepare_smoke_project(project_root)

            config = load_config(workbook_path)

            self.assertEqual(config.project_root_folder, os.path.abspath(project_root))
            self.assertEqual(config.input_load_folder, os.path.join(os.path.abspath(project_root), "Data_Input"))
            self.assertEqual(config.input_master_folder, os.path.join(os.path.abspath(project_root), "Data_Input"))
            self.assertEqual(config.output_folder, os.path.join(os.path.abspath(project_root), "output"))
            self.assertEqual(config.run_mode, "ModeA")
            self.assertTrue(config.direct_mode)
            self.assertEqual(config.start_month, "2026-01")
            self.assertEqual(config.horizon_months, 1)

    def test_smoke_run_and_output_workbook(self):
        with workspace_tempdir() as tmpdir:
            project_root = os.path.join(tmpdir, "project")
            workspace_dir = os.path.join(tmpdir, "workspace")
            workbook_path = self._prepare_smoke_project(project_root)
            fake_license = self._fake_license_info(project_root)
            runner = CliRunner()

            with patch.dict(
                os.environ,
                {"CAPACITY_OPTIMIZER_WORKSPACE": workspace_dir},
                clear=False,
            ), patch(
                "app.license_validator.validate_license_with_fallback",
                return_value=fake_license,
            ):
                result = runner.invoke(main_cli, ["--input-template", workbook_path])

            self.assertEqual(result.exit_code, 0, msg=result.output)

            output_pattern = os.path.join(project_root, "output", "*ModeA*.xlsx")
            output_files = sorted(glob.glob(output_pattern))
            self.assertTrue(output_files, msg=f"No output workbook found: {output_pattern}")

            output_wb = load_workbook(output_files[-1], data_only=True)
            self.assertIn("Dashboard", output_wb.sheetnames)
            self.assertIn("Allocation_Detail", output_wb.sheetnames)
            self.assertIn("Run_Info", output_wb.sheetnames)
            output_wb.close()

            logs_pattern = os.path.join(workspace_dir, "logs", "optimizer_run_*.log")
            run_logs = sorted(glob.glob(logs_pattern))
            self.assertTrue(run_logs, msg=f"No run log found: {logs_pattern}")
            self._close_logger_handlers()

    def _prepare_smoke_project(self, project_root: str) -> str:
        project_root_abs = os.path.abspath(project_root)
        data_dir = os.path.join(project_root_abs, "Data_Input")
        output_dir = os.path.join(project_root_abs, "output")
        os.makedirs(data_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)

        pd.DataFrame(
            [
                {
                    "Month": "2026-01",
                    "PlannerName": "PlannerA",
                    "Product": "P1",
                    "ProductFamily": "F1",
                    "Plant": "PLT1",
                    "Forecast_Tons": 10.0,
                    "Resource": "WC1",
                    "Scenario": "Base Scenario",
                }
            ]
        ).to_csv(os.path.join(data_dir, "planner1_load.csv"), index=False)

        pd.DataFrame(
            [
                {
                    "Product": "P1",
                    "WorkCenter": "WC1",
                    "Annual_Capacity_Tons": 120.0,
                    "Utilization_Target": 1.0,
                }
            ]
        ).to_csv(os.path.join(data_dir, "master_capacity.csv"), index=False)

        workbook_path = os.path.join(project_root_abs, "Tooling Control Panel", "Capacity_Optimizer_Control.xlsx")
        write_control_workbook(workbook_path, load_dir=data_dir)
        self._set_control_values(workbook_path, project_root_abs)
        return workbook_path

    def _set_control_values(self, workbook_path: str, project_root_abs: str) -> None:
        workbook = load_workbook(workbook_path)
        worksheet = workbook["Control_Panel"]
        rows = {
            worksheet[f"A{row_num}"].value: row_num
            for row_num in range(2, 40)
            if worksheet[f"A{row_num}"].value
        }
        worksheet[f"B{rows['Project_Root_Folder']}"] = project_root_abs
        worksheet[f"B{rows['Input_Load_Folder']}"] = "Data_Input"
        worksheet[f"B{rows['Input_Master_Folder']}"] = "Data_Input"
        worksheet[f"B{rows['Output_Folder']}"] = "output"
        worksheet[f"B{rows['Output_FileName']}"] = "smoke_result.xlsx"
        worksheet[f"B{rows['Scenario_Name']}"] = "Base Scenario"
        worksheet[f"B{rows['Start_Year']}"] = 2026
        worksheet[f"B{rows['Start_Month_Num']}"] = 1
        worksheet[f"B{rows['Horizon_Months']}"] = 1
        worksheet[f"B{rows['Run_Mode']}"] = "ModeA"
        worksheet[f"B{rows['Direct_Mode']}"] = "Yes"
        worksheet[f"B{rows['Verbose']}"] = "No"
        worksheet[f"B{rows['Skip_Validation_Errors']}"] = "No"
        workbook.save(workbook_path)
        workbook.close()

    @staticmethod
    def _fake_license_info(project_root: str) -> SimpleNamespace:
        return SimpleNamespace(
            license_id="LIC-SMOKE-001",
            license_type="trial",
            customer_name="SmokeTest",
            customer_id="SMOKE",
            issue_date="2026-01-01",
            expiry_date="2099-12-31",
            binding_mode="unbound",
            machine_fingerprint="",
            machine_label="",
            note="smoke test",
            features={},
            license_path=os.path.join(project_root, "licenses", "active", "license.json"),
            project_root=os.path.abspath(project_root),
            status="Valid",
        )


if __name__ == "__main__":
    unittest.main()
