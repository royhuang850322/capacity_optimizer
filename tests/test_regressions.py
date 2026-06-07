import os
import shutil
import unittest
import uuid
from contextlib import contextmanager
from types import SimpleNamespace

import pandas as pd
from click.testing import CliRunner
from openpyxl import load_workbook

from app.data_loader import (
    _aggregate_load_records,
    _norm_month,
    _parse_load_df,
    load_direct_mode_a,
    load_direct_mode_a_with_capacity_bases,
)
from app.create_template import main as create_template_main, refresh_control_workbook_license_sheet
from app.data_loader import load_config
from app.load_pressure import build_dashboard_fact_frame, build_pressure_load_frame
from app.main import _validate_input_setup
from app.models import AllocationResult, CapacityRecord, Config, LoadRecord, RoutingRecord
from app.optimizer import _build_demand, run_optimization_mode_a, run_optimization_mode_b
from app.output_writer import write_capacity_basis_results, write_mode_comparison_summary, write_results
from app.validator import ValidationIssue, format_issue_report, validate


TEST_TMP_ROOT = os.path.join(os.path.dirname(__file__), "_tmp")
os.makedirs(TEST_TMP_ROOT, exist_ok=True)


@contextmanager
def workspace_tempdir():
    tmpdir = os.path.join(TEST_TMP_ROOT, f"tmp_{uuid.uuid4().hex}")
    os.mkdir(tmpdir)
    try:
        yield tmpdir
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


class RegressionTests(unittest.TestCase):
    def test_capacity_record_ignores_utilization_target_in_effective_capacity(self):
        record = CapacityRecord(
            product="P1",
            work_center="WC1",
            annual_capacity_tons=1200.0,
            utilization_target=0.5,
        )

        self.assertAlmostEqual(record.monthly_capacity_tons, 100.0)
        self.assertAlmostEqual(record.effective_monthly_capacity_tons, 100.0)

    def test_load_direct_accepts_lowercase_load_headers(self):
        with workspace_tempdir() as tmpdir:
            pd.DataFrame(
                [
                    {
                        "month": "2025-01",
                        "plannername": "P1",
                        "product": "0001",
                        "productfamily": "F1",
                        "plant": "A",
                        "forecast_tons": 10,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "planner1_load.csv"), index=False)
            pd.DataFrame(
                [
                    {
                        "Product": "1",
                        "WorkCenter": "WC1",
                        "Annual_Capacity_Tons": 120,
                        "Utilization_Target": 1,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "master_capacity.csv"), index=False)

            loads, capacities, routings = load_direct_mode_a(tmpdir, tmpdir)

        self.assertEqual(len(loads), 1)
        self.assertEqual(loads[0].product, "1")
        self.assertEqual(len(capacities), 1)
        self.assertEqual(routings, [])

    def test_load_direct_mode_a_uses_master_capacity_even_when_master_routing_exists(self):
        with workspace_tempdir() as tmpdir:
            pd.DataFrame(
                [
                    {
                        "month": "2025-01",
                        "plannername": "P1",
                        "product": "P1",
                        "productfamily": "F1",
                        "plant": "A",
                        "forecast_tons": 10,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "planner1_load.csv"), index=False)
            pd.DataFrame(
                [
                    {
                        "Product": "P1",
                        "WorkCenter": "WC1",
                        "Annual_Capacity_Tons": 100,
                        "Utilization_Target": 1,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "master_capacity.csv"), index=False)
            pd.DataFrame(
                [
                    {
                        "Product": "P1",
                        "Resource": "WC1",
                        "Max Capacity Ton": 240.0,
                        "Designed Capacity Ton": 180.0,
                        "EligibleFalg": 0.5,
                        "Router Type": "Primary",
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "master_routing.csv"), index=False)

            _loads, capacities, routings = load_direct_mode_a(tmpdir, tmpdir)

        self.assertEqual(routings, [])
        self.assertEqual(len(capacities), 1)
        self.assertEqual(capacities[0].annual_capacity_tons, 100.0)
        self.assertAlmostEqual(capacities[0].utilization_target, 1.0)

    def test_load_direct_ignores_utilization_target_values(self):
        with workspace_tempdir() as tmpdir:
            pd.DataFrame(
                [
                    {
                        "Month": "2025-01",
                        "PlannerName": "P1",
                        "Product": "P1",
                        "ProductFamily": "F1",
                        "Plant": "A",
                        "Forecast_Tons": 10,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "planner1_load.csv"), index=False)
            pd.DataFrame(
                [
                    {
                        "Product": "P1",
                        "WorkCenter": "WC1",
                        "Annual_Capacity_Tons": 120,
                        "Utilization_Target": 0.25,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "master_capacity.csv"), index=False)

            _loads, capacities, _routings = load_direct_mode_a(tmpdir, tmpdir)

        self.assertEqual(len(capacities), 1)
        self.assertAlmostEqual(capacities[0].utilization_target, 1.0)
        self.assertAlmostEqual(capacities[0].effective_monthly_capacity_tons, 10.0)

    def test_load_direct_mode_a_with_capacity_bases_prefers_dual_capacity_columns(self):
        with workspace_tempdir() as tmpdir:
            pd.DataFrame(
                [
                    {
                        "Month": "2025-01",
                        "PlannerName": "P1",
                        "Product": "P1",
                        "ProductFamily": "F1",
                        "Plant": "A",
                        "Forecast_Tons": 10,
                        "Resource": "WC1",
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "planner1_load.csv"), index=False)
            pd.DataFrame(
                [
                    {
                        "Product": "P1",
                        "WorkCenter": "WC1",
                        "Annual_Max_Capacity_Tons": 240.0,
                        "Annual_Planned_Capacity_Tons": 120.0,
                        "Annual_Capacity_Tons": 999.0,
                        "Utilization_Target": 1.0,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "master_capacity.csv"), index=False)

            _loads, capacities_by_basis, _routings = load_direct_mode_a_with_capacity_bases(tmpdir, tmpdir)

        self.assertAlmostEqual(capacities_by_basis["Max"][0].annual_capacity_tons, 240.0)
        self.assertAlmostEqual(capacities_by_basis["Planned"][0].annual_capacity_tons, 120.0)

    def test_mode_a_max_vs_planner_capacity_can_change_internal_allocation(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        max_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        planner_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
        ]

        max_results = run_optimization_mode_a(
            months=["2025-01"],
            loads=loads,
            capacities=max_capacities,
        )
        planner_results = run_optimization_mode_a(
            months=["2025-01"],
            loads=loads,
            capacities=planner_capacities,
        )

        max_internal = sum(row.allocated_tons for row in max_results if row.allocation_type == "Internal")
        planner_internal = sum(row.allocated_tons for row in planner_results if row.allocation_type == "Internal")
        max_unmet = max(row.unmet_tons for row in max_results if row.product == "P1")
        planner_unmet = max(row.unmet_tons for row in planner_results if row.product == "P1")

        self.assertAlmostEqual(max_internal, 100.0, places=4)
        self.assertAlmostEqual(planner_internal, 60.0, places=4)
        self.assertAlmostEqual(max_unmet, 0.0, places=4)
        self.assertAlmostEqual(planner_unmet, 40.0, places=4)

    def test_mode_a_setup_triggers_above_one_ton_and_consumes_capacity(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=95.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                setup_hours=74.4,
                setup_reference_annual_capacity_tons=1200.0,
            ),
        ]

        results = run_optimization_mode_a(["2025-01"], loads, capacities)

        internal = [row for row in results if row.allocation_type == "Internal"][0]
        unmet = [row for row in results if row.allocation_type == "Unmet"][0]
        self.assertAlmostEqual(internal.allocated_tons, 90.0)
        self.assertTrue(internal.setup_applied)
        self.assertAlmostEqual(internal.setup_hours, 74.4)
        self.assertAlmostEqual(internal.setup_equivalent_tons_by_max, 10.0)
        self.assertAlmostEqual(internal.capacity_used_tons, 100.0)
        self.assertAlmostEqual(unmet.unmet_tons, 5.0)

    def test_mode_a_setup_does_not_trigger_at_one_ton(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=1.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                setup_hours=744.0,
                setup_reference_annual_capacity_tons=1200.0,
            ),
        ]

        results = run_optimization_mode_a(["2025-01"], loads, capacities)

        internal = [row for row in results if row.allocation_type == "Internal"][0]
        self.assertAlmostEqual(internal.allocated_tons, 1.0)
        self.assertFalse(internal.setup_applied)
        self.assertAlmostEqual(internal.setup_equivalent_tons_by_max, 0.0)

    def test_validate_ignores_utilization_target_range(self):
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=120.0,
                utilization_target=1.5,
            )
        ]

        issues = validate([], capacities, [], mode="ModeA")

        self.assertNotIn(("ERROR", "CapacityUtilRange"), {(i.severity, i.check) for i in issues})

    def test_validate_rejects_missing_setup_hours(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
            )
        ]
        capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]

        issues = validate(loads, capacities, [], mode="ModeA")

        self.assertIn("SetupHoursRequired", {issue.check for issue in issues})

    def test_validate_rejects_capacity_routing_setup_mismatch(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
            )
        ]
        baseline_capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                setup_hours=2.0,
                source_file="master_capacity.csv",
                row_num=2,
            ),
        ]
        routing_capacities = [
            *baseline_capacities,
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                setup_hours=2.0002,
                source_file="master_routing.csv",
                row_num=2,
            ),
        ]
        routings = [
            RoutingRecord(
                product="P1",
                product_family="F1",
                work_center="WC1",
                priority=1,
                eligible_flag=True,
                route_type="Primary",
                max_capacity_tons=1200.0,
                planned_capacity_tons=1200.0,
                setup_hours=2.0002,
                source_file="master_routing.csv",
                row_num=2,
            )
        ]

        issues = validate(
            loads,
            baseline_capacities,
            routings,
            mode="ModeB",
            routing_capacities=routing_capacities,
        )

        self.assertIn("CapacityRoutingSetupMismatch", {issue.check for issue in issues})

    def test_parse_load_preserves_negative_tons_for_validation(self):
        loads = _parse_load_df(
            pd.DataFrame(
                [
                    {
                        "Month": "2025-01",
                        "PlannerName": "P1",
                        "Product": "0001",
                        "ProductFamily": "F1",
                        "Plant": "A",
                        "Forecast_Tons": -5,
                    }
                ]
            ),
            source_file="demo.csv",
        )

        issues = validate(loads, [], [], mode="ModeA")

        self.assertEqual(loads[0].forecast_tons, -5.0)

    def test_norm_month_accepts_mon_dash_yy_format(self):
        self.assertEqual(_norm_month("Jan-26"), "2026-01")
        self.assertEqual(_norm_month("feb-26"), "2026-02")
        self.assertEqual(_norm_month("September-2026"), "2026-09")

    def test_parse_load_normalizes_mon_dash_yy_months_for_validation(self):
        loads = _parse_load_df(
            pd.DataFrame(
                [
                    {
                        "Month": "Jan-26",
                        "PlannerName": "P1",
                        "Product": "P1",
                        "ProductFamily": "F1",
                        "Plant": "A",
                        "Forecast_Tons": 10,
                        "Resource": "WC1",
                    },
                    {
                        "Month": "Feb-26",
                        "PlannerName": "P1",
                        "Product": "P1",
                        "ProductFamily": "F1",
                        "Plant": "A",
                        "Forecast_Tons": 12,
                        "Resource": "WC1",
                    },
                ]
            )
        )
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=120.0,
                utilization_target=1.0,
            )
        ]

        issues = validate(loads, capacities, [], mode="ModeA")

        self.assertEqual([record.month for record in loads], ["2026-01", "2026-02"])
        self.assertNotIn(("ERROR", "LoadMonthFormat"), {(i.severity, i.check) for i in issues})

    def test_negative_duplicates_are_not_aggregated_away(self):
        records = [
            LoadRecord(
                month="2025-01",
                planner_name="P1",
                product="1",
                product_family="F1",
                plant="A",
                forecast_tons=10.0,
                scenario="Case90",
            ),
            LoadRecord(
                month="2025-01",
                planner_name="P1",
                product="1",
                product_family="F1",
                plant="A",
                forecast_tons=-5.0,
                scenario="Case90",
            ),
        ]

        aggregated = _aggregate_load_records(records)

        self.assertEqual(len(aggregated), 2)
        self.assertEqual(sorted(record.forecast_tons for record in aggregated), [-5.0, 10.0])

    def test_aggregate_keeps_multi_plant_load_nodes_separate(self):
        aggregated = _aggregate_load_records(
            [
                LoadRecord(
                    month="2025-01",
                    planner_name="P1",
                    product="1",
                    product_family="F1",
                    plant="N029",
                    forecast_tons=3.0,
                    scenario="Case90",
                ),
                LoadRecord(
                    month="2025-01",
                    planner_name="P1",
                    product="1",
                    product_family="F1",
                    plant="C447",
                    forecast_tons=2.0,
                    scenario="Case90",
                ),
                LoadRecord(
                    month="2025-02",
                    planner_name="P1",
                    product="1",
                    product_family="F1",
                    plant="Z100",
                    forecast_tons=4.0,
                    scenario="Case90",
                ),
            ]
        )

        demand, node_meta = _build_demand(aggregated)

        self.assertEqual(len(aggregated), 3)
        self.assertEqual(
            demand[("2025-01", "1", "N029", "")],
            3.0,
        )
        self.assertEqual(
            demand[("2025-01", "1", "C447", "")],
            2.0,
        )
        self.assertEqual(node_meta[("2025-01", "1", "N029", "")], ("F1", "P1"))

    def test_format_issue_report_compacts_repeated_warnings(self):
        issues = [
            ValidationIssue(
                severity="WARNING",
                check="LoadZeroTons",
                detail=f"row {idx}",
            )
            for idx in range(1, 8)
        ]

        lines = format_issue_report(issues, warning_example_limit=3)
        report = "\n".join(lines)

        self.assertIn("[LoadZeroTons] 7 occurrence(s)", report)
        self.assertIn("e.g. row 1", report)
        self.assertIn("... and 4 more", report)

    def test_mode_b_capacity_only_product_stays_in_stage1(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]

        results, toller_products = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=baseline_capacities,
            routings=[],
        )

        internal_rows = [row for row in results if row.allocation_type == "Internal"]
        self.assertEqual(len(internal_rows), 1)
        self.assertEqual(internal_rows[0].allocation_source, "Capacity_Base")
        self.assertAlmostEqual(internal_rows[0].allocated_tons, 100.0, places=4)
        self.assertEqual(toller_products, set())
        self.assertFalse(any(row.allocation_source == "Routing_Reroute" for row in results))

    def test_mode_b_reroutes_only_stage1_residual_with_routing_only_resource(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
        ]
        routing_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        routings = [
            RoutingRecord(
                product="P1",
                product_family="F1",
                work_center="WC2",
                priority=1,
                eligible_flag=True,
                route_type="Primary",
            ),
        ]

        results, _ = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=routing_capacities,
            routings=routings,
        )

        by_source = {
            source: sum(row.allocated_tons for row in results if row.allocation_source == source)
            for source in {"Capacity_Base", "Routing_Reroute"}
        }
        final_unmet = max(row.unmet_tons for row in results if row.product == "P1")

        self.assertAlmostEqual(by_source["Capacity_Base"], 60.0, places=4)
        self.assertAlmostEqual(by_source["Routing_Reroute"], 40.0, places=4)
        self.assertAlmostEqual(final_unmet, 0.0, places=4)
        reroute_rows = [row for row in results if row.allocation_source == "Routing_Reroute"]
        self.assertEqual({row.work_center for row in reroute_rows}, {"WC2"})
        self.assertTrue(all(row.residual_after_capacity_tons == 40.0 for row in results if row.product == "P1"))
        self.assertTrue(all(row.residual_after_routing_tons == 0.0 for row in results if row.product == "P1"))

    def test_mode_b_toller_classifies_final_residual_as_outsourced(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
        ]
        routings = [
            RoutingRecord(
                product="P1",
                product_family="F1",
                work_center="TOL1",
                priority=3,
                eligible_flag=True,
                route_type="Toller",
            ),
        ]

        results, toller_products = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=baseline_capacities,
            routings=routings,
        )

        outsourced_rows = [row for row in results if row.allocation_type == "Outsourced"]
        self.assertEqual(toller_products, {"P1"})
        self.assertEqual(len(outsourced_rows), 1)
        self.assertEqual(outsourced_rows[0].allocation_source, "Toller")
        self.assertAlmostEqual(outsourced_rows[0].outsourced_tons, 40.0, places=4)

    def test_mode_b_without_internal_routing_or_toller_ends_as_unmet(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
        ]

        results, _ = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=baseline_capacities,
            routings=[],
        )

        unmet_rows = [row for row in results if row.allocation_type == "Unmet"]
        self.assertEqual(len(unmet_rows), 1)
        self.assertEqual(unmet_rows[0].allocation_source, "Unmet")
        self.assertAlmostEqual(unmet_rows[0].unmet_tons, 40.0, places=4)

    def test_validate_rejects_product_with_multiple_family_values(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
            ),
            LoadRecord(
                month="2025-02",
                planner_name="PlannerA",
                product="P1",
                product_family="F2",
                plant="PLT1",
                forecast_tons=12.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=120.0, utilization_target=1.0),
        ]

        issues = validate(loads, capacities, [], mode="ModeA")

        self.assertIn(("ERROR", "ProductFamilyConflict"), {(i.severity, i.check) for i in issues})

    def test_mode_b_max_vs_planner_share_stage1_but_differ_in_stage2(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
        ]
        max_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        planner_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=240.0, utilization_target=1.0),
        ]
        routings = [
            RoutingRecord(
                product="P1",
                product_family="F1",
                work_center="WC2",
                priority=1,
                eligible_flag=True,
                route_type="Primary",
            ),
        ]

        max_results, _ = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=max_capacities,
            routings=routings,
        )
        planner_results, _ = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=planner_capacities,
            routings=routings,
        )

        def _stage1_tons(rows):
            return sum(row.allocated_tons for row in rows if row.allocation_source == "Capacity_Base")

        def _stage2_tons(rows):
            return sum(row.allocated_tons for row in rows if row.allocation_source == "Routing_Reroute")

        def _final_unmet(rows):
            return max(row.unmet_tons for row in rows if row.product == "P1")

        self.assertAlmostEqual(_stage1_tons(max_results), 60.0, places=4)
        self.assertAlmostEqual(_stage1_tons(planner_results), 60.0, places=4)
        self.assertAlmostEqual(_stage2_tons(max_results), 40.0, places=4)
        self.assertAlmostEqual(_stage2_tons(planner_results), 20.0, places=4)
        self.assertAlmostEqual(_final_unmet(max_results), 0.0, places=4)
        self.assertAlmostEqual(_final_unmet(planner_results), 20.0, places=4)

    def test_mode_b_global_routing_moves_flexible_product_to_alternative(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerNoAlt",
                product="P_NOALT",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="EV2",
            ),
            LoadRecord(
                month="2025-01",
                planner_name="PlannerAlt",
                product="P_ALT",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="EV2",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P_NOALT", work_center="EV2", annual_capacity_tons=1200.0, utilization_target=1.0),
            CapacityRecord(product="P_ALT", work_center="EV2", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        routing_capacities = [
            *baseline_capacities,
            CapacityRecord(product="P_ALT", work_center="SA3", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        routings = [
            RoutingRecord(
                product="P_ALT",
                product_family="F1",
                work_center="SA3",
                priority=1,
                eligible_flag=True,
                route_type="Alternative",
            ),
        ]

        results, _ = run_optimization_mode_b(
            months=["2025-01"],
            loads=loads,
            baseline_capacities=baseline_capacities,
            routing_capacities=routing_capacities,
            routings=routings,
        )

        allocated = {
            (row.product, row.work_center, row.allocation_source): row.allocated_tons
            for row in results
            if row.allocation_type == "Internal"
        }
        final_unmet = {
            row.product: row.unmet_tons
            for row in results
            if row.allocation_type == "Unmet"
        }

        self.assertAlmostEqual(allocated[("P_NOALT", "EV2", "Capacity_Base")], 100.0, places=4)
        self.assertAlmostEqual(allocated[("P_ALT", "SA3", "Routing_Reroute")], 100.0, places=4)
        self.assertEqual(final_unmet, {})

    def test_capacity_effective_windows_are_selected_by_month(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=80.0,
                resource_group_owner="WC1",
            ),
            LoadRecord(
                month="2025-02",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=80.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=600.0,
                utilization_target=1.0,
                effective_from="99999",
                effective_to="99999",
            ),
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from="2025-01-31",
                effective_to="2025-01-31",
            ),
        ]

        results = run_optimization_mode_a(
            months=["2025-01", "2025-02"],
            loads=loads,
            capacities=capacities,
        )

        internal_by_month = {
            row.month: row.allocated_tons
            for row in results
            if row.allocation_type == "Internal"
        }
        unmet_by_month = {
            row.month: row.unmet_tons
            for row in results
            if row.allocation_type == "Unmet"
        }

        self.assertAlmostEqual(internal_by_month["2025-01"], 80.0, places=4)
        self.assertAlmostEqual(internal_by_month["2025-02"], 50.0, places=4)
        self.assertAlmostEqual(unmet_by_month["2025-02"], 30.0, places=4)

    def test_validate_rejects_overlapping_capacity_windows(self):
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from="2027-01-01",
                effective_to="2027-03-01",
                source_file="master_capacity.csv",
                row_num=2,
            ),
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from="2027-03-31",
                effective_to="2027-04-30",
                source_file="master_capacity.csv",
                row_num=3,
            ),
        ]

        issues = validate([], capacities, [], mode="ModeA")
        overlap_details = [issue.detail for issue in issues if issue.check == "CapacityEffectiveWindowOverlap"]

        self.assertEqual(len(overlap_details), 1)
        self.assertIn("row 2", overlap_details[0])
        self.assertIn("row 3", overlap_details[0])

    def test_validate_rejects_duplicate_capacity_default_rows(self):
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from="99999",
                effective_to="99999",
                row_num=2,
            ),
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=900.0,
                utilization_target=1.0,
                effective_from="99999",
                effective_to="99999",
                row_num=3,
            ),
        ]

        issues = validate([], capacities, [], mode="ModeA")

        self.assertIn(("ERROR", "CapacityDefaultDuplicate"), {(i.severity, i.check) for i in issues})

    def test_capacity_default_window_accepts_nan_blank_and_excel_numeric_marker(self):
        loads = [
            LoadRecord(
                month="2025-02",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
            ),
            LoadRecord(
                month="2025-02",
                planner_name="PlannerB",
                product="P2",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from=float("nan"),
                effective_to=float("nan"),
            ),
            CapacityRecord(
                product="P2",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from="99999.0",
                effective_to="99999.0",
            ),
        ]

        issues = validate(loads, capacities, [], mode="ModeA")

        self.assertNotIn(("ERROR", "CapacityEffectiveWindowInvalid"), {(i.severity, i.check) for i in issues})
        self.assertNotIn(("ERROR", "CapacityEffectiveMissing"), {(i.severity, i.check) for i in issues})

    def test_validate_rejects_missing_effective_capacity_for_planner_month(self):
        loads = [
            LoadRecord(
                month="2025-02",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
                source_file="planner.csv",
                row_num=2,
            ),
        ]
        capacities = [
            CapacityRecord(
                product="P1",
                work_center="WC1",
                annual_capacity_tons=1200.0,
                utilization_target=1.0,
                effective_from="2025-01-01",
                effective_to="2025-01-31",
            ),
        ]

        issues = validate(loads, capacities, [], mode="ModeA")
        missing_details = [issue.detail for issue in issues if issue.check == "CapacityEffectiveMissing"]

        self.assertEqual(len(missing_details), 1)
        self.assertIn("planner.csv row 2", missing_details[0])
        self.assertIn("Month='2025-02'", missing_details[0])

    def test_create_template_and_load_control_config(self):
        with workspace_tempdir() as tmpdir:
            project_root = os.path.join(tmpdir, "portable_root")
            template_path = os.path.join(project_root, "Archive", "legacy_excel_control_panel", "Capacity_Optimizer_Control.xlsx")
            runner = CliRunner()
            result = runner.invoke(create_template_main, ["--out", template_path])

            self.assertEqual(result.exit_code, 0, msg=result.output)

            workbook = load_workbook(template_path)
            self.assertIn("Deployment_Steps", workbook.sheetnames)
            self.assertIn("License", workbook.sheetnames)
            self.assertEqual(workbook.active.title, "Deployment_Steps")
            deployment_ws = workbook["Deployment_Steps"]
            deployment_text = "\n".join(
                str(cell.value)
                for row in deployment_ws.iter_rows()
                for cell in row
                if cell.value
            )
            self.assertIn("license.json", deployment_text)
            self.assertIn(r"runtime\get_machine_fingerprint.bat", deployment_text)
            self.assertIn("unbound", deployment_text.lower())
            instructions_text = "\n".join(
                str(cell.value)
                for row in workbook["Instructions"].iter_rows()
                for cell in row
                if cell.value
            )
            self.assertIn("license.json", instructions_text)
            self.assertIn("unbound", instructions_text.lower())
            self.assertIn(r"runtime\setup_requirements.bat", instructions_text)
            self.assertIn("python -m app.main", instructions_text)
            license_ws = workbook["License"]
            self.assertEqual(license_ws["A1"].value, "Current License")
            self.assertEqual(license_ws["A5"].value, "License_Status")
            self.assertEqual(license_ws["B5"].value, "Not configured")
            ws = workbook["Control_Panel"]
            row_by_parameter = {
                ws[f"A{row_num}"].value: row_num
                for row_num in range(2, 30)
                if ws[f"A{row_num}"].value
            }
            ws[f"B{row_by_parameter['Start_Year']}"] = 2027
            ws[f"B{row_by_parameter['Start_Month_Num']}"] = 4
            ws[f"B{row_by_parameter['Run_Mode']}"] = "Both"
            ws[f"B{row_by_parameter['Verbose']}"] = "Yes"
            ws[f"B{row_by_parameter['Skip_Validation_Errors']}"] = "No"
            workbook.save(template_path)
            workbook.close()

            config = load_config(template_path)

        self.assertEqual(config.start_month, "2027-04")
        self.assertEqual(config.run_mode, "Both")
        self.assertTrue(config.verbose)
        self.assertFalse(config.skip_validation_errors)
        self.assertEqual(config.project_root_folder, project_root)
        self.assertEqual(config.input_load_folder, os.path.join(project_root, "Data_Input"))
        self.assertEqual(config.input_master_folder, os.path.join(project_root, "Data_Input"))
        self.assertEqual(config.output_folder, os.path.join(project_root, "output"))

    def test_refresh_control_workbook_license_sheet_updates_license_values(self):
        with workspace_tempdir() as tmpdir:
            project_root = os.path.join(tmpdir, "portable_root")
            template_path = os.path.join(project_root, "Archive", "legacy_excel_control_panel", "Capacity_Optimizer_Control.xlsx")
            runner = CliRunner()
            result = runner.invoke(create_template_main, ["--out", template_path])
            self.assertEqual(result.exit_code, 0, msg=result.output)

            refresh_control_workbook_license_sheet(
                template_path,
                project_root=project_root,
                license_info=SimpleNamespace(
                    status="Valid",
                    license_id="LIC-DUPONT-COMM-2026-0002",
                    license_type="commercial",
                    customer_name="DuPont",
                    issue_date="2026-03-31",
                    expiry_date="2027-03-30",
                    binding_mode="machine_locked",
                    machine_label="DUPONT-PC01",
                    license_path=os.path.join(project_root, "licenses", "active", "license.json"),
                    note="Annual commercial license",
                ),
            )

            workbook = load_workbook(template_path, data_only=True)
            license_ws = workbook["License"]
            sheet_values = {
                license_ws[f"A{row_num}"].value: license_ws[f"B{row_num}"].value
                for row_num in range(5, 17)
                if license_ws[f"A{row_num}"].value
            }
            workbook.close()

        self.assertEqual(sheet_values["License_Status"], "Valid")
        self.assertEqual(sheet_values["License_Name"], "LIC-DUPONT-COMM-2026-0002")
        self.assertEqual(sheet_values["License_Mode"], "Commercial / Machine Locked")
        self.assertEqual(sheet_values["Licensed_To"], "DuPont")
        self.assertEqual(sheet_values["Issue_Date"], "2026-03-31")
        self.assertEqual(sheet_values["Expiry_Date"], "2027-03-30")
        self.assertEqual(sheet_values["Machine_Name"], "DUPONT-PC01")

    def test_write_results_creates_excel_report_sheets(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="demo.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="Both",
            )
            loads = [
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerA",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=100.0,
                    resource_group_owner="WC1",
                )
            ]
            capacities = [
                CapacityRecord(
                    product="P1",
                    work_center="WC1",
                    annual_capacity_tons=1200.0,
                    utilization_target=0.5,
                )
            ]
            routings = [
                RoutingRecord(
                    product="P1",
                    product_family=None,
                    work_center="WC1",
                    priority=1,
                    eligible_flag=True,
                    route_type="Primary",
                ),
                RoutingRecord(
                    product="P1",
                    product_family=None,
                    work_center="TOL1",
                    priority=3,
                    eligible_flag=True,
                    route_type="Toller",
                ),
            ]
            results = [
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Internal",
                    work_center="WC1",
                    route_type="Primary",
                    priority=1,
                    demand_tons=100.0,
                    allocated_tons=60.0,
                    outsourced_tons=0.0,
                    unmet_tons=20.0,
                    capacity_share_pct=60.0,
                    allocation_source="Capacity_Base",
                    residual_after_capacity_tons=40.0,
                    residual_after_routing_tons=20.0,
                ),
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Outsourced",
                    work_center="[UNALLOCATED]",
                    route_type="Toller",
                    priority=3,
                    demand_tons=100.0,
                    allocated_tons=0.0,
                    outsourced_tons=20.0,
                    unmet_tons=20.0,
                    capacity_share_pct=0.0,
                    allocation_source="Toller",
                    residual_after_capacity_tons=40.0,
                    residual_after_routing_tons=20.0,
                ),
            ]
            issues = [ValidationIssue(severity="WARNING", check="DemoWarning", detail="Example")]

            out_path = write_results(
                results=results,
                loads=loads,
                capacities=capacities,
                routings=routings,
                config=config,
                issues=issues,
                months=["2025-01", "2025-02"],
                mode="ModeB",
                toller_products={"P1"},
                metrics_by_mode={
                    "ModeA": {
                        "total_internal_allocated": 50.0,
                        "total_outsourced": 0.0,
                        "total_unmet": 50.0,
                        "service_level": 50.0,
                        "scenario_name": "Baseline",
                    },
                    "ModeB": {
                        "total_internal_allocated": 60.0,
                        "total_outsourced": 20.0,
                        "total_unmet": 20.0,
                        "service_level": 80.0,
                        "scenario_name": "Baseline",
                    },
                },
            )

            workbook = load_workbook(out_path)

            expected_sheets = {
                "Dashboard",
                "Monthly_Trend",
                "Bottleneck",
                "WC_Heatmap",
                "Product_Risk",
                "Allocation_Detail",
                "Unmet_Attribution_Detail",
                "Planner_Result_Summary",
                "Validation_Issues",
                "Run_Info",
            }
            self.assertTrue(expected_sheets.issubset(set(workbook.sheetnames)))
            self.assertNotIn("_Dashboard_Fact", workbook.sheetnames)
            self.assertEqual(
                [ws.title for ws in workbook.worksheets if ws.sheet_state != "visible"],
                ["_Dashboard_Helper", "Validation_Issues", "Run_Info"],
            )
            dashboard_ws = workbook["Dashboard"]
            self.assertEqual(dashboard_ws["O3"].value, "WorkCenter Filter")
            self.assertEqual(dashboard_ws["O4"].value, "Selection Mode")
            self.assertEqual(dashboard_ws["P4"].value, "All")
            self.assertIsNone(dashboard_ws["X1"].value)
            self.assertNotIn("Allocation_Summary", workbook.sheetnames)
            self.assertNotIn("Outsource_Summary", workbook.sheetnames)
            self.assertNotIn("Unmet_Summary", workbook.sheetnames)
            self.assertNotIn("Planner_Product_Month", workbook.sheetnames)
            self.assertNotIn("Binary_Feasibility", workbook.sheetnames)
            monthly_ws = workbook["Monthly_Trend"]
            monthly_merges = {str(rng) for rng in monthly_ws.merged_cells.ranges}
            self.assertIn("A1:H1", monthly_merges)
            self.assertIn("A2:H2", monthly_merges)
            self.assertIn("A3:G3", monthly_merges)
            self.assertGreaterEqual(len(monthly_ws.tables), 3)
            self.assertIsNone(monthly_ws.freeze_panes)
            bottleneck_ws = workbook["Bottleneck"]
            bottleneck_merges = {str(rng) for rng in bottleneck_ws.merged_cells.ranges}
            self.assertIn("A1:D1", bottleneck_merges)
            self.assertIn("A2:D2", bottleneck_merges)
            self.assertGreaterEqual(len(bottleneck_ws.tables), 2)
            self.assertIsNone(bottleneck_ws.freeze_panes)
            heatmap_ws = workbook["WC_Heatmap"]
            heatmap_rows = list(heatmap_ws.iter_rows(min_row=4, values_only=False))
            demand_row = next(
                row
                for row in heatmap_rows
                if row[0].value == "WC1" and row[1].value == "Demand"
            )
            load_row = next(
                row
                for row in heatmap_rows
                if row[0].value == "WC1" and row[1].value == "Load%"
            )
            self.assertIsNone(demand_row[2].fill.patternType)
            self.assertEqual(load_row[2].fill.patternType, "solid")
            detail_ws = workbook["Allocation_Detail"]
            detail_headers = [detail_ws.cell(3, idx).value for idx in range(1, detail_ws.max_column + 1)]
            detail_capacity_idx = detail_headers.index("CapacityShare_Pct") + 1
            self.assertIn("Allocation_Source", detail_headers)
            self.assertIn("Residual_After_Capacity_Tons", detail_headers)
            self.assertIn("Residual_After_Routing_Tons", detail_headers)
            self.assertAlmostEqual(detail_ws.cell(4, detail_capacity_idx).value, 0.6, places=6)
            self.assertGreaterEqual(len(detail_ws.tables), 1)
            self.assertIsNone(detail_ws.freeze_panes)
            attribution_ws = workbook["Unmet_Attribution_Detail"]
            attribution_headers = [attribution_ws.cell(3, idx).value for idx in range(1, attribution_ws.max_column + 1)]
            self.assertIn("Attributed_WorkCenter", attribution_headers)
            self.assertIn("Attributed_Unmet_Tons", attribution_headers)
            self.assertGreaterEqual(len(attribution_ws.tables), 1)
            self.assertNotIn("WC_Load_Pct", workbook.sheetnames)
            workbook.close()

    def test_write_results_stores_ten_decimal_data_without_changing_display_format(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="precision.xlsx",
                scenario_name="Precision",
                start_month="2025-01",
                horizon_months=1,
                run_mode="ModeA",
            )
            loads = [
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerA",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=0.123456789012,
                    resource_group_owner="WC1",
                )
            ]
            capacities = [
                CapacityRecord(
                    product="P1",
                    work_center="WC1",
                    annual_capacity_tons=10000.0,
                    utilization_target=1.0,
                )
            ]
            results = [
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Internal",
                    work_center="WC1",
                    route_type="Primary",
                    priority=1,
                    demand_tons=0.123456789012,
                    allocated_tons=0.123456789012,
                    outsourced_tons=0.0,
                    unmet_tons=0.0,
                    capacity_share_pct=0.01481481468144,
                )
            ]

            out_path = write_results(
                results=results,
                loads=loads,
                capacities=capacities,
                routings=[],
                config=config,
                issues=[],
                months=["2025-01"],
                mode="ModeA",
            )

            workbook = load_workbook(out_path, data_only=False)
            detail_ws = workbook["Allocation_Detail"]
            headers = [detail_ws.cell(3, idx).value for idx in range(1, detail_ws.max_column + 1)]
            allocated_idx = headers.index("Allocated_Tons") + 1
            capacity_idx = headers.index("CapacityShare_Pct") + 1

            self.assertAlmostEqual(detail_ws.cell(4, allocated_idx).value, 0.123456789, places=10)
            self.assertEqual(detail_ws.cell(4, allocated_idx).number_format, "#,##0.0")
            self.assertAlmostEqual(detail_ws.cell(4, capacity_idx).value, 0.0001481481, places=10)
            self.assertEqual(detail_ws.cell(4, capacity_idx).number_format, "0.0%")
            workbook.close()

    def test_write_results_keeps_merged_planner_labels_without_split_back(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="demo.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="ModeA",
            )
            loads = [
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerA",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=60.0,
                    resource_group_owner="WC1",
                ),
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerB",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=40.0,
                    resource_group_owner="WC1",
                ),
            ]
            capacities = [
                CapacityRecord(
                    product="P1",
                    work_center="WC1",
                    annual_capacity_tons=1200.0,
                    utilization_target=0.5,
                )
            ]
            results = [
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Internal",
                    work_center="WC1",
                    route_type="Primary",
                    priority=1,
                    demand_tons=100.0,
                    allocated_tons=60.0,
                    outsourced_tons=0.0,
                    unmet_tons=20.0,
                    capacity_share_pct=60.0,
                ),
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Unmet",
                    work_center="[UNALLOCATED]",
                    route_type="N/A",
                    priority=99,
                    demand_tons=100.0,
                    allocated_tons=0.0,
                    outsourced_tons=0.0,
                    unmet_tons=40.0,
                    capacity_share_pct=0.0,
                ),
            ]

            out_path = write_results(
                results=results,
                loads=loads,
                capacities=capacities,
                routings=[],
                config=config,
                issues=[],
                months=["2025-01", "2025-02"],
                mode="ModeA",
                toller_products=set(),
                metrics_by_mode={
                    "ModeA": {
                        "total_internal_allocated": 60.0,
                        "total_outsourced": 0.0,
                        "total_unmet": 40.0,
                        "service_level": 60.0,
                        "scenario_name": "Baseline",
                    },
                },
            )

            workbook = load_workbook(out_path, data_only=True)
            detail_ws = workbook["Allocation_Detail"]
            detail_headers = [detail_ws.cell(3, idx).value for idx in range(1, detail_ws.max_column + 1)]
            self.assertIn("PlannerName", detail_headers)
            self.assertNotIn("Allocation_Source", detail_headers)
            self.assertNotIn("Residual_After_Capacity_Tons", detail_headers)
            self.assertNotIn("Residual_After_Routing_Tons", detail_headers)
            attribution_ws = workbook["Unmet_Attribution_Detail"]
            attribution_headers = [attribution_ws.cell(3, idx).value for idx in range(1, attribution_ws.max_column + 1)]
            attributed_idx = attribution_headers.index("Attributed_Unmet_Tons") + 1
            planner_idx = attribution_headers.index("PlannerName") + 1
            source_idx = attribution_headers.index("Source_Resource") + 1
            attribution_rows = [
                (
                    attribution_ws.cell(row, planner_idx).value,
                    attribution_ws.cell(row, source_idx).value,
                    attribution_ws.cell(row, attributed_idx).value,
                )
                for row in range(4, attribution_ws.max_row + 1)
                if attribution_ws.cell(row, planner_idx).value
            ]

            planner_ws = workbook["Planner_Result_Summary"]
            planner_rows = list(planner_ws.iter_rows(min_row=3, values_only=True))
            header = planner_rows[0]
            planner_idx = header.index("PlannerName")
            demand_idx = header.index("Demand_Tons")
            internal_idx = header.index("Internal_Tons")
            unmet_idx = header.index("Unmet_Tons")
            data_rows = [
                row
                for row in planner_rows[1:]
                if row[planner_idx] not in (None, "") and isinstance(row[demand_idx], (int, float))
            ]

            summary = {
                row[planner_idx]: {
                    "demand": row[demand_idx],
                    "internal": row[internal_idx],
                    "unmet": row[unmet_idx],
                }
                for row in data_rows
            }
            workbook.close()

        self.assertEqual(set(summary), {"PlannerA | PlannerB"})
        self.assertAlmostEqual(summary["PlannerA | PlannerB"]["demand"], 100.0, places=4)
        self.assertAlmostEqual(summary["PlannerA | PlannerB"]["internal"], 60.0, places=4)
        self.assertAlmostEqual(summary["PlannerA | PlannerB"]["unmet"], 40.0, places=4)
        self.assertEqual(attribution_rows, [("PlannerA | PlannerB", "WC1", 40.0)])

    def test_write_results_localizes_visible_report_strings_for_chinese(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="zh_demo.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="ModeB",
                language="zh",
            )
            loads = [
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerA",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=100.0,
                    resource_group_owner="WC1",
                ),
            ]
            capacities = [
                CapacityRecord(
                    product="P1",
                    work_center="WC1",
                    annual_capacity_tons=1200.0,
                    utilization_target=1.0,
                ),
            ]
            results = [
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Internal",
                    work_center="WC1",
                    route_type="Primary",
                    priority=1,
                    demand_tons=100.0,
                    allocated_tons=60.0,
                    outsourced_tons=0.0,
                    unmet_tons=40.0,
                    capacity_share_pct=60.0,
                    allocation_source="Capacity_Base",
                    residual_after_capacity_tons=40.0,
                    residual_after_routing_tons=40.0,
                ),
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Unmet",
                    work_center="[UNALLOCATED]",
                    route_type="N/A",
                    priority=99,
                    demand_tons=100.0,
                    allocated_tons=0.0,
                    outsourced_tons=0.0,
                    unmet_tons=40.0,
                    capacity_share_pct=0.0,
                    allocation_source="Unmet",
                    residual_after_capacity_tons=40.0,
                    residual_after_routing_tons=40.0,
                ),
            ]

            out_path = write_results(
                results=results,
                loads=loads,
                capacities=capacities,
                routings=[],
                config=config,
                issues=[],
                months=["2025-01", "2025-02"],
                mode="ModeB",
            )

            workbook = load_workbook(out_path, data_only=True)
            self.assertIn("未满足回挂明细", workbook.sheetnames)
            dashboard_ws = workbook["仪表板"]
            self.assertEqual(dashboard_ws["A2"].value, "场景：基准 | 模式：模式B | 滚动月份：2")
            detail_ws = workbook["分配明细"]
            detail_headers = [detail_ws.cell(3, idx).value for idx in range(1, detail_ws.max_column + 1)]
            self.assertIn("分配来源", detail_headers)
            self.assertEqual(detail_ws["B4"].value, "PlannerA")
            attribution_ws = workbook["未满足回挂明细"]
            self.assertIn("来源资源工作中心", attribution_ws["A2"].value)
            workbook.close()

    def test_write_capacity_basis_results_creates_dual_basis_workbook_for_mode_b(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="basis_demo.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="ModeB",
            )
            loads = [
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerA",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=100.0,
                    resource_group_owner="WC1",
                )
            ]
            basis_capacities = {
                "Max": [CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0)],
                "Planned": [CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=900.0, utilization_target=1.0)],
            }
            basis_results = {
                "Max": [
                    AllocationResult(
                        month="2025-01",
                        product="P1",
                        product_family="F1",
                        plant="PLT1",
                        source_resource="WC1",
                        allocation_type="Internal",
                        work_center="WC1",
                        route_type="Capacity",
                        priority=1,
                        demand_tons=100.0,
                        allocated_tons=80.0,
                        outsourced_tons=0.0,
                        unmet_tons=20.0,
                        capacity_share_pct=80.0,
                        allocation_source="Capacity_Base",
                        residual_after_capacity_tons=20.0,
                        residual_after_routing_tons=20.0,
                    ),
                    AllocationResult(
                        month="2025-01",
                        product="P1",
                        product_family="F1",
                        plant="PLT1",
                        source_resource="WC1",
                        allocation_type="Unmet",
                        work_center="[UNALLOCATED]",
                        route_type="N/A",
                        priority=99,
                        demand_tons=100.0,
                        allocated_tons=0.0,
                        outsourced_tons=0.0,
                        unmet_tons=20.0,
                        capacity_share_pct=0.0,
                        allocation_source="Unmet",
                        residual_after_capacity_tons=20.0,
                        residual_after_routing_tons=20.0,
                    ),
                ],
                "Planned": [
                    AllocationResult(
                        month="2025-01",
                        product="P1",
                        product_family="F1",
                        plant="PLT1",
                        source_resource="WC1",
                        allocation_type="Internal",
                        work_center="WC1",
                        route_type="Capacity",
                        priority=1,
                        demand_tons=100.0,
                        allocated_tons=60.0,
                        outsourced_tons=0.0,
                        unmet_tons=40.0,
                        capacity_share_pct=60.0,
                        allocation_source="Capacity_Base",
                        residual_after_capacity_tons=40.0,
                        residual_after_routing_tons=40.0,
                    ),
                    AllocationResult(
                        month="2025-01",
                        product="P1",
                        product_family="F1",
                        plant="PLT1",
                        source_resource="WC1",
                        allocation_type="Unmet",
                        work_center="[UNALLOCATED]",
                        route_type="N/A",
                        priority=99,
                        demand_tons=100.0,
                        allocated_tons=0.0,
                        outsourced_tons=0.0,
                        unmet_tons=40.0,
                        capacity_share_pct=0.0,
                        allocation_source="Unmet",
                        residual_after_capacity_tons=40.0,
                        residual_after_routing_tons=40.0,
                    ),
                ],
            }

            out_path = write_capacity_basis_results(
                basis_results=basis_results,
                loads=loads,
                basis_capacities=basis_capacities,
                routings=[],
                config=config,
                issues=[],
                months=["2025-01", "2025-02"],
                mode="ModeB",
                toller_products_by_basis={"Max": set(), "Planned": set()},
            )

            workbook = load_workbook(out_path)
            expected_sheets = {
                "Dashboard",
                "Monthly_Trend",
                "Bottleneck",
                "WC_Heatmap",
                "Product_Risk",
                "Planner_Result_Summary",
                "Allocation_Detail",
                "Unmet_Attribution_Detail",
                "Validation_Issues",
                "Run_Info",
            }
            self.assertTrue(expected_sheets.issubset(set(workbook.sheetnames)))
            self.assertNotIn("_Dashboard_Fact", workbook.sheetnames)
            self.assertNotIn("Allocation_Summary", workbook.sheetnames)
            self.assertNotIn("Outsource_Summary", workbook.sheetnames)
            self.assertNotIn("Unmet_Summary", workbook.sheetnames)
            self.assertNotIn("Planner_Product_Month", workbook.sheetnames)
            self.assertNotIn("Binary_Feasibility", workbook.sheetnames)
            self.assertNotIn("WC_Load_Pct", workbook.sheetnames)
            self.assertEqual(
                [ws.title for ws in workbook.worksheets if ws.sheet_state != "visible"],
                ["_Dashboard_Helper", "Validation_Issues", "Run_Info"],
            )
            detail_ws = workbook["Allocation_Detail"]
            detail_headers = [detail_ws.cell(3, idx).value for idx in range(1, detail_ws.max_column + 1)]
            self.assertIn("Capacity_Basis", detail_headers)
            self.assertIn("Allocation_Source", detail_headers)
            self.assertIn("Residual_After_Capacity_Tons", detail_headers)
            self.assertIn("Residual_After_Routing_Tons", detail_headers)
            attribution_ws = workbook["Unmet_Attribution_Detail"]
            attribution_headers = [attribution_ws.cell(3, idx).value for idx in range(1, attribution_ws.max_column + 1)]
            self.assertIn("Capacity_Basis", attribution_headers)
            monthly_ws = workbook["Monthly_Trend"]
            monthly_merges = {str(rng) for rng in monthly_ws.merged_cells.ranges}
            self.assertIn("A1:L1", monthly_merges)
            self.assertIn("A2:L2", monthly_merges)
            self.assertGreaterEqual(len(monthly_ws.tables), 2)
            self.assertIsNone(monthly_ws.freeze_panes)
            dashboard_ws = workbook["Dashboard"]
            self.assertEqual(dashboard_ws["O3"].value, "WorkCenter Filter")
            self.assertEqual(dashboard_ws["R4"].value, "All")
            self.assertIsNone(dashboard_ws["X1"].value)
            workbook.close()

    def test_write_capacity_basis_results_creates_dual_basis_workbook_for_mode_a(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="basis_modea.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="ModeA",
            )
            loads = [
                LoadRecord(
                    month="2025-01",
                    planner_name="PlannerA",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    forecast_tons=100.0,
                    resource_group_owner="WC1",
                )
            ]
            basis_capacities = {
                "Max": [CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0)],
                "Planned": [CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0)],
            }
            basis_results = {
                "Max": run_optimization_mode_a(["2025-01"], loads, basis_capacities["Max"]),
                "Planned": run_optimization_mode_a(["2025-01"], loads, basis_capacities["Planned"]),
            }

            out_path = write_capacity_basis_results(
                basis_results=basis_results,
                loads=loads,
                basis_capacities=basis_capacities,
                routings=[],
                config=config,
                issues=[],
                months=["2025-01", "2025-02"],
                mode="ModeA",
                toller_products_by_basis={"Max": set(), "Planned": set()},
                unmet_capacities_by_basis=basis_capacities,
            )

            workbook = load_workbook(out_path)
            self.assertIn("Dashboard", workbook.sheetnames)
            self.assertIn("Monthly_Trend", workbook.sheetnames)
            self.assertIn("Allocation_Detail", workbook.sheetnames)
            self.assertIn("Unmet_Attribution_Detail", workbook.sheetnames)
            self.assertNotIn("Shared Stage 1 baseline", str(workbook["Dashboard"]["A2"].value))
            helper_ws = workbook["_Dashboard_Helper"]
            self.assertEqual(helper_ws["A1"].value, "Capacity_Basis")
            helper_values = {helper_ws["A2"].value, helper_ws["A3"].value}
            self.assertIn("Max", helper_values)
            self.assertIn("Planned", helper_values)
            self.assertIn('"Max"', str(workbook["Dashboard"]["A6"].value))
            self.assertIn('"Planned"', str(workbook["Dashboard"]["C6"].value))
            self.assertIn('"Planned"', str(workbook["Dashboard"]["J6"].value))
            detail_headers = [workbook["Allocation_Detail"].cell(3, idx).value for idx in range(1, workbook["Allocation_Detail"].max_column + 1)]
            self.assertIn("Capacity_Basis", detail_headers)
            workbook.close()

    def test_allocation_detail_sorts_by_basis_month_plant_workcenter_product(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="sorted_detail.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="ModeA",
            )

            def result(month: str, plant: str, wc: str, product: str) -> AllocationResult:
                return AllocationResult(
                    month=month,
                    planner_name="PlannerA",
                    product=product,
                    product_family="F1",
                    plant=plant,
                    source_resource=wc,
                    allocation_type="Internal",
                    work_center=wc,
                    route_type="Primary",
                    priority=1,
                    demand_tons=1.0,
                    allocated_tons=1.0,
                    outsourced_tons=0.0,
                    unmet_tons=0.0,
                    capacity_share_pct=1.0,
                )

            basis_results = {
                "Max": [
                    result("2025-02", "PLT1", "WC2", "P2"),
                    result("2025-01", "PLT2", "WC1", "P1"),
                    result("2025-01", "PLT1", "WC2", "P2"),
                    result("2025-01", "PLT1", "WC1", "P2"),
                    result("2025-01", "PLT1", "WC1", "P1"),
                ],
                "Planned": [
                    result("2025-01", "PLT1", "WC1", "P1"),
                ],
            }
            basis_capacities = {
                basis: [
                    CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0),
                    CapacityRecord(product="P2", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0),
                    CapacityRecord(product="P2", work_center="WC2", annual_capacity_tons=1200.0, utilization_target=1.0),
                ]
                for basis in ("Max", "Planned")
            }

            out_path = write_capacity_basis_results(
                basis_results=basis_results,
                loads=[],
                basis_capacities=basis_capacities,
                routings=[],
                config=config,
                issues=[],
                months=["2025-01", "2025-02"],
                mode="ModeA",
                toller_products_by_basis={"Max": set(), "Planned": set()},
                unmet_capacities_by_basis=basis_capacities,
            )

            workbook = load_workbook(out_path)
            detail_ws = workbook["Allocation_Detail"]
            headers = [detail_ws.cell(3, idx).value for idx in range(1, detail_ws.max_column + 1)]
            indices = {header: headers.index(header) + 1 for header in ["Capacity_Basis", "Month", "Plant", "WorkCenter", "Product"]}
            observed = [
                tuple(detail_ws.cell(row, indices[column]).value for column in ["Capacity_Basis", "Month", "Plant", "WorkCenter", "Product"])
                for row in range(4, 10)
            ]
            self.assertEqual(
                observed,
                [
                    ("Max", "2025-01", "PLT1", "WC1", "P1"),
                    ("Max", "2025-01", "PLT1", "WC1", "P2"),
                    ("Max", "2025-01", "PLT1", "WC2", "P2"),
                    ("Max", "2025-01", "PLT2", "WC1", "P1"),
                    ("Max", "2025-02", "PLT1", "WC2", "P2"),
                    ("Planned", "2025-01", "PLT1", "WC1", "P1"),
                ],
            )
            workbook.close()

    def test_load_direct_rejects_unrecognized_planner_files(self):
        with workspace_tempdir() as tmpdir:
            pd.DataFrame(
                [
                    {
                        "Month": "2025-01",
                        "PlannerName": "P1",
                        "Product": "0001",
                        "ProductFamily": "F1",
                        "Plant": "A",
                        "Forecast_Tons": 10,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "wrong_name.csv"), index=False)
            pd.DataFrame(
                [
                    {
                        "Product": "1",
                        "WorkCenter": "WC1",
                        "Annual_Capacity_Tons": 120,
                        "Utilization_Target": 1,
                    }
                ]
            ).to_csv(os.path.join(tmpdir, "master_capacity.csv"), index=False)

            with self.assertRaises(FileNotFoundError):
                load_direct_mode_a(tmpdir, tmpdir)

    def test_validate_input_setup_rejects_missing_output_parent(self):
        with workspace_tempdir() as tmpdir:
            os.makedirs(os.path.join(tmpdir, "Data_Input"))
            config = Config(
                project_root_folder=tmpdir,
                input_load_folder=os.path.join(tmpdir, "Data_Input"),
                input_master_folder=os.path.join(tmpdir, "Data_Input"),
                output_folder=os.path.join(tmpdir, "missing", "output"),
                output_file_name="demo.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="Both",
            )

            with self.assertRaises(FileNotFoundError):
                _validate_input_setup(config, ["ModeA", "ModeB"])

    def test_write_mode_comparison_summary_creates_standalone_workbook(self):
        with workspace_tempdir() as tmpdir:
            config = Config(
                input_load_folder=tmpdir,
                input_master_folder=tmpdir,
                output_folder=tmpdir,
                output_file_name="demo.xlsx",
                scenario_name="Baseline",
                start_month="2025-01",
                horizon_months=2,
                run_mode="Both",
            )
            mode_a_results = [
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Internal",
                    work_center="WC1",
                    route_type="Primary",
                    priority=1,
                    demand_tons=100.0,
                    allocated_tons=50.0,
                    outsourced_tons=0.0,
                    unmet_tons=50.0,
                    capacity_share_pct=50.0,
                ),
            ]
            mode_b_results = [
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Internal",
                    work_center="WC1",
                    route_type="Primary",
                    priority=1,
                    demand_tons=100.0,
                    allocated_tons=60.0,
                    outsourced_tons=0.0,
                    unmet_tons=40.0,
                    capacity_share_pct=60.0,
                ),
                AllocationResult(
                    month="2025-01",
                    product="P1",
                    product_family="F1",
                    plant="PLT1",
                    source_resource="WC1",
                    allocation_type="Outsourced",
                    work_center="[UNALLOCATED]",
                    route_type="Toller",
                    priority=3,
                    demand_tons=100.0,
                    allocated_tons=0.0,
                    outsourced_tons=20.0,
                    unmet_tons=20.0,
                    capacity_share_pct=0.0,
                ),
            ]
            mode_loads = {
                "ModeA": [
                    LoadRecord(
                        month="2025-01",
                        planner_name="PlannerA",
                        product="P1",
                        product_family="F1",
                        plant="PLT1",
                        forecast_tons=100.0,
                        resource_group_owner="WC1",
                    )
                ],
                "ModeB": [
                    LoadRecord(
                        month="2025-01",
                        planner_name="PlannerA",
                        product="P1",
                        product_family="F1",
                        plant="PLT1",
                        forecast_tons=100.0,
                        resource_group_owner="WC1",
                    )
                ],
            }
            mode_capacities = {
                "ModeA": [
                    CapacityRecord(
                        product="P1",
                        work_center="WC1",
                        annual_capacity_tons=1200.0,
                        utilization_target=0.5,
                    )
                ],
                "ModeB": [
                    CapacityRecord(
                        product="P1",
                        work_center="WC1",
                        annual_capacity_tons=1200.0,
                        utilization_target=0.5,
                    )
                ],
            }
            mode_routings = {
                "ModeA": [],
                "ModeB": [
                    RoutingRecord(
                        product="P1",
                        product_family=None,
                        work_center="WC1",
                        priority=1,
                        eligible_flag=True,
                        route_type="Primary",
                    ),
                    RoutingRecord(
                        product="P1",
                        product_family=None,
                        work_center="TOL1",
                        priority=3,
                        eligible_flag=True,
                        route_type="Toller",
                    ),
                ],
            }

            out_path = write_mode_comparison_summary(
                mode_results={"ModeA": mode_a_results, "ModeB": mode_b_results},
                mode_loads=mode_loads,
                mode_capacities=mode_capacities,
                mode_routings=mode_routings,
                config=config,
                months=["2025-01", "2025-02"],
                metrics_by_mode={
                    "ModeA": {
                        "total_demand": 100.0,
                        "total_internal_allocated": 50.0,
                        "total_outsourced": 0.0,
                        "total_unmet": 50.0,
                        "service_level": 50.0,
                        "result_rows": 1,
                        "months": 2,
                        "scenario_name": "Baseline",
                    },
                    "ModeB": {
                        "total_demand": 100.0,
                        "total_internal_allocated": 60.0,
                        "total_outsourced": 20.0,
                        "total_unmet": 20.0,
                        "service_level": 80.0,
                        "result_rows": 2,
                        "months": 2,
                        "scenario_name": "Baseline",
                    },
                },
            )

            workbook = load_workbook(out_path)
            expected_sheets = {
                "Executive_Comparison",
                "Monthly_Trend_Compare",
                "Bottleneck_Compare",
                "WC_Heatmap_Compare",
                "Product_Risk_Compare",
                "Unmet_Attribution_Detail",
                "Run_Info",
            }
            self.assertRegex(
                os.path.basename(out_path),
                r"^Summary of Mode A and Mode B_\d{8}_\d{6}\.xlsx$",
            )
            self.assertTrue(expected_sheets.issubset(set(workbook.sheetnames)))
            self.assertNotIn("_Dashboard_Fact", workbook.sheetnames)
            self.assertNotIn("_ModeA_Cap_Fact", workbook.sheetnames)
            self.assertNotIn("_ModeB_Cap_Fact", workbook.sheetnames)
            self.assertNotIn("Planner_Compare", workbook.sheetnames)
            monthly_compare_ws = workbook["Monthly_Trend_Compare"]
            monthly_compare_merges = {str(rng) for rng in monthly_compare_ws.merged_cells.ranges}
            self.assertIn("A1:L1", monthly_compare_merges)
            self.assertIn("A2:L2", monthly_compare_merges)
            self.assertGreaterEqual(len(monthly_compare_ws.tables), 1)
            self.assertIsNone(monthly_compare_ws.freeze_panes)
            self.assertEqual(
                [ws.title for ws in workbook.worksheets if ws.sheet_state != "visible"],
                ["_Dashboard_Helper", "Run_Info"],
            )
            attribution_ws = workbook["Unmet_Attribution_Detail"]
            attribution_headers = [attribution_ws.cell(3, idx).value for idx in range(1, attribution_ws.max_column + 1)]
            self.assertIn("Mode", attribution_headers)
            executive_ws = workbook["Executive_Comparison"]
            self.assertEqual(executive_ws["O3"].value, "WorkCenter Filter")
            self.assertEqual(executive_ws["P4"].value, "All")
            self.assertIsNone(executive_ws["X1"].value)
            workbook.close()

    def test_validate_allows_planner_product_on_multiple_resources(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=10.0,
                resource_group_owner="WC1",
            ),
            LoadRecord(
                month="2025-02",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=12.0,
                resource_group_owner="WC2",
            ),
        ]
        capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=120.0, utilization_target=1.0),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=120.0, utilization_target=1.0),
        ]

        issues = validate(loads, capacities, [], mode="ModeA")

        self.assertNotIn(("ERROR", "LoadPlannerProductMultiResource"), {(i.severity, i.check) for i in issues})

    def test_mode_a_pressure_load_assigns_unmet_by_source_resource(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=60.0,
                resource_group_owner="WC1",
            ),
            LoadRecord(
                month="2025-01",
                planner_name="PlannerB",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=40.0,
                resource_group_owner="WC2",
            ),
        ]
        capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=0.5),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=1200.0, utilization_target=0.5),
        ]
        results = [
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Internal",
                work_center="WC1",
                route_type="Capacity",
                priority=1,
                demand_tons=60.0,
                allocated_tons=50.0,
                outsourced_tons=0.0,
                unmet_tons=10.0,
                capacity_share_pct=100.0,
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Unmet",
                work_center="[UNALLOCATED]",
                route_type="N/A",
                priority=99,
                demand_tons=60.0,
                allocated_tons=0.0,
                outsourced_tons=0.0,
                unmet_tons=10.0,
                capacity_share_pct=0.0,
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC2",
                allocation_type="Unmet",
                work_center="[UNALLOCATED]",
                route_type="N/A",
                priority=99,
                demand_tons=40.0,
                allocated_tons=0.0,
                outsourced_tons=0.0,
                unmet_tons=40.0,
                capacity_share_pct=0.0,
                planner_name="PlannerB",
            ),
        ]

        wc_load_df = build_pressure_load_frame("ModeA", results, loads, capacities, [], ["2025-01"])
        by_wc = {row["WorkCenter"]: row["2025-01"] for _, row in wc_load_df.iterrows()}

        self.assertAlmostEqual(by_wc["WC1"], 0.6, places=6)
        self.assertAlmostEqual(by_wc["WC2"], 0.4, places=6)

    def test_mode_b_pressure_load_uses_lp_for_no_routing_unmet(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=0.5),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=1200.0, utilization_target=0.5),
        ]
        results = [
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Internal",
                work_center="WC1",
                route_type="Capacity",
                priority=1,
                demand_tons=100.0,
                allocated_tons=90.0,
                outsourced_tons=0.0,
                unmet_tons=10.0,
                capacity_share_pct=180.0,
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Unmet",
                work_center="[UNALLOCATED]",
                route_type="N/A",
                priority=99,
                demand_tons=100.0,
                allocated_tons=0.0,
                outsourced_tons=0.0,
                unmet_tons=10.0,
                capacity_share_pct=0.0,
            ),
        ]

        wc_load_df = build_pressure_load_frame("ModeB", results, loads, capacities, [], ["2025-01"])
        by_wc = {row["WorkCenter"]: row["2025-01"] for _, row in wc_load_df.iterrows()}

        self.assertAlmostEqual(by_wc["WC1"], 1.0, places=6)
        self.assertNotIn("WC2", by_wc)

    def test_mode_b_final_unmet_returns_to_baseline_capacity_workcenter(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        baseline_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
        ]
        report_capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=720.0, utilization_target=1.0),
            CapacityRecord(product="P1", work_center="WC2", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        routings = [
            RoutingRecord(
                product="P1",
                product_family="F1",
                work_center="WC2",
                priority=1,
                eligible_flag=True,
                route_type="Primary",
            ),
        ]
        results = [
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Internal",
                work_center="WC1",
                route_type="Capacity",
                priority=1,
                demand_tons=100.0,
                allocated_tons=60.0,
                outsourced_tons=0.0,
                unmet_tons=20.0,
                capacity_share_pct=100.0,
                allocation_source="Capacity_Base",
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Internal",
                work_center="WC2",
                route_type="Primary",
                priority=1,
                demand_tons=100.0,
                allocated_tons=20.0,
                outsourced_tons=0.0,
                unmet_tons=20.0,
                capacity_share_pct=20.0,
                allocation_source="Routing_Reroute",
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Unmet",
                work_center="[UNALLOCATED]",
                route_type="N/A",
                priority=99,
                demand_tons=100.0,
                allocated_tons=0.0,
                outsourced_tons=0.0,
                unmet_tons=20.0,
                capacity_share_pct=0.0,
                allocation_source="Unmet",
            ),
        ]

        wc_load_df = build_pressure_load_frame(
            "ModeB",
            results,
            loads,
            report_capacities,
            routings,
            ["2025-01"],
            unmet_capacities=baseline_capacities,
        )
        by_wc = {row["WorkCenter"]: row["2025-01"] for _, row in wc_load_df.iterrows()}

        self.assertAlmostEqual(by_wc["WC1"], 4.0 / 3.0, places=6)
        self.assertAlmostEqual(by_wc["WC2"], 0.2, places=6)

    def test_dashboard_fact_frame_assigns_mode_b_outsourced_to_toller(self):
        loads = [
            LoadRecord(
                month="2025-01",
                planner_name="PlannerA",
                product="P1",
                product_family="F1",
                plant="PLT1",
                forecast_tons=100.0,
                resource_group_owner="WC1",
            ),
        ]
        capacities = [
            CapacityRecord(product="P1", work_center="WC1", annual_capacity_tons=1200.0, utilization_target=1.0),
        ]
        routings = [
            RoutingRecord(
                product="P1",
                product_family=None,
                work_center="WC1",
                priority=1,
                eligible_flag=True,
                route_type="Primary",
            ),
            RoutingRecord(
                product="P1",
                product_family=None,
                work_center="TOL1",
                priority=3,
                eligible_flag=True,
                route_type="Toller",
            ),
        ]
        results = [
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Internal",
                work_center="WC1",
                route_type="Primary",
                priority=1,
                demand_tons=100.0,
                allocated_tons=60.0,
                outsourced_tons=0.0,
                unmet_tons=20.0,
                capacity_share_pct=60.0,
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Outsourced",
                work_center="[OUTSOURCED]",
                route_type="Toller",
                priority=99,
                demand_tons=100.0,
                allocated_tons=0.0,
                outsourced_tons=20.0,
                unmet_tons=20.0,
                capacity_share_pct=0.0,
            ),
            AllocationResult(
                month="2025-01",
                product="P1",
                product_family="F1",
                plant="PLT1",
                source_resource="WC1",
                allocation_type="Unmet",
                work_center="[UNALLOCATED]",
                route_type="N/A",
                priority=99,
                demand_tons=100.0,
                allocated_tons=0.0,
                outsourced_tons=0.0,
                unmet_tons=20.0,
                capacity_share_pct=0.0,
            ),
        ]

        fact_df = build_dashboard_fact_frame("ModeB", results, loads, capacities, routings)
        by_wc = {
            row["WorkCenter"]: row
            for _, row in fact_df.iterrows()
        }

        self.assertAlmostEqual(by_wc["WC1"]["Internal_Tons"], 60.0, places=4)
        self.assertAlmostEqual(by_wc["WC1"]["Unmet_Tons"], 20.0, places=4)
        self.assertAlmostEqual(by_wc["TOL1"]["Outsourced_Tons"], 20.0, places=4)


if __name__ == "__main__":
    unittest.main()
