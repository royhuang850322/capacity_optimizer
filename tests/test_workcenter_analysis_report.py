import os
import shutil
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.i18n import localize_column_name, localize_sheet_name, localize_value
from app.workcenter_analysis_report import (
    ReportValidationError,
    find_latest_mode_report,
    generate_workcenter_analysis_report,
    load_mode_report_context,
    resolve_mode_report_selection,
)


TEST_TMP_ROOT = os.path.join(os.path.dirname(__file__), "_tmp")
os.makedirs(TEST_TMP_ROOT, exist_ok=True)


@contextmanager
def workspace_tempdir():
    tmpdir = os.path.join(TEST_TMP_ROOT, f"tmp_{uuid.uuid4().hex}")
    os.mkdir(tmpdir)
    try:
        yield tmpdir
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _write_mode_report(path: Path, input_dir: Path, mode: str) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    detail_ws = workbook.create_sheet(localize_sheet_name("zh", "Allocation_Detail"))
    detail_headers = [
        "Capacity_Basis",
        localize_column_name("zh", "Month"),
        localize_column_name("zh", "PlannerName"),
        localize_column_name("zh", "Product"),
        localize_column_name("zh", "ProductFamily"),
        localize_column_name("zh", "Plant"),
        localize_column_name("zh", "AllocationType"),
        localize_column_name("zh", "WorkCenter"),
        localize_column_name("zh", "Demand_Tons"),
        localize_column_name("zh", "Allocated_Tons"),
        localize_column_name("zh", "Outsourced_Tons"),
        localize_column_name("zh", "Unmet_Tons"),
        localize_column_name("zh", "CapacityShare_Pct"),
        localize_column_name("zh", "Allocation_Source"),
        localize_column_name("zh", "Residual_After_Capacity_Tons"),
        localize_column_name("zh", "Residual_After_Routing_Tons"),
        localize_column_name("zh", "RouteType"),
        localize_column_name("zh", "Priority"),
        localize_column_name("zh", "Year"),
    ]
    for col_index, header in enumerate(detail_headers, start=1):
        detail_ws.cell(3, col_index).value = header

    if mode == "ModeA":
        detail_rows = [
            [localize_value("zh", "Max"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 60.0, 0.0, 40.0, 100.0, "", 40.0, 40.0, localize_value("zh", "Capacity"), 1, 2027],
            [localize_value("zh", "Planned"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 55.0, 0.0, 45.0, 100.0, "", 45.0, 45.0, localize_value("zh", "Capacity"), 1, 2027],
        ]
    else:
        detail_rows = [
            [localize_value("zh", "Max"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 60.0, 0.0, 0.0, 100.0, localize_value("zh", "Capacity_Base"), 40.0, 20.0, localize_value("zh", "Capacity"), 1, 2027],
            [localize_value("zh", "Max"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Internal"), "WC2", 100.0, 20.0, 0.0, 0.0, 100.0, localize_value("zh", "Routing_Reroute"), 40.0, 20.0, localize_value("zh", "Primary"), 1, 2027],
            [localize_value("zh", "Max"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Unmet"), "[UNALLOCATED]", 100.0, 0.0, 0.0, 20.0, 0.0, localize_value("zh", "Unmet"), 40.0, 20.0, localize_value("zh", "N/A"), 99, 2027],
            [localize_value("zh", "Planned"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 55.0, 0.0, 0.0, 100.0, localize_value("zh", "Capacity_Base"), 45.0, 20.0, localize_value("zh", "Capacity"), 1, 2027],
            [localize_value("zh", "Planned"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Internal"), "WC2", 100.0, 15.0, 0.0, 0.0, 100.0, localize_value("zh", "Routing_Reroute"), 45.0, 20.0, localize_value("zh", "Primary"), 1, 2027],
            [localize_value("zh", "Planned"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Outsourced"), "TOL1", 100.0, 0.0, 10.0, 0.0, 0.0, localize_value("zh", "Toller"), 45.0, 10.0, localize_value("zh", "Toller"), 3, 2027],
            [localize_value("zh", "Planned"), "2027-01", "PlannerA", "P1", "F1", "PLT1", localize_value("zh", "Unmet"), "[UNALLOCATED]", 100.0, 0.0, 0.0, 20.0, 0.0, localize_value("zh", "Unmet"), 45.0, 20.0, localize_value("zh", "N/A"), 99, 2027],
        ]

    for row_index, row in enumerate(detail_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            detail_ws.cell(row_index, col_index).value = value

    unmet_ws = workbook.create_sheet(localize_sheet_name("zh", "Unmet_Attribution_Detail"))
    unmet_headers = [
        "Capacity_Basis",
        localize_column_name("zh", "Month"),
        localize_column_name("zh", "PlannerName"),
        localize_column_name("zh", "Product"),
        localize_column_name("zh", "ProductFamily"),
        localize_column_name("zh", "Plant"),
        localize_column_name("zh", "Owner_WorkCenter"),
        localize_column_name("zh", "Capacity_Candidate_WorkCenters"),
        localize_column_name("zh", "Attributed_WorkCenter"),
        localize_column_name("zh", "Reference_Demand_Tons"),
        localize_column_name("zh", "Product_Unmet_Tons"),
        localize_column_name("zh", "Attributed_Unmet_Tons"),
        localize_column_name("zh", "Attribution_Rule"),
    ]
    for col_index, header in enumerate(unmet_headers, start=1):
        unmet_ws.cell(3, col_index).value = header

    if mode == "ModeA":
        unmet_rows = [
            [localize_value("zh", "Max"), "2027-01", "PlannerA", "P1", "F1", "PLT1", "WC1", "WC1", "WC1", 100.0, 40.0, 40.0, "模式A - 按计划员归属工作中心回挂"],
            [localize_value("zh", "Planned"), "2027-01", "PlannerA", "P1", "F1", "PLT1", "WC1", "WC1", "WC1", 100.0, 45.0, 45.0, "模式A - 按计划员归属工作中心回挂"],
        ]
    else:
        unmet_rows = [
            [localize_value("zh", "Max"), "2027-01", "", "P1", "F1", "PLT1", "", "WC1", "WC1", 100.0, 20.0, 20.0, "模式B - 回到基础产能工作中心"],
            [localize_value("zh", "Planned"), "2027-01", "", "P1", "F1", "PLT1", "", "WC1", "WC1", 100.0, 20.0, 20.0, "模式B - 回到基础产能工作中心"],
        ]
    for row_index, row in enumerate(unmet_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            unmet_ws.cell(row_index, col_index).value = value

    run_info_ws = workbook.create_sheet(localize_sheet_name("zh", "Run_Info"))
    run_info_ws["A1"] = localize_column_name("zh", "Capacity_Basis")
    run_info_ws["B1"] = "参数"
    run_info_ws["C1"] = "值"
    run_info_rows = [
        [localize_value("zh", "Max"), localize_column_name("zh", "Scenario_Name"), "基准"],
        [localize_value("zh", "Max"), localize_column_name("zh", "Input_Load_Folder"), str(input_dir)],
        [localize_value("zh", "Max"), localize_column_name("zh", "Input_Master_Folder"), str(input_dir)],
        [localize_value("zh", "Max"), localize_column_name("zh", "Output_Folder"), str(path.parent)],
        [localize_value("zh", "Max"), localize_column_name("zh", "Run_Timestamp"), "2026-05-04 09:00:00"],
        [localize_value("zh", "Planned"), localize_column_name("zh", "Scenario_Name"), "基准"],
    ]
    for row_index, row in enumerate(run_info_rows, start=2):
        for col_index, value in enumerate(row, start=1):
            run_info_ws.cell(row_index, col_index).value = value

    workbook.save(path)
    workbook.close()


class WorkCenterAnalysisReportTests(unittest.TestCase):
    def test_find_latest_mode_report_and_manual_resolution(self):
        with workspace_tempdir() as tmpdir:
            output_dir = Path(tmpdir)
            older = output_dir / "capacity_result_ModeA_Baseline_20260101_000000.xlsx"
            newer = output_dir / "capacity_result_ModeA_Baseline_20260102_000000.xlsx"
            modeb = output_dir / "capacity_result_ModeB_Baseline_20260103_000000.xlsx"
            temp_lock = output_dir / "~$capacity_result_ModeA_Baseline_20260104_000000.xlsx"
            for path in (older, newer, modeb, temp_lock):
                path.write_bytes(b"x")
            os.utime(older, (1, 1))
            os.utime(newer, (2, 2))
            os.utime(modeb, (3, 3))
            os.utime(temp_lock, (4, 4))

            latest = find_latest_mode_report(output_dir, "ModeA")
            selection = resolve_mode_report_selection(
                output_dir=output_dir,
                manual_report_path=older.name,
                use_latest_report=False,
                report_mode="ModeA",
            )

        self.assertEqual(latest, newer)
        self.assertEqual(selection.selected_path, older.resolve())
        self.assertEqual(selection.latest_path, newer)
        self.assertFalse(selection.is_latest)

    def test_load_mode_report_context_rejects_wrong_mode(self):
        with workspace_tempdir() as tmpdir:
            workspace_root = Path(tmpdir) / "workspace"
            input_dir = workspace_root / "Data_Input"
            output_dir = workspace_root / "output"
            input_dir.mkdir(parents=True)
            output_dir.mkdir(parents=True)
            report_path = output_dir / "capacity_result_ModeB_Baseline_20260504_090000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeB")

            with self.assertRaises(ReportValidationError):
                load_mode_report_context(report_path, expected_mode="ModeA")

    def test_generate_workcenter_analysis_report_modeb(self):
        with workspace_tempdir() as tmpdir:
            workspace_root = Path(tmpdir) / "workspace"
            input_dir = workspace_root / "Data_Input"
            output_dir = workspace_root / "output"
            input_dir.mkdir(parents=True)
            output_dir.mkdir(parents=True)
            report_path = output_dir / "capacity_result_ModeB_Baseline_20260504_090000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeB")

            pd.DataFrame(
                [
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P1", "ProductFamily": "F1", "Plant": "PLT1", "Forecast_Tons": 100.0, "ScenarioVersion": "Baseline"},
                ]
            ).to_csv(input_dir / "planner1_load.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "Product Family": "F1", "Resource": "WC1", "Annual Max Capacity Tons": 720.0, "Annual Planned Capacity Tons": 660.0, "Utilization Target": 1.0},
                ]
            ).to_csv(input_dir / "master_capacity.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "Product Family": "F1", "Resource": "WC2", "Max Capacity Ton": 20.0, "Planned Capacity Ton": 15.0, "EligibleFalg": "Y", "Router Type": "Primary"},
                    {"Product": "P1", "Product Family": "F1", "Resource": "TOL1", "Max Capacity Ton": 10.0, "Planned Capacity Ton": 10.0, "EligibleFalg": "Y", "Router Type": "Toller"},
                ]
            ).to_csv(input_dir / "master_routing.csv", index=False)

            output_path = generate_workcenter_analysis_report(
                report_path=report_path,
                workcenters=["WC1", "WC2"],
                report_mode="ModeB",
                output_name="wc_demo.xlsx",
            )

            workbook = load_workbook(output_path, read_only=True, data_only=True)

        self.assertTrue(output_path.name.startswith("wc_demo_"))
        self.assertEqual(workbook.sheetnames, ["总览", "1_WC1", "2_WC2"])
        summary_ws = workbook["总览"]
        self.assertEqual(summary_ws["A1"].value, "ModeB 工作中心分析报告")
        self.assertEqual(summary_ws["A5"].value, "工作中心")
        self.assertEqual(summary_ws["B5"].value, "报告模式")
        self.assertEqual(summary_ws["A6"].value, "WC1")
        wc2_ws = workbook["2_WC2"]
        values = [cell for row in wc2_ws.iter_rows(values_only=True) for cell in row if cell is not None]
        self.assertIn("Routing 填充", "".join(str(value) for value in values))
        self.assertIn("master_routing.csv", {str(value) for value in values})
        workbook.close()

    def test_generate_workcenter_analysis_report_modea_shows_no_routing_logic(self):
        with workspace_tempdir() as tmpdir:
            workspace_root = Path(tmpdir) / "workspace"
            input_dir = workspace_root / "Data_Input"
            output_dir = workspace_root / "output"
            input_dir.mkdir(parents=True)
            output_dir.mkdir(parents=True)
            report_path = output_dir / "capacity_result_ModeA_Baseline_20260504_090000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeA")

            pd.DataFrame(
                [
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P1", "ProductFamily": "F1", "Plant": "PLT1", "Forecast_Tons": 100.0, "ScenarioVersion": "Baseline"},
                ]
            ).to_csv(input_dir / "planner1_load.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "Product Family": "F1", "Resource": "WC1", "Annual Max Capacity Tons": 720.0, "Annual Planned Capacity Tons": 660.0, "Utilization Target": 1.0},
                ]
            ).to_csv(input_dir / "master_capacity.csv", index=False)

            output_path = generate_workcenter_analysis_report(
                report_path=report_path,
                workcenters=["WC1"],
                report_mode="ModeA",
                output_name="wc_modea.xlsx",
            )

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            ws = workbook["1_WC1"]
            values = [str(value) for row in ws.iter_rows(values_only=True) for value in row if value is not None]

        self.assertIn("ModeA 不使用 routing。", values)
        self.assertIn("最终 unmet 按 planner resource_group_owner 回挂到该工作中心。", values)
        workbook.close()
