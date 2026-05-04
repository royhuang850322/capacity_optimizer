import os
import shutil
import unittest
import uuid
from contextlib import contextmanager
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.customer_case_launcher import guess_workspace_root, is_capacity_optimizer_workspace
from app.i18n import localize_column_name, localize_sheet_name, localize_value
from app.modeb_customer_case_report import (
    ReportValidationError,
    find_latest_mode_report,
    generate_modeb_customer_case_report,
    infer_workspace_root_from_report,
    load_mode_report_context,
    resolve_mode_report_selection,
)
from app.runtime_paths import build_runtime_paths


TEST_TMP_ROOT = os.path.join(os.path.dirname(__file__), "_tmp")
os.makedirs(TEST_TMP_ROOT, exist_ok=True)


@contextmanager
def workspace_tempdir():
    tmpdir = os.path.join(TEST_TMP_ROOT, f"tmp_{uuid.uuid4().hex}")
    os.mkdir(tmpdir)
    try:
        yield tmpdir
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _write_mode_report(path: Path, input_dir: Path, mode: str, product_name: str = "P1") -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    detail_ws = workbook.create_sheet(localize_sheet_name("zh", "Allocation_Detail"))
    detail_headers = [
        "Capacity_Basis",
        localize_column_name("zh", "Month"),
        localize_column_name("zh", "PlannerName"),
        localize_column_name("zh", "Product"),
        localize_column_name("zh", "ProductFamily"),
        localize_column_name("zh", "Plant"),
        localize_column_name("zh", "AllocationType"),
        localize_column_name("zh", "WorkCenter"),
        localize_column_name("zh", "Demand_Tons"),
        localize_column_name("zh", "Allocated_Tons"),
        localize_column_name("zh", "Outsourced_Tons"),
        localize_column_name("zh", "Unmet_Tons"),
        localize_column_name("zh", "CapacityShare_Pct"),
        "Allocation_Source",
        "Residual_After_Capacity_Tons",
        "Residual_After_Routing_Tons",
        localize_column_name("zh", "RouteType"),
        localize_column_name("zh", "Priority"),
        localize_column_name("zh", "Year"),
    ]
    for col_index, header in enumerate(detail_headers, start=1):
        detail_ws.cell(3, col_index).value = header

    if mode == "ModeA":
        detail_rows = [
            ["Max", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 60.0, 0.0, 0.0, 60.0, "", 40.0, 40.0, "Capacity", 1, 2027],
            ["Max", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Unmet"), "[UNALLOCATED]", 100.0, 0.0, 0.0, 40.0, 0.0, "", 40.0, 40.0, "N/A", 99, 2027],
            ["Planned", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 55.0, 0.0, 0.0, 55.0, "", 45.0, 45.0, "Capacity", 1, 2027],
            ["Planned", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Unmet"), "[UNALLOCATED]", 100.0, 0.0, 0.0, 45.0, 0.0, "", 45.0, 45.0, "N/A", 99, 2027],
            ["Planned", "2027-01", "PlannerA", "P2", "F2", "PLT2", localize_value("zh", "Internal"), "WC2", 90.0, 90.0, 0.0, 0.0, 100.0, "", 0.0, 0.0, "Capacity", 1, 2027],
        ]
    else:
        detail_rows = [
            ["Max", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 60.0, 0.0, 0.0, 60.0, "Capacity_Base", 40.0, 40.0, "Capacity", 1, 2027],
            ["Max", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Outsourced"), "TOL1", 100.0, 0.0, 40.0, 0.0, 0.0, "Toller", 40.0, 0.0, "Toller", 1, 2027],
            ["Max", "2027-01", "PlannerA", "P2", "F2", "PLT2", localize_value("zh", "Internal"), "WC2", 90.0, 90.0, 0.0, 0.0, 100.0, "Capacity_Base", 0.0, 0.0, "Capacity", 1, 2027],
            ["Planned", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Internal"), "WC1", 100.0, 60.0, 0.0, 0.0, 60.0, "Capacity_Base", 40.0, 40.0, "Capacity", 1, 2027],
            ["Planned", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Internal"), "WC3", 100.0, 30.0, 0.0, 0.0, 30.0, "Routing_Reroute", 40.0, 10.0, "Primary", 1, 2027],
            ["Planned", "2027-01", "PlannerA", product_name, "F1", "PLT1", localize_value("zh", "Unmet"), "[UNALLOCATED]", 100.0, 0.0, 0.0, 10.0, 0.0, "Unmet", 40.0, 10.0, "N/A", 99, 2027],
            ["Planned", "2027-01", "PlannerA", "P2", "F2", "PLT2", localize_value("zh", "Internal"), "WC2", 90.0, 90.0, 0.0, 0.0, 100.0, "Capacity_Base", 0.0, 0.0, "Capacity", 1, 2027],
        ]
    for row_index, row in enumerate(detail_rows, start=4):
        for col_index, value in enumerate(row, start=1):
            detail_ws.cell(row_index, col_index).value = value

    run_info_ws = workbook.create_sheet(localize_sheet_name("zh", "Run_Info"))
    run_info_ws["A1"] = "Capacity_Basis"
    run_info_ws["B1"] = "参数"
    run_info_ws["C1"] = "值"
    run_info_rows = [
        ["Max", "Scenario_Name", "Baseline"],
        ["Max", "Input_Load_Folder", str(input_dir)],
        ["Max", "Input_Master_Folder", str(input_dir)],
        ["Max", "Output_Folder", str(input_dir)],
        ["Max", "Run_Timestamp", "2026-05-02 13:00:00"],
        ["Planned", "Scenario_Name", "Baseline"],
    ]
    for row_index, row in enumerate(run_info_rows, start=2):
        for col_index, value in enumerate(row, start=1):
            run_info_ws.cell(row_index, col_index).value = value

    workbook.save(path)
    workbook.close()


class ModeBCustomerCaseReportTests(unittest.TestCase):
    def test_find_latest_modeb_report_ignores_temp_files(self):
        with workspace_tempdir() as tmpdir:
            output_dir = Path(tmpdir)
            older = output_dir / "capacity_result_ModeB_Baseline_20260101_000000.xlsx"
            newer = output_dir / "capacity_result_ModeB_Baseline_20260102_000000.xlsx"
            temp_lock = output_dir / "~$capacity_result_ModeB_Baseline_20260103_000000.xlsx"
            modea = output_dir / "capacity_result_ModeA_Baseline_20260104_000000.xlsx"
            for path in (older, newer, temp_lock, modea):
                path.write_bytes(b"x")
            os.utime(older, (1, 1))
            os.utime(newer, (2, 2))
            os.utime(temp_lock, (3, 3))
            os.utime(modea, (4, 4))

            latest = find_latest_mode_report(output_dir, "ModeB")
            selection = resolve_mode_report_selection(
                output_dir=output_dir,
                manual_report_path=older.name,
                use_latest_report=False,
                report_mode="ModeB",
            )

        self.assertEqual(latest, newer)
        self.assertEqual(selection.selected_path, older.resolve())
        self.assertEqual(selection.latest_path, newer)
        self.assertFalse(selection.is_latest)

    def test_load_modeb_report_context_accepts_localized_workbook(self):
        with workspace_tempdir() as tmpdir:
            input_dir = Path(tmpdir) / "Data_Input"
            input_dir.mkdir()
            report_path = Path(tmpdir) / "capacity_result_ModeB_Baseline_20260502_130000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeB")

            context = load_mode_report_context(report_path, expected_mode="ModeB")

        self.assertEqual(context.scenario_name, "Baseline")
        self.assertEqual(context.available_bases, ("Max", "Planned"))
        self.assertEqual(len(context.detail_rows), 7)
        self.assertEqual(context.input_load_folder, input_dir.resolve())

    def test_load_mode_report_context_rejects_wrong_mode(self):
        with workspace_tempdir() as tmpdir:
            input_dir = Path(tmpdir) / "Data_Input"
            input_dir.mkdir()
            report_path = Path(tmpdir) / "capacity_result_ModeB_Baseline_20260502_130000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeB")

            with self.assertRaises(ReportValidationError):
                load_mode_report_context(report_path, expected_mode="ModeA")

    def test_generate_modeb_customer_case_report_creates_requested_product_sheets(self):
        with workspace_tempdir() as tmpdir:
            input_dir = Path(tmpdir) / "Data_Input"
            input_dir.mkdir()
            report_path = Path(tmpdir) / "capacity_result_ModeB_Baseline_20260502_130000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeB")

            pd.DataFrame(
                [
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P1", "ProductFamily": "F1", "Plant": "PLT1", "Forecast_Tons": 100.0, "ScenarioVersion": "Baseline"},
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P2", "ProductFamily": "F2", "Plant": "PLT2", "Forecast_Tons": 90.0, "ScenarioVersion": "Baseline"},
                ]
            ).to_csv(input_dir / "planner1_load.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "WorkCenter": "WC1", "Annual_Max_Capacity_Tons": 720.0, "Annual_Planned_Capacity_Tons": 720.0, "Utilization_Target": 1.0},
                    {"Product": "P2", "WorkCenter": "WC2", "Annual_Max_Capacity_Tons": 1080.0, "Annual_Planned_Capacity_Tons": 1080.0, "Utilization_Target": 1.0},
                ]
            ).to_csv(input_dir / "master_capacity.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "Resource": "WC3", "Max Capacity Ton": 30.0, "Planned Capacity Ton": 30.0, "EligibleFalg": "Y", "Router Type": "Primary"},
                    {"Product": "P1", "Resource": "TOL1", "Max Capacity Ton": 40.0, "Planned Capacity Ton": 40.0, "EligibleFalg": "Y", "Router Type": "Toller"},
                ]
            ).to_csv(input_dir / "master_routing.csv", index=False)

            output_path = generate_modeb_customer_case_report(
                report_path=report_path,
                products=["P1", "P2"],
                output_dir=tmpdir,
                output_name="product_demo.xlsx",
            )

            workbook = load_workbook(output_path, read_only=True, data_only=True)

        self.assertTrue(output_path.name.startswith("product_demo_"))
        self.assertEqual(workbook.sheetnames, ["总览", "1_P1", "2_P2"])
        summary_ws = workbook["总览"]
        self.assertEqual(summary_ws["A1"].value, "ModeB 产品分析报告")
        self.assertEqual(summary_ws["A5"].value, "产品")
        self.assertEqual(summary_ws["B5"].value, "案例类型")
        product_ws = workbook["1_P1"]
        self.assertEqual(product_ws["A1"].value, "P1 - ModeB 产品分析")
        self.assertEqual(product_ws["A5"].value, "Capacity_Basis")
        self.assertEqual(product_ws["B5"].value, "案例类型")
        workbook.close()

    def test_generate_modea_product_analysis_report(self):
        with workspace_tempdir() as tmpdir:
            input_dir = Path(tmpdir) / "Data_Input"
            input_dir.mkdir()
            report_path = Path(tmpdir) / "capacity_result_ModeA_Baseline_20260502_130000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeA")

            pd.DataFrame(
                [
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P1", "ProductFamily": "F1", "Plant": "PLT1", "Forecast_Tons": 100.0, "ScenarioVersion": "Baseline"},
                ]
            ).to_csv(input_dir / "planner1_load.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "WorkCenter": "WC1", "Annual_Max_Capacity_Tons": 720.0, "Annual_Planned_Capacity_Tons": 660.0, "Utilization_Target": 1.0},
                ]
            ).to_csv(input_dir / "master_capacity.csv", index=False)

            output_path = generate_modeb_customer_case_report(
                report_path=report_path,
                products=["P1"],
                report_mode="ModeA",
                output_dir=tmpdir,
                output_name="modea_product_demo.xlsx",
            )

            workbook = load_workbook(output_path, read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames, ["总览", "1_P1"])
        self.assertEqual(workbook["总览"]["A1"].value, "ModeA 产品分析报告")
        self.assertEqual(workbook["1_P1"]["A1"].value, "P1 - ModeA 产品分析")
        values = [str(value) for row in workbook["1_P1"].iter_rows(values_only=True) for value in row if value is not None]
        self.assertIn("ModeA 不使用 routing", values)
        workbook.close()

    def test_generate_product_analysis_sanitizes_sheet_title(self):
        with workspace_tempdir() as tmpdir:
            input_dir = Path(tmpdir) / "Data_Input"
            input_dir.mkdir()
            report_path = Path(tmpdir) / "capacity_result_ModeB_Baseline_20260502_130000.xlsx"
            _write_mode_report(report_path, input_dir, "ModeB", product_name="P/1")

            pd.DataFrame(
                [
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P/1", "ProductFamily": "F1", "Plant": "PLT1", "Forecast_Tons": 100.0, "ScenarioVersion": "Baseline"},
                ]
            ).to_csv(input_dir / "planner1_load.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P/1", "WorkCenter": "WC1", "Annual_Max_Capacity_Tons": 720.0, "Annual_Planned_Capacity_Tons": 720.0, "Utilization_Target": 1.0},
                ]
            ).to_csv(input_dir / "master_capacity.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P/1", "Resource": "WC3", "Max Capacity Ton": 30.0, "Planned Capacity Ton": 30.0, "EligibleFalg": "Y", "Router Type": "Primary"},
                ]
            ).to_csv(input_dir / "master_routing.csv", index=False)

            output_path = generate_modeb_customer_case_report(
                report_path=report_path,
                products=["P/1"],
                report_mode="ModeB",
                output_dir=tmpdir,
                output_name="slash_product_demo.xlsx",
            )

            workbook = load_workbook(output_path, read_only=True, data_only=True)

        self.assertEqual(workbook.sheetnames, ["总览", "1_P_1"])
        workbook.close()

    def test_generate_modeb_customer_case_report_falls_back_to_report_workspace_root(self):
        with workspace_tempdir() as tmpdir:
            workspace_root = Path(tmpdir) / "CapacityOptimizerWorkspace"
            input_dir = workspace_root / "Data_Input"
            output_dir = workspace_root / "output"
            input_dir.mkdir(parents=True)
            output_dir.mkdir(parents=True)
            report_path = output_dir / "capacity_result_ModeB_Baseline_20260502_130000.xlsx"

            _write_mode_report(report_path, Path(tmpdir) / "missing_input", "ModeB")

            pd.DataFrame(
                [
                    {"Month": "2027-01", "PlannerName": "PlannerA", "Product": "P1", "ProductFamily": "F1", "Plant": "PLT1", "Forecast_Tons": 100.0, "ScenarioVersion": "Baseline"},
                ]
            ).to_csv(input_dir / "planner1_load.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "WorkCenter": "WC1", "Annual_Max_Capacity_Tons": 720.0, "Annual_Planned_Capacity_Tons": 720.0, "Utilization_Target": 1.0},
                ]
            ).to_csv(input_dir / "master_capacity.csv", index=False)
            pd.DataFrame(
                [
                    {"Product": "P1", "Resource": "WC3", "Max Capacity Ton": 30.0, "Planned Capacity Ton": 30.0, "EligibleFalg": "Y", "Router Type": "Primary"},
                ]
            ).to_csv(input_dir / "master_routing.csv", index=False)

            output_path = generate_modeb_customer_case_report(
                report_path=report_path,
                products=["P1"],
                output_name="shared_workspace_demo.xlsx",
            )

            self.assertEqual(output_path.parent, output_dir)
            self.assertTrue(output_path.exists())

    def test_infer_workspace_root_from_report_and_workspace_guessing(self):
        with workspace_tempdir() as tmpdir:
            install_dir = Path(tmpdir) / "ModeBCompanion"
            workspace_root = Path(tmpdir) / "CapacityOptimizerWorkspace"
            (workspace_root / "Data_Input").mkdir(parents=True)
            (workspace_root / "output").mkdir(parents=True)
            (workspace_root / "CapacityOptimizer.exe").write_bytes(b"x")
            report_path = workspace_root / "output" / "capacity_result_ModeB_Baseline_20260502_130000.xlsx"
            report_path.write_bytes(b"x")

            runtime_paths = build_runtime_paths(
                install_dir=install_dir,
                bundled_resources_dir=install_dir,
                user_workspace_dir=install_dir,
                frozen=True,
            )

            guessed = guess_workspace_root(runtime_paths, workspace_root)

            self.assertTrue(is_capacity_optimizer_workspace(workspace_root))
            self.assertEqual(infer_workspace_root_from_report(report_path), workspace_root.resolve())
            self.assertEqual(guessed, workspace_root.resolve())
