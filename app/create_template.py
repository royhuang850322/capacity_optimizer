"""
Create the Excel control workbook used by the Capacity Optimizer.
"""
from __future__ import annotations

import json
import os
from datetime import datetime

import click
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.data_loader import discover_planner_scenarios
from app.runtime_paths import ensure_workspace_dirs, resolve_runtime_paths


HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
CFG_FILL = PatternFill("solid", fgColor="D9E1F2")
VAL_FILL = PatternFill("solid", fgColor="FFFFFF")
MIGRATION_FILL = PatternFill("solid", fgColor="FCE4D6")
MIGRATION_VAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(size=14, bold=True, color="1F4E79")
BTN_FILL = PatternFill("solid", fgColor="2F75B5")
BTN_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DEFAULT_PROJECT_ROOT = ".."
RUNTIME_PATHS = resolve_runtime_paths()
DEFAULT_LOAD_DIR = str(RUNTIME_PATHS.workspace_input_dir)
DEFAULT_OUT = str(RUNTIME_PATHS.control_workbook_path)


@click.command()
@click.option("--out", default=DEFAULT_OUT, show_default=True, help="Output path for the control workbook.")
def main(out: str) -> None:
    write_control_workbook(out)
    click.echo(f"Control workbook written to {out}")


def write_control_workbook(out: str, load_dir: str | None = None) -> None:
    ensure_workspace_dirs()
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    scenario_options = _scenario_options(load_dir or DEFAULT_LOAD_DIR)
    project_root = _default_project_root_for_workbook(out)

    workbook = Workbook()
    workbook.remove(workbook.active)

    lists_ws = workbook.create_sheet("Lists")
    _create_lists(lists_ws, scenario_options)

    deployment_ws = workbook.create_sheet("Deployment_Steps")
    _create_deployment_steps(deployment_ws)

    instructions_ws = workbook.create_sheet("Instructions")
    _create_instructions(instructions_ws)

    license_ws = workbook.create_sheet("License")
    _create_license_sheet(license_ws, _inspect_current_license(project_root))

    control_ws = workbook.create_sheet("Control_Panel")
    _create_control_panel(control_ws, lists_ws, scenario_options)

    lists_ws.sheet_state = "hidden"
    workbook.active = workbook.index(deployment_ws)
    workbook.save(out)


def refresh_control_workbook_license_sheet(
    workbook_path: str,
    *,
    project_root: str | None = None,
    license_info=None,
) -> None:
    workbook = load_workbook(workbook_path)
    try:
        sheet_index = workbook.index(workbook["License"]) if "License" in workbook.sheetnames else workbook.index(workbook["Control_Panel"])
        if "License" in workbook.sheetnames:
            workbook.remove(workbook["License"])
        license_ws = workbook.create_sheet("License", sheet_index)
        effective_project_root = project_root or _default_project_root_for_workbook(workbook_path)
        _create_license_sheet(
            license_ws,
            _inspect_current_license(effective_project_root, license_info=license_info),
        )
        workbook.save(workbook_path)
    finally:
        workbook.close()


def _default_project_root_for_workbook(workbook_path: str) -> str:
    workbook_dir = os.path.dirname(os.path.abspath(workbook_path))
    return os.path.abspath(os.path.join(workbook_dir, DEFAULT_PROJECT_ROOT))


def _inspect_current_license(project_root: str, license_info=None) -> dict[str, str]:
    active_path = os.path.join(project_root, "licenses", "active", "license.json")
    legacy_path = os.path.join(project_root, "license.json")
    existing_path = next((path for path in (active_path, legacy_path) if os.path.exists(path)), active_path)

    if license_info is not None:
        return _license_display_payload(
            status=getattr(license_info, "status", "Valid"),
            license_id=getattr(license_info, "license_id", ""),
            license_type=getattr(license_info, "license_type", ""),
            customer_name=getattr(license_info, "customer_name", ""),
            issue_date=getattr(license_info, "issue_date", ""),
            expiry_date=getattr(license_info, "expiry_date", ""),
            binding_mode=getattr(license_info, "binding_mode", ""),
            machine_label=getattr(license_info, "machine_label", ""),
            license_path=getattr(license_info, "license_path", existing_path),
            note=getattr(license_info, "note", ""),
            message="License validated successfully.",
        )

    try:
        from app.license_validator import LicenseValidationError, validate_license
    except Exception as exc:
        return _license_display_payload(
            status="Unavailable",
            license_path=existing_path,
            message=f"License validation module unavailable: {exc}",
        )

    try:
        info = validate_license(project_root)
        return _license_display_payload(
            status=info.status,
            license_id=info.license_id,
            license_type=info.license_type,
            customer_name=info.customer_name,
            issue_date=info.issue_date,
            expiry_date=info.expiry_date,
            binding_mode=info.binding_mode,
            machine_label=info.machine_label,
            license_path=info.license_path,
            note=info.note,
            message="License validated successfully.",
        )
    except LicenseValidationError as exc:
        raw_payload = _best_effort_license_payload(existing_path)
        status = "Not configured" if not os.path.exists(existing_path) else "Invalid"
        return _license_display_payload(
            status=status,
            license_id=raw_payload.get("license_id", ""),
            license_type=raw_payload.get("license_type", ""),
            customer_name=raw_payload.get("customer_name", ""),
            issue_date=raw_payload.get("issue_date", ""),
            expiry_date=raw_payload.get("expiry_date", ""),
            binding_mode=raw_payload.get("binding_mode", ""),
            machine_label=raw_payload.get("machine_label", ""),
            license_path=existing_path,
            note=raw_payload.get("note", ""),
            message=str(exc).splitlines()[0] if str(exc).strip() else "License is not configured.",
        )


def _best_effort_license_payload(path: str) -> dict[str, str]:
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    return {str(key): str(value).strip() for key, value in payload.items() if value is not None}


def _license_display_payload(
    *,
    status: str,
    license_id: str = "",
    license_type: str = "",
    customer_name: str = "",
    issue_date: str = "",
    expiry_date: str = "",
    binding_mode: str = "",
    machine_label: str = "",
    license_path: str = "",
    note: str = "",
    message: str = "",
) -> dict[str, str]:
    return {
        "License_Status": status or "Unknown",
        "License_Name": license_id or "",
        "License_Mode": _friendly_license_mode(license_type, binding_mode),
        "License_Type": license_type or "",
        "Binding_Mode": binding_mode or "",
        "Licensed_To": customer_name or "",
        "Issue_Date": issue_date or "",
        "Expiry_Date": expiry_date or "",
        "Machine_Name": machine_label or ("Not bound" if binding_mode == "unbound" else ""),
        "License_File_Path": license_path or "",
        "Note": note or "",
        "Message": message or "",
    }


def _friendly_license_mode(license_type: str, binding_mode: str) -> str:
    license_label = {
        "trial": "Trial",
        "commercial": "Commercial",
    }.get(str(license_type or "").strip().lower(), str(license_type or "").strip().title())
    binding_label = {
        "unbound": "Unbound",
        "machine_locked": "Machine Locked",
    }.get(str(binding_mode or "").strip().lower(), str(binding_mode or "").strip().replace("_", " ").title())
    if license_label and binding_label:
        return f"{license_label} / {binding_label}"
    return license_label or binding_label


def _create_license_sheet(worksheet, license_display: dict[str, str]) -> None:
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 68

    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "Current License"
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    worksheet["A1"].fill = HDR_FILL
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].border = BORDER
    worksheet.row_dimensions[1].height = 24

    worksheet.merge_cells("A2:C2")
    worksheet["A2"] = (
        "This sheet shows the currently detected license for this tool copy. "
        "Values refresh when the workbook is generated and again when the optimizer can save updates back to this file."
    )
    worksheet["A2"].font = Font(color="7F6000", bold=True, size=10)
    worksheet["A2"].fill = MIGRATION_VAL_FILL
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet["A2"].border = BORDER
    worksheet.row_dimensions[2].height = 34

    _write_header(worksheet, 4, ["Parameter", "Value", "Description"])
    rows = [
        ("License_Status", license_display.get("License_Status", ""), "Current validation status of the active license"),
        ("License_Name", license_display.get("License_Name", ""), "Current license ID / license name"),
        ("License_Mode", license_display.get("License_Mode", ""), "Friendly label such as Trial / Unbound or Commercial / Machine Locked"),
        ("License_Type", license_display.get("License_Type", ""), "License type from license.json"),
        ("Binding_Mode", license_display.get("Binding_Mode", ""), "Whether the license is unbound or machine locked"),
        ("Licensed_To", license_display.get("Licensed_To", ""), "Customer name recorded in the license"),
        ("Issue_Date", license_display.get("Issue_Date", ""), "License issue date"),
        ("Expiry_Date", license_display.get("Expiry_Date", ""), "License expiry date"),
        ("Machine_Name", license_display.get("Machine_Name", ""), "Licensed machine label when applicable"),
        ("License_File_Path", license_display.get("License_File_Path", ""), "Detected license.json location"),
        ("Note", license_display.get("Note", ""), "Note stored inside the license file"),
        ("Message", license_display.get("Message", ""), "Last license inspection message"),
    ]
    for row_num, (label, value, description) in enumerate(rows, start=5):
        worksheet[f"A{row_num}"] = label
        worksheet[f"B{row_num}"] = value
        worksheet[f"C{row_num}"] = description
        worksheet[f"A{row_num}"].font = Font(bold=True, size=10)
        worksheet[f"A{row_num}"].fill = CFG_FILL
        worksheet[f"B{row_num}"].fill = VAL_FILL
        worksheet[f"C{row_num}"].fill = VAL_FILL
        for col in ("A", "B", "C"):
            worksheet[f"{col}{row_num}"].border = BORDER
            worksheet[f"{col}{row_num}"].alignment = Alignment(wrap_text=True, vertical="top")
        worksheet[f"C{row_num}"].font = Font(italic=True, color="666666", size=9)

    worksheet.freeze_panes = "A5"


def _create_deployment_steps(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 34
    worksheet.column_dimensions["C"].width = 88
    worksheet.column_dimensions["D"].width = 72

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "Deployment Steps - copy the tool, install dependencies, activate the license, and run"
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    worksheet["A1"].fill = HDR_FILL
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].border = BORDER
    worksheet.row_dimensions[1].height = 24

    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = (
        "Recommended portable setup: keep Capacity_Optimizer_Control.xlsx inside 'Tooling Control Panel', "
        "then use Project_Root_Folder = '..', Input_Load_Folder = 'Data_Input', "
        "Input_Master_Folder = 'Data_Input', Output_Folder = 'output', and place 'license.json' in 'licenses\\active\\'. "
        "中文说明见右侧列。"
    )
    worksheet["A2"].font = Font(color="7F6000", bold=True, size=10)
    worksheet["A2"].fill = MIGRATION_VAL_FILL
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet["A2"].border = BORDER
    worksheet.row_dimensions[2].height = 34

    _write_header(worksheet, 4, ["Step", "Action", "What To Do", "中文说明"])
    steps = [
        (
            "Step 1",
            "Copy the full folder",
            "Copy the entire 'capacity_optimizer' folder to the target computer. Do not copy only the Excel file.",
            "把整个 `capacity_optimizer` 文件夹完整复制到目标电脑，不要只复制 Excel 文件。",
        ),
        (
            "Step 2",
            "Install Python",
            "Install Python on the target computer and confirm 'python --version' works in Command Prompt or PowerShell.",
            "在目标电脑安装 Python，并确认在命令行里执行 `python --version` 可以正常返回版本号。",
        ),
        (
            "Step 3",
            "Run the setup batch file",
            "Double-click 'runtime\\setup_requirements.bat' to install the required Python packages automatically.",
            "双击 `runtime\\setup_requirements.bat`，自动安装所需 Python 依赖。",
        ),
        (
            "Step 4",
            "Get the machine fingerprint",
            "If RSCP already gave you a trial/unbound license, skip to Step 6. Otherwise double-click 'runtime\\get_machine_fingerprint.bat'. It creates a timestamped fingerprint file under 'licenses\\requests\\'.",
            "如果 RSCP 已经直接给了试用版或 unbound 授权，可以跳到 Step 6。否则双击 `runtime\\get_machine_fingerprint.bat`，它会在 `licenses\\requests\\` 下生成带时间戳的机器指纹文件。",
        ),
        (
            "Step 5",
            "Request the license file",
            "For a machine-locked license, send the fingerprint file from 'licenses\\requests\\' to RSCP and request the signed 'license.json' for this computer. For a short-term trial, RSCP can issue an unbound license directly.",
            "如果要机绑授权，把 `licenses\\requests\\` 下生成的机器指纹文件发给 RSCP，并申请这台电脑专用的签名授权文件 `license.json`。如果只是短期试用，RSCP 也可以直接发 unbound 授权。",
        ),
        (
            "Step 6",
            "Place the license file",
            "Copy the signed 'license.json' into 'licenses\\active\\license.json'. The legacy project-root 'license.json' path is still supported.",
            "把签名后的 `license.json` 放到 `licenses\\active\\license.json`。旧的项目根目录 `license.json` 方式仍然兼容。",
        ),
        (
            "Step 7",
            "Open the control workbook",
            "Open 'Tooling Control Panel\\Capacity_Optimizer_Control.xlsx'. Save the workbook after any edits, then go to the 'Control_Panel' sheet after reading this page.",
            "打开 `Tooling Control Panel\\Capacity_Optimizer_Control.xlsx`。每次修改后先保存，再切换到 `Control_Panel`。",
        ),
        (
            "Step 8",
            "Check migration settings",
            "In Control_Panel, first check Project_Root_Folder, Input_Load_Folder, Input_Master_Folder, and Output_Folder. If the workbook stays in 'Tooling Control Panel', keep '..', 'Data_Input', 'Data_Input', and 'output'.",
            "先检查 `Project_Root_Folder`、`Input_Load_Folder`、`Input_Master_Folder`、`Output_Folder`。如果 workbook 仍然放在 `Tooling Control Panel` 目录中，默认保持 `.. / Data_Input / Data_Input / output` 即可。",
        ),
        (
            "Step 9",
            "Check run settings",
            "Then review Scenario_Name, Start_Year, Start_Month_Num, Horizon_Months, Run_Mode, Verbose, and Skip_Validation_Errors.",
            "再检查运行参数，例如 `Scenario_Name`、起始年月、`Horizon_Months`、`Run_Mode`、`Verbose`、`Skip_Validation_Errors`。",
        ),
        (
            "Step 10",
            "Run the tool",
            "Press Ctrl+S to save first. Then click 'Run Optimizer' in Control_Panel, or run: python -m app.main --input-template \"Tooling Control Panel/Capacity_Optimizer_Control.xlsx\". If required packages or license files are missing, the run batch file will stop and tell you what to do next.",
            "先按 `Ctrl+S` 保存，再在 `Control_Panel` 点击 `Run Optimizer`，或者在命令行运行：`python -m app.main --input-template \"Tooling Control Panel/Capacity_Optimizer_Control.xlsx\"`。如果依赖或授权文件缺失，运行批处理会先停止并明确告诉你下一步该做什么。",
        ),
        (
            "Step 11",
            "Review output",
            "Open the 'output' folder and check the generated Excel reports. If Run_Mode = Both, the comparison workbook is also generated.",
            "打开 `output` 文件夹查看结果。如果 `Run_Mode = Both`，还会额外生成一份 ModeA 和 ModeB 的对比报告。",
        ),
        (
            "Step 12",
            "If the run fails",
            "Check that the four migration settings point to valid folders, master_capacity.csv and planner files exist, Python packages were installed successfully, and a valid license is present under 'licenses\\active\\license.json' or the legacy project-root path. If needed, run 'runtime\\setup_requirements.bat' again or regenerate the machine fingerprint.",
            "如果运行失败，优先检查四个迁移路径是否正确，`master_capacity.csv` 和 planner 文件是否存在，依赖包是否安装成功，以及 `licenses\\active\\license.json` 或旧项目根目录下是否有有效授权。必要时重新运行 `runtime\\setup_requirements.bat` 或重新生成机器指纹。",
        ),
    ]

    for row_num, (step, action, detail, detail_cn) in enumerate(steps, start=5):
        worksheet[f"A{row_num}"] = step
        worksheet[f"B{row_num}"] = action
        worksheet[f"C{row_num}"] = detail
        worksheet[f"D{row_num}"] = detail_cn
        worksheet[f"A{row_num}"].fill = MIGRATION_FILL
        worksheet[f"A{row_num}"].font = Font(bold=True, color="7F6000")
        worksheet[f"B{row_num}"].font = Font(bold=True)
        worksheet[f"D{row_num}"].fill = SUBHDR_FILL
        worksheet[f"D{row_num}"].font = Font(color="1F1F1F", size=9)
        for col in ("A", "B", "C", "D"):
            worksheet[f"{col}{row_num}"].border = BORDER
            worksheet[f"{col}{row_num}"].alignment = Alignment(wrap_text=True, vertical="top")

    footer_row = 5 + len(steps) + 2
    worksheet.merge_cells(f"A{footer_row}:D{footer_row}")
    worksheet[f"A{footer_row}"] = "Where to continue: open the 'Control_Panel' sheet and start with the Migration Setup section."
    worksheet[f"A{footer_row}"].font = Font(color="1F4E79", bold=True, size=10)
    worksheet[f"A{footer_row}"].fill = SUBHDR_FILL
    worksheet[f"A{footer_row}"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet[f"A{footer_row}"].border = BORDER
    worksheet.freeze_panes = "A5"


def _create_instructions(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 110
    lines = [
        ("CHEMICAL CAPACITY OPTIMIZER - EXCEL WORKFLOW", True),
        ("", False),
        ("1. Run runtime\\setup_requirements.bat on a new computer.", False),
        ("2. For a short trial, ask RSCP for an unbound trial license and place it in licenses\\active\\license.json.", False),
        ("3. For a machine-locked license, run runtime\\get_machine_fingerprint.bat and send the file from licenses\\requests\\ to RSCP.", False),
        ("4. Place the signed license.json in licenses\\active\\license.json.", False),
        ("5. Fill in the Control_Panel sheet.", False),
        ("6. Planner and master data stay in CSV/Excel files on disk; the Python tool reads them directly.", False),
        ("7. Run: python -m app.main --input-template \"Tooling Control Panel/Capacity_Optimizer_Control.xlsx\"", False),
        ("   Or click the Run button inside Control_Panel.", False),
        ("8. The tool writes one Excel result workbook per mode, with dashboard/report sheets included.", False),
        ("", False),
        ("KEY SETTINGS", True),
        ("Project_Root_Folder: main project folder. Keep '..' when the workbook stays inside Tooling Control Panel.", False),
        ("Input_Load_Folder: planner folder name or path. Relative values are resolved from Project_Root_Folder.", False),
        ("Input_Master_Folder: master-data folder name or path. Relative values are resolved from Project_Root_Folder.", False),
        ("Output_Folder: output folder name or path. Relative values are resolved from Project_Root_Folder.", False),
        ("Output_FileName: base output file name; mode and timestamp are appended automatically.", False),
        ("Scenario_Name: planner scenario to load. If your input folder changes, type the scenario manually if needed.", False),
        ("Start_Year / Start_Month_Num: first planning bucket.", False),
        ("Horizon_Months: number of months to optimize.", False),
        ("Run_Mode: ModeA, ModeB, or Both.", False),
        ("Direct_Mode: keep 'Yes' for the folder-based CSV workflow.", False),
        ("Verbose: prints solver detail in the command window.", False),
        ("Skip_Validation_Errors: only use 'Yes' when you intentionally want a forced run.", False),
        ("", False),
        ("LICENSE", True),
        ("license.json: recommended location is licenses\\active\\license.json; the legacy project-root path is also supported.", False),
        ("Trial / unbound license: no machine fingerprint required; RSCP can issue it directly.", False),
        ("runtime\\get_machine_fingerprint.bat: creates a timestamped fingerprint file under licenses\\requests\\ for RSCP to issue a machine-locked license.", False),
        ("Run_Info: each output workbook records license status, license ID, customer, expiry date, and binding mode.", False),
        ("", False),
        ("RESULT WORKBOOK CONTENT", True),
        ("Dashboard, Monthly_Trend, Bottleneck, WC_Heatmap, Product_Risk", False),
        ("Allocation_Detail, Allocation_Summary, Outsource_Summary, Unmet_Summary, WC_Load_Pct", False),
        ("Binary_Feasibility, Validation_Issues, Run_Info", False),
    ]

    for row_num, (text, bold) in enumerate(lines, start=1):
        cell = worksheet.cell(row_num, 1, text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(
            bold=bold,
            size=10 if bold else 9,
            color="1F4E79" if bold else "000000",
        )


def _create_control_panel(worksheet, lists_ws, scenario_options: list[str]) -> None:
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 54
    worksheet.column_dimensions["C"].width = 64

    _write_header(worksheet, 1, ["Parameter", "Value", "Description / Example"])

    worksheet.merge_cells("A2:C2")
    worksheet["A2"] = "Migration Setup - update these cells first after copying the tool to another computer"
    worksheet["A2"].font = Font(bold=True, color="FFFFFF", size=11)
    worksheet["A2"].fill = BTN_FILL
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"].border = BORDER
    worksheet.row_dimensions[2].height = 22

    worksheet.merge_cells("A3:C3")
    worksheet["A3"] = (
        "Recommended portable setup: keep Project_Root_Folder = '..', "
        "Input folders = 'Data_Input', Output_Folder = 'output', and place 'license.json' in 'licenses\\active\\'. "
        "See 'Deployment_Steps' for the full setup checklist."
    )
    worksheet["A3"].font = Font(color="7F6000", bold=True, size=10)
    worksheet["A3"].fill = MIGRATION_VAL_FILL
    worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet["A3"].border = BORDER
    worksheet.row_dimensions[3].height = 32

    now = datetime.now()
    default_scenario = scenario_options[0] if scenario_options else "Base Scenario"

    migration_rows = [
        ("Project_Root_Folder", DEFAULT_PROJECT_ROOT, "Project root folder; '..' means one level above this workbook"),
        ("Input_Load_Folder", "Data_Input", "Planner folder name or full path"),
        ("Input_Master_Folder", "Data_Input", "Master-data folder name or full path"),
        ("Output_Folder", "output", "Result folder name or full path"),
    ]
    run_rows = [
        ("Output_FileName", "capacity_result.xlsx", "Base name for result workbooks"),
        ("Scenario_Name", default_scenario, "Planner Scenario value; type manually if your data folder changes"),
        ("Start_Year", str(now.year), "First planning year"),
        ("Start_Month_Num", str(now.month), "First planning month number (1-12)"),
        ("Start_Month", f"{now.year}-{now.month:02d}", "Optional legacy field; Python derives this from year/month first"),
        ("Horizon_Months", "60", "Number of monthly buckets"),
        ("Run_Mode", "Both", "ModeA, ModeB, or Both"),
        ("Direct_Mode", "Yes", "Use direct folder-based CSV/Excel inputs"),
        ("Verbose", "No", "Print solver detail in the command window"),
        ("Skip_Validation_Errors", "No", "Run even when validation finds errors"),
        ("Run_Timestamp", "", "Filled by Python at runtime"),
        ("Notes", "", "Free text"),
    ]

    value_rows: dict[str, int] = {}
    current_row = 4
    for key, value, description in migration_rows:
        row_num = current_row
        value_rows[key] = row_num
        key_cell = worksheet.cell(row_num, 1, key)
        val_cell = worksheet.cell(row_num, 2, value)
        desc_cell = worksheet.cell(row_num, 3, description)

        key_cell.font = Font(bold=True, size=10)
        key_cell.fill = MIGRATION_FILL
        key_cell.border = BORDER

        val_cell.fill = MIGRATION_VAL_FILL
        val_cell.border = BORDER
        val_cell.alignment = Alignment(horizontal="left")
        val_cell.font = Font(bold=True, color="7F6000")

        desc_cell.font = Font(italic=True, color="7F6000", size=9)
        desc_cell.fill = MIGRATION_VAL_FILL
        desc_cell.border = BORDER
        desc_cell.alignment = Alignment(wrap_text=True)
        current_row += 1

    worksheet.merge_cells(f"A{current_row}:C{current_row}")
    worksheet[f"A{current_row}"] = "Run Settings"
    worksheet[f"A{current_row}"].font = Font(bold=True, color="FFFFFF", size=11)
    worksheet[f"A{current_row}"].fill = HDR_FILL
    worksheet[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet[f"A{current_row}"].border = BORDER
    worksheet.row_dimensions[current_row].height = 20
    current_row += 1

    for key, value, description in run_rows:
        row_num = current_row
        value_rows[key] = row_num
        key_cell = worksheet.cell(row_num, 1, key)
        val_cell = worksheet.cell(row_num, 2, value)
        desc_cell = worksheet.cell(row_num, 3, description)

        key_cell.font = Font(bold=True, size=10)
        key_cell.fill = CFG_FILL
        key_cell.border = BORDER

        val_cell.fill = VAL_FILL
        val_cell.border = BORDER
        val_cell.alignment = Alignment(horizontal="left")

        desc_cell.font = Font(italic=True, color="666666", size=9)
        desc_cell.border = BORDER
        desc_cell.alignment = Alignment(wrap_text=True)
        current_row += 1

    worksheet.freeze_panes = "B4"
    worksheet[f"A{current_row + 1}"] = "Run command"
    worksheet[f"A{current_row + 1}"].font = Font(bold=True, color="1F4E79", size=11)
    worksheet[f"B{current_row + 1}"] = 'python -m app.main --input-template "Tooling Control Panel/Capacity_Optimizer_Control.xlsx"'
    worksheet[f"B{current_row + 1}"].alignment = Alignment(wrap_text=True)

    worksheet[f"A{current_row + 3}"] = "Quick actions"
    worksheet[f"A{current_row + 3}"].font = Font(bold=True, color="1F4E79", size=11)
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 3}",
        bottom_right=f"C{current_row + 4}",
        label="Setup Dependencies",
        target=r"..\runtime\setup_requirements.bat",
    )
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 6}",
        bottom_right=f"C{current_row + 7}",
        label="Get Machine Fingerprint",
        target=r"..\runtime\get_machine_fingerprint.bat",
    )
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 9}",
        bottom_right=f"C{current_row + 10}",
        label="Run Optimizer",
        target=r"..\runtime\run_optimizer.bat",
    )
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 12}",
        bottom_right=f"C{current_row + 13}",
        label="Open Output Folder",
        target=r"..\output",
    )
    worksheet[f"A{current_row + 15}"] = "Button note"
    worksheet[f"A{current_row + 15}"].font = Font(bold=True, color="1F4E79", size=11)
    worksheet[f"B{current_row + 15}"] = (
        "On a new computer, click Setup Dependencies first. If RSCP gave you a trial/unbound license, "
        "place license.json in licenses\\active and continue. Otherwise click Get Machine Fingerprint, "
        "send the generated file from licenses\\requests to RSCP, place the returned license.json in licenses\\active, "
        "save the workbook (Ctrl+S), and click Run Optimizer."
    )
    worksheet[f"B{current_row + 15}"].alignment = Alignment(wrap_text=True)

    _add_list_validation(worksheet, f"B{value_rows['Run_Mode']}", lists_ws, "$A$2:$A$4")
    _add_list_validation(worksheet, f"B{value_rows['Direct_Mode']}", lists_ws, "$B$2:$B$3")
    _add_list_validation(worksheet, f"B{value_rows['Verbose']}", lists_ws, "$B$2:$B$3")
    _add_list_validation(worksheet, f"B{value_rows['Skip_Validation_Errors']}", lists_ws, "$B$2:$B$3")
    _add_list_validation(worksheet, f"B{value_rows['Start_Month_Num']}", lists_ws, "$C$2:$C$13")
    if lists_ws["D2"].value:
        last_scenario_row = max(2, lists_ws.max_row)
        _add_list_validation(
            worksheet,
            f"B{value_rows['Scenario_Name']}",
            lists_ws,
            f"$D$2:$D${last_scenario_row}",
        )


def _create_lists(worksheet, scenario_options: list[str]) -> None:
    worksheet["A1"] = "Run_Mode"
    worksheet["A2"] = "ModeA"
    worksheet["A3"] = "ModeB"
    worksheet["A4"] = "Both"

    worksheet["B1"] = "Yes_No"
    worksheet["B2"] = "Yes"
    worksheet["B3"] = "No"

    worksheet["C1"] = "Month_Number"
    for month in range(1, 13):
        worksheet.cell(month + 1, 3, month)

    worksheet["D1"] = "Scenario_Name"
    for row_num, scenario in enumerate(scenario_options, start=2):
        worksheet.cell(row_num, 4, scenario)

    for column in range(1, 5):
        worksheet.column_dimensions[get_column_letter(column)].width = 22


def _scenario_options(load_dir: str) -> list[str]:
    if os.path.isdir(load_dir):
        try:
            options = discover_planner_scenarios(load_dir)
            if options:
                return options
        except Exception:
            pass
    return ["Base Scenario", "Baseline", "Case90"]


def _write_header(worksheet, row: int, labels: list[str]) -> None:
    for col_num, label in enumerate(labels, start=1):
        cell = worksheet.cell(row, col_num)
        cell.value = label
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    worksheet.row_dimensions[row].height = 20


def _add_action_button(worksheet, top_left: str, bottom_right: str, label: str, target: str) -> None:
    worksheet.merge_cells(f"{top_left}:{bottom_right}")
    cell = worksheet[top_left]
    cell.value = label
    cell.hyperlink = target
    cell.fill = BTN_FILL
    cell.font = BTN_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

    start_col = worksheet[top_left].column
    end_col = worksheet[bottom_right].column
    start_row = worksheet[top_left].row
    end_row = worksheet[bottom_right].row
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = 22
        for col in range(start_col, end_col + 1):
            worksheet.cell(row, col).border = BORDER


def _add_list_validation(worksheet, target_ref: str, lists_ws, source_range: str) -> None:
    validation = DataValidation(
        type="list",
        formula1=f"=Lists!{source_range}",
        allow_blank=True,
    )
    worksheet.add_data_validation(validation)
    validation.add(worksheet[target_ref])


if __name__ == "__main__":
    main()
