"""
Write optimisation results to an Excel workbook.

The workbook includes both raw result sheets and presentation-ready analysis
sheets so the Excel output can replace the former Streamlit views.
"""
from __future__ import annotations

import math
import os
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.load_pressure import (
    build_capacity_compare_heatmap_frames,
    build_dashboard_fact_frame,
    build_pressure_load_frame,
    build_pressure_tons_frame,
    build_raw_capacity_map,
    compute_display_capacity_share_pct,
    _summarize_heatmap_months_to_years,
)
from app.models import (
    AllocationResult,
    CapacityRecord,
    Config,
    LoadRecord,
    RoutingRecord,
    ValidationIssue,
)
from app.result_analysis import (
    build_result_analysis,
)


HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
MODEA_FILL = PatternFill("solid", fgColor="2F75B5")
MODEB_FILL = PatternFill("solid", fgColor="ED7D31")
DELTA_FILL = PatternFill("solid", fgColor="D9E2F3")
SUMMARY_FILL = PatternFill("solid", fgColor="EAF2F8")
ALT_ROW_FILL = PatternFill("solid", fgColor="E2F0D9")
TITLE_FONT = Font(color="1F1F1F", bold=True, size=14)
NOTE_FONT = Font(color="666666", italic=True, size=9)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ERR_FILL = PatternFill("solid", fgColor="FFCCCC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
RISK_HDR_FILL = PatternFill("solid", fgColor="C00000")
CARD_BLUE_FILL = PatternFill("solid", fgColor="DCE6F1")
CARD_GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
CARD_ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
CARD_RED_FILL = PatternFill("solid", fgColor="FDE9D9")
CARD_TEAL_FILL = PatternFill("solid", fgColor="D9EAF2")
CARD_GREY_FILL = PatternFill("solid", fgColor="E7E6E6")
CARD_NEUTRAL_FILL = PatternFill("solid", fgColor="F7F9FC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT = "0.0%"
TONS_FMT = "#,##0.0"
INT_FMT = "#,##0"
FILTER_SLOTS = 8
DASHBOARD_LAST_COL = 20


def _metric_card_palette(metric_name: str) -> tuple[PatternFill, PatternFill]:
    mapping = {
        "Total demand": (HDR_FILL, CARD_BLUE_FILL),
        "Internal allocated": (PatternFill("solid", fgColor="548235"), CARD_GREEN_FILL),
        "Outsourced": (PatternFill("solid", fgColor="C55A11"), CARD_ORANGE_FILL),
        "Residual unmet": (PatternFill("solid", fgColor="C00000"), CARD_RED_FILL),
        "Service level": (PatternFill("solid", fgColor="0F6A7A"), CARD_TEAL_FILL),
        "Selected workcenters": (PatternFill("solid", fgColor="5B6573"), CARD_GREY_FILL),
    }
    return mapping.get(metric_name, (HDR_FILL, CARD_NEUTRAL_FILL))


def build_output_path(config: Config, mode: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(config.output_file_name)[0]
    parts = [base, mode]
    scenario_segment = _sanitize_filename_segment(config.scenario_name)
    if scenario_segment:
        parts.append(scenario_segment)
    parts.append(ts)
    filename = "_".join(parts) + ".xlsx"
    return os.path.join(config.output_folder, filename)


def _build_planner_demand_map(loads: Optional[List[LoadRecord]]) -> dict[tuple[str, str], list[tuple[str, float]]]:
    planner_demand: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    if not loads:
        return {}

    for load in loads:
        tons = max(float(load.forecast_tons or 0.0), 0.0)
        if tons <= 0:
            continue
        key = (str(load.month), str(load.product))
        planner = str(load.planner_name or "").strip()
        planner_demand[key][planner] += tons

    return {
        key: [(planner, totals[planner]) for planner in sorted(totals, key=str.casefold)]
        for key, totals in planner_demand.items()
    }


def _split_value_by_planner(total: float, planner_weights: list[tuple[str, float]]) -> dict[str, float]:
    if not planner_weights:
        return {}

    total_units = int(round(max(float(total or 0.0), 0.0) * 10000))
    active_weights = [(planner, weight) for planner, weight in planner_weights if weight > 0]
    if total_units <= 0 or not active_weights:
        return {planner: 0.0 for planner, _weight in planner_weights}

    total_weight = sum(weight for _planner, weight in active_weights)
    allocations: list[list[Any]] = []
    assigned_units = 0
    for planner, weight in planner_weights:
        exact_units = total_units * weight / total_weight if weight > 0 else 0.0
        base_units = math.floor(exact_units)
        allocations.append([planner, base_units, exact_units - base_units])
        assigned_units += base_units

    residual_units = total_units - assigned_units
    allocations.sort(key=lambda item: (-item[2], str(item[0]).casefold()))
    for idx in range(residual_units):
        allocations[idx % len(allocations)][1] += 1

    allocations.sort(key=lambda item: str(item[0]).casefold())
    return {planner: units / 10000.0 for planner, units, _fraction in allocations}


def _plannerize_results(
    results: List[AllocationResult],
    loads: Optional[List[LoadRecord]],
) -> List[AllocationResult]:
    planner_demand = _build_planner_demand_map(loads)
    if not planner_demand:
        return [
            AllocationResult(
                month=result.month,
                product=result.product,
                product_family=result.product_family,
                plant=result.plant,
                allocation_type=result.allocation_type,
                work_center=result.work_center,
                route_type=result.route_type,
                priority=result.priority,
                demand_tons=result.demand_tons,
                allocated_tons=result.allocated_tons,
                outsourced_tons=result.outsourced_tons,
                unmet_tons=result.unmet_tons,
                capacity_share_pct=result.capacity_share_pct,
                planner_name=result.planner_name,
            )
            for result in results
        ]

    plannerized: list[AllocationResult] = []
    for result in results:
        key = (result.month, result.product)
        planner_weights = planner_demand.get(key)
        if not planner_weights:
            plannerized.append(
                AllocationResult(
                    month=result.month,
                    product=result.product,
                    product_family=result.product_family,
                    plant=result.plant,
                    allocation_type=result.allocation_type,
                    work_center=result.work_center,
                    route_type=result.route_type,
                    priority=result.priority,
                    demand_tons=result.demand_tons,
                    allocated_tons=result.allocated_tons,
                    outsourced_tons=result.outsourced_tons,
                    unmet_tons=result.unmet_tons,
                    capacity_share_pct=result.capacity_share_pct,
                    planner_name=result.planner_name,
                )
            )
            continue

        demand_split = _split_value_by_planner(result.demand_tons, planner_weights)
        allocated_split = _split_value_by_planner(result.allocated_tons, planner_weights)
        outsourced_split = _split_value_by_planner(result.outsourced_tons, planner_weights)
        unmet_split = _split_value_by_planner(result.unmet_tons, planner_weights)
        cap_share_split = _split_value_by_planner(result.capacity_share_pct, planner_weights)

        for planner_name, _weight in planner_weights:
            planner_allocated = round(allocated_split.get(planner_name, 0.0), 4)
            planner_outsourced = round(outsourced_split.get(planner_name, 0.0), 4)
            planner_unmet = round(unmet_split.get(planner_name, 0.0), 4)

            if result.allocation_type == "Internal" and planner_allocated <= 0.0:
                continue
            if result.allocation_type == "Outsourced" and planner_outsourced <= 0.0:
                continue
            if result.allocation_type == "Unmet" and planner_unmet <= 0.0:
                continue

            plannerized.append(
                AllocationResult(
                    month=result.month,
                    product=result.product,
                    product_family=result.product_family,
                    plant=result.plant,
                    allocation_type=result.allocation_type,
                    work_center=result.work_center,
                    route_type=result.route_type,
                    priority=result.priority,
                    demand_tons=round(demand_split.get(planner_name, 0.0), 4),
                    allocated_tons=planner_allocated,
                    outsourced_tons=planner_outsourced,
                    unmet_tons=planner_unmet,
                    capacity_share_pct=round(cap_share_split.get(planner_name, 0.0), 4),
                    planner_name=planner_name,
                )
            )

    return plannerized


def write_results(
    results: List[AllocationResult],
    loads: Optional[List[LoadRecord]],
    capacities: Optional[List[CapacityRecord]],
    routings: Optional[List[RoutingRecord]],
    config: Config,
    issues: List[ValidationIssue],
    months: List[str],
    mode: str = "ModeA",
    toller_products: Optional[set] = None,
    metrics_by_mode: Optional[dict[str, dict[str, Any]]] = None,
    dashboard_facts_by_mode: Optional[dict[str, pd.DataFrame]] = None,
) -> str:
    """
    Write a complete Excel report workbook.

    The workbook contains:
    - report sheets replacing the previous Streamlit views
    - raw output sheets for audit/detail
    """
    os.makedirs(config.output_folder, exist_ok=True)
    out_path = build_output_path(config, mode)

    wb = Workbook()
    wb.remove(wb.active)
    _enable_formula_recalc(wb)

    artifact = _build_mode_artifact(
        results,
        loads,
        capacities,
        routings,
        config,
        months,
        mode,
        dashboard_fact=(dashboard_facts_by_mode or {}).get(mode),
    )
    df_detail = artifact["df_detail"]
    wc_load_df = artifact["wc_load_df"]
    run_info_df = artifact["run_info_df"]
    analysis = artifact["analysis"]
    preview_metrics = artifact["metrics"]

    all_metrics = dict(metrics_by_mode or {})
    all_metrics.setdefault(mode, preview_metrics)

    _write_dashboard(
        wb,
        mode,
        analysis,
        preview_metrics,
        all_metrics,
        dashboard_facts_by_mode or {mode: artifact["dashboard_fact"]},
        issues=issues,
    )
    _write_monthly_analysis(wb, analysis)
    _write_bottleneck_analysis(wb, analysis)
    _write_wc_heatmap(wb, analysis, artifact["wc_tons_df"], months)
    _write_product_risk_analysis(wb, analysis)

    _write_detail(wb, df_detail)
    _write_planner_summary(wb, analysis)
    _write_run_info(wb, run_info_df)

    wb.save(out_path)
    return out_path


def write_mode_comparison_summary(
    mode_results: dict[str, List[AllocationResult]],
    config: Config,
    months: List[str],
    metrics_by_mode: Optional[dict[str, dict[str, Any]]] = None,
    mode_loads: Optional[dict[str, List[LoadRecord]]] = None,
    mode_capacities: Optional[dict[str, List[CapacityRecord]]] = None,
    mode_routings: Optional[dict[str, List[RoutingRecord]]] = None,
    dashboard_facts_by_mode: Optional[dict[str, pd.DataFrame]] = None,
    capacity_basis_payloads_by_mode: Optional[dict[str, dict[str, Any]]] = None,
) -> str:
    """
    Write a standalone comparison workbook when both ModeA and ModeB are run.
    """
    required_modes = {"ModeA", "ModeB"}
    if not required_modes.issubset(mode_results):
        raise ValueError("Comparison workbook requires both ModeA and ModeB results.")

    os.makedirs(config.output_folder, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(config.output_folder, f"Summary of Mode A and Mode B_{ts}.xlsx")

    artifacts = {
        mode: _build_mode_artifact(
            mode_results[mode],
            mode_loads.get(mode) if mode_loads else None,
            mode_capacities.get(mode) if mode_capacities else None,
            mode_routings.get(mode) if mode_routings else None,
            config,
            months,
            mode,
            dashboard_fact=(dashboard_facts_by_mode or {}).get(mode),
        )
        for mode in ("ModeA", "ModeB")
    }
    comparison_metrics = dict(metrics_by_mode or {})
    for mode in ("ModeA", "ModeB"):
        comparison_metrics.setdefault(mode, artifacts[mode]["metrics"])

    wb = Workbook()
    wb.remove(wb.active)
    _enable_formula_recalc(wb)

    _write_executive_comparison(wb, artifacts, comparison_metrics)
    _write_monthly_comparison(wb, artifacts)
    _write_bottleneck_comparison(wb, artifacts)
    _write_heatmap_comparison(wb, artifacts)
    _write_product_risk_comparison(wb, artifacts)
    if capacity_basis_payloads_by_mode:
        _write_summary_capacity_basis_pages(
            wb,
            capacity_basis_payloads_by_mode=capacity_basis_payloads_by_mode,
            config=config,
            months=months,
        )
    _write_comparison_run_info(wb, config, comparison_metrics, os.path.basename(out_path))

    wb.save(out_path)
    return out_path


def write_capacity_basis_results(
    basis_results: dict[str, List[AllocationResult]],
    loads: Optional[List[LoadRecord]],
    basis_capacities: dict[str, List[CapacityRecord]],
    routings: Optional[List[RoutingRecord]],
    config: Config,
    issues: List[ValidationIssue],
    months: List[str],
    mode: str = "ModeA",
    toller_products_by_basis: Optional[dict[str, set[str]]] = None,
) -> str:
    os.makedirs(config.output_folder, exist_ok=True)
    out_path = build_output_path(config, mode)

    wb = Workbook()
    wb.remove(wb.active)
    _enable_formula_recalc(wb)

    basis_labels = ("Max", "Planner")
    artifacts = {
        basis: _build_mode_artifact(
            basis_results[basis],
            loads,
            basis_capacities[basis],
            routings,
            config,
            months,
            basis,
        )
        for basis in basis_labels
    }
    combined_issues = list(issues)

    _write_capacity_basis_dashboard(wb, mode, artifacts, issues=combined_issues)
    _write_capacity_basis_monthly_analysis(wb, mode, artifacts)
    _write_capacity_basis_bottleneck_analysis(wb, mode, artifacts)
    _write_capacity_basis_heatmap(
        wb,
        mode=mode,
        basis_results=basis_results,
        basis_capacities=basis_capacities,
        loads=loads or [],
        routings=routings or [],
        months=months,
    )
    _write_capacity_basis_product_risk(wb, mode, artifacts)
    _write_capacity_basis_planner_summary(wb, mode, artifacts)

    detail_df = _concat_basis_detail_frames(artifacts)
    run_info_df = _concat_basis_run_info_frames(artifacts)

    _write_detail(wb, detail_df)
    _write_run_info(wb, run_info_df)

    wb.save(out_path)
    return out_path


def _results_to_df(
    results: List[AllocationResult],
    raw_capacity_map: dict[tuple[str, str], float],
) -> pd.DataFrame:
    rows = []
    columns = [
        "Month",
        "PlannerName",
        "Product",
        "ProductFamily",
        "Plant",
        "AllocationType",
        "WorkCenter",
        "RouteType",
        "Priority",
        "Demand_Tons",
        "Allocated_Tons",
        "Outsourced_Tons",
        "Unmet_Tons",
        "CapacityShare_Pct",
    ]
    for result in results:
        rows.append(
            {
                "Month": result.month,
                "PlannerName": result.planner_name,
                "Product": result.product,
                "ProductFamily": result.product_family,
                "Plant": result.plant,
                "AllocationType": result.allocation_type,
                "WorkCenter": result.work_center,
                "RouteType": result.route_type,
                "Priority": result.priority,
                "Demand_Tons": result.demand_tons,
                "Allocated_Tons": result.allocated_tons,
                "Outsourced_Tons": result.outsourced_tons,
                "Unmet_Tons": result.unmet_tons,
                "CapacityShare_Pct": compute_display_capacity_share_pct(
                    product=result.product,
                    work_center=result.work_center,
                    allocated_tons=result.allocated_tons,
                    raw_capacity_map=raw_capacity_map,
                ) / 100.0,
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _build_mode_artifact(
    results: List[AllocationResult],
    loads: Optional[List[LoadRecord]],
    capacities: Optional[List[CapacityRecord]],
    routings: Optional[List[RoutingRecord]],
    config: Config,
    months: List[str],
    mode: str,
    dashboard_fact: Optional[pd.DataFrame] = None,
) -> dict[str, Any]:
    capacities = capacities or []
    routings = routings or []
    raw_capacity_map = build_raw_capacity_map(capacities)
    plannerized_results = _plannerize_results(results, loads)
    df_detail = _results_to_df(plannerized_results, raw_capacity_map)
    wc_load_df = build_pressure_load_frame(
        mode=mode,
        results=results,
        loads=loads or [],
        capacities=capacities,
        routings=routings,
        months=months,
    )
    wc_tons_df = build_pressure_tons_frame(
        mode=mode,
        results=results,
        loads=loads or [],
        capacities=capacities,
        routings=routings,
        months=months,
    )
    run_info_df = _build_run_info_df(config, mode)
    analysis = build_result_analysis(df_detail, wc_load_df, run_info_df)
    metrics = _build_preview_metrics(mode, analysis, results, months)
    dashboard_fact = dashboard_fact if dashboard_fact is not None else build_dashboard_fact_frame(
        mode=mode,
        results=results,
        loads=loads or [],
        capacities=capacities,
        routings=routings,
    )
    return {
        "df_detail": df_detail,
        "wc_load_df": wc_load_df,
        "wc_tons_df": wc_tons_df,
        "run_info_df": run_info_df,
        "analysis": analysis,
        "metrics": metrics,
        "dashboard_fact": dashboard_fact,
    }


def _build_wc_load_frame(df_detail: pd.DataFrame, months: List[str]) -> pd.DataFrame:
    internal_df = df_detail[df_detail["AllocationType"] == "Internal"]
    if internal_df.empty:
        return pd.DataFrame(columns=["WorkCenter", *months])

    pivot = internal_df[internal_df["WorkCenter"] != "[UNALLOCATED]"].pivot_table(
        index="WorkCenter",
        columns="Month",
        values="CapacityShare_Pct",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    return pivot


def _build_run_info_df(config: Config, mode: str) -> pd.DataFrame:
    rows = [
        ("Scenario_Name", config.scenario_name),
        ("Mode", mode),
        ("Start_Month", config.start_month),
        ("Horizon_Months", config.horizon_months),
        ("Input_Load_Folder", config.input_load_folder),
        ("Input_Master_Folder", config.input_master_folder),
        ("Output_Folder", config.output_folder),
        ("Project_Root_Folder", getattr(config, "project_root_folder", "")),
        ("Output_FileName", config.output_file_name),
        ("Run_Timestamp", config.run_timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Run_Mode", getattr(config, "run_mode", mode)),
        ("Direct_Mode", "Yes" if getattr(config, "direct_mode", True) else "No"),
        ("Verbose", "Yes" if getattr(config, "verbose", False) else "No"),
        (
            "Skip_Validation_Errors",
            "Yes" if getattr(config, "skip_validation_errors", False) else "No",
        ),
        ("License_Status", getattr(config, "license_status", "")),
        ("License_ID", getattr(config, "license_id", "")),
        ("License_Type", getattr(config, "license_type", "")),
        ("Licensed_To", getattr(config, "licensed_to", "")),
        ("License_Expiry", getattr(config, "license_expiry", "")),
        ("License_Binding_Mode", getattr(config, "license_binding_mode", "")),
        ("License_Machine_Label", getattr(config, "license_machine_label", "")),
        ("Notes", config.notes or ""),
        ("Tool_Version", "1.1.3"),
    ]
    return pd.DataFrame(rows, columns=["Parameter", "Value"])


def _build_preview_metrics(
    mode: str,
    analysis: dict[str, Any],
    results: List[AllocationResult],
    months: List[str],
) -> dict[str, Any]:
    monthly_summary = analysis["monthly_summary"]
    total_demand = float(monthly_summary["Demand_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_internal_allocated = float(monthly_summary["Internal_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_outsourced = float(monthly_summary["Outsourced_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_unmet = float(monthly_summary["Unmet_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_supplied = total_internal_allocated + total_outsourced
    service_level = 100.0 * total_supplied / total_demand if total_demand > 0 else 0.0
    return {
        "mode": mode,
        "scenario_name": analysis.get("scenario_name") or "",
        "total_demand": total_demand,
        "total_internal_allocated": total_internal_allocated,
        "total_outsourced": total_outsourced,
        "total_unmet": total_unmet,
        "service_level": service_level,
        "result_rows": len(results),
        "months": len(months),
    }


def _enable_formula_recalc(wb: Workbook) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def _combine_dashboard_facts(dashboard_facts_by_mode: dict[str, pd.DataFrame]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for mode_name, fact_df in (dashboard_facts_by_mode or {}).items():
        if fact_df is None or fact_df.empty:
            continue
        frame = fact_df.copy()
        if "Mode" not in frame.columns:
            frame.insert(0, "Mode", mode_name)
        frames.append(frame)

    if not frames:
        return pd.DataFrame(
            columns=[
                "Mode",
                "Year",
                "WorkCenter",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
            ]
        )

    combined = pd.concat(frames, ignore_index=True)
    for column in ("Demand_Tons", "Internal_Tons", "Outsourced_Tons", "Unmet_Tons", "Supplied_Tons"):
        combined[column] = pd.to_numeric(combined[column], errors="coerce").fillna(0.0)
    combined["Mode"] = combined["Mode"].astype(str)
    if "Year" not in combined.columns:
        combined["Year"] = "All"
    combined["Year"] = combined["Year"].astype(str)
    combined["WorkCenter"] = combined["WorkCenter"].astype(str)
    combined = (
        combined.groupby(["Mode", "Year", "WorkCenter"], as_index=False)[
            ["Demand_Tons", "Internal_Tons", "Outsourced_Tons", "Unmet_Tons", "Supplied_Tons"]
        ]
        .sum()
        .sort_values(
            ["Mode", "Year", "WorkCenter"],
            key=lambda col: col.map(str.casefold) if col.name in {"Mode", "Year", "WorkCenter"} else col,
        )
    )
    return combined


def _get_or_create_dashboard_helper_sheet(
    wb: Workbook,
    sheet_name: str = "_Dashboard_Helper",
):
    if sheet_name in wb.sheetnames:
        helper_ws = wb[sheet_name]
    else:
        helper_ws = wb.create_sheet(sheet_name)
    helper_ws.sheet_state = "hidden"
    helper_ws.sheet_view.showGridLines = False
    return helper_ws


def _next_helper_start_row(ws) -> int:
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value in (None, ""):
        return 1
    return ws.max_row + 3


def _write_dashboard_helper_table(
    wb: Workbook,
    df: pd.DataFrame,
    num_formats: Optional[dict[str, str]] = None,
    helper_sheet_name: str = "_Dashboard_Helper",
) -> dict[str, Any]:
    ws = _get_or_create_dashboard_helper_sheet(wb, helper_sheet_name)
    start_row = _next_helper_start_row(ws)
    layout = _write_table(
        ws,
        df,
        start_row=start_row,
        start_col=1,
        num_formats=num_formats or {},
    )
    return {
        "sheet_name": ws.title,
        "start_row": layout["start_row"],
        "end_row": layout["end_row"],
        "start_col": layout["start_col"],
        "end_col": layout["end_col"],
        "col_index": layout["col_index"],
    }


def _write_dashboard_fact_sheet(
    wb: Workbook,
    dashboard_facts_by_mode: dict[str, pd.DataFrame],
    helper_sheet_name: str = "_Dashboard_Helper",
) -> dict[str, Any]:
    ws = _get_or_create_dashboard_helper_sheet(wb, helper_sheet_name)
    start_row = _next_helper_start_row(ws)
    start_col = 1
    fact_df = _combine_dashboard_facts(dashboard_facts_by_mode)

    if fact_df.empty:
        fact_df = pd.DataFrame(
            [
                {
                    "Mode": "",
                    "Year": "",
                    "WorkCenter": "",
                    "Demand_Tons": 0.0,
                    "Internal_Tons": 0.0,
                    "Outsourced_Tons": 0.0,
                    "Unmet_Tons": 0.0,
                    "Supplied_Tons": 0.0,
                }
            ]
        )

    layout = _write_table(
        ws,
        fact_df,
        start_row=start_row,
        start_col=start_col,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
        },
    )

    workcenters = sorted(
        {
            str(value).strip()
            for value in fact_df["WorkCenter"].tolist()
            if str(value).strip()
        },
        key=str.casefold,
    )
    list_col = layout["end_col"] + 2
    year_list_col = layout["end_col"] + 3
    ws.cell(start_row, list_col).value = "WorkCenter_List"
    ws.cell(start_row, year_list_col).value = "Year_List"
    if not workcenters:
        ws.cell(start_row + 1, list_col).value = ""
        list_end_row = start_row + 1
    else:
        for offset, work_center in enumerate(workcenters, start=start_row + 1):
            ws.cell(offset, list_col).value = work_center
        list_end_row = start_row + len(workcenters)

    years = sorted(
        {
            str(value).strip()
            for value in fact_df.get("Year", pd.Series(dtype=str)).tolist()
            if str(value).strip()
        },
        key=str.casefold,
    )
    ws.cell(start_row + 1, year_list_col).value = "All"
    for offset, year in enumerate(years, start=start_row + 2):
        ws.cell(offset, year_list_col).value = year
    year_list_end_row = max(start_row + 1, start_row + len(years) + 1)
    return {
        "sheet_name": ws.title,
        "data_start_row": layout["start_row"] + 1,
        "data_end_row": layout["end_row"],
        "col_index": layout["col_index"],
        "list_col": list_col,
        "list_start_row": start_row + 1,
        "list_end_row": list_end_row,
        "year_list_col": year_list_col,
        "year_list_start_row": start_row + 1,
        "year_list_end_row": year_list_end_row,
        "workcenters": workcenters,
        "years": ["All", *years],
    }


def _excel_sheet_ref(sheet_name: str) -> str:
    return f"'{sheet_name}'"


def _sheet_range_ref(meta: dict[str, Any], column_name: str) -> str:
    col_num = meta["col_index"][column_name]
    col_letter = get_column_letter(col_num)
    return (
        f"{_excel_sheet_ref(meta['sheet_name'])}!${col_letter}${meta['data_start_row']}:${col_letter}${meta['data_end_row']}"
    )


def _dashboard_filtered_sum_formula(
    meta: dict[str, Any],
    mode_name: str,
    metric_name: str,
    selection_mode_ref: str,
    selected_range_ref: str,
    selected_year_ref: str | None = None,
) -> str:
    value_range = _sheet_range_ref(meta, metric_name)
    mode_range = _sheet_range_ref(meta, "Mode")
    workcenter_range = _sheet_range_ref(meta, "WorkCenter")
    year_range = _sheet_range_ref(meta, "Year") if "Year" in meta["col_index"] else None
    if year_range and selected_year_ref:
        all_formula = (
            f'IF({selected_year_ref}="All",'
            f'SUMIFS({value_range},{mode_range},"{mode_name}"),'
            f'SUMIFS({value_range},{mode_range},"{mode_name}",{year_range},{selected_year_ref}))'
        )
        filtered_formula = (
            f'IF({selected_year_ref}="All",'
            f'SUMPRODUCT(({mode_range}="{mode_name}")*({value_range})*(COUNTIF({selected_range_ref},{workcenter_range})>0)),'
            f'SUMPRODUCT(({mode_range}="{mode_name}")*({year_range}={selected_year_ref})*({value_range})*'
            f'(COUNTIF({selected_range_ref},{workcenter_range})>0)))'
        )
    else:
        all_formula = f'SUMIFS({value_range},{mode_range},"{mode_name}")'
        filtered_formula = (
            f'SUMPRODUCT(({mode_range}="{mode_name}")*({value_range})*'
            f'(COUNTIF({selected_range_ref},{workcenter_range})>0))'
        )
    return f'=IF({selection_mode_ref}="All",{all_formula},{filtered_formula})'


def _dashboard_selected_workcenter_count_formula(
    meta: dict[str, Any],
    selection_mode_ref: str,
    selected_range_ref: str,
) -> str:
    list_col_letter = get_column_letter(meta["list_col"])
    list_range = (
        f"{_excel_sheet_ref(meta['sheet_name'])}!${list_col_letter}${meta['list_start_row']}:${list_col_letter}${meta['list_end_row']}"
    )
    return f'=IF({selection_mode_ref}="All",COUNTA({list_range}),COUNTIF({selected_range_ref},"<>"))'


def _add_dashboard_filter_controls(
    ws,
    meta: dict[str, Any],
    start_row: int,
    start_col: int,
    title: str = "WorkCenter Filter",
) -> dict[str, Any]:
    title_start = ws.cell(start_row, start_col).coordinate
    title_end = ws.cell(start_row, start_col + 3).coordinate
    ws.merge_cells(f"{title_start}:{title_end}")
    title_cell = ws.cell(start_row, start_col)
    title_cell.value = title
    title_cell.fill = HDR_FILL
    title_cell.font = Font(color="FFFFFF", bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = BORDER
    ws.row_dimensions[start_row].height = 24

    # Keep the filter panel compact and readable regardless of dashboard autofit.
    ws.column_dimensions[get_column_letter(start_col)].width = 16
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 20
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 12
    ws.column_dimensions[get_column_letter(start_col + 3)].width = 12

    selection_mode_cell = ws.cell(start_row + 1, start_col + 1)
    ws.cell(start_row + 1, start_col).value = "Selection Mode"
    ws.cell(start_row + 1, start_col).fill = SUBHDR_FILL
    ws.cell(start_row + 1, start_col).font = Font(bold=True, size=11)
    ws.cell(start_row + 1, start_col).border = BORDER
    ws.cell(start_row + 1, start_col).alignment = Alignment(horizontal="center", vertical="center")
    selection_mode_cell.value = "All"
    selection_mode_cell.border = BORDER
    selection_mode_cell.font = Font(size=11)
    selection_mode_cell.fill = CARD_NEUTRAL_FILL
    selection_mode_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row + 1].height = 24

    selection_dv = DataValidation(type="list", formula1='"All,Filtered"', allow_blank=False)
    selection_dv.error = "Selection Mode must be All or Filtered."
    selection_dv.prompt = "Choose All to show the full dashboard, or Filtered to limit it to selected workcenters."
    ws.add_data_validation(selection_dv)
    selection_dv.add(selection_mode_cell)

    year_label_cell = ws.cell(start_row + 1, start_col + 2)
    year_value_cell = ws.cell(start_row + 1, start_col + 3)
    year_label_cell.value = "Year"
    year_label_cell.fill = SUBHDR_FILL
    year_label_cell.font = Font(bold=True, size=11)
    year_label_cell.border = BORDER
    year_label_cell.alignment = Alignment(horizontal="center", vertical="center")
    year_value_cell.value = "All"
    year_value_cell.border = BORDER
    year_value_cell.font = Font(size=11)
    year_value_cell.fill = CARD_NEUTRAL_FILL
    year_value_cell.alignment = Alignment(horizontal="center", vertical="center")

    year_col_letter = get_column_letter(meta["year_list_col"])
    year_formula = (
        f"={_excel_sheet_ref(meta['sheet_name'])}!${year_col_letter}${meta['year_list_start_row']}:"
        f"${year_col_letter}${meta['year_list_end_row']}"
    )
    year_dv = DataValidation(type="list", formula1=year_formula, allow_blank=False)
    year_dv.error = "Select All or a year from the dropdown list."
    year_dv.prompt = "Choose All for the full horizon, or a specific year."
    ws.add_data_validation(year_dv)
    year_dv.add(year_value_cell)

    list_col_letter = get_column_letter(meta["list_col"])
    list_formula = (
        f"={_excel_sheet_ref(meta['sheet_name'])}!${list_col_letter}${meta['list_start_row']}:${list_col_letter}${meta['list_end_row']}"
    )
    workcenter_dv = DataValidation(type="list", formula1=list_formula, allow_blank=True)
    workcenter_dv.error = "Select a workcenter from the dropdown list."
    workcenter_dv.prompt = "Pick one or more workcenters to filter the dashboard."
    ws.add_data_validation(workcenter_dv)

    selected_cells: list[str] = []
    for offset in range(FILTER_SLOTS):
        row_num = start_row + 2 + offset
        label_cell = ws.cell(row_num, start_col)
        value_cell = ws.cell(row_num, start_col + 1)
        label_cell.value = f"WorkCenter {offset + 1}"
        label_cell.fill = SUBHDR_FILL
        label_cell.font = Font(bold=True, size=11)
        label_cell.border = BORDER
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = BORDER
        value_cell.font = Font(size=11)
        value_cell.fill = CARD_NEUTRAL_FILL
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        workcenter_dv.add(value_cell)
        selected_cells.append(value_cell.coordinate)
        ws.row_dimensions[row_num].height = 22

    note_row = start_row + 2 + FILTER_SLOTS
    ws.merge_cells(
        f"{ws.cell(note_row, start_col).coordinate}:{ws.cell(note_row, start_col + 3).coordinate}"
    )
    ws.cell(note_row, start_col).value = (
        "Use All for the full dashboard, or switch to Filtered and choose one or more workcenters. "
        "Year can stay at All or be narrowed to a single year."
    )
    ws.cell(note_row, start_col).font = NOTE_FONT
    ws.cell(note_row, start_col).fill = CARD_NEUTRAL_FILL
    ws.cell(note_row, start_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 36

    selection_col_letter = get_column_letter(start_col + 1)
    year_col_letter = get_column_letter(start_col + 3)
    return {
        "selection_mode_cell": selection_mode_cell.coordinate,
        "selection_mode_ref": f"${selection_col_letter}${start_row + 1}",
        "selected_year_cell": year_value_cell.coordinate,
        "selected_year_ref": f"${year_col_letter}${start_row + 1}",
        "selected_cells": selected_cells,
        "selected_range": f"${selection_col_letter}${start_row + 2}:${selection_col_letter}${start_row + 1 + FILTER_SLOTS}",
    }


def _write_dashboard(
    wb: Workbook,
    mode: str,
    analysis: dict[str, Any],
    preview_metrics: dict[str, Any],
    metrics_by_mode: dict[str, dict[str, Any]],
    dashboard_facts_by_mode: dict[str, pd.DataFrame],
    issues: Optional[List[ValidationIssue]] = None,
) -> None:
    ws = wb.create_sheet("Dashboard")
    scenario = analysis.get("scenario_name") or preview_metrics.get("scenario_name") or "N/A"
    subtitle = (
        f"Scenario: {scenario} | Mode: {mode} | "
        f"Horizon months: {preview_metrics.get('months', 0)}"
    )
    _prepare_dashboard_sheet(ws, f"Executive Summary - {mode}", subtitle)
    fact_meta = _write_dashboard_fact_sheet(wb, dashboard_facts_by_mode)
    filter_refs = _add_dashboard_filter_controls(ws, fact_meta, start_row=3, start_col=15)
    selection_mode_ref = filter_refs["selection_mode_ref"]
    selected_range_ref = filter_refs["selected_range"]
    selected_year_ref = filter_refs["selected_year_ref"]

    demand_cell = _write_single_kpi_card(
        ws,
        top_row=4,
        left_col=1,
        title="Total demand",
        value=_dashboard_filtered_sum_formula(
            fact_meta,
            mode,
            "Demand_Tons",
            selection_mode_ref,
            selected_range_ref,
            selected_year_ref,
        ),
        number_format=TONS_FMT,
    )
    internal_cell = _write_single_kpi_card(
        ws,
        top_row=4,
        left_col=8,
        title="Internal allocated",
        value=_dashboard_filtered_sum_formula(
            fact_meta,
            mode,
            "Internal_Tons",
            selection_mode_ref,
            selected_range_ref,
            selected_year_ref,
        ),
        number_format=TONS_FMT,
    )
    outsource_cell = _write_single_kpi_card(
        ws,
        top_row=8,
        left_col=1,
        title="Outsourced",
        value=_dashboard_filtered_sum_formula(
            fact_meta,
            mode,
            "Outsourced_Tons",
            selection_mode_ref,
            selected_range_ref,
            selected_year_ref,
        ),
        number_format=TONS_FMT,
    )
    unmet_cell = _write_single_kpi_card(
        ws,
        top_row=8,
        left_col=8,
        title="Residual unmet",
        value=_dashboard_filtered_sum_formula(
            fact_meta,
            mode,
            "Unmet_Tons",
            selection_mode_ref,
            selected_range_ref,
            selected_year_ref,
        ),
        number_format=TONS_FMT,
    )
    service_cell = _write_single_kpi_card(
        ws,
        top_row=12,
        left_col=1,
        title="Service level",
        value=f"=IF({demand_cell}=0,0,({internal_cell}+{outsource_cell})/{demand_cell})",
        number_format=PCT_FMT,
    )
    selected_wc_cell = _write_single_kpi_card(
        ws,
        top_row=12,
        left_col=8,
        title="Selected workcenters",
        value=_dashboard_selected_workcenter_count_formula(
            fact_meta,
            selection_mode_ref,
            selected_range_ref,
        ),
        number_format=INT_FMT,
    )

    helper_mix_df = pd.DataFrame(
        {
            "Category": ["Internal", "Outsourced", "Unmet"],
            "Tons": [
                f"='Dashboard'!{internal_cell}",
                f"='Dashboard'!{outsource_cell}",
                f"='Dashboard'!{unmet_cell}",
            ],
        }
    )
    mix_layout = _write_dashboard_helper_table(
        wb,
        helper_mix_df,
        num_formats={"Tons": TONS_FMT},
    )
    helper_ws = wb[mix_layout["sheet_name"]]

    ws.merge_cells("A16:H16")
    ws["A16"] = "Supply Mix"
    ws["A16"].fill = SUMMARY_FILL
    ws["A16"].font = Font(bold=True, color="1F4E79", size=15)
    ws["A16"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 24
    mix_chart = BarChart()
    mix_chart.title = "Supply Mix"
    mix_chart.height = 8.5
    mix_chart.width = 18.5
    mix_chart.add_data(
        Reference(
            helper_ws,
            min_col=mix_layout["col_index"]["Tons"],
            min_row=mix_layout["start_row"],
            max_row=mix_layout["end_row"],
        ),
        titles_from_data=True,
    )
    mix_chart.set_categories(
        Reference(
            helper_ws,
            min_col=mix_layout["col_index"]["Category"],
            min_row=mix_layout["start_row"] + 1,
            max_row=mix_layout["end_row"],
        )
    )
    _style_dashboard_mix_chart(mix_chart, ["548235"])
    ws.add_chart(mix_chart, "A17")

    _write_dashboard_validation_block(ws, issues or [], start_row=40)
    _autofit(ws)
    _set_dashboard_column_layout(ws)


def _write_executive_comparison(
    wb: Workbook,
    artifacts: dict[str, dict[str, Any]],
    metrics_by_mode: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Executive_Comparison")
    fact_meta = _write_dashboard_fact_sheet(
        wb,
        {mode: artifacts[mode]["dashboard_fact"] for mode in ("ModeA", "ModeB")},
    )
    mode_a = metrics_by_mode["ModeA"]
    mode_b = metrics_by_mode["ModeB"]
    scenario = mode_a.get("scenario_name") or mode_b.get("scenario_name") or "N/A"
    subtitle = f"Scenario: {scenario} | Compare ModeA and ModeB with the same WorkCenter and Year filters"
    _prepare_dashboard_sheet(ws, "Summary of Mode A and Mode B", subtitle)
    filter_refs = _add_dashboard_filter_controls(ws, fact_meta, start_row=3, start_col=15)
    selection_mode_ref = filter_refs["selection_mode_ref"]
    selected_range_ref = filter_refs["selected_range"]
    selected_year_ref = filter_refs["selected_year_ref"]

    demand_cells = _write_compare_kpi_card(
        ws,
        top_row=4,
        left_col=1,
        title="Total demand",
        left_label="ModeA",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeA", "Demand_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="ModeB",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeB", "Demand_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=C6-A6",
        number_format=TONS_FMT,
    )
    internal_cells = _write_compare_kpi_card(
        ws,
        top_row=4,
        left_col=8,
        title="Internal allocated",
        left_label="ModeA",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeA", "Internal_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="ModeB",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeB", "Internal_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=J6-H6",
        number_format=TONS_FMT,
    )
    outsourced_cells = _write_compare_kpi_card(
        ws,
        top_row=8,
        left_col=1,
        title="Outsourced",
        left_label="ModeA",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeA", "Outsourced_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="ModeB",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeB", "Outsourced_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=C10-A10",
        number_format=TONS_FMT,
    )
    unmet_cells = _write_compare_kpi_card(
        ws,
        top_row=8,
        left_col=8,
        title="Residual unmet",
        left_label="ModeA",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeA", "Unmet_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="ModeB",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "ModeB", "Unmet_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=J10-H10",
        number_format=TONS_FMT,
    )
    service_cells = _write_compare_kpi_card(
        ws,
        top_row=12,
        left_col=1,
        title="Service level",
        left_label="ModeA",
        left_value=f"=IF({demand_cells['ModeA']}=0,0,({internal_cells['ModeA']}+{outsourced_cells['ModeA']})/{demand_cells['ModeA']})",
        right_label="ModeB",
        right_value=f"=IF({demand_cells['ModeB']}=0,0,({internal_cells['ModeB']}+{outsourced_cells['ModeB']})/{demand_cells['ModeB']})",
        delta_label="Delta",
        delta_value="=J14-H14",
        number_format=PCT_FMT,
    )
    selected_wc_cells = _write_compare_kpi_card(
        ws,
        top_row=12,
        left_col=8,
        title="Selected workcenters",
        left_label="ModeA",
        left_value=_dashboard_selected_workcenter_count_formula(
            fact_meta, selection_mode_ref, selected_range_ref
        ),
        right_label="ModeB",
        right_value=_dashboard_selected_workcenter_count_formula(
            fact_meta, selection_mode_ref, selected_range_ref
        ),
        delta_label="Delta",
        delta_value='=""',
        number_format=INT_FMT,
    )

    mix_df = pd.DataFrame(
        {
            "Category": ["Internal", "Outsourced", "Unmet"],
            "ModeA": [
                f"='Executive_Comparison'!{internal_cells['ModeA']}",
                f"='Executive_Comparison'!{outsourced_cells['ModeA']}",
                f"='Executive_Comparison'!{unmet_cells['ModeA']}",
            ],
            "ModeB": [
                f"='Executive_Comparison'!{internal_cells['ModeB']}",
                f"='Executive_Comparison'!{outsourced_cells['ModeB']}",
                f"='Executive_Comparison'!{unmet_cells['ModeB']}",
            ],
        }
    )
    mix_layout = _write_dashboard_helper_table(
        wb,
        mix_df,
        num_formats={"ModeA": TONS_FMT, "ModeB": TONS_FMT},
    )
    helper_ws = wb[mix_layout["sheet_name"]]
    ws.merge_cells("A16:H16")
    ws["A16"] = "Supply Mix Comparison"
    ws["A16"].fill = SUMMARY_FILL
    ws["A16"].font = Font(bold=True, color="1F4E79", size=15)
    ws["A16"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 24
    chart = BarChart()
    chart.title = "Supply Mix Comparison"
    chart.height = 8.5
    chart.width = 12.5
    chart.add_data(
        Reference(helper_ws, min_col=mix_layout["col_index"]["ModeA"], min_row=mix_layout["start_row"], max_col=mix_layout["col_index"]["ModeB"], max_row=mix_layout["end_row"]),
        titles_from_data=True,
        from_rows=False,
    )
    chart.set_categories(
        Reference(helper_ws, min_col=mix_layout["col_index"]["Category"], min_row=mix_layout["start_row"] + 1, max_row=mix_layout["end_row"])
    )
    _style_dashboard_mix_chart(chart, ["2F75B5", "ED7D31"])
    ws.add_chart(chart, "A17")

    service_df = pd.DataFrame(
        {
            "Mode": ["ModeA", "ModeB"],
            "Service_Level": [
                f"='Executive_Comparison'!{service_cells['ModeA']}",
                f"='Executive_Comparison'!{service_cells['ModeB']}",
            ],
        }
    )
    service_layout = _write_dashboard_helper_table(
        wb,
        service_df,
        num_formats={"Service_Level": PCT_FMT},
    )
    service_chart = BarChart()
    service_chart.title = "Service Level Comparison"
    service_chart.height = 8.5
    service_chart.width = 9.5
    service_chart.varyColors = True
    service_chart.add_data(
        Reference(helper_ws, min_col=service_layout["col_index"]["Service_Level"], min_row=service_layout["start_row"], max_row=service_layout["end_row"]),
        titles_from_data=True,
    )
    service_chart.set_categories(
        Reference(helper_ws, min_col=service_layout["col_index"]["Mode"], min_row=service_layout["start_row"] + 1, max_row=service_layout["end_row"])
    )
    _style_dashboard_service_chart(service_chart, ["2F75B5"])
    ws.merge_cells("L16:T16")
    ws["L16"] = "Service Level Comparison"
    ws["L16"].fill = SUMMARY_FILL
    ws["L16"].font = Font(bold=True, color="1F4E79", size=15)
    ws["L16"].alignment = Alignment(horizontal="center", vertical="center")
    ws.add_chart(service_chart, "L17")
    _autofit(ws)
    _set_dashboard_column_layout(ws)


def _write_monthly_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Monthly_Trend_Compare")
    _write_sheet_title(ws, "Monthly Trend Comparison")
    _prepare_monthly_trend_sheet(
        ws,
        "Compare ModeA and ModeB month by month across demand, supply mix, unmet, and service level.",
    )

    monthly_a = artifacts["ModeA"]["analysis"]["monthly_summary"].copy()
    monthly_b = artifacts["ModeB"]["analysis"]["monthly_summary"].copy()
    monthly_compare = monthly_a.merge(monthly_b, on="Month", how="outer", suffixes=("_ModeA", "_ModeB")).fillna(0.0)
    monthly_compare["Service_Level_Delta"] = monthly_compare["Service_Level_ModeB"] - monthly_compare["Service_Level_ModeA"]
    monthly_compare["Unmet_Delta"] = monthly_compare["Unmet_Tons_ModeB"] - monthly_compare["Unmet_Tons_ModeA"]
    monthly_compare = monthly_compare.sort_values("Month")

    ws.merge_cells("A3:L3")
    ws["A3"] = "Monthly comparison detail"
    ws["A3"].fill = SUMMARY_FILL
    ws["A3"].font = Font(bold=True, color="1F4E79", size=12)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    layout = _write_table(
        ws,
        monthly_compare[
            [
                "Month",
                "Demand_Tons_ModeA",
                "Internal_Tons_ModeA",
                "Internal_Tons_ModeB",
                "Outsourced_Tons_ModeA",
                "Outsourced_Tons_ModeB",
                "Unmet_Tons_ModeA",
                "Unmet_Tons_ModeB",
                "Service_Level_ModeA",
                "Service_Level_ModeB",
                "Service_Level_Delta",
                "Unmet_Delta",
            ]
        ],
        start_row=4,
        start_col=1,
        num_formats={
            "Demand_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeB": TONS_FMT,
            "Outsourced_Tons_ModeA": TONS_FMT,
            "Outsourced_Tons_ModeB": TONS_FMT,
            "Unmet_Tons_ModeA": TONS_FMT,
            "Unmet_Tons_ModeB": TONS_FMT,
            "Service_Level_ModeA": PCT_FMT,
            "Service_Level_ModeB": PCT_FMT,
            "Service_Level_Delta": PCT_FMT,
            "Unmet_Delta": TONS_FMT,
        },
        highlight_positive_cols=["Unmet_Delta"],
        freeze="A5",
        alternating_fill=ALT_ROW_FILL,
    )

    _autofit(ws)
    _set_monthly_trend_column_layout(ws)


def _write_bottleneck_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Bottleneck_Compare")
    _write_sheet_title(ws, "Bottleneck Comparison")

    wc_a = artifacts["ModeA"]["analysis"]["wc_summary"].copy()
    wc_b = artifacts["ModeB"]["analysis"]["wc_summary"].copy()
    wc_compare = wc_a.merge(wc_b, on="WorkCenter", how="outer", suffixes=("_ModeA", "_ModeB")).fillna(0.0)
    wc_compare["PeakLoad_Delta"] = wc_compare["PeakLoadPct_ModeB"] - wc_compare["PeakLoadPct_ModeA"]
    wc_compare["SortKey"] = wc_compare[["PeakLoadPct_ModeA", "PeakLoadPct_ModeB"]].max(axis=1)
    wc_compare = wc_compare.sort_values(["SortKey", "PeakLoadPct_ModeB"], ascending=[False, False]).head(15)

    ws.merge_cells("A2:H2")
    ws["A2"] = "Top bottleneck workcenters"
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(bold=True, color="1F4E79", size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    layout = _write_table(
        ws,
        wc_compare[
            [
                "WorkCenter",
                "AvgLoadPct_ModeA",
                "AvgLoadPct_ModeB",
                "PeakLoadPct_ModeA",
                "PeakLoadPct_ModeB",
                "Over95Months_ModeA",
                "Over95Months_ModeB",
                "PeakLoad_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "AvgLoadPct_ModeA": PCT_FMT,
            "AvgLoadPct_ModeB": PCT_FMT,
            "PeakLoadPct_ModeA": PCT_FMT,
            "PeakLoadPct_ModeB": PCT_FMT,
            "Over95Months_ModeA": INT_FMT,
            "Over95Months_ModeB": INT_FMT,
            "PeakLoad_Delta": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )

    focus_wc = str(wc_compare.iloc[0]["WorkCenter"]) if not wc_compare.empty else ""
    if focus_wc:
        focus_title_row = layout["end_row"] + 3
        ws.merge_cells(f"A{focus_title_row}:D{focus_title_row}")
        ws.cell(focus_title_row, 1).value = f"Focused workcenter comparison: {focus_wc}"
        ws.cell(focus_title_row, 1).fill = SUMMARY_FILL
        ws.cell(focus_title_row, 1).font = Font(bold=True, color="1F4E79", size=12)
        ws.cell(focus_title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[focus_title_row].height = 22
        focus_a = artifacts["ModeA"]["analysis"]["wc_long"]
        focus_b = artifacts["ModeB"]["analysis"]["wc_long"]
        focus_compare = (
            focus_a[focus_a["WorkCenter"] == focus_wc][["Month", "LoadPct"]]
            .rename(columns={"LoadPct": "Load_ModeA"})
            .merge(
                focus_b[focus_b["WorkCenter"] == focus_wc][["Month", "LoadPct"]].rename(columns={"LoadPct": "Load_ModeB"}),
                on="Month",
                how="outer",
            )
            .fillna(0.0)
            .sort_values("Month")
        )
        focus_compare["Load_Delta"] = focus_compare["Load_ModeB"] - focus_compare["Load_ModeA"]
        focus_layout = _write_table(
            ws,
            focus_compare,
            start_row=focus_title_row + 1,
            start_col=1,
            num_formats={"Load_ModeA": PCT_FMT, "Load_ModeB": PCT_FMT, "Load_Delta": PCT_FMT},
            alternating_fill=ALT_ROW_FILL,
        )
        focus_chart = LineChart()
        focus_chart.title = f"{focus_wc} Load Trend Comparison"
        focus_chart.height = 9
        focus_chart.width = 16
        focus_chart.add_data(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Load_ModeA"],
                min_row=focus_layout["start_row"],
                max_col=focus_layout["col_index"]["Load_ModeB"],
                max_row=focus_layout["end_row"],
            ),
            titles_from_data=True,
            from_rows=False,
        )
        focus_chart.set_categories(
            Reference(ws, min_col=focus_layout["col_index"]["Month"], min_row=focus_layout["start_row"] + 1, max_row=focus_layout["end_row"])
        )
        focus_chart.y_axis.numFmt = PCT_FMT
        _apply_chart_palette(focus_chart, ["2F75B5", "ED7D31"])
        chart_title_row = focus_layout["end_row"] + 3
        ws.merge_cells(f"A{chart_title_row}:H{chart_title_row}")
        ws.cell(chart_title_row, 1).value = f"{focus_wc} Load Trend Comparison"
        ws.cell(chart_title_row, 1).fill = SUMMARY_FILL
        ws.cell(chart_title_row, 1).font = Font(bold=True, color="1F4E79", size=12)
        ws.cell(chart_title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[chart_title_row].height = 22
        ws.add_chart(focus_chart, f"A{chart_title_row + 1}")
    _autofit(ws)


def _write_heatmap_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("WC_Heatmap_Compare")
    _write_sheet_title(ws, "WorkCenter Heatmap Comparison - Demand and Load%")

    wc_names = []
    for mode in ("ModeA", "ModeB"):
        summary = artifacts[mode]["analysis"]["wc_summary"]
        wc_names.extend(summary.head(12)["WorkCenter"].tolist())
    wc_names = list(dict.fromkeys(wc_names))

    next_start_row = 2
    for mode in ("ModeA", "ModeB"):
        header_cell = ws.cell(next_start_row, 1)
        header_cell.value = f"{mode} Heatmap"
        header_cell.font = Font(color="FFFFFF", bold=True, size=11)
        header_cell.fill = MODEA_FILL if mode == "ModeA" else MODEB_FILL
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        header_cell.border = BORDER
        ws.row_dimensions[next_start_row].height = 22

        wc_long = artifacts[mode]["analysis"]["wc_long"]
        wc_tons_df = artifacts[mode].get("wc_tons_df", pd.DataFrame())
        if wc_long.empty and wc_tons_df.empty:
            ws.cell(next_start_row + 1, 1).value = "No heatmap data"
            next_start_row += 4
            continue

        if not wc_tons_df.empty:
            months = [column for column in wc_tons_df.columns if column != "WorkCenter"]
        else:
            months = sorted(wc_long["Month"].astype(str).unique().tolist())

        if not wc_long.empty:
            load_pct_frame = wc_long.pivot(index="WorkCenter", columns="Month", values="LoadPct").fillna(0.0)
            load_pct_frame = load_pct_frame.reindex(columns=months, fill_value=0.0)
            load_pct_frame.reset_index(inplace=True)
        else:
            load_pct_frame = pd.DataFrame(columns=["WorkCenter", *months])

        monthly_frame = _merge_single_mode_heatmap_frames(
            demand_tons=wc_tons_df,
            load_pct=load_pct_frame,
            months=months,
        )
        yearly_frame = _summarize_heatmap_months_to_years(monthly_frame, months)

        ws.cell(next_start_row + 1, 1).value = "Yearly summary"
        ws.cell(next_start_row + 1, 1).font = Font(bold=True, color="1F4E79", size=11)
        yearly_view = _prepare_heatmap_display_frame(yearly_frame, wc_names)
        yearly_layout = _write_table(
            ws,
            yearly_view,
            start_row=next_start_row + 2,
            start_col=1,
        )
        _style_capacity_heatmap_block(ws, yearly_layout, yearly_view)

        monthly_start_row = yearly_layout["end_row"] + 4
        ws.cell(monthly_start_row - 1, 1).value = "Monthly detail"
        ws.cell(monthly_start_row - 1, 1).font = Font(bold=True, color="1F4E79", size=11)
        monthly_view = _prepare_heatmap_display_frame(monthly_frame, wc_names)
        monthly_layout = _write_table(
            ws,
            monthly_view,
            start_row=monthly_start_row,
            start_col=1,
        )
        _style_capacity_heatmap_block(ws, monthly_layout, monthly_view)
        next_start_row = monthly_layout["end_row"] + 4

    _write_note(
        ws,
        f"A{next_start_row}",
        "ModeA is shown first and ModeB directly below. Each block now includes a yearly summary and monthly detail, "
        "with Demand rows and Load% rows aligned to the same workcenter ordering.",
    )
    _autofit(ws)


def _write_product_risk_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Product_Risk_Compare")
    _write_sheet_title(ws, "Product Risk Comparison")
    _prepare_product_risk_sheet(ws)

    product_a = artifacts["ModeA"]["analysis"]["product_summary"].copy()
    product_b = artifacts["ModeB"]["analysis"]["product_summary"].copy()
    product_compare = product_a.merge(
        product_b,
        on=["Product", "ProductFamily", "Plant"],
        how="outer",
        suffixes=("_ModeA", "_ModeB"),
    ).fillna(0.0)
    product_compare["Unmet_Delta"] = product_compare["Unmet_Tons_ModeB"] - product_compare["Unmet_Tons_ModeA"]
    product_compare["Service_Level_Delta"] = product_compare["Service_Level_ModeB"] - product_compare["Service_Level_ModeA"]
    product_compare["SortKey"] = product_compare[["Unmet_Tons_ModeA", "Unmet_Tons_ModeB"]].max(axis=1)
    product_compare = product_compare.sort_values(["SortKey", "Unmet_Tons_ModeB"], ascending=[False, False]).head(20)

    display_columns = [
        "Product",
        "ProductFamily",
        "Plant",
        "Unmet_Tons_ModeA",
        "Unmet_Tons_ModeB",
        "Service_Level_ModeA",
        "Service_Level_ModeB",
        "Unmet_Delta",
        "Service_Level_Delta",
        "Demand_Tons_ModeA",
        "Internal_Tons_ModeA",
        "Internal_Tons_ModeB",
        "Outsourced_Tons_ModeA",
        "Outsourced_Tons_ModeB",
    ]
    layout = _write_table(
        ws,
        product_compare[display_columns],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeB": TONS_FMT,
            "Outsourced_Tons_ModeA": TONS_FMT,
            "Outsourced_Tons_ModeB": TONS_FMT,
            "Unmet_Tons_ModeA": TONS_FMT,
            "Unmet_Tons_ModeB": TONS_FMT,
            "Service_Level_ModeA": PCT_FMT,
            "Service_Level_ModeB": PCT_FMT,
            "Unmet_Delta": TONS_FMT,
            "Service_Level_Delta": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )
    _apply_risk_priority_headers(
        ws,
        layout,
        [
            "Unmet_Tons_ModeA",
            "Unmet_Tons_ModeB",
            "Service_Level_ModeA",
            "Service_Level_ModeB",
            "Unmet_Delta",
            "Service_Level_Delta",
        ],
    )
    _autofit(ws)
    _set_product_risk_column_layout(ws)


def _write_planner_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Planner_Compare")
    _write_sheet_title(ws, "Planner Comparison")

    planner_a = artifacts["ModeA"]["analysis"]["planner_summary"].copy()
    planner_b = artifacts["ModeB"]["analysis"]["planner_summary"].copy()
    for frame in (planner_a, planner_b):
        if "PlannerName" in frame.columns:
            frame = frame
    if not planner_a.empty:
        planner_a = planner_a[planner_a["PlannerName"].astype(str).str.strip().ne("")]
    if not planner_b.empty:
        planner_b = planner_b[planner_b["PlannerName"].astype(str).str.strip().ne("")]

    if planner_a.empty and planner_b.empty:
        ws["A3"] = "No planner comparison data is available for this result."
        return

    planner_compare = planner_a.merge(
        planner_b,
        on="PlannerName",
        how="outer",
        suffixes=("_ModeA", "_ModeB"),
    ).fillna(0.0)
    planner_compare["Demand_Delta"] = planner_compare["Demand_Tons_ModeB"] - planner_compare["Demand_Tons_ModeA"]
    planner_compare["Internal_Delta"] = planner_compare["Internal_Tons_ModeB"] - planner_compare["Internal_Tons_ModeA"]
    planner_compare["Outsourced_Delta"] = planner_compare["Outsourced_Tons_ModeB"] - planner_compare["Outsourced_Tons_ModeA"]
    planner_compare["Unmet_Delta"] = planner_compare["Unmet_Tons_ModeB"] - planner_compare["Unmet_Tons_ModeA"]
    planner_compare["Service_Level_Delta"] = planner_compare["Service_Level_ModeB"] - planner_compare["Service_Level_ModeA"]
    planner_compare["SortKey"] = planner_compare[["Demand_Tons_ModeA", "Demand_Tons_ModeB"]].max(axis=1)
    planner_compare = planner_compare.sort_values(["SortKey", "PlannerName"], ascending=[False, True])

    layout = _write_table(
        ws,
        planner_compare[
            [
                "PlannerName",
                "Demand_Tons_ModeA",
                "Demand_Tons_ModeB",
                "Internal_Tons_ModeA",
                "Internal_Tons_ModeB",
                "Outsourced_Tons_ModeA",
                "Outsourced_Tons_ModeB",
                "Unmet_Tons_ModeA",
                "Unmet_Tons_ModeB",
                "Service_Level_ModeA",
                "Service_Level_ModeB",
                "Service_Level_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_ModeA": TONS_FMT,
            "Demand_Tons_ModeB": TONS_FMT,
            "Internal_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeB": TONS_FMT,
            "Outsourced_Tons_ModeA": TONS_FMT,
            "Outsourced_Tons_ModeB": TONS_FMT,
            "Unmet_Tons_ModeA": TONS_FMT,
            "Unmet_Tons_ModeB": TONS_FMT,
            "Service_Level_ModeA": PCT_FMT,
            "Service_Level_ModeB": PCT_FMT,
            "Service_Level_Delta": PCT_FMT,
        },
        freeze="B2",
    )

    service_chart = BarChart()
    service_chart.title = "Planner Service Level Comparison"
    service_chart.y_axis.title = "Service level"
    service_chart.height = 8
    service_chart.width = 12
    service_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["Service_Level_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["Service_Level_ModeB"],
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        ),
        titles_from_data=True,
        from_rows=False,
    )
    service_chart.set_categories(
        Reference(
            ws,
            min_col=layout["col_index"]["PlannerName"],
            min_row=layout["start_row"] + 1,
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        )
    )
    service_chart.y_axis.numFmt = PCT_FMT
    ws.add_chart(service_chart, "N3")

    unmet_chart = BarChart()
    unmet_chart.type = "bar"
    unmet_chart.title = "Planner Residual Unmet Comparison"
    unmet_chart.height = 8
    unmet_chart.width = 12
    unmet_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["Unmet_Tons_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["Unmet_Tons_ModeB"],
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        ),
        titles_from_data=True,
        from_rows=False,
    )
    unmet_chart.set_categories(
        Reference(
            ws,
            min_col=layout["col_index"]["PlannerName"],
            min_row=layout["start_row"] + 1,
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        )
    )
    ws.add_chart(unmet_chart, "N21")

    focus_planner = str(planner_compare.iloc[0]["PlannerName"]) if not planner_compare.empty else ""
    if focus_planner:
        ws["A24"] = f"Focused planner comparison: {focus_planner}"
        ws["A24"].font = Font(bold=True, color="1F4E79", size=11)

        planner_month_a = (
            artifacts["ModeA"]["analysis"]["planner_product_month_summary"]
            .query("PlannerName == @focus_planner")
            .groupby("Month", as_index=False)
            .agg(
                Demand_ModeA=("Demand_Tons", "sum"),
                Internal_ModeA=("Internal_Tons", "sum"),
                Outsourced_ModeA=("Outsourced_Tons", "sum"),
                Unmet_ModeA=("Unmet_Tons", "sum"),
            )
        )
        planner_month_b = (
            artifacts["ModeB"]["analysis"]["planner_product_month_summary"]
            .query("PlannerName == @focus_planner")
            .groupby("Month", as_index=False)
            .agg(
                Demand_ModeB=("Demand_Tons", "sum"),
                Internal_ModeB=("Internal_Tons", "sum"),
                Outsourced_ModeB=("Outsourced_Tons", "sum"),
                Unmet_ModeB=("Unmet_Tons", "sum"),
            )
        )
        planner_month_compare = planner_month_a.merge(planner_month_b, on="Month", how="outer").fillna(0.0).sort_values("Month")
        planner_month_compare["Service_Level_ModeA"] = (
            planner_month_compare["Internal_ModeA"] + planner_month_compare["Outsourced_ModeA"]
        ).div(planner_month_compare["Demand_ModeA"].replace(0, pd.NA)).fillna(0.0)
        planner_month_compare["Service_Level_ModeB"] = (
            planner_month_compare["Internal_ModeB"] + planner_month_compare["Outsourced_ModeB"]
        ).div(planner_month_compare["Demand_ModeB"].replace(0, pd.NA)).fillna(0.0)

        focus_layout = _write_table(
            ws,
            planner_month_compare[
                [
                    "Month",
                    "Demand_ModeA",
                    "Demand_ModeB",
                    "Internal_ModeA",
                    "Internal_ModeB",
                    "Outsourced_ModeA",
                    "Outsourced_ModeB",
                    "Unmet_ModeA",
                    "Unmet_ModeB",
                    "Service_Level_ModeA",
                    "Service_Level_ModeB",
                ]
            ],
            start_row=25,
            start_col=1,
            num_formats={
                "Demand_ModeA": TONS_FMT,
                "Demand_ModeB": TONS_FMT,
                "Internal_ModeA": TONS_FMT,
                "Internal_ModeB": TONS_FMT,
                "Outsourced_ModeA": TONS_FMT,
                "Outsourced_ModeB": TONS_FMT,
                "Unmet_ModeA": TONS_FMT,
                "Unmet_ModeB": TONS_FMT,
                "Service_Level_ModeA": PCT_FMT,
                "Service_Level_ModeB": PCT_FMT,
            },
        )

        focus_chart = LineChart()
        focus_chart.title = f"{focus_planner} Monthly Service Level Comparison"
        focus_chart.height = 7
        focus_chart.width = 12
        focus_chart.add_data(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Service_Level_ModeA"],
                min_row=focus_layout["start_row"],
                max_col=focus_layout["col_index"]["Service_Level_ModeB"],
                max_row=focus_layout["end_row"],
            ),
            titles_from_data=True,
            from_rows=False,
        )
        focus_chart.set_categories(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Month"],
                min_row=focus_layout["start_row"] + 1,
                max_row=focus_layout["end_row"],
            )
        )
        focus_chart.y_axis.numFmt = PCT_FMT
        ws.add_chart(focus_chart, "M25")

    _write_note(
        ws,
        f"A{max(layout['end_row'] + 2, 40)}",
        "Planner comparison is based on planner-level traceability after the product-month optimization result is proportionally split back to planner demand shares.",
    )
    _autofit(ws)


def _write_comparison_run_info(
    wb: Workbook,
    config: Config,
    metrics_by_mode: dict[str, dict[str, Any]],
    workbook_name: str,
) -> None:
    ws = wb.create_sheet("Run_Info")
    _write_sheet_title(ws, "Comparison Run Info")
    info_df = pd.DataFrame(
        [
            ("Workbook_Name", workbook_name),
            ("Scenario_Name", config.scenario_name),
            ("Modes_Included", "ModeA + ModeB"),
            ("Start_Month", config.start_month),
            ("Horizon_Months", config.horizon_months),
            ("Input_Load_Folder", config.input_load_folder),
            ("Input_Master_Folder", config.input_master_folder),
            ("Output_Folder", config.output_folder),
            ("Project_Root_Folder", getattr(config, "project_root_folder", "")),
            ("Run_Timestamp", config.run_timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("License_Status", getattr(config, "license_status", "")),
            ("License_ID", getattr(config, "license_id", "")),
            ("License_Type", getattr(config, "license_type", "")),
            ("Licensed_To", getattr(config, "licensed_to", "")),
            ("License_Expiry", getattr(config, "license_expiry", "")),
            ("License_Binding_Mode", getattr(config, "license_binding_mode", "")),
            ("License_Machine_Label", getattr(config, "license_machine_label", "")),
            ("ModeA_Service_Level", metrics_by_mode["ModeA"]["service_level"] / 100.0),
            ("ModeB_Service_Level", metrics_by_mode["ModeB"]["service_level"] / 100.0),
        ],
        columns=["Parameter", "Value"],
    )
    layout = _write_table(ws, info_df, start_row=3, start_col=1)
    for row_num in range(layout["start_row"] + 1, layout["end_row"] + 1):
        if ws[f"A{row_num}"].value in {"ModeA_Service_Level", "ModeB_Service_Level"}:
            ws[f"B{row_num}"].number_format = PCT_FMT
    _autofit(ws)
    ws.sheet_state = "hidden"


def _write_capacity_basis_dashboard(
    wb: Workbook,
    mode: str,
    artifacts: dict[str, dict[str, Any]],
    issues: Optional[List[ValidationIssue]] = None,
    sheet_name: str = "Dashboard",
) -> None:
    ws = wb.create_sheet(sheet_name)
    fact_meta = _write_dashboard_fact_sheet(
        wb,
        {basis: artifacts[basis]["dashboard_fact"] for basis in ("Max", "Planner")},
    )
    scenario = (
        artifacts["Max"]["analysis"].get("scenario_name")
        or artifacts["Planner"]["analysis"].get("scenario_name")
        or "N/A"
    )
    subtitle = f"Scenario: {scenario} | Mode: {mode} | Capacity basis comparison: Max vs Planner"
    _prepare_dashboard_sheet(ws, f"Executive Summary - {mode} Capacity Comparison", subtitle)
    filter_refs = _add_dashboard_filter_controls(ws, fact_meta, start_row=3, start_col=15)
    selection_mode_ref = filter_refs["selection_mode_ref"]
    selected_range_ref = filter_refs["selected_range"]
    selected_year_ref = filter_refs["selected_year_ref"]

    demand_cells = _write_compare_kpi_card(
        ws,
        top_row=4,
        left_col=1,
        title="Total demand",
        left_label="Max",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "Max", "Demand_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="Planner",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "Planner", "Demand_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=C6-A6",
        number_format=TONS_FMT,
    )
    internal_cells = _write_compare_kpi_card(
        ws,
        top_row=4,
        left_col=8,
        title="Internal allocated",
        left_label="Max",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "Max", "Internal_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="Planner",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "Planner", "Internal_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=J6-H6",
        number_format=TONS_FMT,
    )
    outsourced_cells = _write_compare_kpi_card(
        ws,
        top_row=8,
        left_col=1,
        title="Outsourced",
        left_label="Max",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "Max", "Outsourced_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="Planner",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "Planner", "Outsourced_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=C10-A10",
        number_format=TONS_FMT,
    )
    unmet_cells = _write_compare_kpi_card(
        ws,
        top_row=8,
        left_col=8,
        title="Residual unmet",
        left_label="Max",
        left_value=_dashboard_filtered_sum_formula(
            fact_meta, "Max", "Unmet_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        right_label="Planner",
        right_value=_dashboard_filtered_sum_formula(
            fact_meta, "Planner", "Unmet_Tons", selection_mode_ref, selected_range_ref, selected_year_ref
        ),
        delta_label="Delta",
        delta_value="=J10-H10",
        number_format=TONS_FMT,
    )
    service_cells = _write_compare_kpi_card(
        ws,
        top_row=12,
        left_col=1,
        title="Service level",
        left_label="Max",
        left_value=f"=IF({demand_cells['Max']}=0,0,({internal_cells['Max']}+{outsourced_cells['Max']})/{demand_cells['Max']})",
        right_label="Planner",
        right_value=f"=IF({demand_cells['Planner']}=0,0,({internal_cells['Planner']}+{outsourced_cells['Planner']})/{demand_cells['Planner']})",
        delta_label="Delta",
        delta_value="=C14-A14",
        number_format=PCT_FMT,
    )
    _write_compare_kpi_card(
        ws,
        top_row=12,
        left_col=8,
        title="Selected workcenters",
        left_label="Max",
        left_value=_dashboard_selected_workcenter_count_formula(
            fact_meta, selection_mode_ref, selected_range_ref
        ),
        right_label="Planner",
        right_value=_dashboard_selected_workcenter_count_formula(
            fact_meta, selection_mode_ref, selected_range_ref
        ),
        delta_label="Delta",
        delta_value='=""',
        number_format=INT_FMT,
    )

    mix_df = pd.DataFrame(
        {
            "Category": ["Internal", "Outsourced", "Unmet"],
            "Max": [
                f"='{sheet_name}'!{internal_cells['Max']}",
                f"='{sheet_name}'!{outsourced_cells['Max']}",
                f"='{sheet_name}'!{unmet_cells['Max']}",
            ],
            "Planner": [
                f"='{sheet_name}'!{internal_cells['Planner']}",
                f"='{sheet_name}'!{outsourced_cells['Planner']}",
                f"='{sheet_name}'!{unmet_cells['Planner']}",
            ],
        }
    )
    mix_layout = _write_dashboard_helper_table(
        wb,
        mix_df,
        num_formats={"Max": TONS_FMT, "Planner": TONS_FMT},
    )
    helper_ws = wb[mix_layout["sheet_name"]]
    ws.merge_cells("A16:H16")
    ws["A16"] = "Supply Mix Comparison"
    ws["A16"].fill = SUMMARY_FILL
    ws["A16"].font = Font(bold=True, color="1F4E79", size=15)
    ws["A16"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[16].height = 24
    chart = BarChart()
    chart.title = "Supply Mix Comparison"
    chart.height = 8.5
    chart.width = 12.5
    chart.add_data(
        Reference(
            helper_ws,
            min_col=mix_layout["col_index"]["Max"],
            min_row=mix_layout["start_row"],
            max_col=mix_layout["col_index"]["Planner"],
            max_row=mix_layout["end_row"],
        ),
        titles_from_data=True,
        from_rows=False,
    )
    chart.set_categories(
        Reference(
            helper_ws,
            min_col=mix_layout["col_index"]["Category"],
            min_row=mix_layout["start_row"] + 1,
            max_row=mix_layout["end_row"],
        )
    )
    _style_dashboard_mix_chart(chart, ["2F75B5", "ED7D31"])
    ws.add_chart(chart, "A17")

    service_df = pd.DataFrame(
        {
            "Basis": ["Max", "Planner"],
            "Service_Level": [
                f"='{sheet_name}'!{service_cells['Max']}",
                f"='{sheet_name}'!{service_cells['Planner']}",
            ],
        }
    )
    service_layout = _write_dashboard_helper_table(
        wb,
        service_df,
        num_formats={"Service_Level": PCT_FMT},
    )
    service_chart = BarChart()
    service_chart.title = "Service Level Comparison"
    service_chart.height = 8.5
    service_chart.width = 9.5
    service_chart.add_data(
        Reference(
            helper_ws,
            min_col=service_layout["col_index"]["Service_Level"],
            min_row=service_layout["start_row"],
            max_row=service_layout["end_row"],
        ),
        titles_from_data=True,
    )
    service_chart.set_categories(
        Reference(
            helper_ws,
            min_col=service_layout["col_index"]["Basis"],
            min_row=service_layout["start_row"] + 1,
            max_row=service_layout["end_row"],
        )
    )
    _style_dashboard_service_chart(service_chart, ["2F75B5"])
    ws.merge_cells("L16:T16")
    ws["L16"] = "Service Level Comparison"
    ws["L16"].fill = SUMMARY_FILL
    ws["L16"].font = Font(bold=True, color="1F4E79", size=15)
    ws["L16"].alignment = Alignment(horizontal="center", vertical="center")
    ws.add_chart(service_chart, "L17")
    _write_dashboard_validation_block(ws, issues or [], start_row=40)
    _autofit(ws)
    _set_dashboard_column_layout(ws)


def _write_capacity_basis_monthly_analysis(
    wb: Workbook,
    mode: str,
    artifacts: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Monthly_Trend")
    _write_sheet_title(ws, f"Monthly Trend - {mode} | Max vs Planner")
    _prepare_monthly_trend_sheet(
        ws,
        f"Compare Max and Planner capacity basis month by month for {mode}.",
    )

    monthly_max = artifacts["Max"]["analysis"]["monthly_summary"].copy()
    monthly_planner = artifacts["Planner"]["analysis"]["monthly_summary"].copy()
    monthly_compare = monthly_max.merge(
        monthly_planner,
        on="Month",
        how="outer",
        suffixes=("_Max", "_Planner"),
    ).fillna(0.0).sort_values("Month")
    if monthly_compare.empty:
        ws["A3"] = "No monthly comparison data is available for this result."
        return

    monthly_compare.insert(1, "Year", monthly_compare["Month"].astype(str).str[:4])
    monthly_compare["Internal_Delta"] = monthly_compare["Internal_Tons_Planner"] - monthly_compare["Internal_Tons_Max"]
    monthly_compare["Outsourced_Delta"] = monthly_compare["Outsourced_Tons_Planner"] - monthly_compare["Outsourced_Tons_Max"]
    monthly_compare["Unmet_Delta"] = monthly_compare["Unmet_Tons_Planner"] - monthly_compare["Unmet_Tons_Max"]
    monthly_compare["Service_Level_Delta"] = monthly_compare["Service_Level_Planner"] - monthly_compare["Service_Level_Max"]

    yearly_compare = (
        monthly_compare.groupby("Year", as_index=False)
        .agg(
            Demand_Tons_Max=("Demand_Tons_Max", "sum"),
            Internal_Tons_Max=("Internal_Tons_Max", "sum"),
            Internal_Tons_Planner=("Internal_Tons_Planner", "sum"),
            Outsourced_Tons_Max=("Outsourced_Tons_Max", "sum"),
            Outsourced_Tons_Planner=("Outsourced_Tons_Planner", "sum"),
            Unmet_Tons_Max=("Unmet_Tons_Max", "sum"),
            Unmet_Tons_Planner=("Unmet_Tons_Planner", "sum"),
        )
        .sort_values("Year")
    )
    yearly_compare["Service_Level_Max"] = (
        (yearly_compare["Internal_Tons_Max"] + yearly_compare["Outsourced_Tons_Max"])
        .div(yearly_compare["Demand_Tons_Max"].replace(0, pd.NA))
        .fillna(0.0)
    )
    yearly_compare["Service_Level_Planner"] = (
        (yearly_compare["Internal_Tons_Planner"] + yearly_compare["Outsourced_Tons_Planner"])
        .div(yearly_compare["Demand_Tons_Max"].replace(0, pd.NA))
        .fillna(0.0)
    )
    yearly_compare["Unmet_Delta"] = yearly_compare["Unmet_Tons_Planner"] - yearly_compare["Unmet_Tons_Max"]
    ws.merge_cells("A3:K3")
    ws["A3"] = "Yearly summary"
    ws["A3"].fill = SUMMARY_FILL
    ws["A3"].font = Font(bold=True, color="1F4E79", size=12)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    yearly_layout = _write_table(
        ws,
        yearly_compare[
            [
                "Year",
                "Demand_Tons_Max",
                "Internal_Tons_Max",
                "Internal_Tons_Planner",
                "Outsourced_Tons_Max",
                "Outsourced_Tons_Planner",
                "Unmet_Tons_Max",
                "Unmet_Tons_Planner",
                "Service_Level_Max",
                "Service_Level_Planner",
                "Unmet_Delta",
            ]
        ],
        start_row=4,
        start_col=1,
        num_formats={
            "Demand_Tons_Max": TONS_FMT,
            "Internal_Tons_Max": TONS_FMT,
            "Internal_Tons_Planner": TONS_FMT,
            "Outsourced_Tons_Max": TONS_FMT,
            "Outsourced_Tons_Planner": TONS_FMT,
            "Unmet_Tons_Max": TONS_FMT,
            "Unmet_Tons_Planner": TONS_FMT,
            "Service_Level_Max": PCT_FMT,
            "Service_Level_Planner": PCT_FMT,
            "Unmet_Delta": TONS_FMT,
        },
        highlight_positive_cols=["Unmet_Delta"],
        alternating_fill=ALT_ROW_FILL,
    )

    monthly_start_row = yearly_layout["end_row"] + 4
    ws.merge_cells(f"A{monthly_start_row - 1}:L{monthly_start_row - 1}")
    ws.cell(monthly_start_row - 1, 1).value = "Monthly detail"
    ws.cell(monthly_start_row - 1, 1).fill = SUMMARY_FILL
    ws.cell(monthly_start_row - 1, 1).font = Font(bold=True, color="1F4E79", size=12)
    ws.cell(monthly_start_row - 1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[monthly_start_row - 1].height = 22
    monthly_layout = _write_table(
        ws,
        monthly_compare[
            [
                "Month",
                "Year",
                "Demand_Tons_Max",
                "Internal_Tons_Max",
                "Internal_Tons_Planner",
                "Outsourced_Tons_Max",
                "Outsourced_Tons_Planner",
                "Unmet_Tons_Max",
                "Unmet_Tons_Planner",
                "Service_Level_Max",
                "Service_Level_Planner",
                "Service_Level_Delta",
            ]
        ],
        start_row=monthly_start_row,
        start_col=1,
        num_formats={
            "Demand_Tons_Max": TONS_FMT,
            "Internal_Tons_Max": TONS_FMT,
            "Internal_Tons_Planner": TONS_FMT,
            "Outsourced_Tons_Max": TONS_FMT,
            "Outsourced_Tons_Planner": TONS_FMT,
            "Unmet_Tons_Max": TONS_FMT,
            "Unmet_Tons_Planner": TONS_FMT,
            "Service_Level_Max": PCT_FMT,
            "Service_Level_Planner": PCT_FMT,
            "Service_Level_Delta": PCT_FMT,
        },
        highlight_positive_cols=["Unmet_Tons_Max", "Unmet_Tons_Planner"],
        alternating_fill=ALT_ROW_FILL,
    )
    ws.auto_filter.ref = (
        f"A{monthly_layout['start_row']}:"
        f"{get_column_letter(monthly_layout['end_col'])}{monthly_layout['end_row']}"
    )

    _write_note(
        ws,
        f"A{monthly_layout['end_row'] + 2}",
        "Use the Excel filter on the Year column to narrow this monthly view to a specific year when needed.",
    )
    _autofit(ws)
    _set_monthly_trend_column_layout(ws)


def _write_capacity_basis_bottleneck_analysis(
    wb: Workbook,
    mode: str,
    artifacts: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Bottleneck")
    _write_sheet_title(ws, f"Bottleneck Analysis - {mode} | Max vs Planner")

    wc_max = artifacts["Max"]["analysis"]["wc_summary"].copy()
    wc_planner = artifacts["Planner"]["analysis"]["wc_summary"].copy()
    wc_compare = wc_max.merge(
        wc_planner,
        on="WorkCenter",
        how="outer",
        suffixes=("_Max", "_Planner"),
    ).fillna(0.0)
    if wc_compare.empty:
        ws["A3"] = "No workcenter load data is available for this result."
        return

    wc_compare["PeakLoad_Delta"] = wc_compare["PeakLoadPct_Planner"] - wc_compare["PeakLoadPct_Max"]
    wc_compare["SortKey"] = wc_compare[["PeakLoadPct_Max", "PeakLoadPct_Planner"]].max(axis=1)
    wc_compare = wc_compare.sort_values(["SortKey", "WorkCenter"], ascending=[False, True]).head(15)
    ws.merge_cells("A2:H2")
    ws["A2"] = "Top bottleneck workcenters"
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(bold=True, color="1F4E79", size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    layout = _write_table(
        ws,
        wc_compare[
            [
                "WorkCenter",
                "AvgLoadPct_Max",
                "AvgLoadPct_Planner",
                "PeakLoadPct_Max",
                "PeakLoadPct_Planner",
                "Over95Months_Max",
                "Over95Months_Planner",
                "PeakLoad_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "AvgLoadPct_Max": PCT_FMT,
            "AvgLoadPct_Planner": PCT_FMT,
            "PeakLoadPct_Max": PCT_FMT,
            "PeakLoadPct_Planner": PCT_FMT,
            "PeakLoad_Delta": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )
    _autofit(ws)


def _write_capacity_basis_heatmap(
    wb: Workbook,
    mode: str,
    basis_results: dict[str, List[AllocationResult]],
    basis_capacities: dict[str, List[CapacityRecord]],
    loads: List[LoadRecord],
    routings: List[RoutingRecord],
    months: List[str],
    sheet_name: str = "WC_Heatmap",
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_sheet_title(ws, f"WorkCenter Heatmap - {mode} | Demand, Max Load%, Planner Load%")

    heatmap_frames = build_capacity_compare_heatmap_frames(
        mode=mode,
        basis_results=basis_results,
        basis_capacities=basis_capacities,
        loads=loads,
        routings=routings,
        months=months,
    )
    yearly_df = heatmap_frames["yearly"]
    monthly_df = heatmap_frames["monthly"]
    if yearly_df.empty and monthly_df.empty:
        ws["A3"] = "No heatmap data is available for this result."
        return

    ranking_source = yearly_df[yearly_df["Metric"].isin(["Max Load%", "Planner Load%"])].copy()
    period_cols = [column for column in yearly_df.columns if column not in {"WorkCenter", "Metric"}]
    if not ranking_source.empty and period_cols:
        ranking_source["SortKey"] = ranking_source[period_cols].max(axis=1)
        workcenters = (
            ranking_source.groupby("WorkCenter", as_index=False)["SortKey"]
            .max()
            .sort_values(["SortKey", "WorkCenter"], ascending=[False, True])
            .head(12)["WorkCenter"]
            .tolist()
        )
    else:
        workcenters = list(dict.fromkeys(monthly_df["WorkCenter"].tolist()))[:12]

    ws["A2"] = "Yearly summary"
    ws["A2"].font = Font(bold=True, color="1F4E79", size=11)
    yearly_view = _prepare_heatmap_display_frame(yearly_df, workcenters)
    yearly_layout = _write_table(
        ws,
        yearly_view,
        start_row=3,
        start_col=1,
    )
    _style_capacity_heatmap_block(ws, yearly_layout, yearly_view)

    monthly_start_row = yearly_layout["end_row"] + 4
    ws.cell(monthly_start_row - 1, 1).value = "Monthly detail"
    ws.cell(monthly_start_row - 1, 1).font = Font(bold=True, color="1F4E79", size=11)
    monthly_view = _prepare_heatmap_display_frame(monthly_df, workcenters)
    monthly_layout = _write_table(
        ws,
        monthly_view,
        start_row=monthly_start_row,
        start_col=1,
    )
    _style_capacity_heatmap_block(ws, monthly_layout, monthly_view)

    _write_note(
        ws,
        f"A{monthly_layout['end_row'] + 2}",
        "Demand rows follow the Planner-basis assigned tons. Max Load% and Planner Load% compare that workcenter demand against the two capacity baselines.",
    )
    _autofit(ws)


def _write_capacity_basis_product_risk(
    wb: Workbook,
    mode: str,
    artifacts: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Product_Risk")
    _write_sheet_title(ws, f"Product Risk - {mode} | Max vs Planner")
    _prepare_product_risk_sheet(ws)

    product_max = artifacts["Max"]["analysis"]["product_summary"].copy()
    product_planner = artifacts["Planner"]["analysis"]["product_summary"].copy()
    product_compare = product_max.merge(
        product_planner,
        on=["Product", "ProductFamily", "Plant"],
        how="outer",
        suffixes=("_Max", "_Planner"),
    ).fillna(0.0)
    if product_compare.empty:
        ws["A3"] = "No product comparison data is available for this result."
        return

    product_compare["Unmet_Delta"] = product_compare["Unmet_Tons_Planner"] - product_compare["Unmet_Tons_Max"]
    product_compare["Service_Level_Delta"] = product_compare["Service_Level_Planner"] - product_compare["Service_Level_Max"]
    product_compare["SortKey"] = product_compare[["Unmet_Tons_Max", "Unmet_Tons_Planner"]].max(axis=1)
    product_compare = product_compare.sort_values(["SortKey", "Product"], ascending=[False, True]).head(20)
    display_columns = [
        "Product",
        "ProductFamily",
        "Plant",
        "Unmet_Tons_Max",
        "Unmet_Tons_Planner",
        "Service_Level_Max",
        "Service_Level_Planner",
        "Unmet_Delta",
        "Service_Level_Delta",
        "Demand_Tons_Max",
        "Internal_Tons_Max",
        "Internal_Tons_Planner",
        "Outsourced_Tons_Max",
        "Outsourced_Tons_Planner",
    ]
    layout = _write_table(
        ws,
        product_compare[display_columns],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_Max": TONS_FMT,
            "Internal_Tons_Max": TONS_FMT,
            "Internal_Tons_Planner": TONS_FMT,
            "Outsourced_Tons_Max": TONS_FMT,
            "Outsourced_Tons_Planner": TONS_FMT,
            "Unmet_Tons_Max": TONS_FMT,
            "Unmet_Tons_Planner": TONS_FMT,
            "Service_Level_Max": PCT_FMT,
            "Service_Level_Planner": PCT_FMT,
            "Unmet_Delta": TONS_FMT,
            "Service_Level_Delta": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )
    _apply_risk_priority_headers(
        ws,
        layout,
        [
            "Unmet_Tons_Max",
            "Unmet_Tons_Planner",
            "Service_Level_Max",
            "Service_Level_Planner",
            "Unmet_Delta",
            "Service_Level_Delta",
        ],
    )
    _autofit(ws)
    _set_product_risk_column_layout(ws)


def _write_capacity_basis_planner_summary(
    wb: Workbook,
    mode: str,
    artifacts: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Planner_Result_Summary")
    _prepare_planner_summary_sheet(
        ws,
        title=f"Planner Result Summary - {mode} | Max vs Planner",
        subtitle="Risk-first planner comparison across Max and Planner capacity baselines.",
        compare_mode=True,
    )

    planner_max = artifacts["Max"]["analysis"]["planner_summary"].copy()
    planner_planner = artifacts["Planner"]["analysis"]["planner_summary"].copy()
    planner_compare = planner_max.merge(
        planner_planner,
        on="PlannerName",
        how="outer",
        suffixes=("_Max", "_Planner"),
    ).fillna(0.0)
    if planner_compare.empty:
        ws["A3"] = "No planner summary is available for this result."
        return

    planner_compare["Service_Level_Delta"] = planner_compare["Service_Level_Planner"] - planner_compare["Service_Level_Max"]
    planner_compare = planner_compare.sort_values(["Unmet_Tons_Planner", "PlannerName"], ascending=[False, True])
    display_columns = [
        "PlannerName",
        "Unmet_Tons_Max",
        "Unmet_Tons_Planner",
        "Service_Level_Max",
        "Service_Level_Planner",
        "Demand_Tons_Max",
        "Outsourced_Tons_Max",
        "Outsourced_Tons_Planner",
        "Internal_Tons_Max",
        "Internal_Tons_Planner",
    ]
    layout = _write_table(
        ws,
        planner_compare[display_columns],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_Max": TONS_FMT,
            "Internal_Tons_Max": TONS_FMT,
            "Internal_Tons_Planner": TONS_FMT,
            "Outsourced_Tons_Max": TONS_FMT,
            "Outsourced_Tons_Planner": TONS_FMT,
            "Unmet_Tons_Max": TONS_FMT,
            "Unmet_Tons_Planner": TONS_FMT,
            "Service_Level_Max": PCT_FMT,
            "Service_Level_Planner": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )
    _apply_risk_priority_headers(
        ws,
        layout,
        [
            "Unmet_Tons_Max",
            "Unmet_Tons_Planner",
            "Service_Level_Max",
            "Service_Level_Planner",
        ],
    )
    _write_note(
        ws,
        f"A{layout['end_row'] + 2}",
        "This planner comparison keeps planner traceability while showing how the capacity baseline shifts each planner's internal supply, outsourcing, and residual unmet.",
    )
    _autofit(ws)
    _set_planner_summary_column_layout(ws, compare_mode=True)


def _write_summary_capacity_basis_pages(
    wb: Workbook,
    capacity_basis_payloads_by_mode: dict[str, dict[str, Any]],
    config: Config,
    months: List[str],
) -> None:
    for mode in ("ModeA", "ModeB"):
        payload = capacity_basis_payloads_by_mode.get(mode)
        if not payload:
            continue
        artifacts = {
            basis: _build_mode_artifact(
                payload["basis_results"][basis],
                payload["loads"],
                payload["basis_capacities"][basis],
                payload["routings"],
                config,
                months,
                basis,
            )
            for basis in ("Max", "Planner")
        }
        _write_capacity_basis_dashboard(
            wb,
            mode=mode,
            artifacts=artifacts,
            sheet_name=f"{mode}_Cap_Summary",
        )
        _write_capacity_basis_heatmap(
            wb,
            mode=mode,
            basis_results=payload["basis_results"],
            basis_capacities=payload["basis_capacities"],
            loads=payload["loads"],
            routings=payload["routings"],
            months=months,
            sheet_name=f"{mode}_Cap_Heatmap",
        )


def _concat_basis_detail_frames(artifacts: dict[str, dict[str, Any]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for basis in ("Max", "Planner"):
        frame = artifacts[basis]["df_detail"].copy()
        frame.insert(0, "Capacity_Basis", basis)
        if "Month" in frame.columns and "Year" not in frame.columns:
            frame.insert(2, "Year", frame["Month"].astype(str).str[:4])
        frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _concat_basis_planner_product_month_frames(artifacts: dict[str, dict[str, Any]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for basis in ("Max", "Planner"):
        frame = artifacts[basis]["analysis"].get("planner_product_month_summary", pd.DataFrame()).copy()
        if frame.empty:
            continue
        frame.insert(0, "Capacity_Basis", basis)
        frame.insert(2, "Year", frame["Month"].astype(str).str[:4])
        frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _concat_basis_wc_load_frames(artifacts: dict[str, dict[str, Any]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for basis in ("Max", "Planner"):
        frame = artifacts[basis]["wc_load_df"].copy()
        if frame.empty:
            continue
        frame.insert(0, "Capacity_Basis", basis)
        frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _concat_basis_run_info_frames(artifacts: dict[str, dict[str, Any]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for basis in ("Max", "Planner"):
        frame = artifacts[basis]["run_info_df"].copy()
        if frame.empty:
            continue
        frame.insert(0, "Capacity_Basis", basis)
        frames.append(frame)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _prepare_heatmap_display_frame(df: pd.DataFrame, workcenters: list[str]) -> pd.DataFrame:
    if df.empty:
        return df
    frame = df[df["WorkCenter"].isin(workcenters)].copy()
    if frame.empty:
        return frame
    frame["WorkCenter"] = pd.Categorical(frame["WorkCenter"], categories=workcenters, ordered=True)
    frame = frame.sort_values(["WorkCenter", "Metric"]).reset_index(drop=True)
    frame["WorkCenter"] = frame["WorkCenter"].astype(str)
    return frame


def _merge_single_mode_heatmap_frames(
    demand_tons: pd.DataFrame,
    load_pct: pd.DataFrame,
    months: List[str],
) -> pd.DataFrame:
    workcenters = sorted(
        set(demand_tons.get("WorkCenter", pd.Series(dtype=str)).tolist())
        | set(load_pct.get("WorkCenter", pd.Series(dtype=str)).tolist()),
        key=str.casefold,
    )
    metric_specs = [
        ("Demand", demand_tons),
        ("Load%", load_pct),
    ]

    rows: list[dict[str, object]] = []
    for work_center in workcenters:
        for metric_name, frame in metric_specs:
            frame_row = frame[frame["WorkCenter"] == work_center] if not frame.empty else pd.DataFrame()
            row = {
                "WorkCenter": work_center,
                "Metric": metric_name,
            }
            for month in months:
                row[month] = (
                    float(frame_row.iloc[0][month])
                    if not frame_row.empty and month in frame_row.columns
                    else 0.0
                )
            rows.append(row)
    return pd.DataFrame(rows, columns=["WorkCenter", "Metric", *months])


def _style_capacity_heatmap_block(
    ws,
    layout: dict[str, Any],
    df: pd.DataFrame,
) -> None:
    if df.empty:
        return
    value_columns = [column for column in df.columns if column not in {"WorkCenter", "Metric"}]
    if not value_columns:
        return

    demand_max = float(df[df["Metric"] == "Demand"][value_columns].to_numpy().max()) if (df["Metric"] == "Demand").any() else 0.0
    for row_offset, (_, row) in enumerate(df.iterrows(), start=1):
        excel_row = layout["start_row"] + row_offset
        metric = str(row["Metric"])
        for column in df.columns:
            ws.cell(excel_row, layout["col_index"][column]).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for column in value_columns:
            cell = ws.cell(excel_row, layout["col_index"][column])
            value = float(row[column] or 0.0)
            if metric == "Demand":
                cell.number_format = TONS_FMT
                if demand_max > 0:
                    if value >= demand_max * 0.85:
                        cell.fill = PatternFill("solid", fgColor="C00000")
                    elif value >= demand_max * 0.55:
                        cell.fill = PatternFill("solid", fgColor="F4B183")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
            else:
                cell.number_format = PCT_FMT
                if value >= 1.0:
                    cell.fill = PatternFill("solid", fgColor="C00000")
                elif value >= 0.85:
                    cell.fill = PatternFill("solid", fgColor="F4B183")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")


def _write_planner_product_month_detail(wb: Workbook, planner_product_month_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Planner_Product_Month")
    _write_sheet_title(ws, "Planner Product Month Summary")
    if planner_product_month_df.empty:
        ws["A3"] = "No planner product-month summary is available for this result."
        return
    _write_table(
        ws,
        planner_product_month_df,
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        freeze="D4",
    )
    _autofit(ws)


def _write_monthly_analysis(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Monthly_Trend")
    _write_sheet_title(ws, "Monthly Trend")
    _prepare_monthly_trend_sheet(
        ws,
        "Review monthly demand, internal supply, outsourced volume, unmet demand, and service level.",
    )

    monthly_summary = analysis["monthly_summary"]
    if monthly_summary.empty:
        ws["A3"] = "No monthly summary is available for this result."
        return

    monthly_detail = monthly_summary.copy()
    monthly_detail.insert(1, "Year", monthly_detail["Month"].astype(str).str[:4])
    yearly_summary = (
        monthly_detail.groupby("Year", as_index=False)
        .agg(
            Demand_Tons=("Demand_Tons", "sum"),
            Internal_Tons=("Internal_Tons", "sum"),
            Outsourced_Tons=("Outsourced_Tons", "sum"),
            Unmet_Tons=("Unmet_Tons", "sum"),
            Supplied_Tons=("Supplied_Tons", "sum"),
        )
        .sort_values("Year")
    )
    yearly_summary["Service_Level"] = (
        yearly_summary["Supplied_Tons"]
        .div(yearly_summary["Demand_Tons"].replace(0, pd.NA))
        .fillna(0.0)
    )

    ws.merge_cells("A3:G3")
    ws["A3"] = "Yearly summary"
    ws["A3"].fill = SUMMARY_FILL
    ws["A3"].font = Font(bold=True, color="1F4E79", size=12)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    yearly_layout = _write_table(
        ws,
        yearly_summary[
            [
                "Year",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=4,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        highlight_positive_cols=["Unmet_Tons"],
        alternating_fill=ALT_ROW_FILL,
    )

    monthly_title_row = yearly_layout["end_row"] + 4
    ws.merge_cells(f"A{monthly_title_row}:H{monthly_title_row}")
    ws.cell(monthly_title_row, 1).value = "Monthly detail"
    ws.cell(monthly_title_row, 1).fill = SUMMARY_FILL
    ws.cell(monthly_title_row, 1).font = Font(bold=True, color="1F4E79", size=12)
    ws.cell(monthly_title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[monthly_title_row].height = 22

    monthly_layout = _write_table(
        ws,
        monthly_detail[
            [
                "Month",
                "Year",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=monthly_title_row + 1,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        highlight_positive_cols=["Unmet_Tons"],
        freeze=f"A{monthly_title_row + 2}",
        alternating_fill=ALT_ROW_FILL,
    )

    gap_table = monthly_detail.sort_values(["Unmet_Tons", "Demand_Tons"], ascending=[False, False]).head(8)
    gap_title_row = monthly_layout["end_row"] + 3
    ws.merge_cells(f"A{gap_title_row}:G{gap_title_row}")
    ws.cell(gap_title_row, 1).value = "Highest gap months"
    ws.cell(gap_title_row, 1).fill = SUMMARY_FILL
    ws.cell(gap_title_row, 1).font = Font(bold=True, color="1F4E79", size=12)
    ws.cell(gap_title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[gap_title_row].height = 22
    _write_table(
        ws,
        gap_table[["Month", "Year", "Demand_Tons", "Internal_Tons", "Outsourced_Tons", "Unmet_Tons", "Service_Level"]],
        start_row=gap_title_row + 1,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        highlight_positive_cols=["Unmet_Tons"],
        alternating_fill=ALT_ROW_FILL,
    )
    _autofit(ws)
    _set_monthly_trend_column_layout(ws)


def _write_bottleneck_analysis(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Bottleneck")
    _write_sheet_title(ws, "Bottleneck Analysis")

    wc_summary = analysis["wc_summary"]
    wc_long = analysis["wc_long"]

    if wc_summary.empty or wc_long.empty:
        ws["A3"] = "No work-center load data is available for this result."
        return

    top_wc = wc_summary.head(12).copy()
    ws.merge_cells("A2:F2")
    ws["A2"] = "Top bottleneck workcenters"
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(bold=True, color="1F4E79", size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    top_layout = _write_table(
        ws,
        top_wc[["WorkCenter", "AvgLoadPct", "PeakLoadPct", "Over95Months"]],
        start_row=3,
        start_col=1,
        num_formats={
            "AvgLoadPct": PCT_FMT,
            "PeakLoadPct": PCT_FMT,
            "Over95Months": INT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )

    focus_wc = str(top_wc.iloc[0]["WorkCenter"])
    focus_title_row = top_layout["end_row"] + 3
    ws.merge_cells(f"A{focus_title_row}:C{focus_title_row}")
    ws.cell(focus_title_row, 1).value = f"Focused workcenter: {focus_wc}"
    ws.cell(focus_title_row, 1).fill = SUMMARY_FILL
    ws.cell(focus_title_row, 1).font = Font(bold=True, color="1F4E79", size=12)
    ws.cell(focus_title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[focus_title_row].height = 22

    wc_line_df = wc_long[wc_long["WorkCenter"] == focus_wc].sort_values("Month")
    line_layout = _write_table(
        ws,
        wc_line_df[["Month", "LoadPct"]],
        start_row=focus_title_row + 1,
        start_col=1,
        num_formats={"LoadPct": PCT_FMT},
        alternating_fill=ALT_ROW_FILL,
    )
    wc_line_chart = LineChart()
    wc_line_chart.title = f"{focus_wc} Pressure Load Trend"
    wc_line_chart.y_axis.title = "Pressure load"
    wc_line_chart.height = 9
    wc_line_chart.width = 16
    wc_line_chart.add_data(
        Reference(
            ws,
            min_col=line_layout["col_index"]["LoadPct"],
            min_row=line_layout["start_row"],
            max_row=line_layout["end_row"],
        ),
        titles_from_data=True,
    )
    wc_line_chart.set_categories(
        Reference(
            ws,
            min_col=line_layout["col_index"]["Month"],
            min_row=line_layout["start_row"] + 1,
            max_row=line_layout["end_row"],
        )
    )
    wc_line_chart.y_axis.numFmt = PCT_FMT
    chart_title_row = line_layout["end_row"] + 3
    ws.merge_cells(f"A{chart_title_row}:H{chart_title_row}")
    ws.cell(chart_title_row, 1).value = f"{focus_wc} Pressure Load Trend"
    ws.cell(chart_title_row, 1).fill = SUMMARY_FILL
    ws.cell(chart_title_row, 1).font = Font(bold=True, color="1F4E79", size=12)
    ws.cell(chart_title_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[chart_title_row].height = 22
    ws.add_chart(wc_line_chart, f"A{chart_title_row + 1}")

    _write_note(
        ws,
        f"A{chart_title_row + 19}",
        "Bottleneck metrics and heatmap percentages are based on internal allocation plus assigned unmet, "
        "shown against raw nameplate monthly capacity.",
    )
    _autofit(ws)


def _write_wc_heatmap(
    wb: Workbook,
    analysis: dict[str, Any],
    wc_tons_df: pd.DataFrame,
    months: List[str],
) -> None:
    ws = wb.create_sheet("WC_Heatmap")
    _write_sheet_title(ws, "WorkCenter Heatmap - Demand and Load%")

    wc_long = analysis["wc_long"]
    wc_summary = analysis["wc_summary"]
    if (wc_long.empty or wc_summary.empty) and wc_tons_df.empty:
        ws["A3"] = "No heatmap data is available for this result."
        return

    ranking_source = wc_summary.copy()
    if not ranking_source.empty:
        heatmap_wc_names = ranking_source.head(12)["WorkCenter"].tolist()
    else:
        heatmap_wc_names = list(dict.fromkeys(wc_tons_df.get("WorkCenter", pd.Series(dtype=str)).tolist()))[:12]

    if not wc_long.empty:
        load_pct_frame = wc_long.pivot(index="WorkCenter", columns="Month", values="LoadPct").fillna(0.0)
        load_pct_frame = load_pct_frame.reindex(columns=months, fill_value=0.0)
        load_pct_frame.reset_index(inplace=True)
    else:
        load_pct_frame = pd.DataFrame(columns=["WorkCenter", *months])

    monthly_frame = _merge_single_mode_heatmap_frames(
        demand_tons=wc_tons_df,
        load_pct=load_pct_frame,
        months=months,
    )
    yearly_frame = _summarize_heatmap_months_to_years(monthly_frame, months)

    ws["A2"] = "Yearly summary"
    ws["A2"].font = Font(bold=True, color="1F4E79", size=11)
    yearly_view = _prepare_heatmap_display_frame(yearly_frame, heatmap_wc_names)
    yearly_layout = _write_table(
        ws,
        yearly_view,
        start_row=3,
        start_col=1,
    )
    _style_capacity_heatmap_block(ws, yearly_layout, yearly_view)

    monthly_start_row = yearly_layout["end_row"] + 4
    ws.cell(monthly_start_row - 1, 1).value = "Monthly detail"
    ws.cell(monthly_start_row - 1, 1).font = Font(bold=True, color="1F4E79", size=11)
    monthly_view = _prepare_heatmap_display_frame(monthly_frame, heatmap_wc_names)
    monthly_layout = _write_table(
        ws,
        monthly_view,
        start_row=monthly_start_row,
        start_col=1,
    )
    _style_capacity_heatmap_block(ws, monthly_layout, monthly_view)
    _write_note(
        ws,
        f"A{monthly_layout['end_row'] + 2}",
        "Demand rows show assigned tons by workcenter. Load% rows show internal allocation plus assigned unmet "
        "against raw nameplate monthly capacity, so values may exceed 100%.",
    )
    _autofit(ws)


def _write_product_risk_analysis(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Product_Risk")
    _write_sheet_title(ws, "Product Risk")
    _prepare_product_risk_sheet(ws)

    product_summary = analysis["product_summary"]

    if product_summary.empty:
        ws["A3"] = "No product risk view is available for this result."
        return

    top_products = product_summary.head(20).copy()
    display_columns = [
        "Product",
        "ProductFamily",
        "Plant",
        "Unmet_Tons",
        "Service_Level",
        "Demand_Tons",
        "Outsourced_Tons",
        "Internal_Tons",
        "Supplied_Tons",
    ]
    layout = _write_table(
        ws,
        top_products[display_columns],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )
    _apply_risk_priority_headers(
        ws,
        layout,
        [
            "Unmet_Tons",
            "Service_Level",
            "Demand_Tons",
            "Outsourced_Tons",
            "Internal_Tons",
            "Supplied_Tons",
        ],
    )
    _autofit(ws)
    _set_product_risk_column_layout(ws)


def _write_planner_summary(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Planner_Result_Summary")
    scenario = analysis.get("scenario_name") or "N/A"
    mode_name = analysis.get("mode_name") or "Mode"
    _prepare_planner_summary_sheet(
        ws,
        title="Planner Result Summary",
        subtitle=f"Scenario: {scenario} | Mode: {mode_name} | Risk-first planner roll-up.",
        compare_mode=False,
    )

    planner_summary = analysis.get("planner_summary", pd.DataFrame())
    if planner_summary.empty:
        ws["A3"] = "No planner summary is available for this result."
        return

    display_columns = [
        "PlannerName",
        "Unmet_Tons",
        "Service_Level",
        "Demand_Tons",
        "Outsourced_Tons",
        "Internal_Tons",
    ]
    layout = _write_table(
        ws,
        planner_summary[display_columns],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        alternating_fill=ALT_ROW_FILL,
        freeze="A4",
    )
    _apply_risk_priority_headers(
        ws,
        layout,
        [
            "Unmet_Tons",
            "Service_Level",
        ],
    )
    _write_note(
        ws,
        f"A{layout['end_row'] + 2}",
        "This sheet shows planner-level traceability after the product-month optimization result is split back to planner shares.",
    )
    _autofit(ws)
    _set_planner_summary_column_layout(ws, compare_mode=False)


def _write_planner_product_month_summary(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Planner_Product_Month")
    _write_sheet_title(ws, "Planner Product Month Summary")

    planner_product_month = analysis.get("planner_product_month_summary", pd.DataFrame())
    if planner_product_month.empty:
        ws["A3"] = "No planner product-month summary is available for this result."
        return

    layout = _write_table(
        ws,
        planner_product_month[
            [
                "Month",
                "PlannerName",
                "Product",
                "ProductFamily",
                "Plant",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        freeze="C2",
    )
    _write_note(
        ws,
        f"A{layout['end_row'] + 2}",
        "Use this sheet when you need to trace a planner's monthly demand into internal, outsourced, and unmet outcomes.",
    )
    _autofit(ws)


def _write_detail(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Allocation_Detail")
    detail_df = _reorder_detail_columns(df)
    has_capacity_basis = "Capacity_Basis" in detail_df.columns
    freeze_col_index = len(
        [column for column in detail_df.columns if column in {"Capacity_Basis", "Month", "PlannerName", "Product", "ProductFamily", "Plant", "AllocationType", "WorkCenter"}]
    ) + 1
    freeze_panes = f"{get_column_letter(freeze_col_index)}4"
    _prepare_allocation_detail_sheet(
        ws,
        subtitle="Planner traceability detail by month, product, allocation type, and workcenter.",
        freeze_panes=freeze_panes,
        has_capacity_basis=has_capacity_basis,
    )
    _write_table(
        ws,
        detail_df,
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Allocated_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "CapacityShare_Pct": PCT_FMT,
        },
        freeze=freeze_panes,
        alternating_fill=ALT_ROW_FILL,
    )
    _set_allocation_detail_column_layout(ws, has_capacity_basis=has_capacity_basis)


def _write_allocation_summary(wb: Workbook, df: pd.DataFrame, months: List[str]) -> None:
    ws = wb.create_sheet("Allocation_Summary")
    internal_df = df[df["AllocationType"] == "Internal"]
    if internal_df.empty:
        ws["A1"] = "No data"
        return
    index_cols = ["Product", "ProductFamily", "Plant"]
    if "Capacity_Basis" in internal_df.columns:
        index_cols = ["Capacity_Basis", *index_cols]
    if "PlannerName" in internal_df.columns:
        index_cols = ["PlannerName", *index_cols]
    pivot = internal_df.pivot_table(
        index=index_cols,
        columns="Month",
        values="Allocated_Tons",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    freeze_col = get_column_letter(len(index_cols) + 1)
    _write_table(ws, pivot, tons_cols=pivot.columns[len(index_cols):].tolist(), freeze=f"{freeze_col}2")


def _write_outsource_summary(wb: Workbook, df: pd.DataFrame, months: List[str]) -> None:
    outsource_df = df[df["AllocationType"] == "Outsourced"]
    if outsource_df.empty:
        return
    ws = wb.create_sheet("Outsource_Summary")

    index_cols = ["Product", "ProductFamily", "Plant"]
    if "Capacity_Basis" in outsource_df.columns:
        index_cols = ["Capacity_Basis", *index_cols]
    if "PlannerName" in outsource_df.columns:
        index_cols = ["PlannerName", *index_cols]
    pivot = outsource_df.pivot_table(
        index=index_cols,
        columns="Month",
        values="Outsourced_Tons",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    freeze_col = get_column_letter(len(index_cols) + 1)
    _write_table(
        ws,
        pivot,
        tons_cols=pivot.columns[len(index_cols):].tolist(),
        freeze=f"{freeze_col}2",
        highlight_positive_cols=pivot.columns[len(index_cols):].tolist(),
    )


def _write_unmet_summary(wb: Workbook, df: pd.DataFrame, months: List[str]) -> None:
    ws = wb.create_sheet("Unmet_Summary")
    if df.empty:
        ws["A1"] = "No data"
        return

    index_cols = ["Product", "ProductFamily", "Plant"]
    if "Capacity_Basis" in df.columns:
        index_cols = ["Capacity_Basis", *index_cols]
    if "PlannerName" in df.columns:
        index_cols = ["PlannerName", *index_cols]
    pivot = df.pivot_table(
        index=index_cols,
        columns="Month",
        values="Unmet_Tons",
        aggfunc="max",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    freeze_col = get_column_letter(len(index_cols) + 1)
    _write_table(
        ws,
        pivot,
        tons_cols=pivot.columns[len(index_cols):].tolist(),
        freeze=f"{freeze_col}2",
        highlight_positive_cols=pivot.columns[len(index_cols):].tolist(),
    )
def _write_binary_report(
    wb: Workbook,
    df: pd.DataFrame,
    months: List[str],
    toller_products: set,
) -> None:
    ws = wb.create_sheet("Binary_Feasibility")
    if df.empty:
        ws["A1"] = "No data"
        return

    unmet_pivot = df.pivot_table(
        index=[
            *(
                ["Capacity_Basis"]
                if "Capacity_Basis" in df.columns
                else []
            ),
            "Product",
            "ProductFamily",
            "Plant",
        ],
        columns="Month",
        values="Unmet_Tons",
        aggfunc="max",
        fill_value=0,
    )
    unmet_pivot = unmet_pivot.reindex(columns=[month for month in months if month in unmet_pivot.columns], fill_value=0)
    binary_pivot = (unmet_pivot <= 0.01).astype(int)
    binary_pivot.reset_index(inplace=True)

    if toller_products:
        binary_pivot["Product"] = binary_pivot["Product"].apply(
            lambda product: f"{product} [Toller]" if product in toller_products else product
        )

    layout = _write_table(
        ws,
        binary_pivot,
        num_formats={
            column: "0"
            for column in binary_pivot.columns
            if column not in {"Capacity_Basis", "Product", "ProductFamily", "Plant"}
        },
        freeze=f"{get_column_letter(len([col for col in binary_pivot.columns if col in {'Capacity_Basis', 'Product', 'ProductFamily', 'Plant'}]) + 1)}2",
    )

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    month_cols = [month for month in months if month in binary_pivot.columns]
    for column in month_cols:
        col_num = layout["col_index"][column]
        for row in range(layout["start_row"] + 1, layout["end_row"] + 1):
            cell = ws.cell(row, col_num)
            if cell.value == 1:
                cell.fill = green_fill
            elif cell.value == 0:
                cell.fill = red_fill

    legend_row = layout["end_row"] + 3
    ws.cell(legend_row, 1).value = "Legend"
    ws.cell(legend_row, 1).font = Font(bold=True)
    ws.cell(legend_row + 1, 1).value = "1 = demand fully met"
    ws.cell(legend_row + 1, 1).fill = green_fill
    ws.cell(legend_row + 2, 1).value = "0 = residual unmet demand"
    ws.cell(legend_row + 2, 1).fill = red_fill
    if toller_products:
        ws.cell(legend_row + 3, 1).value = "[Toller] = toller-eligible product"
        ws.cell(legend_row + 3, 1).font = NOTE_FONT
    _autofit(ws)


def _write_validation(wb: Workbook, issues: List[ValidationIssue]) -> None:
    ws = wb.create_sheet("Validation_Issues")
    if not issues:
        _write_table(ws, pd.DataFrame([{"Severity": "OK", "Check": "Validation", "Detail": "No issues found."}]))
        ws["A2"].fill = OK_FILL
        ws["B2"].fill = OK_FILL
        ws["C2"].fill = OK_FILL
        ws.sheet_state = "hidden"
        return

    issue_df = pd.DataFrame(
        [{"Severity": issue.severity, "Check": issue.check, "Detail": issue.detail} for issue in issues]
    )
    layout = _write_table(ws, issue_df)
    for row in range(layout["start_row"] + 1, layout["end_row"] + 1):
        severity = ws.cell(row, layout["col_index"]["Severity"]).value
        fill = ERR_FILL if severity == "ERROR" else WARN_FILL
        for col in range(layout["start_col"], layout["end_col"] + 1):
            ws.cell(row, col).fill = fill
    _autofit(ws)
    ws.sheet_state = "hidden"


def _write_run_info(wb: Workbook, run_info_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Run_Info")
    _write_table(ws, run_info_df)
    _autofit(ws)
    ws.sheet_state = "hidden"


def _write_sheet_title(ws, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT


def _set_dashboard_column_layout(ws) -> None:
    content_width = 13
    gap_width = 3
    for column in ("A", "B", "C", "D", "E", "F", "H", "I", "J", "K", "L", "M", "O", "P", "Q", "R"):
        ws.column_dimensions[column].width = content_width
    for column in ("G", "N"):
        ws.column_dimensions[column].width = gap_width
    for column in ("S", "T"):
        ws.column_dimensions[column].width = content_width


def _prepare_dashboard_sheet(ws, title: str, subtitle: str) -> None:
    ws.sheet_view.showGridLines = False
    _set_dashboard_column_layout(ws)
    ws.merge_cells("A1:T1")
    ws["A1"] = title
    ws["A1"].fill = HDR_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:T2")
    ws["A2"] = subtitle
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(color="44546A", bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"


def _write_single_kpi_card(
    ws,
    top_row: int,
    left_col: int,
    title: str,
    value: Any,
    number_format: str | None = None,
) -> str:
    title_fill, body_fill = _metric_card_palette(title)
    right_col = left_col + 5
    value_row = top_row + 1
    ws.merge_cells(
        f"{ws.cell(top_row, left_col).coordinate}:{ws.cell(top_row, right_col).coordinate}"
    )
    title_cell = ws.cell(top_row, left_col)
    title_cell.value = title
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        f"{ws.cell(value_row, left_col).coordinate}:{ws.cell(value_row + 1, right_col).coordinate}"
    )
    value_cell = ws.cell(value_row, left_col)
    value_cell.value = value
    value_cell.fill = body_fill
    value_cell.font = Font(color="1F1F1F", bold=True, size=20)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        value_cell.number_format = number_format

    for row_num in range(top_row, value_row + 2):
        ws.row_dimensions[row_num].height = 24
        for col_num in range(left_col, right_col + 1):
            ws.cell(row_num, col_num).border = BORDER
    return value_cell.coordinate


def _write_compare_kpi_card(
    ws,
    top_row: int,
    left_col: int,
    title: str,
    left_label: str,
    left_value: Any,
    right_label: str,
    right_value: Any,
    delta_label: str,
    delta_value: Any,
    number_format: str | None = None,
) -> dict[str, str]:
    title_fill, body_fill = _metric_card_palette(title)
    right_col = left_col + 5
    ws.merge_cells(
        f"{ws.cell(top_row, left_col).coordinate}:{ws.cell(top_row, right_col).coordinate}"
    )
    title_cell = ws.cell(top_row, left_col)
    title_cell.value = title
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    groups = [
        (left_col, left_col + 1, left_label, left_value, body_fill),
        (left_col + 2, left_col + 3, right_label, right_value, body_fill),
        (left_col + 4, left_col + 5, delta_label, delta_value, DELTA_FILL),
    ]
    value_refs: dict[str, str] = {}
    for start_col, end_col, label, value, fill in groups:
        ws.merge_cells(
            f"{ws.cell(top_row + 1, start_col).coordinate}:{ws.cell(top_row + 1, end_col).coordinate}"
        )
        label_cell = ws.cell(top_row + 1, start_col)
        label_cell.value = label
        label_cell.fill = fill
        label_cell.font = Font(color="1F1F1F", bold=True, size=10)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(
            f"{ws.cell(top_row + 2, start_col).coordinate}:{ws.cell(top_row + 2, end_col).coordinate}"
        )
        value_cell = ws.cell(top_row + 2, start_col)
        value_cell.value = value
        value_cell.fill = fill
        value_cell.font = Font(color="1F1F1F", bold=True, size=15)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if number_format:
            value_cell.number_format = number_format
        value_refs[label] = value_cell.coordinate

    for row_num in range(top_row, top_row + 3):
        ws.row_dimensions[row_num].height = 24
        for col_num in range(left_col, right_col + 1):
            ws.cell(row_num, col_num).border = BORDER
    return value_refs


def _write_metric_block(
    ws,
    start_row: int,
    start_col: int,
    rows: list[tuple[str, Any, str | None]],
) -> None:
    _write_header_row(ws, ["Metric", "Value"], start_row=start_row, start_col=start_col)
    for offset, (label, value, fmt) in enumerate(rows, start=1):
        label_cell = ws.cell(start_row + offset, start_col)
        value_cell = ws.cell(start_row + offset, start_col + 1)
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(bold=True)
        label_cell.fill = SUBHDR_FILL
        for cell in (label_cell, value_cell):
            cell.border = BORDER
        if fmt:
            value_cell.number_format = fmt


def _write_note(ws, cell_ref: str, text: str) -> None:
    ws[cell_ref] = text
    ws[cell_ref].font = NOTE_FONT
    ws[cell_ref].alignment = Alignment(wrap_text=True, vertical="top")


def _write_dashboard_validation_block(
    ws,
    issues: List[ValidationIssue],
    start_row: int,
    title: str = "Data Validation / Issues",
) -> None:
    if not issues:
        ws.merge_cells(f"A{start_row}:T{start_row}")
        ok_cell = ws[f"A{start_row}"]
        ok_cell.value = "Validation Status: OK - No data issues found."
        ok_cell.fill = OK_FILL
        ok_cell.font = Font(bold=True, color="1F1F1F", size=11)
        ok_cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, 21):
            ws.cell(start_row, col).border = BORDER
        ws.row_dimensions[start_row].height = 24
        return

    ws.merge_cells(f"A{start_row}:T{start_row}")
    title_cell = ws[f"A{start_row}"]
    title_cell.value = title
    title_cell.fill = SUMMARY_FILL
    title_cell.font = Font(bold=True, color="1F4E79", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22

    header_row = start_row + 1
    _write_header_row(ws, ["Severity", "Check", "Detail"], start_row=header_row, start_col=1)
    ws.merge_cells(f"C{header_row}:T{header_row}")
    ws[f"C{header_row}"].alignment = Alignment(horizontal="center", vertical="center")

    for offset, issue in enumerate(issues, start=1):
        row_num = header_row + offset
        severity = str(issue.severity or "").strip().upper()
        fill = ERR_FILL if severity == "ERROR" else WARN_FILL

        severity_cell = ws.cell(row_num, 1)
        severity_cell.value = severity
        severity_cell.fill = fill
        severity_cell.font = Font(bold=True, size=10)
        severity_cell.alignment = Alignment(horizontal="center", vertical="center")
        severity_cell.border = BORDER

        check_cell = ws.cell(row_num, 2)
        check_cell.value = issue.check
        check_cell.fill = fill
        check_cell.font = Font(size=10)
        check_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        check_cell.border = BORDER

        ws.merge_cells(f"C{row_num}:T{row_num}")
        detail_cell = ws.cell(row_num, 3)
        detail_cell.value = issue.detail
        detail_cell.fill = fill
        detail_cell.font = Font(size=10)
        detail_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col in range(3, 21):
            ws.cell(row_num, col).border = BORDER
        ws.row_dimensions[row_num].height = 30


def _apply_chart_palette(chart, colors: list[str]) -> None:
    for series, color in zip(chart.series, colors):
        try:
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color
        except Exception:
            continue


def _style_dashboard_mix_chart(chart, colors: list[str]) -> None:
    chart.style = 10
    chart.type = "bar"
    chart.gapWidth = 75
    chart.overlap = 0
    chart.title = None
    chart.y_axis.title = None
    chart.legend.position = "b"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    _apply_chart_palette(chart, colors)


def _style_dashboard_service_chart(chart, colors: list[str] | None = None) -> None:
    chart.style = 11
    chart.gapWidth = 65
    chart.title = None
    chart.y_axis.title = None
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.y_axis.numFmt = PCT_FMT
    try:
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
    except Exception:
        pass
    if colors:
        _apply_chart_palette(chart, colors)


def _set_monthly_trend_column_layout(ws) -> None:
    widths = {
        "A": 12,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 13,
        "H": 13,
        "I": 13,
        "J": 13,
        "K": 13,
        "L": 13,
        "M": 13,
        "N": 13,
        "O": 13,
        "P": 13,
        "Q": 3,
        "R": 3,
        "S": 13,
        "T": 13,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _prepare_monthly_trend_sheet(ws, subtitle: str) -> None:
    ws.sheet_view.showGridLines = False
    _set_monthly_trend_column_layout(ws)
    ws.merge_cells("A1:T1")
    ws["A1"] = ws["A1"].value
    ws["A1"].fill = HDR_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:T2")
    ws["A2"] = subtitle
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(color="44546A", bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24


def _style_monthly_balance_chart(chart) -> None:
    chart.style = 10
    chart.title = None
    chart.y_axis.title = None
    chart.gapWidth = 55
    chart.legend.position = "b"
    _apply_chart_palette(chart, ["548235", "ED7D31", "C00000", "2F75B5"])


def _style_monthly_service_chart(chart, colors: list[str], show_legend: bool) -> None:
    chart.style = 10
    chart.title = None
    chart.y_axis.title = None
    chart.y_axis.numFmt = PCT_FMT
    try:
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
    except Exception:
        pass
    chart.legend.position = "b"
    if not show_legend:
        chart.legend = None
    _apply_chart_palette(chart, colors)


def _style_monthly_unmet_chart(chart, colors: list[str]) -> None:
    chart.style = 10
    chart.title = None
    chart.y_axis.title = None
    chart.gapWidth = 60
    chart.legend.position = "b"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    _apply_chart_palette(chart, colors)


def _set_product_risk_column_layout(ws) -> None:
    widths = {
        "A": 24,
        "B": 18,
        "C": 14,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _prepare_product_risk_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    _set_product_risk_column_layout(ws)
    ws.freeze_panes = "A4"


def _apply_risk_priority_headers(
    ws,
    layout: dict[str, Any],
    priority_columns: list[str],
) -> None:
    header_row = layout["start_row"]
    for column_name in priority_columns:
        column_index = layout["col_index"].get(column_name)
        if not column_index:
            continue
        cell = ws.cell(header_row, column_index)
        cell.fill = RISK_HDR_FILL
        cell.font = Font(color="FFFFFF", bold=True, size=10)


def _set_planner_summary_column_layout(ws, compare_mode: bool) -> None:
    widths = {
        "A": 18,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
    }
    if compare_mode:
        widths = {
            "A": 18,
            "B": 14,
            "C": 16,
            "D": 15,
            "E": 17,
            "F": 15,
            "G": 15,
            "H": 17,
            "I": 15,
            "J": 15,
        }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _prepare_planner_summary_sheet(ws, title: str, subtitle: str, compare_mode: bool) -> None:
    ws.sheet_view.showGridLines = False
    _set_planner_summary_column_layout(ws, compare_mode=compare_mode)
    end_col = "J" if compare_mode else "F"
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].fill = HDR_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(color="44546A", bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A4"


def _set_allocation_detail_column_layout(ws, has_capacity_basis: bool) -> None:
    widths = {
        "A": 14,
        "B": 12,
        "C": 18,
        "D": 22,
        "E": 18,
        "F": 14,
        "G": 14,
        "H": 22,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 10,
    }
    if not has_capacity_basis:
        widths = {
            "A": 12,
            "B": 18,
            "C": 22,
            "D": 18,
            "E": 14,
            "F": 14,
            "G": 22,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 14,
            "M": 14,
            "N": 10,
        }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _prepare_allocation_detail_sheet(ws, subtitle: str, freeze_panes: str, has_capacity_basis: bool) -> None:
    ws.sheet_view.showGridLines = False
    _set_allocation_detail_column_layout(ws, has_capacity_basis=has_capacity_basis)
    ws.merge_cells("A1:O1" if has_capacity_basis else "A1:N1")
    ws["A1"] = "Allocation Detail"
    ws["A1"].fill = HDR_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:O2" if has_capacity_basis else "A2:N2")
    ws["A2"] = subtitle
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(color="44546A", bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = freeze_panes


def _reorder_detail_columns(df: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "Capacity_Basis",
        "Month",
        "PlannerName",
        "Product",
        "ProductFamily",
        "Plant",
        "AllocationType",
        "WorkCenter",
        "Demand_Tons",
        "Allocated_Tons",
        "Outsourced_Tons",
        "Unmet_Tons",
        "CapacityShare_Pct",
        "RouteType",
        "Priority",
    ]
    ordered = [column for column in preferred if column in df.columns]
    trailing = [column for column in df.columns if column not in ordered]
    return df[ordered + trailing].copy()


def _write_header_row(ws, headers: list[str], start_row: int = 1, start_col: int = 1) -> None:
    for index, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + index)
        cell.value = header
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[start_row].height = 24


def _write_table(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    num_formats: Optional[dict[str, str]] = None,
    tons_cols: Optional[list[str]] = None,
    freeze: Optional[str] = None,
    highlight_positive_cols: Optional[list[str]] = None,
    highlight_over_100_pct: Optional[list[str]] = None,
    alternating_fill: Optional[PatternFill] = None,
) -> dict[str, Any]:
    num_formats = num_formats or {}
    tons_cols = tons_cols or []
    highlight_positive_cols = highlight_positive_cols or []
    highlight_over_100_pct = highlight_over_100_pct or []

    headers = list(df.columns)
    if not headers:
        return {
            "start_row": start_row,
            "end_row": start_row,
            "start_col": start_col,
            "end_col": start_col,
            "col_index": {},
        }

    _write_header_row(ws, headers, start_row=start_row, start_col=start_col)
    col_index = {header: start_col + offset for offset, header in enumerate(headers)}

    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        row_num = start_row + row_offset
        for header, value in zip(headers, row):
            cell = ws.cell(row_num, col_index[header])
            cell.value = value if not _is_nan(value) else None
            cell.border = BORDER
            if alternating_fill and row_offset % 2 == 0:
                cell.fill = alternating_fill

            fmt = num_formats.get(header)
            if fmt is None and header in tons_cols:
                fmt = TONS_FMT
            if fmt:
                cell.number_format = fmt

            if header in highlight_positive_cols and isinstance(value, (int, float)) and value > 0:
                cell.fill = ERR_FILL
            if header in highlight_over_100_pct and isinstance(value, (int, float)) and value > 1.0:
                cell.fill = ERR_FILL

    end_row = start_row + len(df)
    end_col = start_col + len(headers) - 1
    if freeze:
        ws.freeze_panes = freeze
    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
        "col_index": col_index,
    }


def _autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column, max_len in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 10), 42)


def _is_nan(value: Any) -> bool:
    try:
        import math

        return math.isnan(value)
    except Exception:
        return False


def _sanitize_filename_segment(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r'[<>:"/\\|?*]+', "-", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("._-")
