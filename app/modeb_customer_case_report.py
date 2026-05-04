"""Generate product analysis workbooks from existing ModeB outputs."""
from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.capacity_basis import PLANNED_BASIS, normalize_capacity_basis
from app.data_loader import load_direct_mode_a_with_capacity_bases, load_direct_mode_b_with_capacity_bases
from app.i18n import localize_column_name, localize_sheet_name, localize_value
from app.models import LoadRecord
from app.runtime_paths import RuntimePaths, resolve_runtime_paths, with_workspace_dir


DEFAULT_OUTPUT_NAME = "product_analysis.xlsx"
DEFAULT_MAX_PRODUCTS = 10
SUPPORTED_REPORT_MODES = ("ModeA", "ModeB")

THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=15, bold=True, color="1F1F1F")
SECTION_FONT = Font(size=11, bold=True, color="1F1F1F")

SUMMARY_HEADERS = {
    "Product": "产品",
    "Case_Description": "案例类型",
    "Capacity_Basis": "Capacity_Basis",
    "Demand_Tons": "需求吨位",
    "Capacity_Base_Tons": "Capacity_Base 吨位",
    "Routing_Reroute_Tons": "Routing_Reroute 吨位",
    "Outsourced_Tons": "外协吨位",
    "Unmet_Tons": "未满足吨位",
    "Planner_Source_File": "Planner 来源文件",
    "Master_Capacity_WC": "master_capacity 工作中心",
    "Master_Routing_WC": "master_routing 内部路径工作中心",
    "Toller_WC": "Toller 工作中心",
}

TABLE_HEADER_LABELS = {
    **SUMMARY_HEADERS,
    "Source_File": "来源文件",
    "Planner_Name": "Planner 名称",
    "Product_Family": "产品族",
    "Plant": "工厂",
    "Scenario": "场景",
    "Month_Span": "月份范围",
    "Record_Count": "记录数",
    "WorkCenter": "工作中心",
    "Annual_Max_Capacity_Tons": "Annual Max 产能吨位",
    "Annual_Planned_Capacity_Tons": "Annual Planned 产能吨位",
    "Monthly_Max_Capacity_Tons": "Monthly Max 产能吨位",
    "Monthly_Planned_Capacity_Tons": "Monthly Planned 产能吨位",
    "Utilization_Target": "Utilization_Target",
    "Route_Type": "路径类型",
    "Max_Capacity_Tons": "Max 产能吨位",
    "Planned_Capacity_Tons": "Planned 产能吨位",
    "Eligible_Flag": "Eligible 标记",
    "Allocation_Source": "分配来源",
    "Internal_Tons": "内部承接吨位",
    "Share_of_Product_Demand_Pct": "占该产品需求比例(%)",
    "WC_Role": "工作中心角色",
    "Focus_Product_Tons": "当前产品吨位",
    "Total_WC_Internal_Tons": "该 WC 总内部吨位",
    "Focus_Product_Share_of_WC_Pct": "当前产品占该 WC 比例(%)",
    "Other_Product": "同 WC 其他产品",
    "Other_Product_Internal_Tons": "其他产品内部吨位",
    "Focus_Product_Outsourced_Tons": "当前产品外协吨位",
    "Other_Product_Outsourced_Tons": "其他产品外协吨位",
    "Routing_Definition": "Routing 定义",
}

ROUTE_TYPE_LABELS = {
    "Primary": "主路径",
    "Alternative": "替代路径",
    "Toller": "外协路径",
}

ALLOCATION_SOURCE_LABELS = {
    "ModeA_Internal": "ModeA 内部分配",
    "Capacity_Base": "Capacity_Base",
    "Routing_Reroute": "Routing_Reroute",
    "ModeB_Internal": "ModeB 内部分配",
    "Unmet": "Unmet",
}

WC_ROLE_LABELS = {
    "ModeA_Internal_WC": "ModeA 内部分配 WC",
    "Capacity_Base": "ModeB Capacity_Base WC",
    "Routing_Reroute": "ModeB Routing_Reroute WC",
    "ModeB_Toller_Route": "ModeB Toller 路径",
}

DETAIL_CANONICAL_COLUMNS = (
    "Capacity_Basis",
    "Month",
    "PlannerName",
    "Product",
    "ProductFamily",
    "Plant",
    "AllocationType",
    "WorkCenter",
    "Demand_Tons",
    "Allocated_Tons",
    "Outsourced_Tons",
    "Unmet_Tons",
    "CapacityShare_Pct",
    "Allocation_Source",
    "Residual_After_Capacity_Tons",
    "Residual_After_Routing_Tons",
    "RouteType",
    "Priority",
    "Year",
)

CANONICAL_VALUE_ALIASES = {
    "Internal": {localize_value("zh", "Internal"), "Internal"},
    "Outsourced": {localize_value("zh", "Outsourced"), "Outsourced"},
    "Unmet": {localize_value("zh", "Unmet"), "Unmet"},
    "ModeA": {localize_value("zh", "ModeA"), "ModeA"},
    "ModeB": {localize_value("zh", "ModeB"), "ModeB"},
    "Both": {localize_value("zh", "Both"), "Both"},
}


class ReportValidationError(ValueError):
    """Raised when the selected workbook cannot be analyzed as a ModeB report."""


@dataclass(frozen=True)
class ModeReportSelection:
    selected_path: Path
    latest_path: Path | None
    is_latest: bool


@dataclass(frozen=True)
class ModeReportContext:
    report_path: Path
    report_mode: str
    scenario_name: str
    run_timestamp: str | None
    input_load_folder: Path | None
    input_master_folder: Path | None
    output_folder: Path | None
    available_bases: tuple[str, ...]
    detail_rows: tuple[dict[str, Any], ...]


ModeBReportSelection = ModeReportSelection
ModeBReportContext = ModeReportContext


@dataclass(frozen=True)
class SupportingInputData:
    load_folder: Path
    master_folder: Path
    loads: tuple[LoadRecord, ...]
    raw_capacity_rows: tuple[dict[str, str], ...]
    raw_routing_rows: tuple[dict[str, str], ...]


@dataclass
class ProductMetrics:
    demand: float = 0.0
    internal: float = 0.0
    outsourced: float = 0.0
    unmet: float = 0.0
    capacity_base: float = 0.0
    routing_reroute: float = 0.0


class _TableNameFactory:
    def __init__(self) -> None:
        self._counter = 0

    def next(self, prefix: str) -> str:
        self._counter += 1
        safe_prefix = "".join(ch for ch in prefix if ch.isalnum()) or "Table"
        # Excel table names cannot look like cell references such as CAP14.
        return f"Tbl{safe_prefix[:17]}{self._counter}"


def _label(header: str) -> str:
    return TABLE_HEADER_LABELS.get(header, header)


def _route_type_label(value: str) -> str:
    return ROUTE_TYPE_LABELS.get(value, value)


def _allocation_source_label(value: str) -> str:
    return ALLOCATION_SOURCE_LABELS.get(value, value)


def _wc_role_label(value: str) -> str:
    return WC_ROLE_LABELS.get(value, value)


def _safe_pct(numerator: float, denominator: float) -> float:
    if abs(denominator) <= 1e-9:
        return 0.0
    return 100.0 * numerator / denominator


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _canonical_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    for canonical, aliases in CANONICAL_VALUE_ALIASES.items():
        if text in aliases:
            return canonical
    return text


def _header_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for canonical in DETAIL_CANONICAL_COLUMNS:
        aliases[canonical] = canonical
        aliases[localize_column_name("zh", canonical)] = canonical
    return aliases


def _find_sheet_name(sheet_names: Sequence[str], canonical_name: str) -> str | None:
    candidates = {
        canonical_name,
        localize_sheet_name("zh", canonical_name),
    }
    for name in sheet_names:
        if name in candidates:
            return name
    return None


def _trim_headers(values: Sequence[Any]) -> list[str]:
    headers = [str(value).strip() if value is not None else "" for value in values]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _read_detail_rows(workbook_path: Path) -> tuple[tuple[dict[str, Any], ...], tuple[str, ...]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        detail_sheet_name = _find_sheet_name(workbook.sheetnames, "Allocation_Detail")
        if not detail_sheet_name:
            raise ReportValidationError("所选文件缺少 Allocation_Detail/分配明细 sheet，无法作为单模式报告分析。")
        ws = workbook[detail_sheet_name]
        row_iter = ws.iter_rows(min_row=3, values_only=True)
        try:
            header_values = next(row_iter)
        except StopIteration as exc:
            raise ReportValidationError("所选文件的分配明细表为空，无法继续分析。") from exc
        raw_headers = _trim_headers(list(header_values))
        if not raw_headers:
            raise ReportValidationError("所选文件的分配明细表头为空，无法继续分析。")
        aliases = _header_alias_map()
        headers = [aliases.get(header, header) for header in raw_headers]
        rows: list[dict[str, Any]] = []
        for values in row_iter:
            values = list(values[: len(headers)])
            if not any(value is not None and str(value).strip() != "" for value in values):
                continue
            row = {header: _canonical_value(value) for header, value in zip(headers, values)}
            if not str(row.get("Product", "")).strip():
                continue
            rows.append(row)
        return tuple(rows), tuple(headers)
    finally:
        workbook.close()


def _read_run_info(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        run_info_sheet_name = _find_sheet_name(workbook.sheetnames, "Run_Info")
        if not run_info_sheet_name:
            return {}
        ws = workbook[run_info_sheet_name]
        row_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(row_iter)
        except StopIteration:
            return {}
        first_header = str((first_row[0] if first_row else "") or "").strip()
        basis_aware = first_header == "Capacity_Basis"
        info: dict[str, Any] = {}
        for row in row_iter:
            if basis_aware:
                basis = row[0] if len(row) > 0 else None
                parameter = row[1] if len(row) > 1 else None
                value = row[2] if len(row) > 2 else None
                if parameter in (None, ""):
                    continue
                parameter_text = str(parameter).strip()
                if parameter_text and parameter_text not in info:
                    info[parameter_text] = value
                if basis not in (None, "") and parameter_text:
                    info.setdefault("by_basis", {}).setdefault(normalize_capacity_basis(str(basis)), {})[parameter_text] = value
            else:
                parameter = row[0] if len(row) > 0 else None
                value = row[1] if len(row) > 1 else None
                if parameter in (None, ""):
                    continue
                parameter_text = str(parameter).strip()
                if parameter_text and parameter_text not in info:
                    info[parameter_text] = value
        return info
    finally:
        workbook.close()


def _looks_like_mode_output_name(path: Path, report_mode: str) -> bool:
    name = path.name.lower()
    return name.startswith(f"capacity_result_{report_mode.lower()}_") and name.endswith(".xlsx")


def find_latest_mode_report(output_dir: str | Path, report_mode: str) -> Path | None:
    output_path = Path(output_dir)
    if not output_path.exists():
        return None
    candidates = [
        path
        for path in output_path.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and _looks_like_mode_output_name(path, report_mode)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def find_latest_modeb_report(output_dir: str | Path) -> Path | None:
    return find_latest_mode_report(output_dir, "ModeB")


def resolve_mode_report_selection(
    *,
    output_dir: str | Path,
    manual_report_path: str | Path | None,
    use_latest_report: bool,
    report_mode: str,
) -> ModeReportSelection:
    latest_path = find_latest_mode_report(output_dir, report_mode)
    if use_latest_report:
        if latest_path is None:
            raise FileNotFoundError(f"当前 output 目录下没有可用的 {report_mode} 输出报告。")
        return ModeReportSelection(selected_path=latest_path, latest_path=latest_path, is_latest=True)

    manual_text = str(manual_report_path or "").strip()
    if not manual_text:
        raise FileNotFoundError(f"请输入 {report_mode} 报告文件路径或文件名。")

    raw_path = Path(manual_text).expanduser()
    search_candidates = [raw_path]
    if not raw_path.is_absolute():
        search_candidates.insert(0, Path(output_dir) / raw_path)

    resolved_path: Path | None = None
    for candidate in search_candidates:
        if candidate.exists():
            resolved_path = candidate.resolve()
            break
    if resolved_path is None:
        raise FileNotFoundError(f"未找到指定的报告文件：{manual_text}")

    is_latest = latest_path is not None and resolved_path == latest_path.resolve()
    return ModeReportSelection(selected_path=resolved_path, latest_path=latest_path, is_latest=is_latest)

def resolve_modeb_report_selection(
    *,
    output_dir: str | Path,
    manual_report_path: str | Path | None,
    use_latest_report: bool,
) -> ModeBReportSelection:
    return resolve_mode_report_selection(
        output_dir=output_dir,
        manual_report_path=manual_report_path,
        use_latest_report=use_latest_report,
        report_mode="ModeB",
    )


def _infer_report_mode_from_name(report_file: Path) -> str | None:
    name = report_file.name.lower()
    if name.startswith("capacity_result_modea_"):
        return "ModeA"
    if name.startswith("capacity_result_modeb_"):
        return "ModeB"
    if name.startswith("summary of mode a and mode b_"):
        return "Both"
    return None


def _infer_report_mode_from_detail_rows(detail_rows: Sequence[dict[str, Any]]) -> str:
    for row in detail_rows:
        source = str(row.get("Allocation_Source") or "").strip()
        route_type = str(row.get("RouteType") or "").strip()
        if source in {"Capacity_Base", "Routing_Reroute", "Toller"}:
            return "ModeB"
        if route_type in {"Primary", "Alternative", "Toller"}:
            return "ModeB"
    return "ModeA"


def load_mode_report_context(report_path: str | Path, *, expected_mode: str | None = None) -> ModeReportContext:
    report_file = Path(report_path).expanduser().resolve()
    if not report_file.exists():
        raise FileNotFoundError(f"报告文件不存在：{report_file}")
    if report_file.name.startswith("~$"):
        raise ReportValidationError("当前选择的是 Excel 临时锁文件，请选择正式的 .xlsx 报告。")
    detail_rows, detail_headers = _read_detail_rows(report_file)
    required_headers = {"Product", "AllocationType", "WorkCenter", "Demand_Tons"}
    if not required_headers.issubset(set(detail_headers)):
        raise ReportValidationError("所选文件不是有效的单模式结果报告：分配明细缺少关键列。")

    run_info = _read_run_info(report_file)
    actual_mode = _infer_report_mode_from_name(report_file) or _infer_report_mode_from_detail_rows(detail_rows)
    if actual_mode == "Both":
        raise ReportValidationError("该工具只支持 ModeA 或 ModeB 单报告，不支持 Both 汇总报告。")
    if expected_mode and actual_mode != expected_mode:
        raise ReportValidationError(f"当前文件属于 {actual_mode}，与所选的 {expected_mode} 不一致。")
    available_bases = sorted(
        {
            normalize_capacity_basis(str(row.get("Capacity_Basis") or PLANNED_BASIS).strip())
            for row in detail_rows
            if str(row.get("Capacity_Basis") or "").strip()
        }
    ) or [PLANNED_BASIS]

    return ModeReportContext(
        report_path=report_file,
        report_mode=actual_mode,
        scenario_name=str(run_info.get("Scenario_Name") or "").strip(),
        run_timestamp=str(run_info.get("Run_Timestamp") or "").strip() or None,
        input_load_folder=Path(run_info["Input_Load_Folder"]).resolve() if run_info.get("Input_Load_Folder") else None,
        input_master_folder=Path(run_info["Input_Master_Folder"]).resolve() if run_info.get("Input_Master_Folder") else None,
        output_folder=Path(run_info["Output_Folder"]).resolve() if run_info.get("Output_Folder") else None,
        available_bases=tuple(available_bases),
        detail_rows=detail_rows,
    )


def load_modeb_report_context(report_path: str | Path) -> ModeBReportContext:
    return load_mode_report_context(report_path, expected_mode="ModeB")


def infer_workspace_root_from_report(report_path: str | Path) -> Path | None:
    report_file = Path(report_path).expanduser().resolve()
    parent = report_file.parent
    if parent.name.lower() == "output":
        return parent.parent.resolve()
    return None


def _resolve_input_folder(candidate: Path | None, fallback: Path) -> Path:
    if candidate and candidate.exists():
        return candidate
    return fallback


def _find_tabular_file(folder: Path, stem: str, *, required: bool = True) -> Path | None:
    for ext in (".xlsx", ".xls", ".csv"):
        path = folder / f"{stem}{ext}"
        if path.exists():
            return path
    if required:
        raise FileNotFoundError(f"未找到 {stem} 文件：{folder}")
    return None


def _read_tabular_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(path, header=0)
    elif suffix == ".csv":
        df = pd.read_csv(path, encoding="utf-8-sig")
    else:
        raise ValueError(f"不支持的文件类型：{path}")
    df.columns = [str(column).strip() for column in df.columns]
    return df.fillna("").to_dict(orient="records")


def load_supporting_input_data(
    context: ModeReportContext,
    runtime_paths: RuntimePaths | None = None,
) -> SupportingInputData:
    resolved_paths = runtime_paths or resolve_runtime_paths()
    default_input_dir = resolved_paths.workspace_input_dir
    load_folder = _resolve_input_folder(context.input_load_folder, default_input_dir)
    master_folder = _resolve_input_folder(context.input_master_folder, default_input_dir)
    scenario_name = context.scenario_name or "Baseline"

    if context.report_mode == "ModeA":
        loads, _capacity_bases, _ = load_direct_mode_a_with_capacity_bases(
            str(load_folder),
            str(master_folder),
            selected_scenario=scenario_name,
        )
        raw_routing_rows: list[dict[str, Any]] = []
    else:
        loads, _baseline_capacities_by_basis, _capacities_by_basis, _routings = load_direct_mode_b_with_capacity_bases(
            str(load_folder),
            str(master_folder),
            selected_scenario=scenario_name,
        )
        routing_path = _find_tabular_file(master_folder, "master_routing", required=False)
        raw_routing_rows = _read_tabular_rows(routing_path) if routing_path else []
    capacity_path = _find_tabular_file(master_folder, "master_capacity", required=True)
    raw_capacity_rows = _read_tabular_rows(capacity_path)
    return SupportingInputData(
        load_folder=load_folder,
        master_folder=master_folder,
        loads=tuple(loads),
        raw_capacity_rows=tuple(raw_capacity_rows),
        raw_routing_rows=tuple(raw_routing_rows),
    )


def _basis_key(row: dict[str, Any]) -> str:
    raw_basis = str(row.get("Capacity_Basis") or PLANNED_BASIS).strip() or PLANNED_BASIS
    return normalize_capacity_basis(raw_basis)


def _product_metrics(detail_rows: Sequence[dict[str, Any]]) -> dict[tuple[str, str], ProductMetrics]:
    metrics: dict[tuple[str, str], ProductMetrics] = defaultdict(ProductMetrics)
    seen_demand: set[tuple[str, str, str, str, str]] = set()
    for row in detail_rows:
        basis = _basis_key(row)
        product = str(row.get("Product") or "").strip()
        metric = metrics[(basis, product)]
        demand_key = (
            basis,
            str(row.get("Month") or "").strip(),
            str(row.get("PlannerName") or "").strip(),
            product,
            str(row.get("Plant") or "").strip(),
        )
        if demand_key not in seen_demand:
            metric.demand += float(row.get("Demand_Tons") or 0.0)
            seen_demand.add(demand_key)
        allocation_type = str(row.get("AllocationType") or "").strip()
        if allocation_type == "Internal":
            metric.internal += float(row.get("Allocated_Tons") or 0.0)
            source = str(row.get("Allocation_Source") or "").strip()
            if source in {"", "Capacity_Base", "ModeA_Internal"}:
                metric.capacity_base += float(row.get("Allocated_Tons") or 0.0)
            elif source == "Routing_Reroute":
                metric.routing_reroute += float(row.get("Allocated_Tons") or 0.0)
        elif allocation_type == "Outsourced":
            metric.outsourced += float(row.get("Outsourced_Tons") or 0.0)
        elif allocation_type == "Unmet":
            metric.unmet += float(row.get("Unmet_Tons") or 0.0)
    return metrics


def _internal_by_product_wc_and_source(detail_rows: Sequence[dict[str, Any]]) -> dict[tuple[str, str, str, str], float]:
    totals: dict[tuple[str, str, str, str], float] = defaultdict(float)
    for row in detail_rows:
        if str(row.get("AllocationType") or "").strip() != "Internal":
            continue
        basis = _basis_key(row)
        product = str(row.get("Product") or "").strip()
        source = str(row.get("Allocation_Source") or "").strip() or "ModeA_Internal"
        wc = str(row.get("WorkCenter") or "").strip()
        totals[(basis, product, source, wc)] += float(row.get("Allocated_Tons") or 0.0)
    return totals


def _wc_internal_totals(detail_rows: Sequence[dict[str, Any]]) -> dict[tuple[str, str], dict[str, float]]:
    totals: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in detail_rows:
        if str(row.get("AllocationType") or "").strip() != "Internal":
            continue
        basis = _basis_key(row)
        wc = str(row.get("WorkCenter") or "").strip()
        product = str(row.get("Product") or "").strip()
        totals[(basis, wc)][product] += float(row.get("Allocated_Tons") or 0.0)
    return totals


def _toller_wc_by_product(routing_rows: Sequence[dict[str, str]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for row in routing_rows:
        route_type = str(row.get("Router Type", "")).strip().lower()
        product = str(row.get("Product", "")).strip()
        wc = str(row.get("Resource", "")).strip()
        if product and wc and route_type == "toller":
            mapping[product] = wc
    return mapping


def _outsourced_by_basis_toller_wc(
    detail_rows: Sequence[dict[str, Any]],
    toller_wc_lookup: dict[str, str],
) -> dict[tuple[str, str], dict[str, float]]:
    totals: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in detail_rows:
        if str(row.get("AllocationType") or "").strip() != "Outsourced":
            continue
        basis = _basis_key(row)
        product = str(row.get("Product") or "").strip()
        wc = toller_wc_lookup.get(product)
        if not wc:
            continue
        totals[(basis, wc)][product] += float(row.get("Outsourced_Tons") or 0.0)
    return totals


def _planner_input_rows(loads: Sequence[LoadRecord], product: str) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str, str, str], dict[str, Any]] = {}
    product_loads = [load for load in loads if load.product == product]
    if not product_loads:
        return []

    months = sorted({load.month for load in product_loads})
    month_span = f"{months[0]} -> {months[-1]}" if months else ""
    for load in product_loads:
        key = (
            str(load.source_file or ""),
            str(load.planner_name or ""),
            str(load.product_family or ""),
            str(load.plant or ""),
            str(load.resource_group_owner or ""),
            str(load.scenario_version or load.scenario or ""),
        )
        row = grouped.setdefault(
            key,
            {
                "Source_File": key[0],
                "Planner_Name": key[1],
                "Product_Family": key[2],
                "Plant": key[3],
                "WorkCenter": key[4],
                "Scenario": key[5],
                "Month_Span": month_span,
                "Demand_Tons": 0.0,
                "Record_Count": 0,
            },
        )
        row["Demand_Tons"] += float(load.forecast_tons or 0.0)
        row["Record_Count"] += 1
    rows = list(grouped.values())
    rows.sort(key=lambda row: (row["Source_File"], row["Planner_Name"], row["Plant"]))
    return rows


def _master_capacity_rows(raw_capacity_rows: Sequence[dict[str, str]], product: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in raw_capacity_rows:
        if str(row.get("Product", "")).strip() != product:
            continue
        annual_max_capacity = float(
            row.get("Annual Max Capacity Tons", row.get("Annual Capacity Tons", 0)) or 0.0
        )
        annual_planned_capacity = float(
            row.get("Annual Planned Capacity Tons", row.get("Annual Planner Capacity Tons", row.get("Annual Capacity Tons", 0))) or 0.0
        )
        rows.append(
            {
                "Source_File": "master_capacity.csv",
                "WorkCenter": str(row.get("WorkCenter", row.get("Resource", ""))).strip(),
                "Annual_Max_Capacity_Tons": annual_max_capacity,
                "Annual_Planned_Capacity_Tons": annual_planned_capacity,
                "Monthly_Max_Capacity_Tons": annual_max_capacity / 12.0,
                "Monthly_Planned_Capacity_Tons": annual_planned_capacity / 12.0,
                "Utilization_Target": str(row.get("Utilization Target", "")).strip(),
            }
        )
    return rows


def _master_routing_rows(raw_routing_rows: Sequence[dict[str, str]], product: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in raw_routing_rows:
        if str(row.get("Product", "")).strip() != product:
            continue
        rows.append(
            {
                "Source_File": "master_routing.csv",
                "WorkCenter": str(row.get("WorkCenter", row.get("Resource", ""))).strip(),
                "Route_Type": _route_type_label(str(row.get("Router Type", "")).strip()),
                "Max_Capacity_Tons": float(row.get("Max Capacity Ton", 0) or 0.0),
                "Planned_Capacity_Tons": float(
                    row.get("Planned Capacity Ton", row.get("Planner Capacity Ton", 0)) or 0.0
                ),
                "Eligible_Flag": str(row.get("EligibleFalg", "")).strip(),
            }
        )
    return rows


def _case_type_label(report_mode: str, metric: ProductMetrics) -> str:
    if report_mode == "ModeA":
        if metric.unmet <= 0:
            return "全部由 capacity 消化"
        if metric.capacity_base > 0:
            return "capacity 吃掉一部分，剩余进入 unmet"
        return "没有可用 capacity，全部进入 unmet"
    if metric.routing_reroute <= 0 and metric.outsourced <= 0 and metric.unmet <= 0:
        return "全部由 baseline capacity 消化"
    if metric.routing_reroute > 0 and metric.outsourced <= 0 and metric.unmet <= 0:
        return "baseline + routing 完全消化"
    if metric.outsourced > 0 and metric.unmet <= 0:
        return "baseline + toller 完全消化"
    if metric.routing_reroute > 0 and metric.unmet > 0:
        return "baseline + routing 后仍有 unmet"
    if metric.outsourced > 0 and metric.unmet > 0:
        return "baseline + toller 后仍有 unmet"
    return "只有 baseline，剩余进入 unmet"


def _overview_rows(
    report_mode: str,
    products: Sequence[str],
    bases: Sequence[str],
    loads: Sequence[LoadRecord],
    raw_capacity_rows: Sequence[dict[str, str]],
    raw_routing_rows: Sequence[dict[str, str]],
    metrics: dict[tuple[str, str], ProductMetrics],
) -> list[dict[str, Any]]:
    planner_source_lookup: dict[str, str] = {}
    for product in products:
        source_files = sorted({str(load.source_file or "") for load in loads if load.product == product and load.source_file})
        planner_source_lookup[product] = " | ".join(source_files) or "（无）"

    rows: list[dict[str, Any]] = []
    for product in products:
        cap_wcs = sorted(
            {
                str(row.get("WorkCenter", row.get("Resource", ""))).strip()
                for row in raw_capacity_rows
                if str(row.get("Product", "")).strip() == product
            }
        )
        routing_wcs = sorted(
            {
                str(row.get("WorkCenter", row.get("Resource", ""))).strip()
                for row in raw_routing_rows
                if str(row.get("Product", "")).strip() == product
                and str(row.get("Router Type", "")).strip().lower() != "toller"
            }
        )
        toller_wcs = sorted(
            {
                str(row.get("WorkCenter", row.get("Resource", ""))).strip()
                for row in raw_routing_rows
                if str(row.get("Product", "")).strip() == product
                and str(row.get("Router Type", "")).strip().lower() == "toller"
            }
        )
        for basis in bases:
            metric = metrics[(basis, product)]
            rows.append(
                {
                    "Product": product,
                    "Case_Description": _case_type_label(report_mode, metric),
                    "Capacity_Basis": basis,
                    "Demand_Tons": metric.demand,
                    "Capacity_Base_Tons": metric.capacity_base,
                    "Routing_Reroute_Tons": metric.routing_reroute,
                    "Outsourced_Tons": metric.outsourced,
                    "Unmet_Tons": metric.unmet,
                    "Planner_Source_File": planner_source_lookup.get(product, "（无）"),
                    "Master_Capacity_WC": " | ".join(cap_wcs) or "（无）",
                    "Master_Routing_WC": " | ".join(routing_wcs) if report_mode == "ModeB" else "（不适用）",
                    "Toller_WC": " | ".join(toller_wcs) if report_mode == "ModeB" else "（不适用）",
                }
            )
    return rows


def _basis_summary_rows(
    report_mode: str,
    product: str,
    bases: Sequence[str],
    metrics: dict[tuple[str, str], ProductMetrics],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for basis in bases:
        metric = metrics[(basis, product)]
        rows.append(
            {
                "Capacity_Basis": basis,
                "Case_Description": _case_type_label(report_mode, metric),
                "Demand_Tons": metric.demand,
                "Capacity_Base_Tons": metric.capacity_base,
                "Routing_Reroute_Tons": metric.routing_reroute,
                "Outsourced_Tons": metric.outsourced,
                "Unmet_Tons": metric.unmet,
            }
        )
    return rows


def _allocation_rows(
    *,
    basis: str,
    product: str,
    internal_totals: dict[tuple[str, str, str, str], float],
    demand_tons: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for (row_basis, row_product, source, wc), tons in sorted(
        internal_totals.items(),
        key=lambda item: (item[0][0], item[0][2], -item[1], item[0][3]),
    ):
        if row_basis != basis or row_product != product or tons <= 0:
            continue
        rows.append(
            {
                "Allocation_Source": _allocation_source_label(source),
                "WorkCenter": wc,
                "Internal_Tons": tons,
                "Share_of_Product_Demand_Pct": _safe_pct(tons, demand_tons),
            }
        )
    return rows


def _other_products_on_wc_rows(
    *,
    basis: str,
    focus_product: str,
    wc_roles: Sequence[tuple[str, str]],
    wc_internal_totals: dict[tuple[str, str], dict[str, float]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for wc_role, wc in wc_roles:
        wc_products = wc_internal_totals.get((basis, wc), {})
        total_wc_tons = sum(wc_products.values())
        focus_tons = wc_products.get(focus_product, 0.0)
        other_rows = [
            (product, tons)
            for product, tons in sorted(wc_products.items(), key=lambda item: (-item[1], item[0]))
            if product != focus_product and tons > 0
        ]
        if not other_rows:
            rows.append(
                {
                    "WC_Role": _wc_role_label(wc_role),
                    "WorkCenter": wc,
                    "Focus_Product_Tons": focus_tons,
                    "Total_WC_Internal_Tons": total_wc_tons,
                    "Focus_Product_Share_of_WC_Pct": _safe_pct(focus_tons, total_wc_tons),
                    "Other_Product": "（无）",
                    "Other_Product_Internal_Tons": 0.0,
                }
            )
            continue
        for other_product, other_tons in other_rows:
            rows.append(
                {
                    "WC_Role": _wc_role_label(wc_role),
                    "WorkCenter": wc,
                    "Focus_Product_Tons": focus_tons,
                    "Total_WC_Internal_Tons": total_wc_tons,
                    "Focus_Product_Share_of_WC_Pct": _safe_pct(focus_tons, total_wc_tons),
                    "Other_Product": other_product,
                    "Other_Product_Internal_Tons": other_tons,
                }
            )
    return rows


def _other_products_on_toller_rows(
    *,
    basis: str,
    focus_product: str,
    toller_wc: str | None,
    focus_outsourced_tons: float,
    outsourced_by_toller_wc: dict[tuple[str, str], dict[str, float]],
) -> list[dict[str, Any]]:
    if not toller_wc:
        return []
    wc_products = outsourced_by_toller_wc.get((basis, toller_wc), {})
    other_rows = [
        (product, tons)
        for product, tons in sorted(wc_products.items(), key=lambda item: (-item[1], item[0]))
        if product != focus_product and tons > 0
    ]
    if not other_rows:
        return [
            {
                "WC_Role": _wc_role_label("ModeB_Toller_Route"),
                "WorkCenter": toller_wc,
                "Focus_Product_Outsourced_Tons": focus_outsourced_tons,
                "Other_Product": "（无）",
                "Other_Product_Outsourced_Tons": 0.0,
            }
        ]
    return [
        {
            "WC_Role": _wc_role_label("ModeB_Toller_Route"),
            "WorkCenter": toller_wc,
            "Focus_Product_Outsourced_Tons": focus_outsourced_tons,
            "Other_Product": other_product,
            "Other_Product_Outsourced_Tons": other_tons,
        }
        for other_product, other_tons in other_rows
    ]


def _style_title(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_note(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = NOTE_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_section(ws, row: int, text: str, end_col: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_table(
    *,
    ws,
    start_row: int,
    rows: Sequence[dict[str, Any]],
    table_name: str,
    pct_columns: Iterable[str] | None = None,
    ton_columns: Iterable[str] | None = None,
) -> int:
    if not rows:
        ws.cell(start_row, 1).value = "（无）"
        return start_row

    pct_columns = set(pct_columns or [])
    ton_columns = set(ton_columns or [])
    headers = list(rows[0].keys())

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_index)
        cell.value = _label(header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for row_offset, row_data in enumerate(rows, start=1):
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(start_row + row_offset, col_index)
            value = row_data.get(header)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if header in ton_columns and isinstance(value, (int, float)):
                cell.number_format = '#,##0.0'
            elif header in pct_columns and isinstance(value, (int, float)):
                cell.number_format = '0.0'

    end_row = start_row + len(rows)
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return end_row


def _autofit(ws) -> None:
    max_widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            max_widths[cell.column] = max(max_widths[cell.column], len(value))
    for col_index, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 12), 36)


def _normalize_products(products: Sequence[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for product in products:
        text = str(product or "").strip()
        if not text or text in seen:
            continue
        normalized.append(text)
        seen.add(text)
    if not normalized:
        raise ReportValidationError("请至少输入 1 个产品号。")
    if len(normalized) > DEFAULT_MAX_PRODUCTS:
        raise ReportValidationError(f"最多只支持 {DEFAULT_MAX_PRODUCTS} 个产品号。")
    return normalized


def _sheet_name_for_product(index: int, product: str) -> str:
    safe = "".join("_" if ch in '[]:*?/\\\\' else ch for ch in product)
    return f"{index}_{safe}"[:31]


def _build_output_path(output_dir: Path, output_name: str) -> Path:
    base_name = Path(output_name or DEFAULT_OUTPUT_NAME).name
    if not base_name.lower().endswith(".xlsx"):
        base_name = f"{base_name}.xlsx"
    stem = Path(base_name).stem or Path(DEFAULT_OUTPUT_NAME).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / f"{stem}_{timestamp}.xlsx"


def generate_modeb_customer_case_report(
    *,
    report_path: str | Path,
    products: Sequence[str],
    report_mode: str = "ModeB",
    output_dir: str | Path | None = None,
    output_name: str = DEFAULT_OUTPUT_NAME,
    runtime_paths: RuntimePaths | None = None,
    latest_report_path: str | Path | None = None,
) -> Path:
    if report_mode not in SUPPORTED_REPORT_MODES:
        raise ReportValidationError(f"不支持的报告模式：{report_mode}")
    normalized_products = _normalize_products(products)
    context = load_mode_report_context(report_path, expected_mode=report_mode)
    if runtime_paths is None:
        base_paths = resolve_runtime_paths()
        inferred_workspace_root = infer_workspace_root_from_report(context.report_path) or base_paths.user_workspace_dir
        resolved_paths = with_workspace_dir(base_paths, inferred_workspace_root)
    else:
        resolved_paths = runtime_paths
    supporting_input = load_supporting_input_data(context, resolved_paths)
    output_folder = Path(output_dir) if output_dir is not None else resolved_paths.outputs_dir
    output_folder.mkdir(parents=True, exist_ok=True)

    report_products = {str(row.get("Product") or "").strip() for row in context.detail_rows}
    missing_products = [product for product in normalized_products if product not in report_products]
    if missing_products:
        raise ReportValidationError(f"以下产品不在所选 {context.report_mode} 报告中：{', '.join(missing_products)}")

    metrics = _product_metrics(context.detail_rows)
    internal_totals = _internal_by_product_wc_and_source(context.detail_rows)
    wc_internal_totals = _wc_internal_totals(context.detail_rows)
    toller_lookup = _toller_wc_by_product(supporting_input.raw_routing_rows)
    outsourced_by_toller_wc = _outsourced_by_basis_toller_wc(context.detail_rows, toller_lookup)
    table_names = _TableNameFactory()

    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_ws = workbook.create_sheet("总览")
    _style_title(summary_ws, 1, f"{context.report_mode} 产品分析报告", 14)
    note_lines = [
        f"分析对象报告：{context.report_path}",
        f"场景：{context.scenario_name or '（未写入）'}",
        f"报告时间：{context.run_timestamp or '（未写入）'}",
        f"读取 load 数据目录：{supporting_input.load_folder}",
        f"读取 master 数据目录：{supporting_input.master_folder}",
    ]
    if latest_report_path:
        latest_path = Path(latest_report_path).expanduser().resolve()
        if latest_path != context.report_path:
            note_lines.append(f"注意：当前选择的不是最新 {context.report_mode} 报告。最新文件：{latest_path}")
    _style_note(summary_ws, 2, "\n".join(note_lines), 14)

    overview_rows = _overview_rows(
        context.report_mode,
        normalized_products,
        context.available_bases,
        supporting_input.loads,
        supporting_input.raw_capacity_rows,
        supporting_input.raw_routing_rows,
        metrics,
    )
    next_row = _write_section(summary_ws, 4, "产品总览", 14)
    next_row = _write_table(
        ws=summary_ws,
        start_row=next_row,
        rows=overview_rows,
        table_name=table_names.next("Summary"),
        ton_columns={
            "Demand_Tons",
            "Capacity_Base_Tons",
            "Routing_Reroute_Tons",
            "Outsourced_Tons",
            "Unmet_Tons",
        },
    ) + 2
    _write_section(summary_ws, next_row, "阅读方法", 14)
    summary_ws.cell(next_row + 1, 1).value = "1. 先看产品总览，确认该产品在当前模式下由 capacity、routing、外协、unmet 各承担了多少。"
    summary_ws.cell(next_row + 2, 1).value = "2. 再看单个产品页，确认需求来自哪个 planner 文件，以及 master_capacity / master_routing 给了哪些工作中心。"
    summary_ws.cell(next_row + 3, 1).value = "3. 最后看对应 Capacity_Basis 的工作中心上下文，理解该产品为什么会落到这些 WC，以及这些 WC 上还有哪些其他产品。"

    for sheet_index, product in enumerate(normalized_products, start=1):
        sheet_name = _sheet_name_for_product(sheet_index, product)
        ws = workbook.create_sheet(sheet_name)
        _style_title(ws, 1, f"{product} - {context.report_mode} 产品分析", 12)
        _style_note(
            ws,
            2,
            f"这张表把这个产品在 {context.report_mode} 报告里的结果，与对应的 planner 输入、master_capacity、master_routing 和工作中心上下文连在一起，方便直接解释。",
            12,
        )

        planner_rows = _planner_input_rows(supporting_input.loads, product)
        capacity_rows = _master_capacity_rows(supporting_input.raw_capacity_rows, product)
        routing_rows = _master_routing_rows(supporting_input.raw_routing_rows, product)

        current_row = _write_section(ws, 4, "产品摘要", 12)
        current_row = _write_table(
            ws=ws,
            start_row=current_row,
            rows=_basis_summary_rows(context.report_mode, product, context.available_bases, metrics),
            table_name=table_names.next(f"Summary{sheet_index}"),
            ton_columns={
                "Demand_Tons",
                "Capacity_Base_Tons",
                "Routing_Reroute_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
            },
        ) + 2

        current_row = _write_section(ws, current_row, "需求来源（Planner 输入）", 12)
        current_row = _write_table(
            ws=ws,
            start_row=current_row,
            rows=planner_rows or [{"Source_File": "（无）", "Planner_Name": "（无）", "Product_Family": "（无）", "Plant": "（无）", "WorkCenter": "（无）", "Scenario": context.scenario_name or "（未写入）", "Month_Span": "（无）", "Demand_Tons": 0.0, "Record_Count": 0}],
            table_name=table_names.next(f"Planner{sheet_index}"),
            ton_columns={"Demand_Tons"},
        ) + 2

        current_row = _write_section(ws, current_row, "这个产品用到的 master_capacity 行", 12)
        current_row = _write_table(
            ws=ws,
            start_row=current_row,
            rows=capacity_rows or [{
                "Source_File": "master_capacity.csv",
                "WorkCenter": "（无）",
                "Annual_Max_Capacity_Tons": 0.0,
                "Annual_Planned_Capacity_Tons": 0.0,
                "Monthly_Max_Capacity_Tons": 0.0,
                "Monthly_Planned_Capacity_Tons": 0.0,
                "Utilization_Target": "",
            }],
            table_name=table_names.next(f"Cap{sheet_index}"),
            ton_columns={
                "Annual_Max_Capacity_Tons",
                "Annual_Planned_Capacity_Tons",
                "Monthly_Max_Capacity_Tons",
                "Monthly_Planned_Capacity_Tons",
            },
        ) + 2

        current_row = _write_section(ws, current_row, "这个产品用到的 master_routing 行", 12)
        current_row = _write_table(
            ws=ws,
            start_row=current_row,
            rows=(
                routing_rows
                if context.report_mode == "ModeB"
                else [{"Source_File": "master_routing.csv", "WorkCenter": "（不适用）", "Route_Type": "ModeA 不使用 routing", "Max_Capacity_Tons": 0.0, "Planned_Capacity_Tons": 0.0, "Eligible_Flag": ""}]
            ) or [{"Source_File": "master_routing.csv", "WorkCenter": "（无）", "Route_Type": "没有 product-level routing", "Max_Capacity_Tons": 0.0, "Planned_Capacity_Tons": 0.0, "Eligible_Flag": ""}],
            table_name=table_names.next(f"Route{sheet_index}"),
            ton_columns={"Max_Capacity_Tons", "Planned_Capacity_Tons"},
        ) + 2

        for basis in context.available_bases:
            basis_metric = metrics[(basis, product)]
            current_row = _write_section(ws, current_row, f"{basis} 分配路径", 12)
            allocation_rows = _allocation_rows(
                basis=basis,
                product=product,
                internal_totals=internal_totals,
                demand_tons=basis_metric.demand,
            )
            current_row = _write_table(
                ws=ws,
                start_row=current_row,
                rows=allocation_rows or [{"Allocation_Source": "（无）", "WorkCenter": "（无）", "Internal_Tons": 0.0, "Share_of_Product_Demand_Pct": 0.0}],
                table_name=table_names.next(f"Alloc{sheet_index}{basis}"),
                ton_columns={"Internal_Tons"},
                pct_columns={"Share_of_Product_Demand_Pct"},
            ) + 2

            wc_roles = [(row["Allocation_Source"], row["WorkCenter"]) for row in allocation_rows]
            current_row = _write_section(ws, current_row, f"{basis} 工作中心中的其他产品", 12)
            current_row = _write_table(
                ws=ws,
                start_row=current_row,
                rows=_other_products_on_wc_rows(
                    basis=basis,
                    focus_product=product,
                    wc_roles=wc_roles,
                    wc_internal_totals=wc_internal_totals,
                )
                or [{
                    "WC_Role": "（无）",
                    "WorkCenter": "（无）",
                    "Focus_Product_Tons": 0.0,
                    "Total_WC_Internal_Tons": 0.0,
                    "Focus_Product_Share_of_WC_Pct": 0.0,
                    "Other_Product": "（无）",
                    "Other_Product_Internal_Tons": 0.0,
                }],
                table_name=table_names.next(f"WC{sheet_index}{basis}"),
                ton_columns={"Focus_Product_Tons", "Total_WC_Internal_Tons", "Other_Product_Internal_Tons"},
                pct_columns={"Focus_Product_Share_of_WC_Pct"},
            ) + 2

            if context.report_mode == "ModeB" and basis_metric.outsourced > 0:
                current_row = _write_section(ws, current_row, f"{basis} Toller 路径上的其他产品", 12)
                current_row = _write_table(
                    ws=ws,
                    start_row=current_row,
                    rows=_other_products_on_toller_rows(
                        basis=basis,
                        focus_product=product,
                        toller_wc=toller_lookup.get(product),
                        focus_outsourced_tons=basis_metric.outsourced,
                        outsourced_by_toller_wc=outsourced_by_toller_wc,
                    ) or [{
                        "WC_Role": _wc_role_label("ModeB_Toller_Route"),
                        "WorkCenter": toller_lookup.get(product) or "（无）",
                        "Focus_Product_Outsourced_Tons": basis_metric.outsourced,
                        "Other_Product": "（无）",
                        "Other_Product_Outsourced_Tons": 0.0,
                    }],
                    table_name=table_names.next(f"Tol{sheet_index}{basis}"),
                    ton_columns={"Focus_Product_Outsourced_Tons", "Other_Product_Outsourced_Tons"},
                ) + 2

            if context.report_mode == "ModeA":
                _style_note(
                    ws,
                    current_row,
                    f"{context.report_mode} 不使用 routing。{basis} 下所有内部吨位都直接来自 master_capacity；若仍有剩余需求，则最终进入 unmet。",
                    12,
                )
                current_row += 2
            elif not routing_rows:
                _style_note(
                    ws,
                    current_row,
                    f"{product} 在 master_routing.csv 里没有 product-level routing 定义，所以 {basis} 只能沿用 baseline capacity 的分配逻辑。",
                    12,
                )
                current_row += 2
            elif basis_metric.routing_reroute <= 0 and basis_metric.outsourced <= 0 and basis_metric.unmet > 0:
                _style_note(
                    ws,
                    current_row,
                    f"{basis} 下虽然定义了 routing，但相关工作中心在承接其他产品后没有足够余量，因此剩余需求仍然保留为 unmet。",
                    12,
                )
                current_row += 2

        _autofit(ws)

    _autofit(summary_ws)
    output_path = _build_output_path(output_folder, output_name)
    workbook.save(output_path)
    workbook.close()
    return output_path


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a product analysis workbook from an existing ModeA or ModeB output report.")
    parser.add_argument("--report", required=True, help="Path to the ModeA or ModeB output workbook.")
    parser.add_argument("--mode", choices=SUPPORTED_REPORT_MODES, default="ModeB")
    parser.add_argument("--product", action="append", dest="products", default=[], help="Product code to include. Repeat up to 10 times.")
    parser.add_argument("--output-dir", default=str(resolve_runtime_paths().outputs_dir))
    parser.add_argument("--output-name", default=DEFAULT_OUTPUT_NAME)
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    output_path = generate_modeb_customer_case_report(
        report_path=args.report,
        products=args.products,
        report_mode=args.mode,
        output_dir=args.output_dir,
        output_name=args.output_name,
    )
    print(f"产品分析 workbook 已生成：{output_path}")


if __name__ == "__main__":
    main()
