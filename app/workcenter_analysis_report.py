"""Generate workcenter analysis workbooks from existing ModeA or ModeB outputs."""
from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from numbers import Real
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.capacity_basis import MAX_BASIS, PLANNED_BASIS, normalize_capacity_basis
from app.data_loader import (
    load_direct_mode_a_with_capacity_bases,
    load_direct_mode_b_with_capacity_bases,
)
from app.i18n import localize_column_name, localize_sheet_name, localize_value
from app.models import CapacityRecord, LoadRecord
from app.runtime_paths import RuntimePaths, resolve_runtime_paths, with_workspace_dir


DEFAULT_OUTPUT_NAME = "workcenter_analysis.xlsx"
DEFAULT_MAX_WORKCENTERS = 10
REPORT_DATA_DECIMALS = 10
SUPPORTED_REPORT_MODES = ("ModeA", "ModeB")

THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=15, bold=True, color="1F1F1F")
SECTION_FONT = Font(size=11, bold=True, color="1F1F1F")


def _round_report_numeric(value: Any) -> Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, Real):
        return round(float(value), REPORT_DATA_DECIMALS)
    return value


DETAIL_CANONICAL_COLUMNS = (
    "Capacity_Basis",
    "Month",
    "PlannerName",
    "Product",
    "ProductFamily",
    "Plant",
    "AllocationType",
    "WorkCenter",
    "Demand_Tons",
    "Allocated_Tons",
    "Outsourced_Tons",
    "Unmet_Tons",
    "CapacityShare_Pct",
    "Allocation_Source",
    "Residual_After_Capacity_Tons",
    "Residual_After_Routing_Tons",
    "RouteType",
    "Priority",
    "Year",
)

UNMET_CANONICAL_COLUMNS = (
    "Capacity_Basis",
    "Month",
    "PlannerName",
    "Product",
    "ProductFamily",
    "Plant",
    "Owner_WorkCenter",
    "Capacity_Candidate_WorkCenters",
    "Attributed_WorkCenter",
    "Reference_Demand_Tons",
    "Product_Unmet_Tons",
    "Attributed_Unmet_Tons",
    "Attribution_Rule",
)

RUN_INFO_PARAMETERS = (
    "Scenario_Name",
    "Input_Load_Folder",
    "Input_Master_Folder",
    "Output_Folder",
    "Run_Timestamp",
    "Run_Mode",
)

TABLE_HEADER_LABELS = {
    "WorkCenter": "工作中心",
    "Report_Mode": "报告模式",
    "Capacity_Basis": "产能口径",
    "Capacity_Product_Count": "Capacity 产品数",
    "Capacity_Internal_Tons": "Capacity 内部吨位",
    "Routing_Internal_Product_Count": "Routing 内部产品数",
    "Routing_Internal_Tons": "Routing 内部吨位",
    "Routing_Outsourced_Product_Count": "Routing 外协产品数",
    "Routing_Outsourced_Tons": "Routing 外协吨位",
    "Unmet_Product_Count": "Unmet 产品数",
    "Unmet_Tons": "Unmet 回挂吨位",
    "Source_Table": "来源定义表",
    "Source_File": "来源文件",
    "Product": "产品",
    "Product_Family": "产品族",
    "Route_Type": "路径类型",
    "Eligible_Flag": "Eligible 标记",
    "Annual_Max_Capacity_Tons": "Annual Max 产能吨位",
    "Annual_Planned_Capacity_Tons": "Annual Planned 产能吨位",
    "Monthly_Max_Capacity_Tons": "Monthly Max 产能吨位",
    "Monthly_Planned_Capacity_Tons": "Monthly Planned 产能吨位",
    "Max_Capacity_Tons": "Max 路径吨位",
    "Planned_Capacity_Tons": "Planned 路径吨位",
    "Planner_Names": "计划员",
    "Planner_Source_Files": "需求来源文件",
    "Report_Table": "来源报表",
    "Covered_Months": "覆盖月份数",
    "Month_Span": "月份范围",
    "Monthly_Capacity_Tons": "月产能吨位",
    "Capacity_Window_Tons": "覆盖期产能吨位",
    "Filled_Tons": "填充吨位",
    "Share_of_Window_Capacity_Pct": "占覆盖期产能(%)",
    "Allocation_Logic": "分配逻辑",
    "Result_Type": "结果类型",
    "Monthly_Routing_Capacity_Tons": "月路径吨位",
    "Routing_Window_Tons": "覆盖期路径吨位",
    "Routing_Candidate_WorkCenters": "候选工作中心",
    "Attributed_Unmet_Tons": "回挂未满足吨位",
    "Attribution_Logic": "回挂逻辑",
}


BASIS_ALIASES = {
    MAX_BASIS: {MAX_BASIS, localize_value("zh", MAX_BASIS), "最大产能"},
    PLANNED_BASIS: {
        PLANNED_BASIS,
        "Planner",
        localize_value("zh", PLANNED_BASIS),
        localize_value("zh", "Planner"),
        "计划产能",
    },
}

VALUE_ALIASES = {
    "Internal": {localize_value("zh", "Internal"), "Internal", "内部"},
    "Outsourced": {localize_value("zh", "Outsourced"), "Outsourced", "外协"},
    "Unmet": {localize_value("zh", "Unmet"), "Unmet", "未满足"},
    "Capacity_Base": {localize_value("zh", "Capacity_Base"), "Capacity_Base", "基础产能分配"},
    "Routing_Reroute": {localize_value("zh", "Routing_Reroute"), "Routing_Reroute", "路径重分配"},
    "Toller": {localize_value("zh", "Toller"), "Toller", "外协路径"},
    "Capacity": {"Capacity", "产能分配"},
    "Primary": {"Primary", "主路径"},
    "Alternative": {"Alternative", "替代路径"},
    "N/A": {"N/A", "不适用"},
    "ModeA": {localize_value("zh", "ModeA"), "ModeA", "模式A"},
    "ModeB": {localize_value("zh", "ModeB"), "ModeB", "模式B"},
    "Both": {localize_value("zh", "Both"), "Both", "同时运行"},
}


class ReportValidationError(ValueError):
    """Raised when the selected workbook cannot be analyzed."""


@dataclass(frozen=True)
class ModeReportSelection:
    selected_path: Path
    latest_path: Path | None
    is_latest: bool


@dataclass(frozen=True)
class ModeReportContext:
    report_path: Path
    report_mode: str
    scenario_name: str
    run_timestamp: str | None
    input_load_folder: Path | None
    input_master_folder: Path | None
    output_folder: Path | None
    available_bases: tuple[str, ...]
    detail_rows: tuple[dict[str, Any], ...]
    unmet_rows: tuple[dict[str, Any], ...]


@dataclass(frozen=True)
class SupportingInputData:
    load_folder: Path
    master_folder: Path
    capacity_source_path: Path
    routing_source_path: Path | None
    loads: tuple[LoadRecord, ...]
    raw_capacity_rows: tuple[dict[str, Any], ...]
    raw_routing_rows: tuple[dict[str, Any], ...]
    capacity_monthly_by_basis: dict[str, dict[tuple[str, str], float]]
    routing_monthly_by_basis: dict[str, dict[tuple[str, str, str], float]]


class _TableNameFactory:
    def __init__(self) -> None:
        self._counter = 0

    def next(self, prefix: str) -> str:
        self._counter += 1
        safe_prefix = "".join(ch for ch in prefix if ch.isalnum()) or "Table"
        return f"Tbl{safe_prefix[:17]}{self._counter}"


def _label(header: str) -> str:
    return TABLE_HEADER_LABELS.get(header, header)


def _canonical_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    for canonical, aliases in VALUE_ALIASES.items():
        if text in aliases:
            return canonical
    return text


def _canonical_basis(value: Any) -> str:
    text = str(value or "").strip()
    for canonical, aliases in BASIS_ALIASES.items():
        if text in aliases:
            return canonical
    return normalize_capacity_basis(text or PLANNED_BASIS)


def _find_sheet_name(sheet_names: Sequence[str], canonical_name: str) -> str | None:
    candidates = {
        canonical_name,
        localize_sheet_name("zh", canonical_name),
    }
    for name in sheet_names:
        if name in candidates:
            return name
    return None


def _trim_headers(values: Sequence[Any]) -> list[str]:
    headers = [str(value).strip() if value is not None else "" for value in values]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def _detail_header_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for canonical in DETAIL_CANONICAL_COLUMNS:
        aliases[canonical] = canonical
        aliases[localize_column_name("zh", canonical)] = canonical
    return aliases


def _unmet_header_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for canonical in UNMET_CANONICAL_COLUMNS:
        aliases[canonical] = canonical
        aliases[localize_column_name("zh", canonical)] = canonical
    return aliases


def _run_info_alias_map() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for canonical in RUN_INFO_PARAMETERS:
        aliases[canonical] = canonical
        aliases[localize_column_name("zh", canonical)] = canonical
    return aliases


def _read_sheet_rows(
    workbook_path: Path,
    *,
    canonical_sheet_name: str,
    header_alias_map: dict[str, str],
) -> tuple[tuple[dict[str, Any], ...], tuple[str, ...]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_name = _find_sheet_name(workbook.sheetnames, canonical_sheet_name)
        if not sheet_name:
            if canonical_sheet_name == "Unmet_Attribution_Detail":
                return tuple(), tuple()
            raise ReportValidationError(f"所选文件缺少 {canonical_sheet_name} sheet。")
        ws = workbook[sheet_name]
        row_iter = ws.iter_rows(min_row=3, values_only=True)
        try:
            header_values = next(row_iter)
        except StopIteration as exc:
            raise ReportValidationError(f"{canonical_sheet_name} 为空，无法继续分析。") from exc
        raw_headers = _trim_headers(list(header_values))
        if not raw_headers:
            return tuple(), tuple()
        headers = [header_alias_map.get(header, header) for header in raw_headers]
        rows: list[dict[str, Any]] = []
        for values in row_iter:
            values = list(values[: len(headers)])
            if not any(value is not None and str(value).strip() != "" for value in values):
                continue
            row = {header: _canonical_value(value) for header, value in zip(headers, values)}
            if not str(row.get("Product", "")).strip() and canonical_sheet_name != "Run_Info":
                continue
            if "Capacity_Basis" in row:
                row["Capacity_Basis"] = _canonical_basis(row["Capacity_Basis"])
            rows.append(row)
        return tuple(rows), tuple(headers)
    finally:
        workbook.close()


def _read_run_info(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        run_info_sheet_name = _find_sheet_name(workbook.sheetnames, "Run_Info")
        if not run_info_sheet_name:
            return {}
        ws = workbook[run_info_sheet_name]
        row_iter = ws.iter_rows(values_only=True)
        try:
            first_row = next(row_iter)
        except StopIteration:
            return {}
        first_header = str((first_row[0] if first_row else "") or "").strip()
        basis_aware = first_header in {"Capacity_Basis", localize_column_name("zh", "Capacity_Basis")}
        aliases = _run_info_alias_map()
        info: dict[str, Any] = {}
        for row in row_iter:
            if basis_aware:
                basis = row[0] if len(row) > 0 else None
                parameter = row[1] if len(row) > 1 else None
                value = row[2] if len(row) > 2 else None
            else:
                basis = None
                parameter = row[0] if len(row) > 0 else None
                value = row[1] if len(row) > 1 else None
            if parameter in (None, ""):
                continue
            parameter_text = aliases.get(str(parameter).strip(), str(parameter).strip())
            if parameter_text and parameter_text not in info:
                info[parameter_text] = value
            if basis not in (None, "") and parameter_text:
                basis_key = _canonical_basis(basis)
                info.setdefault("by_basis", {}).setdefault(basis_key, {})[parameter_text] = value
        return info
    finally:
        workbook.close()


def _looks_like_mode_output_name(path: Path, report_mode: str) -> bool:
    name = path.name.lower()
    return name.startswith(f"capacity_result_{report_mode.lower()}_") and name.endswith(".xlsx")


def find_latest_mode_report(output_dir: str | Path, report_mode: str) -> Path | None:
    output_path = Path(output_dir)
    if not output_path.exists():
        return None
    candidates = [
        path
        for path in output_path.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and _looks_like_mode_output_name(path, report_mode)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def resolve_mode_report_selection(
    *,
    output_dir: str | Path,
    manual_report_path: str | Path | None,
    use_latest_report: bool,
    report_mode: str,
) -> ModeReportSelection:
    latest_path = find_latest_mode_report(output_dir, report_mode)
    if use_latest_report:
        if latest_path is None:
            raise FileNotFoundError(f"当前 output 目录下没有可用的 {report_mode} 报告。")
        return ModeReportSelection(selected_path=latest_path, latest_path=latest_path, is_latest=True)

    manual_text = str(manual_report_path or "").strip()
    if not manual_text:
        raise FileNotFoundError(f"请输入 {report_mode} 报告文件路径或文件名。")

    raw_path = Path(manual_text).expanduser()
    search_candidates = [raw_path]
    if not raw_path.is_absolute():
        search_candidates.insert(0, Path(output_dir) / raw_path)

    resolved_path: Path | None = None
    for candidate in search_candidates:
        if candidate.exists():
            resolved_path = candidate.resolve()
            break
    if resolved_path is None:
        raise FileNotFoundError(f"未找到指定的报告文件：{manual_text}")

    is_latest = latest_path is not None and resolved_path == latest_path.resolve()
    return ModeReportSelection(selected_path=resolved_path, latest_path=latest_path, is_latest=is_latest)


def _infer_report_mode_from_name(report_file: Path) -> str | None:
    name = report_file.name.lower()
    if name.startswith("capacity_result_modea_"):
        return "ModeA"
    if name.startswith("capacity_result_modeb_"):
        return "ModeB"
    if name.startswith("summary of mode a and mode b_"):
        return "Both"
    return None


def _infer_report_mode_from_detail_rows(detail_rows: Sequence[dict[str, Any]]) -> str:
    for row in detail_rows:
        source = str(row.get("Allocation_Source") or "").strip()
        route_type = str(row.get("RouteType") or "").strip()
        if source in {"Capacity_Base", "Routing_Reroute", "Toller"}:
            return "ModeB"
        if route_type in {"Primary", "Alternative", "Toller"}:
            return "ModeB"
    return "ModeA"


def load_mode_report_context(report_path: str | Path, *, expected_mode: str | None = None) -> ModeReportContext:
    report_file = Path(report_path).expanduser().resolve()
    if not report_file.exists():
        raise FileNotFoundError(f"报告文件不存在：{report_file}")
    if report_file.name.startswith("~$"):
        raise ReportValidationError("当前选择的是 Excel 临时锁文件，请选择正式的 .xlsx 报告。")

    detail_rows, detail_headers = _read_sheet_rows(
        report_file,
        canonical_sheet_name="Allocation_Detail",
        header_alias_map=_detail_header_alias_map(),
    )
    required_headers = {"Product", "AllocationType", "WorkCenter", "Demand_Tons"}
    if not required_headers.issubset(set(detail_headers)):
        raise ReportValidationError("所选文件不是有效的单模式结果报告：分配明细缺少关键列。")

    unmet_rows, _ = _read_sheet_rows(
        report_file,
        canonical_sheet_name="Unmet_Attribution_Detail",
        header_alias_map=_unmet_header_alias_map(),
    )
    run_info = _read_run_info(report_file)

    actual_mode = _infer_report_mode_from_name(report_file) or _infer_report_mode_from_detail_rows(detail_rows)
    if actual_mode == "Both":
        raise ReportValidationError("该工具只支持 ModeA 或 ModeB 单报告，不支持 Both 汇总报告。")
    if expected_mode and actual_mode != expected_mode:
        raise ReportValidationError(f"当前文件属于 {actual_mode}，与所选的 {expected_mode} 不一致。")

    available_bases = sorted(
        {
            _canonical_basis(row.get("Capacity_Basis"))
            for row in [*detail_rows, *unmet_rows]
            if str(row.get("Capacity_Basis") or "").strip()
        },
        key=lambda value: (value != MAX_BASIS, value.casefold()),
    )
    if not available_bases:
        available_bases = [PLANNED_BASIS]

    return ModeReportContext(
        report_path=report_file,
        report_mode=actual_mode,
        scenario_name=str(run_info.get("Scenario_Name") or "").strip(),
        run_timestamp=str(run_info.get("Run_Timestamp") or "").strip() or None,
        input_load_folder=Path(run_info["Input_Load_Folder"]).resolve() if run_info.get("Input_Load_Folder") else None,
        input_master_folder=Path(run_info["Input_Master_Folder"]).resolve() if run_info.get("Input_Master_Folder") else None,
        output_folder=Path(run_info["Output_Folder"]).resolve() if run_info.get("Output_Folder") else None,
        available_bases=tuple(available_bases),
        detail_rows=detail_rows,
        unmet_rows=unmet_rows,
    )


def infer_workspace_root_from_report(report_path: str | Path) -> Path | None:
    report_file = Path(report_path).expanduser().resolve()
    parent = report_file.parent
    if parent.name.lower() == "output":
        return parent.parent.resolve()
    return None


def _resolve_input_folder(candidate: Path | None, fallback: Path) -> Path:
    if candidate and candidate.exists():
        return candidate
    return fallback


def _canonical_scenario_name(value: str | None) -> str:
    text = str(value or "").strip()
    mapping = {
        "基准": "Baseline",
        "Baseline": "Baseline",
        "扩展": "Expansion",
        "Expansion": "Expansion",
        "精益": "Lean",
        "Lean": "Lean",
    }
    return mapping.get(text, text or "Baseline")


def _find_tabular_file(folder: Path, stem: str, *, required: bool = True) -> Path | None:
    for ext in (".xlsx", ".xls", ".csv"):
        path = folder / f"{stem}{ext}"
        if path.exists():
            return path
    if required:
        raise FileNotFoundError(f"未找到 {stem} 文件：{folder}")
    return None


def _read_tabular_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(path, header=0)
    elif suffix == ".csv":
        df = pd.read_csv(path, encoding="utf-8-sig")
    else:
        raise ValueError(f"不支持的文件类型：{path}")
    df.columns = [str(column).strip() for column in df.columns]
    return df.fillna("").to_dict(orient="records")


def _capacity_monthly_by_basis(
    capacity_bases: dict[str, Sequence[CapacityRecord]],
) -> dict[str, dict[tuple[str, str], float]]:
    payload: dict[str, dict[tuple[str, str], float]] = {}
    for basis, records in capacity_bases.items():
        payload[basis] = {
            (record.product, record.work_center): float(record.monthly_capacity_tons)
            for record in records
        }
    return payload


def _routing_monthly_by_basis(raw_rows: Sequence[dict[str, Any]]) -> dict[str, dict[tuple[str, str, str], float]]:
    payload = {
        MAX_BASIS: {},
        PLANNED_BASIS: {},
    }
    for row in raw_rows:
        product = str(row.get("Product") or "").strip()
        work_center = str(row.get("WorkCenter") or row.get("Resource") or "").strip()
        route_type = str(row.get("Router Type") or row.get("Route Type") or row.get("Route_Type") or "").strip()
        if not product or not work_center:
            continue
        max_tons = _to_float(row.get("Max Capacity Ton"))
        planned_tons = _to_float(row.get("Planned Capacity Ton") or row.get("Planner Capacity Ton"))
        payload[MAX_BASIS][(product, work_center, route_type or "Primary")] = max_tons
        payload[PLANNED_BASIS][(product, work_center, route_type or "Primary")] = planned_tons
    return payload


def load_supporting_input_data(
    context: ModeReportContext,
    runtime_paths: RuntimePaths | None = None,
) -> SupportingInputData:
    resolved_paths = runtime_paths or resolve_runtime_paths()
    default_input_dir = resolved_paths.workspace_input_dir
    load_folder = _resolve_input_folder(context.input_load_folder, default_input_dir)
    master_folder = _resolve_input_folder(context.input_master_folder, default_input_dir)
    scenario_name = _canonical_scenario_name(context.scenario_name)

    if context.report_mode == "ModeA":
        loads, capacity_bases, _ = load_direct_mode_a_with_capacity_bases(
            str(load_folder),
            str(master_folder),
            selected_scenario=scenario_name,
        )
        raw_routing_rows: list[dict[str, Any]] = []
    else:
        loads, baseline_capacity_bases, _merged_capacities, _routings = load_direct_mode_b_with_capacity_bases(
            str(load_folder),
            str(master_folder),
            selected_scenario=scenario_name,
        )
        capacity_bases = baseline_capacity_bases
        routing_path = _find_tabular_file(master_folder, "master_routing", required=True)
        raw_routing_rows = _read_tabular_rows(routing_path)

    capacity_path = _find_tabular_file(master_folder, "master_capacity", required=True)
    raw_capacity_rows = _read_tabular_rows(capacity_path)
    routing_path = _find_tabular_file(master_folder, "master_routing", required=False)

    return SupportingInputData(
        load_folder=load_folder,
        master_folder=master_folder,
        capacity_source_path=capacity_path,
        routing_source_path=routing_path,
        loads=tuple(loads),
        raw_capacity_rows=tuple(raw_capacity_rows),
        raw_routing_rows=tuple(raw_routing_rows),
        capacity_monthly_by_basis=_capacity_monthly_by_basis(capacity_bases),
        routing_monthly_by_basis=_routing_monthly_by_basis(raw_routing_rows),
    )


def _to_float(value: Any) -> float:
    try:
        text = str(value).strip()
        if not text:
            return 0.0
        return float(text)
    except Exception:
        return 0.0


def _normalize_workcenters(workcenters: Sequence[str]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for work_center in workcenters:
        text = str(work_center or "").strip()
        if not text or text in seen:
            continue
        normalized.append(text)
        seen.add(text)
    if not normalized:
        raise ReportValidationError("请至少输入 1 个工作中心。")
    if len(normalized) > DEFAULT_MAX_WORKCENTERS:
        raise ReportValidationError(f"最多只支持 {DEFAULT_MAX_WORKCENTERS} 个工作中心。")
    return normalized


def _basis_key(row: dict[str, Any]) -> str:
    return _canonical_basis(row.get("Capacity_Basis") or PLANNED_BASIS)


def _sheet_name_for_workcenter(index: int, work_center: str) -> str:
    safe = "".join("_" if ch in '[]:*?/\\\\' else ch for ch in work_center)
    return f"{index}_{safe}"[:31]


def _build_output_path(output_dir: Path, output_name: str) -> Path:
    base_name = Path(output_name or DEFAULT_OUTPUT_NAME).name
    if not base_name.lower().endswith(".xlsx"):
        base_name = f"{base_name}.xlsx"
    stem = Path(base_name).stem or Path(DEFAULT_OUTPUT_NAME).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_dir / f"{stem}_{timestamp}.xlsx"


def _month_range(months: Iterable[str]) -> str:
    ordered = sorted({str(month).strip() for month in months if str(month).strip()})
    if not ordered:
        return ""
    return ordered[0] if len(ordered) == 1 else f"{ordered[0]} ~ {ordered[-1]}"


def _unique_join(values: Iterable[str]) -> str:
    ordered = sorted({str(value).strip() for value in values if str(value).strip()}, key=str.casefold)
    return " | ".join(ordered)


def _planner_sources_by_product(loads: Sequence[LoadRecord]) -> dict[str, dict[str, set[str]]]:
    payload: dict[str, dict[str, set[str]]] = defaultdict(lambda: {"planner_names": set(), "source_files": set()})
    for load in loads:
        entry = payload[load.product]
        if load.planner_name:
            entry["planner_names"].add(load.planner_name)
        if load.source_file:
            entry["source_files"].add(load.source_file)
    return payload


def _capacity_definition_rows(
    raw_capacity_rows: Sequence[dict[str, Any]],
    work_center: str,
    *,
    source_name: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in raw_capacity_rows:
        resource = str(row.get("WorkCenter") or row.get("Resource") or "").strip()
        if resource != work_center:
            continue
        annual_max = _to_float(row.get("Annual Max Capacity Tons") or row.get("Annual_Max_Capacity_Tons"))
        annual_planned = _to_float(
            row.get("Annual Planned Capacity Tons")
            or row.get("Annual Planner Capacity Tons")
            or row.get("Annual_Planned_Capacity_Tons")
            or row.get("Annual_Planner_Capacity_Tons")
        )
        rows.append(
            {
                "Source_Table": source_name,
                "Product": str(row.get("Product") or "").strip(),
                "Product_Family": str(row.get("Product Family") or row.get("ProductFamily") or "").strip(),
                "WorkCenter": resource,
                "Annual_Max_Capacity_Tons": annual_max,
                "Annual_Planned_Capacity_Tons": annual_planned,
                "Monthly_Max_Capacity_Tons": annual_max / 12.0,
                "Monthly_Planned_Capacity_Tons": annual_planned / 12.0,
            }
        )
    return rows


def _routing_definition_rows(
    raw_routing_rows: Sequence[dict[str, Any]],
    work_center: str,
    *,
    source_name: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in raw_routing_rows:
        resource = str(row.get("WorkCenter") or row.get("Resource") or "").strip()
        if resource != work_center:
            continue
        rows.append(
            {
                "Source_Table": source_name,
                "Product": str(row.get("Product") or "").strip(),
                "Product_Family": str(row.get("Product Family") or row.get("ProductFamily") or "").strip(),
                "WorkCenter": resource,
                "Route_Type": str(row.get("Router Type") or row.get("Route Type") or row.get("Route_Type") or "").strip(),
                "Eligible_Flag": str(row.get("EligibleFalg") or row.get("Eligible Flag") or row.get("Eligible_Flag") or "").strip(),
                "Max_Capacity_Tons": _to_float(row.get("Max Capacity Ton")),
                "Planned_Capacity_Tons": _to_float(row.get("Planned Capacity Ton") or row.get("Planner Capacity Ton")),
            }
        )
    return rows


def _capacity_logic(report_mode: str) -> str:
    if report_mode == "ModeA":
        return "ModeA 直接按 master_capacity 的产品-工作中心产能做内部填充，不使用 routing。"
    return "ModeB Stage 1 先按 master_capacity 的基础产能做内部填充。"


def _routing_logic(allocation_source: str) -> str:
    if allocation_source == "Routing_Reroute":
        return "ModeB Stage 2 把 Capacity 后剩余需求按 master_routing 的 product-level routing 重分配到该工作中心。"
    return "ModeB Stage 3 若 master_routing 定义了 Toller 路径，则把剩余需求转到该外协工作中心。"


def _unmet_logic(report_mode: str) -> str:
    if report_mode == "ModeA":
        return "最终 unmet 按 planner resource_group_owner 回挂到该工作中心。"
    return "最终 unmet 按基础产能工作中心回挂到该工作中心。"


def _capacity_fill_rows(
    *,
    report_mode: str,
    basis: str,
    work_center: str,
    detail_rows: Sequence[dict[str, Any]],
    planner_lookup: dict[str, dict[str, set[str]]],
    capacity_monthly_by_basis: dict[str, dict[tuple[str, str], float]],
    source_name: str,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        if _basis_key(row) != basis:
            continue
        if str(row.get("AllocationType") or "").strip() != "Internal":
            continue
        if str(row.get("WorkCenter") or "").strip() != work_center:
            continue
        source = str(row.get("Allocation_Source") or "").strip()
        if report_mode == "ModeB" and source != "Capacity_Base":
            continue
        grouped[str(row.get("Product") or "").strip()].append(row)

    rows: list[dict[str, Any]] = []
    for product, product_rows in sorted(grouped.items(), key=lambda item: item[0].casefold()):
        months = {str(row.get("Month") or "").strip() for row in product_rows}
        planners = {str(row.get("PlannerName") or "").strip() for row in product_rows if str(row.get("PlannerName") or "").strip()}
        monthly_capacity = capacity_monthly_by_basis.get(basis, {}).get((product, work_center), 0.0)
        filled_tons = sum(_to_float(row.get("Allocated_Tons")) for row in product_rows)
        window_tons = monthly_capacity * len(months)
        lookup = planner_lookup.get(product, {"planner_names": set(), "source_files": set()})
        rows.append(
            {
                "Product": product,
                "Product_Family": str(product_rows[0].get("ProductFamily") or "").strip(),
                "Planner_Names": _unique_join(planners or lookup["planner_names"]),
                "Planner_Source_Files": _unique_join(lookup["source_files"]),
                "Report_Table": "Allocation_Detail",
                "Source_Table": source_name,
                "Covered_Months": len(months),
                "Month_Span": _month_range(months),
                "Monthly_Capacity_Tons": monthly_capacity,
                "Capacity_Window_Tons": window_tons,
                "Filled_Tons": filled_tons,
                "Share_of_Window_Capacity_Pct": _pct(filled_tons, window_tons),
                "Allocation_Logic": _capacity_logic(report_mode),
            }
        )
    return rows


def _routing_fill_rows(
    *,
    basis: str,
    work_center: str,
    detail_rows: Sequence[dict[str, Any]],
    planner_lookup: dict[str, dict[str, set[str]]],
    routing_monthly_by_basis: dict[str, dict[tuple[str, str, str], float]],
    source_name: str,
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        if _basis_key(row) != basis:
            continue
        if str(row.get("WorkCenter") or "").strip() != work_center:
            continue
        allocation_type = str(row.get("AllocationType") or "").strip()
        source = str(row.get("Allocation_Source") or "").strip()
        route_type = str(row.get("RouteType") or "").strip() or "Primary"
        if allocation_type == "Internal" and source == "Routing_Reroute":
            grouped[(str(row.get("Product") or "").strip(), allocation_type, route_type)].append(row)
        elif allocation_type == "Outsourced":
            grouped[(str(row.get("Product") or "").strip(), allocation_type, route_type)].append(row)

    rows: list[dict[str, Any]] = []
    for (product, allocation_type, route_type), product_rows in sorted(grouped.items(), key=lambda item: (item[0][0].casefold(), item[0][1], item[0][2])):
        months = {str(row.get("Month") or "").strip() for row in product_rows}
        planners = {str(row.get("PlannerName") or "").strip() for row in product_rows if str(row.get("PlannerName") or "").strip()}
        monthly_capacity = routing_monthly_by_basis.get(basis, {}).get((product, work_center, route_type), 0.0)
        if monthly_capacity <= 0:
            monthly_capacity = routing_monthly_by_basis.get(basis, {}).get((product, work_center, "Primary"), 0.0)
        filled_tons = (
            sum(_to_float(row.get("Allocated_Tons")) for row in product_rows)
            if allocation_type == "Internal"
            else sum(_to_float(row.get("Outsourced_Tons")) for row in product_rows)
        )
        window_tons = monthly_capacity * len(months)
        lookup = planner_lookup.get(product, {"planner_names": set(), "source_files": set()})
        rows.append(
            {
                "Product": product,
                "Product_Family": str(product_rows[0].get("ProductFamily") or "").strip(),
                "Planner_Names": _unique_join(planners or lookup["planner_names"]),
                "Planner_Source_Files": _unique_join(lookup["source_files"]),
                "Report_Table": "Allocation_Detail",
                "Source_Table": source_name,
                "Result_Type": allocation_type,
                "Route_Type": route_type,
                "Covered_Months": len(months),
                "Month_Span": _month_range(months),
                "Monthly_Routing_Capacity_Tons": monthly_capacity,
                "Routing_Window_Tons": window_tons,
                "Filled_Tons": filled_tons,
                "Share_of_Window_Capacity_Pct": _pct(filled_tons, window_tons),
                "Allocation_Logic": _routing_logic(
                    "Routing_Reroute" if allocation_type == "Internal" else "Toller"
                ),
            }
        )
    return rows


def _unmet_fill_rows(
    *,
    report_mode: str,
    basis: str,
    work_center: str,
    unmet_rows: Sequence[dict[str, Any]],
    planner_lookup: dict[str, dict[str, set[str]]],
    capacity_monthly_by_basis: dict[str, dict[tuple[str, str], float]],
    source_name: str,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in unmet_rows:
        if _basis_key(row) != basis:
            continue
        if str(row.get("Attributed_WorkCenter") or "").strip() != work_center:
            continue
        grouped[str(row.get("Product") or "").strip()].append(row)

    rows: list[dict[str, Any]] = []
    for product, product_rows in sorted(grouped.items(), key=lambda item: item[0].casefold()):
        months = {str(row.get("Month") or "").strip() for row in product_rows}
        planners = {str(row.get("PlannerName") or "").strip() for row in product_rows if str(row.get("PlannerName") or "").strip()}
        monthly_capacity = capacity_monthly_by_basis.get(basis, {}).get((product, work_center), 0.0)
        unmet_tons = sum(_to_float(row.get("Attributed_Unmet_Tons")) for row in product_rows)
        window_tons = monthly_capacity * len(months)
        lookup = planner_lookup.get(product, {"planner_names": set(), "source_files": set()})
        rows.append(
            {
                "Product": product,
                "Product_Family": str(product_rows[0].get("ProductFamily") or "").strip(),
                "Planner_Names": _unique_join(planners or lookup["planner_names"]),
                "Planner_Source_Files": _unique_join(lookup["source_files"]),
                "Report_Table": "Unmet_Attribution_Detail",
                "Source_Table": source_name,
                "Routing_Candidate_WorkCenters": _unique_join(
                    candidate
                    for row in product_rows
                    for candidate in str(row.get("Capacity_Candidate_WorkCenters") or "").split("|")
                ),
                "Covered_Months": len(months),
                "Month_Span": _month_range(months),
                "Monthly_Capacity_Tons": monthly_capacity,
                "Capacity_Window_Tons": window_tons,
                "Attributed_Unmet_Tons": unmet_tons,
                "Share_of_Window_Capacity_Pct": _pct(unmet_tons, window_tons),
                "Attribution_Logic": _unmet_logic(report_mode),
            }
        )
    return rows


def _pct(numerator: float, denominator: float) -> float:
    if abs(denominator) <= 1e-9:
        return 0.0
    return 100.0 * numerator / denominator


def _summary_row(
    *,
    work_center: str,
    report_mode: str,
    basis: str,
    capacity_rows: Sequence[dict[str, Any]],
    routing_rows: Sequence[dict[str, Any]],
    unmet_rows: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    routing_internal_rows = [row for row in routing_rows if row.get("Result_Type") == "Internal"]
    routing_outsourced_rows = [row for row in routing_rows if row.get("Result_Type") == "Outsourced"]
    return {
        "WorkCenter": work_center,
        "Report_Mode": report_mode,
        "Capacity_Basis": basis,
        "Capacity_Product_Count": len(capacity_rows),
        "Capacity_Internal_Tons": _round_report_numeric(sum(_to_float(row.get("Filled_Tons")) for row in capacity_rows)),
        "Routing_Internal_Product_Count": len(routing_internal_rows),
        "Routing_Internal_Tons": _round_report_numeric(sum(_to_float(row.get("Filled_Tons")) for row in routing_internal_rows)),
        "Routing_Outsourced_Product_Count": len(routing_outsourced_rows),
        "Routing_Outsourced_Tons": _round_report_numeric(sum(_to_float(row.get("Filled_Tons")) for row in routing_outsourced_rows)),
        "Unmet_Product_Count": len(unmet_rows),
        "Unmet_Tons": _round_report_numeric(sum(_to_float(row.get("Attributed_Unmet_Tons")) for row in unmet_rows)),
    }


def _style_title(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_note(ws, row: int, text: str, end_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = NOTE_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_section(ws, row: int, text: str, end_col: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1)
    cell.value = text
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_table(
    *,
    ws,
    start_row: int,
    rows: Sequence[dict[str, Any]],
    table_name: str,
    pct_columns: Iterable[str] | None = None,
    ton_columns: Iterable[str] | None = None,
) -> int:
    if not rows:
        ws.cell(start_row, 1).value = "（无）"
        return start_row

    pct_columns = set(pct_columns or [])
    ton_columns = set(ton_columns or [])
    headers = list(rows[0].keys())

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_index)
        cell.value = _label(header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for row_offset, row_data in enumerate(rows, start=1):
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(start_row + row_offset, col_index)
            value = row_data.get(header)
            stored_value = _round_report_numeric(value)
            cell.value = stored_value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if header in ton_columns and isinstance(stored_value, (int, float)):
                cell.number_format = "#,##0.0"
            elif header in pct_columns and isinstance(stored_value, (int, float)):
                cell.number_format = "0.0"

    end_row = start_row + len(rows)
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return end_row


def _autofit(ws) -> None:
    max_widths: dict[int, int] = defaultdict(int)
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            max_widths[cell.column] = max(max_widths[cell.column], len(value))
    for col_index, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 12), 42)


def generate_workcenter_analysis_report(
    *,
    report_path: str | Path,
    workcenters: Sequence[str],
    report_mode: str,
    output_dir: str | Path | None = None,
    output_name: str = DEFAULT_OUTPUT_NAME,
    runtime_paths: RuntimePaths | None = None,
    latest_report_path: str | Path | None = None,
) -> Path:
    if report_mode not in SUPPORTED_REPORT_MODES:
        raise ReportValidationError(f"不支持的报告模式：{report_mode}")
    normalized_workcenters = _normalize_workcenters(workcenters)
    context = load_mode_report_context(report_path, expected_mode=report_mode)
    if runtime_paths is None:
        base_paths = resolve_runtime_paths()
        inferred_workspace_root = infer_workspace_root_from_report(context.report_path) or base_paths.user_workspace_dir
        resolved_paths = with_workspace_dir(base_paths, inferred_workspace_root)
    else:
        resolved_paths = runtime_paths
    supporting_input = load_supporting_input_data(context, resolved_paths)
    output_folder = Path(output_dir) if output_dir is not None else resolved_paths.outputs_dir
    output_folder.mkdir(parents=True, exist_ok=True)

    planner_lookup = _planner_sources_by_product(supporting_input.loads)
    capacity_source_name = supporting_input.capacity_source_path.name
    routing_source_name = supporting_input.routing_source_path.name if supporting_input.routing_source_path else "master_routing"

    workbook = Workbook()
    workbook.remove(workbook.active)
    table_names = _TableNameFactory()

    summary_ws = workbook.create_sheet("总览")
    _style_title(summary_ws, 1, f"{context.report_mode} 工作中心分析报告", 14)
    note_lines = [
        f"分析对象报告：{context.report_path}",
        f"场景：{context.scenario_name or '（未写入）'}",
        f"报告时间：{context.run_timestamp or '（未写入）'}",
        f"读取 load 数据目录：{supporting_input.load_folder}",
        f"读取 master 数据目录：{supporting_input.master_folder}",
        "占比产能 = 该工作中心/产品在覆盖月份中的吨位合计 ÷ (定义月产能 × 覆盖月份数)。",
    ]
    if latest_report_path:
        latest_path = Path(latest_report_path).expanduser().resolve()
        if latest_path != context.report_path:
            note_lines.append(f"注意：当前选择的不是最新 {context.report_mode} 报告。最新文件：{latest_path}")
    _style_note(summary_ws, 2, "\n".join(note_lines), 14)

    summary_rows: list[dict[str, Any]] = []
    for work_center in normalized_workcenters:
        for basis in context.available_bases:
            capacity_rows = _capacity_fill_rows(
                report_mode=context.report_mode,
                basis=basis,
                work_center=work_center,
                detail_rows=context.detail_rows,
                planner_lookup=planner_lookup,
                capacity_monthly_by_basis=supporting_input.capacity_monthly_by_basis,
                source_name=capacity_source_name,
            )
            routing_rows = (
                _routing_fill_rows(
                    basis=basis,
                    work_center=work_center,
                    detail_rows=context.detail_rows,
                    planner_lookup=planner_lookup,
                    routing_monthly_by_basis=supporting_input.routing_monthly_by_basis,
                    source_name=routing_source_name,
                )
                if context.report_mode == "ModeB"
                else []
            )
            unmet_rows = _unmet_fill_rows(
                report_mode=context.report_mode,
                basis=basis,
                work_center=work_center,
                unmet_rows=context.unmet_rows,
                planner_lookup=planner_lookup,
                capacity_monthly_by_basis=supporting_input.capacity_monthly_by_basis,
                source_name=capacity_source_name,
            )
            summary_rows.append(
                _summary_row(
                    work_center=work_center,
                    report_mode=context.report_mode,
                    basis=basis,
                    capacity_rows=capacity_rows,
                    routing_rows=routing_rows,
                    unmet_rows=unmet_rows,
                )
            )

    next_row = _write_section(summary_ws, 4, "工作中心总览", 14)
    next_row = _write_table(
        ws=summary_ws,
        start_row=next_row,
        rows=summary_rows,
        table_name=table_names.next("Summary"),
        ton_columns={
            "Capacity_Internal_Tons",
            "Routing_Internal_Tons",
            "Routing_Outsourced_Tons",
            "Unmet_Tons",
        },
    ) + 2
    _write_section(summary_ws, next_row, "阅读方法", 14)
    summary_ws.cell(next_row + 1, 1).value = "1. 先看工作中心总览，确认该 WC 在 Capacity、Routing、Unmet 三条路径上的总量。"
    summary_ws.cell(next_row + 2, 1).value = "2. 再看单个工作中心页，先看该 WC 在 master_capacity / master_routing 里的输入定义。"
    summary_ws.cell(next_row + 3, 1).value = "3. 然后按 Capacity、Routing、Unmet 三段查看有哪些产品进入该 WC、对应吨位和分配逻辑。"

    for index, work_center in enumerate(normalized_workcenters, start=1):
        ws = workbook.create_sheet(_sheet_name_for_workcenter(index, work_center))
        _style_title(ws, 1, f"{work_center} - {context.report_mode} 工作中心分析", 14)
        _style_note(
            ws,
            2,
            "本页按工作中心视角拆解该 WC 在报告中的 Capacity 填充、Routing 填充和最终 Unmet 回挂。",
            14,
        )
        current_row = _write_section(ws, 4, "输入定义", 14)
        capacity_def_rows = _capacity_definition_rows(
            supporting_input.raw_capacity_rows,
            work_center,
            source_name=capacity_source_name,
        )
        current_row = _write_table(
            ws=ws,
            start_row=current_row,
            rows=capacity_def_rows or [{
                "Source_Table": capacity_source_name,
                "Product": "（无）",
                "Product_Family": "",
                "WorkCenter": work_center,
                "Annual_Max_Capacity_Tons": 0.0,
                "Annual_Planned_Capacity_Tons": 0.0,
                "Monthly_Max_Capacity_Tons": 0.0,
                "Monthly_Planned_Capacity_Tons": 0.0,
            }],
            table_name=table_names.next(f"CapDef{index}"),
            ton_columns={
                "Annual_Max_Capacity_Tons",
                "Annual_Planned_Capacity_Tons",
                "Monthly_Max_Capacity_Tons",
                "Monthly_Planned_Capacity_Tons",
            },
        ) + 2

        routing_def_rows = _routing_definition_rows(
            supporting_input.raw_routing_rows,
            work_center,
            source_name=routing_source_name,
        )
        current_row = _write_section(ws, current_row, "Routing 定义", 14)
        current_row = _write_table(
            ws=ws,
            start_row=current_row,
            rows=routing_def_rows or [{
                "Source_Table": routing_source_name,
                "Product": "（无）",
                "Product_Family": "",
                "WorkCenter": work_center,
                "Route_Type": "",
                "Eligible_Flag": "",
                "Max_Capacity_Tons": 0.0,
                "Planned_Capacity_Tons": 0.0,
            }],
            table_name=table_names.next(f"RouteDef{index}"),
            ton_columns={"Max_Capacity_Tons", "Planned_Capacity_Tons"},
        ) + 2

        for basis in context.available_bases:
            capacity_rows = _capacity_fill_rows(
                report_mode=context.report_mode,
                basis=basis,
                work_center=work_center,
                detail_rows=context.detail_rows,
                planner_lookup=planner_lookup,
                capacity_monthly_by_basis=supporting_input.capacity_monthly_by_basis,
                source_name=capacity_source_name,
            )
            routing_rows = (
                _routing_fill_rows(
                    basis=basis,
                    work_center=work_center,
                    detail_rows=context.detail_rows,
                    planner_lookup=planner_lookup,
                    routing_monthly_by_basis=supporting_input.routing_monthly_by_basis,
                    source_name=routing_source_name,
                )
                if context.report_mode == "ModeB"
                else []
            )
            unmet_rows = _unmet_fill_rows(
                report_mode=context.report_mode,
                basis=basis,
                work_center=work_center,
                unmet_rows=context.unmet_rows,
                planner_lookup=planner_lookup,
                capacity_monthly_by_basis=supporting_input.capacity_monthly_by_basis,
                source_name=capacity_source_name,
            )

            current_row = _write_section(ws, current_row, f"{basis} - Capacity 内部填充", 14)
            current_row = _write_table(
                ws=ws,
                start_row=current_row,
                rows=capacity_rows or [{
                    "Product": "（无）",
                    "Product_Family": "",
                    "Planner_Names": "",
                    "Planner_Source_Files": "",
                    "Report_Table": "Allocation_Detail",
                    "Source_Table": capacity_source_name,
                    "Covered_Months": 0,
                    "Month_Span": "",
                    "Monthly_Capacity_Tons": 0.0,
                    "Capacity_Window_Tons": 0.0,
                    "Filled_Tons": 0.0,
                    "Share_of_Window_Capacity_Pct": 0.0,
                    "Allocation_Logic": _capacity_logic(context.report_mode),
                }],
                table_name=table_names.next(f"CapFill{index}{basis}"),
                ton_columns={"Monthly_Capacity_Tons", "Capacity_Window_Tons", "Filled_Tons"},
                pct_columns={"Share_of_Window_Capacity_Pct"},
            ) + 2

            current_row = _write_section(ws, current_row, f"{basis} - Routing 填充", 14)
            routing_placeholder = [{
                "Product": "（无）",
                "Product_Family": "",
                "Planner_Names": "",
                "Planner_Source_Files": "",
                "Report_Table": "Allocation_Detail",
                "Source_Table": routing_source_name,
                "Result_Type": "",
                "Route_Type": "",
                "Covered_Months": 0,
                "Month_Span": "",
                "Monthly_Routing_Capacity_Tons": 0.0,
                "Routing_Window_Tons": 0.0,
                "Filled_Tons": 0.0,
                "Share_of_Window_Capacity_Pct": 0.0,
                "Allocation_Logic": "ModeA 不使用 routing。" if context.report_mode == "ModeA" else _routing_logic("Routing_Reroute"),
            }]
            current_row = _write_table(
                ws=ws,
                start_row=current_row,
                rows=routing_rows or routing_placeholder,
                table_name=table_names.next(f"RouteFill{index}{basis}"),
                ton_columns={"Monthly_Routing_Capacity_Tons", "Routing_Window_Tons", "Filled_Tons"},
                pct_columns={"Share_of_Window_Capacity_Pct"},
            ) + 2

            current_row = _write_section(ws, current_row, f"{basis} - Unmet 回挂", 14)
            current_row = _write_table(
                ws=ws,
                start_row=current_row,
                rows=unmet_rows or [{
                    "Product": "（无）",
                    "Product_Family": "",
                    "Planner_Names": "",
                    "Planner_Source_Files": "",
                    "Report_Table": "Unmet_Attribution_Detail",
                    "Source_Table": capacity_source_name,
                    "Routing_Candidate_WorkCenters": "",
                    "Covered_Months": 0,
                    "Month_Span": "",
                    "Monthly_Capacity_Tons": 0.0,
                    "Capacity_Window_Tons": 0.0,
                    "Attributed_Unmet_Tons": 0.0,
                    "Share_of_Window_Capacity_Pct": 0.0,
                    "Attribution_Logic": _unmet_logic(context.report_mode),
                }],
                table_name=table_names.next(f"Unmet{index}{basis}"),
                ton_columns={"Monthly_Capacity_Tons", "Capacity_Window_Tons", "Attributed_Unmet_Tons"},
                pct_columns={"Share_of_Window_Capacity_Pct"},
            ) + 2

        _autofit(ws)

    _autofit(summary_ws)
    output_path = _build_output_path(output_folder, output_name)
    workbook.save(output_path)
    workbook.close()
    return output_path


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a workcenter analysis workbook from an existing ModeA or ModeB output report.")
    parser.add_argument("--report", required=True, help="Path to the ModeA or ModeB output workbook.")
    parser.add_argument("--mode", choices=SUPPORTED_REPORT_MODES, required=True)
    parser.add_argument("--workcenter", action="append", dest="workcenters", default=[], help="Workcenter to include. Repeat up to 10 times.")
    parser.add_argument("--output-dir", default=str(resolve_runtime_paths().outputs_dir))
    parser.add_argument("--output-name", default=DEFAULT_OUTPUT_NAME)
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    output_path = generate_workcenter_analysis_report(
        report_path=args.report,
        workcenters=args.workcenters,
        report_mode=args.mode,
        output_dir=args.output_dir,
        output_name=args.output_name,
    )
    print(f"工作中心分析 workbook 已生成：{output_path}")


if __name__ == "__main__":
    main()
