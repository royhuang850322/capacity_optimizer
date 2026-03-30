"""
Write optimisation results to an Excel workbook.

The workbook includes both raw result sheets and presentation-ready analysis
sheets so the Excel output can replace the former Streamlit views.
"""
from __future__ import annotations

import math
import os
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import AllocationResult, Config, LoadRecord, ValidationIssue
from result_analysis import (
    build_executive_insights,
    build_mode_comparison_frame,
    build_result_analysis,
)


HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
MODEA_FILL = PatternFill("solid", fgColor="2F75B5")
MODEB_FILL = PatternFill("solid", fgColor="ED7D31")
DELTA_FILL = PatternFill("solid", fgColor="D9E2F3")
SUMMARY_FILL = PatternFill("solid", fgColor="EAF2F8")
TITLE_FONT = Font(color="1F1F1F", bold=True, size=14)
NOTE_FONT = Font(color="666666", italic=True, size=9)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ERR_FILL = PatternFill("solid", fgColor="FFCCCC")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT_FMT = "0.0%"
TONS_FMT = "#,##0.0"
INT_FMT = "#,##0"


def build_output_path(config: Config, mode: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(config.output_file_name)[0]
    parts = [base, mode]
    scenario_segment = _sanitize_filename_segment(config.scenario_name)
    if scenario_segment:
        parts.append(scenario_segment)
    parts.append(ts)
    filename = "_".join(parts) + ".xlsx"
    return os.path.join(config.output_folder, filename)


def _build_planner_demand_map(loads: Optional[List[LoadRecord]]) -> dict[tuple[str, str], list[tuple[str, float]]]:
    planner_demand: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: defaultdict(float))
    if not loads:
        return {}

    for load in loads:
        tons = max(float(load.forecast_tons or 0.0), 0.0)
        if tons <= 0:
            continue
        key = (str(load.month), str(load.product))
        planner = str(load.planner_name or "").strip()
        planner_demand[key][planner] += tons

    return {
        key: [(planner, totals[planner]) for planner in sorted(totals, key=str.casefold)]
        for key, totals in planner_demand.items()
    }


def _split_value_by_planner(total: float, planner_weights: list[tuple[str, float]]) -> dict[str, float]:
    if not planner_weights:
        return {}

    total_units = int(round(max(float(total or 0.0), 0.0) * 10000))
    active_weights = [(planner, weight) for planner, weight in planner_weights if weight > 0]
    if total_units <= 0 or not active_weights:
        return {planner: 0.0 for planner, _weight in planner_weights}

    total_weight = sum(weight for _planner, weight in active_weights)
    allocations: list[list[Any]] = []
    assigned_units = 0
    for planner, weight in planner_weights:
        exact_units = total_units * weight / total_weight if weight > 0 else 0.0
        base_units = math.floor(exact_units)
        allocations.append([planner, base_units, exact_units - base_units])
        assigned_units += base_units

    residual_units = total_units - assigned_units
    allocations.sort(key=lambda item: (-item[2], str(item[0]).casefold()))
    for idx in range(residual_units):
        allocations[idx % len(allocations)][1] += 1

    allocations.sort(key=lambda item: str(item[0]).casefold())
    return {planner: units / 10000.0 for planner, units, _fraction in allocations}


def _plannerize_results(
    results: List[AllocationResult],
    loads: Optional[List[LoadRecord]],
) -> List[AllocationResult]:
    planner_demand = _build_planner_demand_map(loads)
    if not planner_demand:
        return [
            AllocationResult(
                month=result.month,
                product=result.product,
                product_family=result.product_family,
                plant=result.plant,
                allocation_type=result.allocation_type,
                work_center=result.work_center,
                route_type=result.route_type,
                priority=result.priority,
                demand_tons=result.demand_tons,
                allocated_tons=result.allocated_tons,
                outsourced_tons=result.outsourced_tons,
                unmet_tons=result.unmet_tons,
                capacity_share_pct=result.capacity_share_pct,
                planner_name=result.planner_name,
            )
            for result in results
        ]

    plannerized: list[AllocationResult] = []
    for result in results:
        key = (result.month, result.product)
        planner_weights = planner_demand.get(key)
        if not planner_weights:
            plannerized.append(
                AllocationResult(
                    month=result.month,
                    product=result.product,
                    product_family=result.product_family,
                    plant=result.plant,
                    allocation_type=result.allocation_type,
                    work_center=result.work_center,
                    route_type=result.route_type,
                    priority=result.priority,
                    demand_tons=result.demand_tons,
                    allocated_tons=result.allocated_tons,
                    outsourced_tons=result.outsourced_tons,
                    unmet_tons=result.unmet_tons,
                    capacity_share_pct=result.capacity_share_pct,
                    planner_name=result.planner_name,
                )
            )
            continue

        demand_split = _split_value_by_planner(result.demand_tons, planner_weights)
        allocated_split = _split_value_by_planner(result.allocated_tons, planner_weights)
        outsourced_split = _split_value_by_planner(result.outsourced_tons, planner_weights)
        unmet_split = _split_value_by_planner(result.unmet_tons, planner_weights)
        cap_share_split = _split_value_by_planner(result.capacity_share_pct, planner_weights)

        for planner_name, _weight in planner_weights:
            planner_allocated = round(allocated_split.get(planner_name, 0.0), 4)
            planner_outsourced = round(outsourced_split.get(planner_name, 0.0), 4)
            planner_unmet = round(unmet_split.get(planner_name, 0.0), 4)

            if result.allocation_type == "Internal" and planner_allocated <= 0.0:
                continue
            if result.allocation_type == "Outsourced" and planner_outsourced <= 0.0:
                continue
            if result.allocation_type == "Unmet" and planner_unmet <= 0.0:
                continue

            plannerized.append(
                AllocationResult(
                    month=result.month,
                    product=result.product,
                    product_family=result.product_family,
                    plant=result.plant,
                    allocation_type=result.allocation_type,
                    work_center=result.work_center,
                    route_type=result.route_type,
                    priority=result.priority,
                    demand_tons=round(demand_split.get(planner_name, 0.0), 4),
                    allocated_tons=planner_allocated,
                    outsourced_tons=planner_outsourced,
                    unmet_tons=planner_unmet,
                    capacity_share_pct=round(cap_share_split.get(planner_name, 0.0), 4),
                    planner_name=planner_name,
                )
            )

    return plannerized


def write_results(
    results: List[AllocationResult],
    loads: Optional[List[LoadRecord]],
    config: Config,
    issues: List[ValidationIssue],
    months: List[str],
    mode: str = "ModeA",
    toller_products: Optional[set] = None,
    metrics_by_mode: Optional[dict[str, dict[str, Any]]] = None,
) -> str:
    """
    Write a complete Excel report workbook.

    The workbook contains:
    - report sheets replacing the previous Streamlit views
    - raw output sheets for audit/detail
    """
    os.makedirs(config.output_folder, exist_ok=True)
    out_path = build_output_path(config, mode)

    wb = Workbook()
    wb.remove(wb.active)

    artifact = _build_mode_artifact(results, loads, config, months, mode)
    df_detail = artifact["df_detail"]
    wc_load_df = artifact["wc_load_df"]
    run_info_df = artifact["run_info_df"]
    analysis = artifact["analysis"]
    preview_metrics = artifact["metrics"]

    all_metrics = dict(metrics_by_mode or {})
    all_metrics.setdefault(mode, preview_metrics)

    _write_dashboard(wb, mode, analysis, preview_metrics, all_metrics)
    _write_monthly_analysis(wb, analysis)
    _write_bottleneck_analysis(wb, analysis)
    _write_wc_heatmap(wb, analysis)
    _write_product_risk_analysis(wb, analysis)

    _write_detail(wb, df_detail)
    _write_planner_summary(wb, analysis)
    _write_planner_product_month_summary(wb, analysis)
    _write_allocation_summary(wb, df_detail, months)
    _write_outsource_summary(wb, df_detail, months)
    _write_unmet_summary(wb, df_detail, months)
    _write_wc_load(wb, wc_load_df)
    _write_binary_report(wb, df_detail, months, toller_products or set())
    _write_validation(wb, issues)
    _write_run_info(wb, run_info_df)

    wb.save(out_path)
    return out_path


def write_mode_comparison_summary(
    mode_results: dict[str, List[AllocationResult]],
    config: Config,
    months: List[str],
    metrics_by_mode: Optional[dict[str, dict[str, Any]]] = None,
    mode_loads: Optional[dict[str, List[LoadRecord]]] = None,
) -> str:
    """
    Write a standalone comparison workbook when both ModeA and ModeB are run.
    """
    required_modes = {"ModeA", "ModeB"}
    if not required_modes.issubset(mode_results):
        raise ValueError("Comparison workbook requires both ModeA and ModeB results.")

    os.makedirs(config.output_folder, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(config.output_folder, f"Summary of Mode A and Mode B_{ts}.xlsx")

    artifacts = {
        mode: _build_mode_artifact(mode_results[mode], mode_loads.get(mode) if mode_loads else None, config, months, mode)
        for mode in ("ModeA", "ModeB")
    }
    comparison_metrics = dict(metrics_by_mode or {})
    for mode in ("ModeA", "ModeB"):
        comparison_metrics.setdefault(mode, artifacts[mode]["metrics"])

    wb = Workbook()
    wb.remove(wb.active)

    _write_executive_comparison(wb, artifacts, comparison_metrics)
    _write_monthly_comparison(wb, artifacts)
    _write_bottleneck_comparison(wb, artifacts)
    _write_heatmap_comparison(wb, artifacts)
    _write_product_risk_comparison(wb, artifacts)
    _write_planner_comparison(wb, artifacts)
    _write_comparison_run_info(wb, config, comparison_metrics, os.path.basename(out_path))

    wb.save(out_path)
    return out_path


def _results_to_df(results: List[AllocationResult]) -> pd.DataFrame:
    rows = []
    columns = [
        "Month",
        "PlannerName",
        "Product",
        "ProductFamily",
        "Plant",
        "AllocationType",
        "WorkCenter",
        "RouteType",
        "Priority",
        "Demand_Tons",
        "Allocated_Tons",
        "Outsourced_Tons",
        "Unmet_Tons",
        "CapacityShare_Pct",
    ]
    for result in results:
        rows.append(
            {
                "Month": result.month,
                "PlannerName": result.planner_name,
                "Product": result.product,
                "ProductFamily": result.product_family,
                "Plant": result.plant,
                "AllocationType": result.allocation_type,
                "WorkCenter": result.work_center,
                "RouteType": result.route_type,
                "Priority": result.priority,
                "Demand_Tons": result.demand_tons,
                "Allocated_Tons": result.allocated_tons,
                "Outsourced_Tons": result.outsourced_tons,
                "Unmet_Tons": result.unmet_tons,
                "CapacityShare_Pct": result.capacity_share_pct / 100.0,
            }
        )
    return pd.DataFrame(rows, columns=columns)


def _build_mode_artifact(
    results: List[AllocationResult],
    loads: Optional[List[LoadRecord]],
    config: Config,
    months: List[str],
    mode: str,
) -> dict[str, Any]:
    plannerized_results = _plannerize_results(results, loads)
    df_detail = _results_to_df(plannerized_results)
    wc_load_df = _build_wc_load_frame(df_detail, months)
    run_info_df = _build_run_info_df(config, mode)
    analysis = build_result_analysis(df_detail, wc_load_df, run_info_df)
    metrics = _build_preview_metrics(mode, analysis, results, months)
    return {
        "df_detail": df_detail,
        "wc_load_df": wc_load_df,
        "run_info_df": run_info_df,
        "analysis": analysis,
        "metrics": metrics,
    }


def _build_wc_load_frame(df_detail: pd.DataFrame, months: List[str]) -> pd.DataFrame:
    internal_df = df_detail[df_detail["AllocationType"] == "Internal"]
    if internal_df.empty:
        return pd.DataFrame(columns=["WorkCenter", *months])

    pivot = internal_df[internal_df["WorkCenter"] != "[UNALLOCATED]"].pivot_table(
        index="WorkCenter",
        columns="Month",
        values="CapacityShare_Pct",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    return pivot


def _build_run_info_df(config: Config, mode: str) -> pd.DataFrame:
    rows = [
        ("Scenario_Name", config.scenario_name),
        ("Mode", mode),
        ("Start_Month", config.start_month),
        ("Horizon_Months", config.horizon_months),
        ("Input_Load_Folder", config.input_load_folder),
        ("Input_Master_Folder", config.input_master_folder),
        ("Output_Folder", config.output_folder),
        ("Project_Root_Folder", getattr(config, "project_root_folder", "")),
        ("Output_FileName", config.output_file_name),
        ("Run_Timestamp", config.run_timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Run_Mode", getattr(config, "run_mode", mode)),
        ("Direct_Mode", "Yes" if getattr(config, "direct_mode", True) else "No"),
        ("Verbose", "Yes" if getattr(config, "verbose", False) else "No"),
        (
            "Skip_Validation_Errors",
            "Yes" if getattr(config, "skip_validation_errors", False) else "No",
        ),
        ("License_Status", getattr(config, "license_status", "")),
        ("License_ID", getattr(config, "license_id", "")),
        ("License_Type", getattr(config, "license_type", "")),
        ("Licensed_To", getattr(config, "licensed_to", "")),
        ("License_Expiry", getattr(config, "license_expiry", "")),
        ("License_Binding_Mode", getattr(config, "license_binding_mode", "")),
        ("License_Machine_Label", getattr(config, "license_machine_label", "")),
        ("Notes", config.notes or ""),
        ("Tool_Version", "1.1.0"),
    ]
    return pd.DataFrame(rows, columns=["Parameter", "Value"])


def _build_preview_metrics(
    mode: str,
    analysis: dict[str, Any],
    results: List[AllocationResult],
    months: List[str],
) -> dict[str, Any]:
    monthly_summary = analysis["monthly_summary"]
    total_demand = float(monthly_summary["Demand_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_internal_allocated = float(monthly_summary["Internal_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_outsourced = float(monthly_summary["Outsourced_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_unmet = float(monthly_summary["Unmet_Tons"].sum()) if not monthly_summary.empty else 0.0
    total_supplied = total_internal_allocated + total_outsourced
    service_level = 100.0 * total_supplied / total_demand if total_demand > 0 else 0.0
    return {
        "mode": mode,
        "scenario_name": analysis.get("scenario_name") or "",
        "total_demand": total_demand,
        "total_internal_allocated": total_internal_allocated,
        "total_outsourced": total_outsourced,
        "total_unmet": total_unmet,
        "service_level": service_level,
        "result_rows": len(results),
        "months": len(months),
    }


def _write_dashboard(
    wb: Workbook,
    mode: str,
    analysis: dict[str, Any],
    preview_metrics: dict[str, Any],
    metrics_by_mode: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Dashboard")
    _write_sheet_title(ws, f"Executive Summary - {mode}")

    ws["A2"] = f"Scenario: {analysis.get('scenario_name') or preview_metrics.get('scenario_name') or 'N/A'}"
    ws["A3"] = f"Mode: {mode}"
    ws["A4"] = f"Horizon months: {preview_metrics.get('months', 0)}"
    for cell_ref in ("A2", "A3", "A4"):
        ws[cell_ref].font = Font(color="444444", size=10)

    metric_rows = [
        ("Total demand", preview_metrics["total_demand"], TONS_FMT),
        ("Internal allocated", preview_metrics["total_internal_allocated"], TONS_FMT),
        ("Outsourced", preview_metrics["total_outsourced"], TONS_FMT),
        ("Residual unmet", preview_metrics["total_unmet"], TONS_FMT),
        ("Service level", preview_metrics["service_level"] / 100.0, PCT_FMT),
        ("Result rows", preview_metrics["result_rows"], INT_FMT),
    ]
    _write_metric_block(ws, start_row=6, start_col=1, rows=metric_rows)

    insights = build_executive_insights(mode, preview_metrics, metrics_by_mode, analysis)
    ws["D6"] = "Key conclusions"
    ws["D6"].font = Font(bold=True, color="1F4E79", size=11)
    for offset, line in enumerate(insights, start=1):
        cell = ws.cell(6 + offset, 4)
        cell.value = f"- {line}"
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    supply_mix_df = pd.DataFrame(
        {
            "Category": ["Internal allocated", "Outsourced", "Residual unmet"],
            "Tons": [
                preview_metrics["total_internal_allocated"],
                preview_metrics["total_outsourced"],
                preview_metrics["total_unmet"],
            ],
        }
    )
    mix_layout = _write_table(
        ws,
        supply_mix_df,
        start_row=14,
        start_col=1,
        num_formats={"Tons": TONS_FMT},
    )
    mix_chart = BarChart()
    mix_chart.title = "Supply Mix"
    mix_chart.y_axis.title = "Tons"
    mix_chart.height = 7
    mix_chart.width = 11
    mix_chart.legend = None
    mix_chart.add_data(
        Reference(
            ws,
            min_col=mix_layout["col_index"]["Tons"],
            min_row=mix_layout["start_row"],
            max_row=mix_layout["end_row"],
        ),
        titles_from_data=True,
    )
    mix_chart.set_categories(
        Reference(
            ws,
            min_col=mix_layout["col_index"]["Category"],
            min_row=mix_layout["start_row"] + 1,
            max_row=mix_layout["end_row"],
        )
    )
    mix_chart.dataLabels = DataLabelList()
    mix_chart.dataLabels.showVal = True
    ws.add_chart(mix_chart, "D14")

    comparison_df = build_mode_comparison_frame(metrics_by_mode)
    if not comparison_df.empty:
        comparison_layout = _write_table(
            ws,
            comparison_df,
            start_row=14,
            start_col=8,
            num_formats={"Value": TONS_FMT},
        )
        comp_chart = BarChart()
        comp_chart.title = "Mode Comparison"
        comp_chart.y_axis.title = "Tons"
        comp_chart.height = 7
        comp_chart.width = 12
        data = Reference(
            ws,
            min_col=comparison_layout["col_index"]["Value"],
            min_row=comparison_layout["start_row"],
            max_row=comparison_layout["end_row"],
        )
        cats = Reference(
            ws,
            min_col=comparison_layout["col_index"]["Metric"],
            min_row=comparison_layout["start_row"] + 1,
            max_row=comparison_layout["end_row"],
        )
        comp_chart.add_data(data, titles_from_data=True, from_rows=False)
        comp_chart.set_categories(cats)
        ws.add_chart(comp_chart, "K14")
    else:
        _write_note(
            ws,
            "H14",
            "Mode comparison appears when ModeA and ModeB are run from the same control workbook session.",
        )

    _autofit(ws)


def _write_executive_comparison(
    wb: Workbook,
    artifacts: dict[str, dict[str, Any]],
    metrics_by_mode: dict[str, dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Executive_Comparison")
    ws.sheet_view.showGridLines = False

    mode_a = metrics_by_mode["ModeA"]
    mode_b = metrics_by_mode["ModeB"]
    scenario = mode_a.get("scenario_name") or mode_b.get("scenario_name") or "N/A"
    service_delta_pct = mode_b["service_level"] - mode_a["service_level"]
    unmet_delta_tons = mode_b["total_unmet"] - mode_a["total_unmet"]
    internal_delta_tons = mode_b["total_internal_allocated"] - mode_a["total_internal_allocated"]
    outsourced_delta_tons = mode_b["total_outsourced"] - mode_a["total_outsourced"]

    ws.merge_cells("A1:N1")
    ws["A1"] = "Summary of Mode A and Mode B"
    ws["A1"].fill = HDR_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:N2")
    ws["A2"] = (
        f"Executive comparison workbook | Scenario: {scenario} | "
        "Use this cover page to review the main trade-offs before opening the detailed tabs."
    )
    ws["A2"].fill = SUMMARY_FILL
    ws["A2"].font = Font(color="44546A", bold=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A4:F5")
    ws["A4"] = (
        "MODE A\n"
        "Internal-first baseline\n"
        f"Service level: {mode_a['service_level'] / 100.0:.1%} | "
        f"Internal allocated: {mode_a['total_internal_allocated']:,.1f} tons"
    )
    ws["A4"].fill = MODEA_FILL
    ws["A4"].font = Font(color="FFFFFF", bold=True, size=12)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("H4:N5")
    ws["H4"] = (
        "MODE B\n"
        "Expanded supply option\n"
        f"Service level: {mode_b['service_level'] / 100.0:.1%} | "
        f"Outsourced: {mode_b['total_outsourced']:,.1f} tons"
    )
    ws["H4"].fill = MODEB_FILL
    ws["H4"].font = Font(color="FFFFFF", bold=True, size=12)
    ws["H4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A7:N7")
    ws["A7"] = (
        "Management lens: compare service improvement, outsourced reliance, "
        "and residual unmet reduction across the two operating modes."
    )
    ws["A7"].fill = SUBHDR_FILL
    ws["A7"].font = Font(color="44546A", italic=True, size=10)
    ws["A7"].alignment = Alignment(horizontal="left", vertical="center")

    metric_rows = [
        {
            "Metric": "Total demand",
            "ModeA": mode_a["total_demand"],
            "ModeB": mode_b["total_demand"],
            "Delta (ModeB - ModeA)": mode_b["total_demand"] - mode_a["total_demand"],
        },
        {
            "Metric": "Internal allocated",
            "ModeA": mode_a["total_internal_allocated"],
            "ModeB": mode_b["total_internal_allocated"],
            "Delta (ModeB - ModeA)": internal_delta_tons,
        },
        {
            "Metric": "Outsourced",
            "ModeA": mode_a["total_outsourced"],
            "ModeB": mode_b["total_outsourced"],
            "Delta (ModeB - ModeA)": outsourced_delta_tons,
        },
        {
            "Metric": "Residual unmet",
            "ModeA": mode_a["total_unmet"],
            "ModeB": mode_b["total_unmet"],
            "Delta (ModeB - ModeA)": unmet_delta_tons,
        },
        {
            "Metric": "Service level",
            "ModeA": mode_a["service_level"] / 100.0,
            "ModeB": mode_b["service_level"] / 100.0,
            "Delta (ModeB - ModeA)": service_delta_pct / 100.0,
        },
        {
            "Metric": "Result rows",
            "ModeA": mode_a["result_rows"],
            "ModeB": mode_b["result_rows"],
            "Delta (ModeB - ModeA)": mode_b["result_rows"] - mode_a["result_rows"],
        },
    ]
    metric_df = pd.DataFrame(metric_rows)
    metric_layout = _write_table(
        ws,
        metric_df,
        start_row=9,
        start_col=1,
        num_formats={
            "ModeA": TONS_FMT,
            "ModeB": TONS_FMT,
            "Delta (ModeB - ModeA)": TONS_FMT,
        },
    )
    ws.cell(metric_layout["start_row"], metric_layout["col_index"]["ModeA"]).fill = MODEA_FILL
    ws.cell(metric_layout["start_row"], metric_layout["col_index"]["ModeB"]).fill = MODEB_FILL
    ws.cell(metric_layout["start_row"], metric_layout["col_index"]["Delta (ModeB - ModeA)"]).fill = DELTA_FILL
    for row_num in range(metric_layout["start_row"] + 1, metric_layout["end_row"] + 1):
        label_cell = ws.cell(row_num, metric_layout["start_col"])
        label_cell.fill = SUMMARY_FILL
        label_cell.font = Font(bold=True, color="1F1F1F")

    service_row = metric_layout["start_row"] + 5
    for column in ("B", "C", "D"):
        ws[f"{column}{service_row}"].number_format = PCT_FMT
        ws[f"{column}{service_row}"].fill = OK_FILL
    result_rows_row = metric_layout["start_row"] + 6
    for column in ("B", "C", "D"):
        ws[f"{column}{result_rows_row}"].number_format = INT_FMT

    ws.merge_cells("F9:N9")
    ws["F9"] = "Management conclusion"
    ws["F9"].fill = HDR_FILL
    ws["F9"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["F9"].alignment = Alignment(horizontal="left", vertical="center")
    conclusion_lines = [
        (
            f"Service level: ModeB is {service_delta_pct:+.1f} pts versus ModeA "
            f"({mode_a['service_level'] / 100.0:.1%} -> {mode_b['service_level'] / 100.0:.1%})."
        ),
        (
            f"Residual unmet: ModeB changes unmet demand by {unmet_delta_tons:+,.1f} tons "
            f"({mode_a['total_unmet']:,.1f} -> {mode_b['total_unmet']:,.1f})."
        ),
        f"Internal supply: ModeB changes internal allocation by {internal_delta_tons:+,.1f} tons.",
        f"External reliance: ModeB changes outsourced volume by {outsourced_delta_tons:+,.1f} tons.",
        "Use the detailed tabs to confirm where the differences come from: monthly balance, bottlenecks, heatmap, and product risk.",
    ]
    for offset, line in enumerate(conclusion_lines, start=10):
        ws.merge_cells(f"F{offset}:N{offset}")
        ws[f"F{offset}"] = f"- {line}"
        ws[f"F{offset}"].fill = SUMMARY_FILL
        ws[f"F{offset}"].font = Font(color="1F1F1F", size=10)
        ws[f"F{offset}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[offset].height = 24

    mix_df = pd.DataFrame(
        {
            "Category": ["Internal allocated", "Outsourced", "Residual unmet"],
            "ModeA": [
                mode_a["total_internal_allocated"],
                mode_a["total_outsourced"],
                mode_a["total_unmet"],
            ],
            "ModeB": [
                mode_b["total_internal_allocated"],
                mode_b["total_outsourced"],
                mode_b["total_unmet"],
            ],
        }
    )
    mix_layout = _write_table(ws, mix_df, start_row=18, start_col=1, num_formats={"ModeA": TONS_FMT, "ModeB": TONS_FMT})
    ws.cell(mix_layout["start_row"], mix_layout["col_index"]["ModeA"]).fill = MODEA_FILL
    ws.cell(mix_layout["start_row"], mix_layout["col_index"]["ModeB"]).fill = MODEB_FILL
    chart = BarChart()
    chart.title = "Supply Mix Comparison"
    chart.y_axis.title = "Tons"
    chart.height = 8
    chart.width = 12
    chart.add_data(
        Reference(ws, min_col=mix_layout["col_index"]["ModeA"], min_row=mix_layout["start_row"], max_col=mix_layout["col_index"]["ModeB"], max_row=mix_layout["end_row"]),
        titles_from_data=True,
        from_rows=False,
    )
    chart.set_categories(
        Reference(ws, min_col=mix_layout["col_index"]["Category"], min_row=mix_layout["start_row"] + 1, max_row=mix_layout["end_row"])
    )
    _apply_chart_palette(chart, ["2F75B5", "ED7D31"])
    ws.add_chart(chart, "E18")

    service_df = pd.DataFrame(
        {
            "Mode": ["ModeA", "ModeB"],
            "Service_Level": [
                mode_a["service_level"] / 100.0,
                mode_b["service_level"] / 100.0,
            ],
        }
    )
    service_layout = _write_table(ws, service_df, start_row=18, start_col=10, num_formats={"Service_Level": PCT_FMT})
    service_chart = BarChart()
    service_chart.title = "Service Level Comparison"
    service_chart.y_axis.title = "Service level"
    service_chart.height = 8
    service_chart.width = 9
    service_chart.varyColors = True
    service_chart.add_data(
        Reference(ws, min_col=service_layout["col_index"]["Service_Level"], min_row=service_layout["start_row"], max_row=service_layout["end_row"]),
        titles_from_data=True,
    )
    service_chart.set_categories(
        Reference(ws, min_col=service_layout["col_index"]["Mode"], min_row=service_layout["start_row"] + 1, max_row=service_layout["end_row"])
    )
    service_chart.y_axis.numFmt = PCT_FMT
    ws.add_chart(service_chart, "M18")

    insights = [
        f"ModeB service level is {service_delta_pct:+.1f} pts versus ModeA.",
        f"ModeB changes residual unmet by {unmet_delta_tons:+,.1f} tons.",
        f"ModeB changes internal allocation by {internal_delta_tons:+,.1f} tons.",
        f"ModeB changes outsourced tons by {outsourced_delta_tons:+,.1f} tons.",
    ]
    ws.merge_cells("A33:N33")
    ws["A33"] = "Quick read-out"
    ws["A33"].fill = SUBHDR_FILL
    ws["A33"].font = Font(bold=True, color="1F4E79", size=11)
    ws["A33"].alignment = Alignment(horizontal="left", vertical="center")
    for offset, line in enumerate(insights, start=34):
        ws.merge_cells(f"A{offset}:N{offset}")
        ws[f"A{offset}"] = f"- {line}"
        ws[f"A{offset}"].alignment = Alignment(wrap_text=True)
        ws[f"A{offset}"].font = Font(color="1F1F1F", size=10)
    _autofit(ws)


def _write_monthly_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Monthly_Trend_Compare")
    _write_sheet_title(ws, "Monthly Trend Comparison")

    monthly_a = artifacts["ModeA"]["analysis"]["monthly_summary"].copy()
    monthly_b = artifacts["ModeB"]["analysis"]["monthly_summary"].copy()
    monthly_compare = monthly_a.merge(monthly_b, on="Month", how="outer", suffixes=("_ModeA", "_ModeB")).fillna(0.0)
    monthly_compare["Service_Level_Delta"] = monthly_compare["Service_Level_ModeB"] - monthly_compare["Service_Level_ModeA"]
    monthly_compare["Unmet_Delta"] = monthly_compare["Unmet_Tons_ModeB"] - monthly_compare["Unmet_Tons_ModeA"]
    monthly_compare = monthly_compare.sort_values("Month")

    layout = _write_table(
        ws,
        monthly_compare[
            [
                "Month",
                "Demand_Tons_ModeA",
                "Internal_Tons_ModeA",
                "Internal_Tons_ModeB",
                "Outsourced_Tons_ModeA",
                "Outsourced_Tons_ModeB",
                "Unmet_Tons_ModeA",
                "Unmet_Tons_ModeB",
                "Service_Level_ModeA",
                "Service_Level_ModeB",
                "Service_Level_Delta",
                "Unmet_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeB": TONS_FMT,
            "Outsourced_Tons_ModeA": TONS_FMT,
            "Outsourced_Tons_ModeB": TONS_FMT,
            "Unmet_Tons_ModeA": TONS_FMT,
            "Unmet_Tons_ModeB": TONS_FMT,
            "Service_Level_ModeA": PCT_FMT,
            "Service_Level_ModeB": PCT_FMT,
            "Service_Level_Delta": PCT_FMT,
            "Unmet_Delta": TONS_FMT,
        },
    )

    service_chart = LineChart()
    service_chart.title = "Monthly Service Level Comparison"
    service_chart.y_axis.title = "Service level"
    service_chart.height = 8
    service_chart.width = 14
    service_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["Service_Level_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["Service_Level_ModeB"],
            max_row=layout["end_row"],
        ),
        titles_from_data=True,
        from_rows=False,
    )
    service_chart.set_categories(
        Reference(ws, min_col=layout["col_index"]["Month"], min_row=layout["start_row"] + 1, max_row=layout["end_row"])
    )
    service_chart.y_axis.numFmt = PCT_FMT
    ws.add_chart(service_chart, "N3")

    unmet_chart = BarChart()
    unmet_chart.title = "Monthly Residual Unmet Comparison"
    unmet_chart.y_axis.title = "Tons"
    unmet_chart.height = 8
    unmet_chart.width = 14
    unmet_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["Unmet_Tons_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["Unmet_Tons_ModeB"],
            max_row=layout["end_row"],
        ),
        titles_from_data=True,
        from_rows=False,
    )
    unmet_chart.set_categories(
        Reference(ws, min_col=layout["col_index"]["Month"], min_row=layout["start_row"] + 1, max_row=layout["end_row"])
    )
    ws.add_chart(unmet_chart, "N21")
    _autofit(ws)


def _write_bottleneck_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Bottleneck_Compare")
    _write_sheet_title(ws, "Bottleneck Comparison")

    wc_a = artifacts["ModeA"]["analysis"]["wc_summary"].copy()
    wc_b = artifacts["ModeB"]["analysis"]["wc_summary"].copy()
    wc_compare = wc_a.merge(wc_b, on="WorkCenter", how="outer", suffixes=("_ModeA", "_ModeB")).fillna(0.0)
    wc_compare["PeakLoad_Delta"] = wc_compare["PeakLoadPct_ModeB"] - wc_compare["PeakLoadPct_ModeA"]
    wc_compare["SortKey"] = wc_compare[["PeakLoadPct_ModeA", "PeakLoadPct_ModeB"]].max(axis=1)
    wc_compare = wc_compare.sort_values(["SortKey", "PeakLoadPct_ModeB"], ascending=[False, False]).head(15)

    layout = _write_table(
        ws,
        wc_compare[
            [
                "WorkCenter",
                "AvgLoadPct_ModeA",
                "AvgLoadPct_ModeB",
                "PeakLoadPct_ModeA",
                "PeakLoadPct_ModeB",
                "Over95Months_ModeA",
                "Over95Months_ModeB",
                "PeakLoad_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "AvgLoadPct_ModeA": PCT_FMT,
            "AvgLoadPct_ModeB": PCT_FMT,
            "PeakLoadPct_ModeA": PCT_FMT,
            "PeakLoadPct_ModeB": PCT_FMT,
            "Over95Months_ModeA": INT_FMT,
            "Over95Months_ModeB": INT_FMT,
            "PeakLoad_Delta": PCT_FMT,
        },
    )

    peak_chart = BarChart()
    peak_chart.type = "bar"
    peak_chart.title = "Peak Load by WorkCenter"
    peak_chart.height = 9
    peak_chart.width = 13
    peak_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["PeakLoadPct_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["PeakLoadPct_ModeB"],
            max_row=layout["end_row"],
        ),
        titles_from_data=True,
        from_rows=False,
    )
    peak_chart.set_categories(
        Reference(ws, min_col=layout["col_index"]["WorkCenter"], min_row=layout["start_row"] + 1, max_row=layout["end_row"])
    )
    peak_chart.x_axis.numFmt = PCT_FMT
    ws.add_chart(peak_chart, "J3")

    focus_wc = str(wc_compare.iloc[0]["WorkCenter"]) if not wc_compare.empty else ""
    if focus_wc:
        ws["A22"] = f"Focused workcenter comparison: {focus_wc}"
        ws["A22"].font = Font(bold=True, color="1F4E79", size=11)
        focus_a = artifacts["ModeA"]["analysis"]["wc_long"]
        focus_b = artifacts["ModeB"]["analysis"]["wc_long"]
        focus_compare = (
            focus_a[focus_a["WorkCenter"] == focus_wc][["Month", "LoadPct"]]
            .rename(columns={"LoadPct": "Load_ModeA"})
            .merge(
                focus_b[focus_b["WorkCenter"] == focus_wc][["Month", "LoadPct"]].rename(columns={"LoadPct": "Load_ModeB"}),
                on="Month",
                how="outer",
            )
            .fillna(0.0)
            .sort_values("Month")
        )
        focus_compare["Load_Delta"] = focus_compare["Load_ModeB"] - focus_compare["Load_ModeA"]
        focus_layout = _write_table(
            ws,
            focus_compare,
            start_row=23,
            start_col=1,
            num_formats={"Load_ModeA": PCT_FMT, "Load_ModeB": PCT_FMT, "Load_Delta": PCT_FMT},
        )
        focus_chart = LineChart()
        focus_chart.title = f"{focus_wc} Load Trend Comparison"
        focus_chart.height = 7
        focus_chart.width = 13
        focus_chart.add_data(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Load_ModeA"],
                min_row=focus_layout["start_row"],
                max_col=focus_layout["col_index"]["Load_ModeB"],
                max_row=focus_layout["end_row"],
            ),
            titles_from_data=True,
            from_rows=False,
        )
        focus_chart.set_categories(
            Reference(ws, min_col=focus_layout["col_index"]["Month"], min_row=focus_layout["start_row"] + 1, max_row=focus_layout["end_row"])
        )
        focus_chart.y_axis.numFmt = PCT_FMT
        ws.add_chart(focus_chart, "F23")
    _autofit(ws)


def _write_heatmap_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("WC_Heatmap_Compare")
    _write_sheet_title(ws, "WorkCenter Heatmap Comparison")

    wc_names = []
    for mode in ("ModeA", "ModeB"):
        summary = artifacts[mode]["analysis"]["wc_summary"]
        wc_names.extend(summary.head(12)["WorkCenter"].tolist())
    wc_names = list(dict.fromkeys(wc_names))

    next_start_row = 2
    for mode in ("ModeA", "ModeB"):
        header_cell = ws.cell(next_start_row, 1)
        header_cell.value = f"{mode} Heatmap"
        header_cell.font = Font(color="FFFFFF", bold=True, size=11)
        header_cell.fill = MODEA_FILL if mode == "ModeA" else MODEB_FILL
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        header_cell.border = BORDER
        ws.row_dimensions[next_start_row].height = 22

        wc_long = artifacts[mode]["analysis"]["wc_long"]
        if wc_long.empty:
            ws.cell(next_start_row + 1, 1).value = "No heatmap data"
            next_start_row += 4
            continue

        pivot = wc_long[wc_long["WorkCenter"].isin(wc_names)].pivot(index="WorkCenter", columns="Month", values="LoadPct").fillna(0.0)
        pivot = pivot.reindex(index=wc_names)
        pivot.reset_index(inplace=True)
        layout = _write_table(
            ws,
            pivot,
            start_row=next_start_row + 1,
            start_col=1,
            num_formats={column: PCT_FMT for column in pivot.columns if column != "WorkCenter"},
        )
        if layout["end_row"] > layout["start_row"]:
            ws.conditional_formatting.add(
                f"{get_column_letter(layout['start_col'] + 1)}{layout['start_row'] + 1}:{get_column_letter(layout['end_col'])}{layout['end_row']}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0.0,
                    start_color="FFF2CC",
                    mid_type="num",
                    mid_value=0.85,
                    mid_color="F4B183",
                    end_type="num",
                    end_value=1.0,
                    end_color="C00000",
                ),
            )
        next_start_row = layout["end_row"] + 3

    _write_note(ws, f"A{next_start_row}", "ModeA heatmap is shown first, with ModeB directly below for vertical comparison.")
    _autofit(ws)


def _write_product_risk_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Product_Risk_Compare")
    _write_sheet_title(ws, "Product Risk Comparison")

    product_a = artifacts["ModeA"]["analysis"]["product_summary"].copy()
    product_b = artifacts["ModeB"]["analysis"]["product_summary"].copy()
    product_compare = product_a.merge(
        product_b,
        on=["Product", "ProductFamily", "Plant"],
        how="outer",
        suffixes=("_ModeA", "_ModeB"),
    ).fillna(0.0)
    product_compare["Unmet_Delta"] = product_compare["Unmet_Tons_ModeB"] - product_compare["Unmet_Tons_ModeA"]
    product_compare["Service_Level_Delta"] = product_compare["Service_Level_ModeB"] - product_compare["Service_Level_ModeA"]
    product_compare["SortKey"] = product_compare[["Unmet_Tons_ModeA", "Unmet_Tons_ModeB"]].max(axis=1)
    product_compare = product_compare.sort_values(["SortKey", "Unmet_Tons_ModeB"], ascending=[False, False]).head(20)

    layout = _write_table(
        ws,
        product_compare[
            [
                "Product",
                "ProductFamily",
                "Plant",
                "Demand_Tons_ModeA",
                "Internal_Tons_ModeA",
                "Internal_Tons_ModeB",
                "Outsourced_Tons_ModeA",
                "Outsourced_Tons_ModeB",
                "Unmet_Tons_ModeA",
                "Unmet_Tons_ModeB",
                "Service_Level_ModeA",
                "Service_Level_ModeB",
                "Unmet_Delta",
                "Service_Level_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeB": TONS_FMT,
            "Outsourced_Tons_ModeA": TONS_FMT,
            "Outsourced_Tons_ModeB": TONS_FMT,
            "Unmet_Tons_ModeA": TONS_FMT,
            "Unmet_Tons_ModeB": TONS_FMT,
            "Service_Level_ModeA": PCT_FMT,
            "Service_Level_ModeB": PCT_FMT,
            "Unmet_Delta": TONS_FMT,
            "Service_Level_Delta": PCT_FMT,
        },
    )

    top_products = product_compare.head(12)[["Product", "Unmet_Tons_ModeA", "Unmet_Tons_ModeB"]]
    top_layout = _write_table(
        ws,
        top_products,
        start_row=3,
        start_col=16,
        num_formats={"Unmet_Tons_ModeA": TONS_FMT, "Unmet_Tons_ModeB": TONS_FMT},
    )
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top Product Unmet Comparison"
    chart.height = 8
    chart.width = 11
    chart.add_data(
        Reference(
            ws,
            min_col=top_layout["col_index"]["Unmet_Tons_ModeA"],
            min_row=top_layout["start_row"],
            max_col=top_layout["col_index"]["Unmet_Tons_ModeB"],
            max_row=top_layout["end_row"],
        ),
        titles_from_data=True,
        from_rows=False,
    )
    chart.set_categories(
        Reference(ws, min_col=top_layout["col_index"]["Product"], min_row=top_layout["start_row"] + 1, max_row=top_layout["end_row"])
    )
    ws.add_chart(chart, "S3")

    focus_product = str(product_compare.iloc[0]["Product"]) if not product_compare.empty else ""
    if focus_product:
        ws["A28"] = f"Focused product comparison: {focus_product}"
        ws["A28"].font = Font(bold=True, color="1F4E79", size=11)
        month_a = artifacts["ModeA"]["analysis"]["product_month_summary"]
        month_b = artifacts["ModeB"]["analysis"]["product_month_summary"]
        focus_compare = (
            month_a[month_a["Product"] == focus_product][["Month", "Demand_Tons", "Unmet_Tons", "Service_Level"]]
            .rename(columns={"Demand_Tons": "Demand_ModeA", "Unmet_Tons": "Unmet_ModeA", "Service_Level": "Service_Level_ModeA"})
            .merge(
                month_b[month_b["Product"] == focus_product][["Month", "Demand_Tons", "Unmet_Tons", "Service_Level"]].rename(
                    columns={"Demand_Tons": "Demand_ModeB", "Unmet_Tons": "Unmet_ModeB", "Service_Level": "Service_Level_ModeB"}
                ),
                on="Month",
                how="outer",
            )
            .fillna(0.0)
            .sort_values("Month")
        )
        focus_compare["Unmet_Delta"] = focus_compare["Unmet_ModeB"] - focus_compare["Unmet_ModeA"]
        focus_layout = _write_table(
            ws,
            focus_compare,
            start_row=29,
            start_col=1,
            num_formats={
                "Demand_ModeA": TONS_FMT,
                "Demand_ModeB": TONS_FMT,
                "Unmet_ModeA": TONS_FMT,
                "Unmet_ModeB": TONS_FMT,
                "Service_Level_ModeA": PCT_FMT,
                "Service_Level_ModeB": PCT_FMT,
                "Unmet_Delta": TONS_FMT,
            },
        )
        focus_chart = LineChart()
        focus_chart.title = f"{focus_product} Service Level Comparison"
        focus_chart.height = 7
        focus_chart.width = 13
        focus_chart.add_data(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Service_Level_ModeA"],
                min_row=focus_layout["start_row"],
                max_col=focus_layout["col_index"]["Service_Level_ModeB"],
                max_row=focus_layout["end_row"],
            ),
            titles_from_data=True,
            from_rows=False,
        )
        focus_chart.set_categories(
            Reference(ws, min_col=focus_layout["col_index"]["Month"], min_row=focus_layout["start_row"] + 1, max_row=focus_layout["end_row"])
        )
        focus_chart.y_axis.numFmt = PCT_FMT
        ws.add_chart(focus_chart, "J29")
    _autofit(ws)


def _write_planner_comparison(wb: Workbook, artifacts: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Planner_Compare")
    _write_sheet_title(ws, "Planner Comparison")

    planner_a = artifacts["ModeA"]["analysis"]["planner_summary"].copy()
    planner_b = artifacts["ModeB"]["analysis"]["planner_summary"].copy()
    for frame in (planner_a, planner_b):
        if "PlannerName" in frame.columns:
            frame = frame
    if not planner_a.empty:
        planner_a = planner_a[planner_a["PlannerName"].astype(str).str.strip().ne("")]
    if not planner_b.empty:
        planner_b = planner_b[planner_b["PlannerName"].astype(str).str.strip().ne("")]

    if planner_a.empty and planner_b.empty:
        ws["A3"] = "No planner comparison data is available for this result."
        return

    planner_compare = planner_a.merge(
        planner_b,
        on="PlannerName",
        how="outer",
        suffixes=("_ModeA", "_ModeB"),
    ).fillna(0.0)
    planner_compare["Demand_Delta"] = planner_compare["Demand_Tons_ModeB"] - planner_compare["Demand_Tons_ModeA"]
    planner_compare["Internal_Delta"] = planner_compare["Internal_Tons_ModeB"] - planner_compare["Internal_Tons_ModeA"]
    planner_compare["Outsourced_Delta"] = planner_compare["Outsourced_Tons_ModeB"] - planner_compare["Outsourced_Tons_ModeA"]
    planner_compare["Unmet_Delta"] = planner_compare["Unmet_Tons_ModeB"] - planner_compare["Unmet_Tons_ModeA"]
    planner_compare["Service_Level_Delta"] = planner_compare["Service_Level_ModeB"] - planner_compare["Service_Level_ModeA"]
    planner_compare["SortKey"] = planner_compare[["Demand_Tons_ModeA", "Demand_Tons_ModeB"]].max(axis=1)
    planner_compare = planner_compare.sort_values(["SortKey", "PlannerName"], ascending=[False, True])

    layout = _write_table(
        ws,
        planner_compare[
            [
                "PlannerName",
                "Demand_Tons_ModeA",
                "Demand_Tons_ModeB",
                "Internal_Tons_ModeA",
                "Internal_Tons_ModeB",
                "Outsourced_Tons_ModeA",
                "Outsourced_Tons_ModeB",
                "Unmet_Tons_ModeA",
                "Unmet_Tons_ModeB",
                "Service_Level_ModeA",
                "Service_Level_ModeB",
                "Service_Level_Delta",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons_ModeA": TONS_FMT,
            "Demand_Tons_ModeB": TONS_FMT,
            "Internal_Tons_ModeA": TONS_FMT,
            "Internal_Tons_ModeB": TONS_FMT,
            "Outsourced_Tons_ModeA": TONS_FMT,
            "Outsourced_Tons_ModeB": TONS_FMT,
            "Unmet_Tons_ModeA": TONS_FMT,
            "Unmet_Tons_ModeB": TONS_FMT,
            "Service_Level_ModeA": PCT_FMT,
            "Service_Level_ModeB": PCT_FMT,
            "Service_Level_Delta": PCT_FMT,
        },
        freeze="B2",
    )

    service_chart = BarChart()
    service_chart.title = "Planner Service Level Comparison"
    service_chart.y_axis.title = "Service level"
    service_chart.height = 8
    service_chart.width = 12
    service_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["Service_Level_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["Service_Level_ModeB"],
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        ),
        titles_from_data=True,
        from_rows=False,
    )
    service_chart.set_categories(
        Reference(
            ws,
            min_col=layout["col_index"]["PlannerName"],
            min_row=layout["start_row"] + 1,
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        )
    )
    service_chart.y_axis.numFmt = PCT_FMT
    ws.add_chart(service_chart, "N3")

    unmet_chart = BarChart()
    unmet_chart.type = "bar"
    unmet_chart.title = "Planner Residual Unmet Comparison"
    unmet_chart.height = 8
    unmet_chart.width = 12
    unmet_chart.add_data(
        Reference(
            ws,
            min_col=layout["col_index"]["Unmet_Tons_ModeA"],
            min_row=layout["start_row"],
            max_col=layout["col_index"]["Unmet_Tons_ModeB"],
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        ),
        titles_from_data=True,
        from_rows=False,
    )
    unmet_chart.set_categories(
        Reference(
            ws,
            min_col=layout["col_index"]["PlannerName"],
            min_row=layout["start_row"] + 1,
            max_row=min(layout["end_row"], layout["start_row"] + 12),
        )
    )
    ws.add_chart(unmet_chart, "N21")

    focus_planner = str(planner_compare.iloc[0]["PlannerName"]) if not planner_compare.empty else ""
    if focus_planner:
        ws["A24"] = f"Focused planner comparison: {focus_planner}"
        ws["A24"].font = Font(bold=True, color="1F4E79", size=11)

        planner_month_a = (
            artifacts["ModeA"]["analysis"]["planner_product_month_summary"]
            .query("PlannerName == @focus_planner")
            .groupby("Month", as_index=False)
            .agg(
                Demand_ModeA=("Demand_Tons", "sum"),
                Internal_ModeA=("Internal_Tons", "sum"),
                Outsourced_ModeA=("Outsourced_Tons", "sum"),
                Unmet_ModeA=("Unmet_Tons", "sum"),
            )
        )
        planner_month_b = (
            artifacts["ModeB"]["analysis"]["planner_product_month_summary"]
            .query("PlannerName == @focus_planner")
            .groupby("Month", as_index=False)
            .agg(
                Demand_ModeB=("Demand_Tons", "sum"),
                Internal_ModeB=("Internal_Tons", "sum"),
                Outsourced_ModeB=("Outsourced_Tons", "sum"),
                Unmet_ModeB=("Unmet_Tons", "sum"),
            )
        )
        planner_month_compare = planner_month_a.merge(planner_month_b, on="Month", how="outer").fillna(0.0).sort_values("Month")
        planner_month_compare["Service_Level_ModeA"] = (
            planner_month_compare["Internal_ModeA"] + planner_month_compare["Outsourced_ModeA"]
        ).div(planner_month_compare["Demand_ModeA"].replace(0, pd.NA)).fillna(0.0)
        planner_month_compare["Service_Level_ModeB"] = (
            planner_month_compare["Internal_ModeB"] + planner_month_compare["Outsourced_ModeB"]
        ).div(planner_month_compare["Demand_ModeB"].replace(0, pd.NA)).fillna(0.0)

        focus_layout = _write_table(
            ws,
            planner_month_compare[
                [
                    "Month",
                    "Demand_ModeA",
                    "Demand_ModeB",
                    "Internal_ModeA",
                    "Internal_ModeB",
                    "Outsourced_ModeA",
                    "Outsourced_ModeB",
                    "Unmet_ModeA",
                    "Unmet_ModeB",
                    "Service_Level_ModeA",
                    "Service_Level_ModeB",
                ]
            ],
            start_row=25,
            start_col=1,
            num_formats={
                "Demand_ModeA": TONS_FMT,
                "Demand_ModeB": TONS_FMT,
                "Internal_ModeA": TONS_FMT,
                "Internal_ModeB": TONS_FMT,
                "Outsourced_ModeA": TONS_FMT,
                "Outsourced_ModeB": TONS_FMT,
                "Unmet_ModeA": TONS_FMT,
                "Unmet_ModeB": TONS_FMT,
                "Service_Level_ModeA": PCT_FMT,
                "Service_Level_ModeB": PCT_FMT,
            },
        )

        focus_chart = LineChart()
        focus_chart.title = f"{focus_planner} Monthly Service Level Comparison"
        focus_chart.height = 7
        focus_chart.width = 12
        focus_chart.add_data(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Service_Level_ModeA"],
                min_row=focus_layout["start_row"],
                max_col=focus_layout["col_index"]["Service_Level_ModeB"],
                max_row=focus_layout["end_row"],
            ),
            titles_from_data=True,
            from_rows=False,
        )
        focus_chart.set_categories(
            Reference(
                ws,
                min_col=focus_layout["col_index"]["Month"],
                min_row=focus_layout["start_row"] + 1,
                max_row=focus_layout["end_row"],
            )
        )
        focus_chart.y_axis.numFmt = PCT_FMT
        ws.add_chart(focus_chart, "M25")

    _write_note(
        ws,
        f"A{max(layout['end_row'] + 2, 40)}",
        "Planner comparison is based on planner-level traceability after the product-month optimization result is proportionally split back to planner demand shares.",
    )
    _autofit(ws)


def _write_comparison_run_info(
    wb: Workbook,
    config: Config,
    metrics_by_mode: dict[str, dict[str, Any]],
    workbook_name: str,
) -> None:
    ws = wb.create_sheet("Run_Info")
    _write_sheet_title(ws, "Comparison Run Info")
    info_df = pd.DataFrame(
        [
            ("Workbook_Name", workbook_name),
            ("Scenario_Name", config.scenario_name),
            ("Modes_Included", "ModeA + ModeB"),
            ("Start_Month", config.start_month),
            ("Horizon_Months", config.horizon_months),
            ("Input_Load_Folder", config.input_load_folder),
            ("Input_Master_Folder", config.input_master_folder),
            ("Output_Folder", config.output_folder),
            ("Project_Root_Folder", getattr(config, "project_root_folder", "")),
            ("Run_Timestamp", config.run_timestamp or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("License_Status", getattr(config, "license_status", "")),
            ("License_ID", getattr(config, "license_id", "")),
            ("License_Type", getattr(config, "license_type", "")),
            ("Licensed_To", getattr(config, "licensed_to", "")),
            ("License_Expiry", getattr(config, "license_expiry", "")),
            ("License_Binding_Mode", getattr(config, "license_binding_mode", "")),
            ("License_Machine_Label", getattr(config, "license_machine_label", "")),
            ("ModeA_Service_Level", metrics_by_mode["ModeA"]["service_level"] / 100.0),
            ("ModeB_Service_Level", metrics_by_mode["ModeB"]["service_level"] / 100.0),
        ],
        columns=["Parameter", "Value"],
    )
    layout = _write_table(ws, info_df, start_row=3, start_col=1)
    for row_num in range(layout["start_row"] + 1, layout["end_row"] + 1):
        if ws[f"A{row_num}"].value in {"ModeA_Service_Level", "ModeB_Service_Level"}:
            ws[f"B{row_num}"].number_format = PCT_FMT
    _autofit(ws)


def _write_monthly_analysis(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Monthly_Trend")
    _write_sheet_title(ws, "Monthly Trend")

    monthly_summary = analysis["monthly_summary"]
    if monthly_summary.empty:
        ws["A3"] = "No monthly summary is available for this result."
        return

    monthly_layout = _write_table(
        ws,
        monthly_summary[
            [
                "Month",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
    )

    balance_chart = BarChart()
    balance_chart.grouping = "stacked"
    balance_chart.overlap = 100
    balance_chart.title = "Monthly Supply-Demand Balance"
    balance_chart.y_axis.title = "Tons"
    balance_chart.height = 8
    balance_chart.width = 15
    balance_chart.add_data(
        Reference(
            ws,
            min_col=monthly_layout["col_index"]["Internal_Tons"],
            min_row=monthly_layout["start_row"],
            max_col=monthly_layout["col_index"]["Unmet_Tons"],
            max_row=monthly_layout["end_row"],
        ),
        titles_from_data=True,
    )
    balance_chart.set_categories(
        Reference(
            ws,
            min_col=monthly_layout["col_index"]["Month"],
            min_row=monthly_layout["start_row"] + 1,
            max_row=monthly_layout["end_row"],
        )
    )

    demand_line = LineChart()
    demand_line.add_data(
        Reference(
            ws,
            min_col=monthly_layout["col_index"]["Demand_Tons"],
            min_row=monthly_layout["start_row"],
            max_row=monthly_layout["end_row"],
        ),
        titles_from_data=True,
    )
    balance_chart += demand_line
    ws.add_chart(balance_chart, "J3")

    service_chart = LineChart()
    service_chart.title = "Monthly Service Level"
    service_chart.y_axis.title = "Service level"
    service_chart.height = 7
    service_chart.width = 15
    service_chart.add_data(
        Reference(
            ws,
            min_col=monthly_layout["col_index"]["Service_Level"],
            min_row=monthly_layout["start_row"],
            max_row=monthly_layout["end_row"],
        ),
        titles_from_data=True,
    )
    service_chart.set_categories(
        Reference(
            ws,
            min_col=monthly_layout["col_index"]["Month"],
            min_row=monthly_layout["start_row"] + 1,
            max_row=monthly_layout["end_row"],
        )
    )
    service_chart.y_axis.numFmt = PCT_FMT
    ws.add_chart(service_chart, "J21")

    gap_table = monthly_summary.sort_values(["Unmet_Tons", "Demand_Tons"], ascending=[False, False]).head(8)
    ws["A24"] = "Highest gap months"
    ws["A24"].font = Font(bold=True, color="1F4E79", size=11)
    _write_table(
        ws,
        gap_table[["Month", "Demand_Tons", "Internal_Tons", "Outsourced_Tons", "Unmet_Tons", "Service_Level"]],
        start_row=25,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
    )
    _autofit(ws)


def _write_bottleneck_analysis(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Bottleneck")
    _write_sheet_title(ws, "Bottleneck Analysis")

    wc_summary = analysis["wc_summary"]
    wc_long = analysis["wc_long"]
    detail_df = analysis["detail"]

    if wc_summary.empty or wc_long.empty:
        ws["A3"] = "No work-center load data is available for this result."
        return

    top_wc = wc_summary.head(12).copy()
    top_layout = _write_table(
        ws,
        top_wc[["WorkCenter", "AvgLoadPct", "PeakLoadPct", "MinLoadPct", "StdLoadPct", "Over95Months"]],
        start_row=3,
        start_col=1,
        num_formats={
            "AvgLoadPct": PCT_FMT,
            "PeakLoadPct": PCT_FMT,
            "MinLoadPct": PCT_FMT,
            "StdLoadPct": PCT_FMT,
            "Over95Months": INT_FMT,
        },
    )

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 10
    bar_chart.title = "Top Bottleneck WorkCenters"
    bar_chart.x_axis.title = "Peak load"
    bar_chart.y_axis.title = "WorkCenter"
    bar_chart.height = 8
    bar_chart.width = 12
    bar_chart.legend = None
    bar_chart.add_data(
        Reference(
            ws,
            min_col=top_layout["col_index"]["PeakLoadPct"],
            min_row=top_layout["start_row"],
            max_row=top_layout["end_row"],
        ),
        titles_from_data=True,
    )
    bar_chart.set_categories(
        Reference(
            ws,
            min_col=top_layout["col_index"]["WorkCenter"],
            min_row=top_layout["start_row"] + 1,
            max_row=top_layout["end_row"],
        )
    )
    bar_chart.x_axis.numFmt = PCT_FMT
    ws.add_chart(bar_chart, "H3")

    focus_wc = str(top_wc.iloc[0]["WorkCenter"])
    ws["A20"] = f"Focused workcenter: {focus_wc}"
    ws["A20"].font = Font(bold=True, color="1F4E79", size=11)

    wc_line_df = wc_long[wc_long["WorkCenter"] == focus_wc].sort_values("Month")
    line_layout = _write_table(
        ws,
        wc_line_df[["Month", "LoadPct"]],
        start_row=21,
        start_col=1,
        num_formats={"LoadPct": PCT_FMT},
    )
    wc_line_chart = LineChart()
    wc_line_chart.title = f"{focus_wc} Load Trend"
    wc_line_chart.y_axis.title = "Load"
    wc_line_chart.height = 7
    wc_line_chart.width = 12
    wc_line_chart.add_data(
        Reference(
            ws,
            min_col=line_layout["col_index"]["LoadPct"],
            min_row=line_layout["start_row"],
            max_row=line_layout["end_row"],
        ),
        titles_from_data=True,
    )
    wc_line_chart.set_categories(
        Reference(
            ws,
            min_col=line_layout["col_index"]["Month"],
            min_row=line_layout["start_row"] + 1,
            max_row=line_layout["end_row"],
        )
    )
    wc_line_chart.y_axis.numFmt = PCT_FMT
    ws.add_chart(wc_line_chart, "D21")

    wc_product_mix = (
        detail_df[
            (detail_df["AllocationType"] == "Internal")
            & (detail_df["WorkCenter"] == focus_wc)
        ]
        .groupby(["Product", "ProductFamily"], as_index=False)["Allocated_Tons"]
        .sum()
        .sort_values("Allocated_Tons", ascending=False)
        .head(10)
    )
    if wc_product_mix.empty:
        _write_note(ws, "J21", "No internal product allocation is recorded for the focused workcenter.")
    else:
        mix_layout = _write_table(
            ws,
            wc_product_mix,
            start_row=21,
            start_col=10,
            num_formats={"Allocated_Tons": TONS_FMT},
        )
        wc_mix_chart = BarChart()
        wc_mix_chart.type = "bar"
        wc_mix_chart.title = f"{focus_wc} Product Mix"
        wc_mix_chart.height = 7
        wc_mix_chart.width = 11
        wc_mix_chart.add_data(
            Reference(
                ws,
                min_col=mix_layout["col_index"]["Allocated_Tons"],
                min_row=mix_layout["start_row"],
                max_row=mix_layout["end_row"],
            ),
            titles_from_data=True,
        )
        wc_mix_chart.set_categories(
            Reference(
                ws,
                min_col=mix_layout["col_index"]["Product"],
                min_row=mix_layout["start_row"] + 1,
                max_row=mix_layout["end_row"],
            )
        )
        wc_mix_chart.legend = None
        ws.add_chart(wc_mix_chart, "M21")

    _write_note(
        ws,
        "A38",
        "The full heatmap is written to the separate 'WC_Heatmap' sheet to keep long planning horizons readable.",
    )
    _autofit(ws)


def _write_wc_heatmap(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("WC_Heatmap")
    _write_sheet_title(ws, "WorkCenter Load Heatmap")

    wc_long = analysis["wc_long"]
    wc_summary = analysis["wc_summary"]
    if wc_long.empty or wc_summary.empty:
        ws["A3"] = "No heatmap data is available for this result."
        return

    heatmap_wc_names = wc_summary.head(12)["WorkCenter"].tolist()
    heatmap_df = wc_long[wc_long["WorkCenter"].isin(heatmap_wc_names)].copy()
    pivot = heatmap_df.pivot(index="WorkCenter", columns="Month", values="LoadPct").fillna(0.0)
    pivot = pivot.reindex(index=heatmap_wc_names)
    pivot.reset_index(inplace=True)

    layout = _write_table(
        ws,
        pivot,
        start_row=3,
        start_col=1,
        num_formats={column: PCT_FMT for column in pivot.columns if column != "WorkCenter"},
    )
    if layout["end_row"] > layout["start_row"]:
        ws.conditional_formatting.add(
            f"{get_column_letter(layout['start_col'] + 1)}{layout['start_row'] + 1}:"
            f"{get_column_letter(layout['end_col'])}{layout['end_row']}",
            ColorScaleRule(
                start_type="num",
                start_value=0.0,
                start_color="FFF2CC",
                mid_type="num",
                mid_value=0.85,
                mid_color="F4B183",
                end_type="num",
                end_value=1.0,
                end_color="C00000",
            ),
        )
    _write_note(
        ws,
        "A20",
        "Heatmap colors are based on absolute load percentage so 100% cells stay visually consistent.",
    )
    _autofit(ws)


def _write_product_risk_analysis(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Product_Risk")
    _write_sheet_title(ws, "Product Risk")

    product_summary = analysis["product_summary"]
    product_month_summary = analysis["product_month_summary"]
    detail_df = analysis["detail"]

    if product_summary.empty or product_month_summary.empty:
        ws["A3"] = "No product risk view is available for this result."
        return

    top_products = product_summary.head(20).copy()
    _write_table(
        ws,
        top_products[
            [
                "Product",
                "ProductFamily",
                "Plant",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
    )

    chart_source = top_products.head(12)
    chart_layout = _write_table(
        ws,
        chart_source[["Product", "Unmet_Tons"]],
        start_row=3,
        start_col=11,
        num_formats={"Unmet_Tons": TONS_FMT},
    )
    unmet_chart = BarChart()
    unmet_chart.type = "bar"
    unmet_chart.title = "Top Products by Residual Unmet"
    unmet_chart.height = 8
    unmet_chart.width = 11
    unmet_chart.legend = None
    unmet_chart.add_data(
        Reference(
            ws,
            min_col=chart_layout["col_index"]["Unmet_Tons"],
            min_row=chart_layout["start_row"],
            max_row=chart_layout["end_row"],
        ),
        titles_from_data=True,
    )
    unmet_chart.set_categories(
        Reference(
            ws,
            min_col=chart_layout["col_index"]["Product"],
            min_row=chart_layout["start_row"] + 1,
            max_row=chart_layout["end_row"],
        )
    )
    ws.add_chart(unmet_chart, "N3")

    focus_product = str(product_summary.iloc[0]["Product"])
    ws["A28"] = f"Focused product: {focus_product}"
    ws["A28"].font = Font(bold=True, color="1F4E79", size=11)

    product_month_df = product_month_summary[product_month_summary["Product"] == focus_product].sort_values("Month")
    month_layout = _write_table(
        ws,
        product_month_df[
            [
                "Month",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Service_Level",
            ]
        ],
        start_row=29,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
    )
    balance_chart = BarChart()
    balance_chart.grouping = "stacked"
    balance_chart.overlap = 100
    balance_chart.title = f"{focus_product} Monthly Balance"
    balance_chart.height = 7
    balance_chart.width = 12
    balance_chart.add_data(
        Reference(
            ws,
            min_col=month_layout["col_index"]["Internal_Tons"],
            min_row=month_layout["start_row"],
            max_col=month_layout["col_index"]["Unmet_Tons"],
            max_row=month_layout["end_row"],
        ),
        titles_from_data=True,
    )
    balance_chart.set_categories(
        Reference(
            ws,
            min_col=month_layout["col_index"]["Month"],
            min_row=month_layout["start_row"] + 1,
            max_row=month_layout["end_row"],
        )
    )
    demand_line = LineChart()
    demand_line.add_data(
        Reference(
            ws,
            min_col=month_layout["col_index"]["Demand_Tons"],
            min_row=month_layout["start_row"],
            max_row=month_layout["end_row"],
        ),
        titles_from_data=True,
    )
    balance_chart += demand_line
    ws.add_chart(balance_chart, "D29")

    product_wc_df = (
        detail_df[
            (detail_df["AllocationType"] == "Internal")
            & (detail_df["Product"] == focus_product)
            & (detail_df["WorkCenter"] != "[UNALLOCATED]")
        ]
        .groupby("WorkCenter", as_index=False)["Allocated_Tons"]
        .sum()
        .sort_values("Allocated_Tons", ascending=False)
        .head(10)
    )
    if product_wc_df.empty:
        _write_note(ws, "J29", "This product has no internal workcenter allocation in the selected result.")
    else:
        product_wc_layout = _write_table(
            ws,
            product_wc_df,
            start_row=29,
            start_col=11,
            num_formats={"Allocated_Tons": TONS_FMT},
        )
        product_wc_chart = BarChart()
        product_wc_chart.type = "bar"
        product_wc_chart.title = f"{focus_product} WorkCenter Mix"
        product_wc_chart.height = 7
        product_wc_chart.width = 11
        product_wc_chart.add_data(
            Reference(
                ws,
                min_col=product_wc_layout["col_index"]["Allocated_Tons"],
                min_row=product_wc_layout["start_row"],
                max_row=product_wc_layout["end_row"],
            ),
            titles_from_data=True,
        )
        product_wc_chart.set_categories(
            Reference(
                ws,
                min_col=product_wc_layout["col_index"]["WorkCenter"],
                min_row=product_wc_layout["start_row"] + 1,
                max_row=product_wc_layout["end_row"],
            )
        )
        product_wc_chart.legend = None
        ws.add_chart(product_wc_chart, "N29")

    _write_note(
        ws,
        "A52",
        "The table at the top of this sheet is the Excel version of the former product risk table.",
    )
    _autofit(ws)


def _write_planner_summary(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Planner_Result_Summary")
    _write_sheet_title(ws, "Planner Result Summary")

    planner_summary = analysis.get("planner_summary", pd.DataFrame())
    if planner_summary.empty:
        ws["A3"] = "No planner summary is available for this result."
        return

    layout = _write_table(
        ws,
        planner_summary[
            [
                "PlannerName",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
    )
    _write_note(
        ws,
        f"A{layout['end_row'] + 2}",
        "This sheet shows planner-level traceability after the product-month optimization result is split back to planner shares.",
    )
    _autofit(ws)


def _write_planner_product_month_summary(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Planner_Product_Month")
    _write_sheet_title(ws, "Planner Product Month Summary")

    planner_product_month = analysis.get("planner_product_month_summary", pd.DataFrame())
    if planner_product_month.empty:
        ws["A3"] = "No planner product-month summary is available for this result."
        return

    layout = _write_table(
        ws,
        planner_product_month[
            [
                "Month",
                "PlannerName",
                "Product",
                "ProductFamily",
                "Plant",
                "Demand_Tons",
                "Internal_Tons",
                "Outsourced_Tons",
                "Unmet_Tons",
                "Supplied_Tons",
                "Service_Level",
            ]
        ],
        start_row=3,
        start_col=1,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Internal_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "Supplied_Tons": TONS_FMT,
            "Service_Level": PCT_FMT,
        },
        freeze="C2",
    )
    _write_note(
        ws,
        f"A{layout['end_row'] + 2}",
        "Use this sheet when you need to trace a planner's monthly demand into internal, outsourced, and unmet outcomes.",
    )
    _autofit(ws)


def _write_detail(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Allocation_Detail")
    _write_table(
        ws,
        df,
        num_formats={
            "Demand_Tons": TONS_FMT,
            "Allocated_Tons": TONS_FMT,
            "Outsourced_Tons": TONS_FMT,
            "Unmet_Tons": TONS_FMT,
            "CapacityShare_Pct": PCT_FMT,
        },
        freeze="C2",
    )


def _write_allocation_summary(wb: Workbook, df: pd.DataFrame, months: List[str]) -> None:
    ws = wb.create_sheet("Allocation_Summary")
    internal_df = df[df["AllocationType"] == "Internal"]
    if internal_df.empty:
        ws["A1"] = "No data"
        return
    index_cols = ["Product", "ProductFamily", "Plant"]
    if "PlannerName" in internal_df.columns:
        index_cols = ["PlannerName", *index_cols]
    pivot = internal_df.pivot_table(
        index=index_cols,
        columns="Month",
        values="Allocated_Tons",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    freeze_col = get_column_letter(len(index_cols) + 1)
    _write_table(ws, pivot, tons_cols=pivot.columns[len(index_cols):].tolist(), freeze=f"{freeze_col}2")


def _write_outsource_summary(wb: Workbook, df: pd.DataFrame, months: List[str]) -> None:
    ws = wb.create_sheet("Outsource_Summary")
    outsource_df = df[df["AllocationType"] == "Outsourced"]
    if outsource_df.empty:
        ws["A1"] = "No data"
        return

    index_cols = ["Product", "ProductFamily", "Plant"]
    if "PlannerName" in outsource_df.columns:
        index_cols = ["PlannerName", *index_cols]
    pivot = outsource_df.pivot_table(
        index=index_cols,
        columns="Month",
        values="Outsourced_Tons",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    freeze_col = get_column_letter(len(index_cols) + 1)
    _write_table(
        ws,
        pivot,
        tons_cols=pivot.columns[len(index_cols):].tolist(),
        freeze=f"{freeze_col}2",
        highlight_positive_cols=pivot.columns[len(index_cols):].tolist(),
    )


def _write_unmet_summary(wb: Workbook, df: pd.DataFrame, months: List[str]) -> None:
    ws = wb.create_sheet("Unmet_Summary")
    if df.empty:
        ws["A1"] = "No data"
        return

    index_cols = ["Product", "ProductFamily", "Plant"]
    if "PlannerName" in df.columns:
        index_cols = ["PlannerName", *index_cols]
    pivot = df.pivot_table(
        index=index_cols,
        columns="Month",
        values="Unmet_Tons",
        aggfunc="max",
        fill_value=0,
    )
    pivot = pivot.reindex(columns=[month for month in months if month in pivot.columns], fill_value=0)
    pivot.reset_index(inplace=True)
    freeze_col = get_column_letter(len(index_cols) + 1)
    _write_table(
        ws,
        pivot,
        tons_cols=pivot.columns[len(index_cols):].tolist(),
        freeze=f"{freeze_col}2",
        highlight_positive_cols=pivot.columns[len(index_cols):].tolist(),
    )


def _write_wc_load(wb: Workbook, wc_load_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("WC_Load_Pct")
    if wc_load_df.empty:
        ws["A1"] = "No data"
        return
    pct_cols = [column for column in wc_load_df.columns if column != "WorkCenter"]
    _write_table(
        ws,
        wc_load_df,
        num_formats={column: PCT_FMT for column in pct_cols},
        freeze="B2",
        highlight_over_100_pct=pct_cols,
    )


def _write_binary_report(
    wb: Workbook,
    df: pd.DataFrame,
    months: List[str],
    toller_products: set,
) -> None:
    ws = wb.create_sheet("Binary_Feasibility")
    if df.empty:
        ws["A1"] = "No data"
        return

    unmet_pivot = df.pivot_table(
        index=["Product", "ProductFamily", "Plant"],
        columns="Month",
        values="Unmet_Tons",
        aggfunc="max",
        fill_value=0,
    )
    unmet_pivot = unmet_pivot.reindex(columns=[month for month in months if month in unmet_pivot.columns], fill_value=0)
    binary_pivot = (unmet_pivot <= 0.01).astype(int)
    binary_pivot.reset_index(inplace=True)

    if toller_products:
        binary_pivot["Product"] = binary_pivot["Product"].apply(
            lambda product: f"{product} [Toller]" if product in toller_products else product
        )

    layout = _write_table(
        ws,
        binary_pivot,
        num_formats={
            column: "0"
            for column in binary_pivot.columns
            if column not in {"Product", "ProductFamily", "Plant"}
        },
        freeze="D2",
    )

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    month_cols = [month for month in months if month in binary_pivot.columns]
    for column in month_cols:
        col_num = layout["col_index"][column]
        for row in range(layout["start_row"] + 1, layout["end_row"] + 1):
            cell = ws.cell(row, col_num)
            if cell.value == 1:
                cell.fill = green_fill
            elif cell.value == 0:
                cell.fill = red_fill

    legend_row = layout["end_row"] + 3
    ws.cell(legend_row, 1).value = "Legend"
    ws.cell(legend_row, 1).font = Font(bold=True)
    ws.cell(legend_row + 1, 1).value = "1 = demand fully met"
    ws.cell(legend_row + 1, 1).fill = green_fill
    ws.cell(legend_row + 2, 1).value = "0 = residual unmet demand"
    ws.cell(legend_row + 2, 1).fill = red_fill
    if toller_products:
        ws.cell(legend_row + 3, 1).value = "[Toller] = toller-eligible product"
        ws.cell(legend_row + 3, 1).font = NOTE_FONT
    _autofit(ws)


def _write_validation(wb: Workbook, issues: List[ValidationIssue]) -> None:
    ws = wb.create_sheet("Validation_Issues")
    if not issues:
        _write_table(ws, pd.DataFrame([{"Severity": "OK", "Check": "Validation", "Detail": "No issues found."}]))
        ws["A2"].fill = OK_FILL
        ws["B2"].fill = OK_FILL
        ws["C2"].fill = OK_FILL
        return

    issue_df = pd.DataFrame(
        [{"Severity": issue.severity, "Check": issue.check, "Detail": issue.detail} for issue in issues]
    )
    layout = _write_table(ws, issue_df)
    for row in range(layout["start_row"] + 1, layout["end_row"] + 1):
        severity = ws.cell(row, layout["col_index"]["Severity"]).value
        fill = ERR_FILL if severity == "ERROR" else WARN_FILL
        for col in range(layout["start_col"], layout["end_col"] + 1):
            ws.cell(row, col).fill = fill
    _autofit(ws)


def _write_run_info(wb: Workbook, run_info_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Run_Info")
    _write_table(ws, run_info_df)
    _autofit(ws)


def _write_sheet_title(ws, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT


def _write_metric_block(
    ws,
    start_row: int,
    start_col: int,
    rows: list[tuple[str, Any, str | None]],
) -> None:
    _write_header_row(ws, ["Metric", "Value"], start_row=start_row, start_col=start_col)
    for offset, (label, value, fmt) in enumerate(rows, start=1):
        label_cell = ws.cell(start_row + offset, start_col)
        value_cell = ws.cell(start_row + offset, start_col + 1)
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(bold=True)
        label_cell.fill = SUBHDR_FILL
        for cell in (label_cell, value_cell):
            cell.border = BORDER
        if fmt:
            value_cell.number_format = fmt


def _write_note(ws, cell_ref: str, text: str) -> None:
    ws[cell_ref] = text
    ws[cell_ref].font = NOTE_FONT
    ws[cell_ref].alignment = Alignment(wrap_text=True, vertical="top")


def _apply_chart_palette(chart, colors: list[str]) -> None:
    for series, color in zip(chart.series, colors):
        try:
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color
        except Exception:
            continue


def _write_header_row(ws, headers: list[str], start_row: int = 1, start_col: int = 1) -> None:
    for index, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + index)
        cell.value = header
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[start_row].height = 24


def _write_table(
    ws,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    num_formats: Optional[dict[str, str]] = None,
    tons_cols: Optional[list[str]] = None,
    freeze: Optional[str] = None,
    highlight_positive_cols: Optional[list[str]] = None,
    highlight_over_100_pct: Optional[list[str]] = None,
) -> dict[str, Any]:
    num_formats = num_formats or {}
    tons_cols = tons_cols or []
    highlight_positive_cols = highlight_positive_cols or []
    highlight_over_100_pct = highlight_over_100_pct or []

    headers = list(df.columns)
    if not headers:
        return {
            "start_row": start_row,
            "end_row": start_row,
            "start_col": start_col,
            "end_col": start_col,
            "col_index": {},
        }

    _write_header_row(ws, headers, start_row=start_row, start_col=start_col)
    col_index = {header: start_col + offset for offset, header in enumerate(headers)}

    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        row_num = start_row + row_offset
        for header, value in zip(headers, row):
            cell = ws.cell(row_num, col_index[header])
            cell.value = value if not _is_nan(value) else None
            cell.border = BORDER

            fmt = num_formats.get(header)
            if fmt is None and header in tons_cols:
                fmt = TONS_FMT
            if fmt:
                cell.number_format = fmt

            if header in highlight_positive_cols and isinstance(value, (int, float)) and value > 0:
                cell.fill = ERR_FILL
            if header in highlight_over_100_pct and isinstance(value, (int, float)) and value > 1.0:
                cell.fill = ERR_FILL

    end_row = start_row + len(df)
    end_col = start_col + len(headers) - 1
    if freeze:
        ws.freeze_panes = freeze
    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
        "col_index": col_index,
    }


def _autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column, max_len in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 10), 42)


def _is_nan(value: Any) -> bool:
    try:
        import math

        return math.isnan(value)
    except Exception:
        return False


def _sanitize_filename_segment(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r'[<>:"/\\|?*]+', "-", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("._-")
