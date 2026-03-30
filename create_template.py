"""
Create the Excel control workbook used by the Capacity Optimizer.
"""
from __future__ import annotations

import os
from datetime import datetime

import click
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from data_loader import discover_planner_scenarios


HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
CFG_FILL = PatternFill("solid", fgColor="D9E1F2")
VAL_FILL = PatternFill("solid", fgColor="FFFFFF")
MIGRATION_FILL = PatternFill("solid", fgColor="FCE4D6")
MIGRATION_VAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(size=14, bold=True, color="1F4E79")
BTN_FILL = PatternFill("solid", fgColor="2F75B5")
BTN_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROOT_DIR = os.path.dirname(__file__)
DEFAULT_LOAD_DIR = os.path.join(ROOT_DIR, "Data_Input")
DEFAULT_OUT = os.path.join(ROOT_DIR, "Tooling Control Panel", "Capacity_Optimizer_Control.xlsx")
DEFAULT_PROJECT_ROOT = ".."


@click.command()
@click.option("--out", default=DEFAULT_OUT, show_default=True, help="Output path for the control workbook.")
def main(out: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    lists_ws = workbook.create_sheet("Lists")
    _create_lists(lists_ws)

    deployment_ws = workbook.create_sheet("Deployment_Steps")
    _create_deployment_steps(deployment_ws)

    instructions_ws = workbook.create_sheet("Instructions")
    _create_instructions(instructions_ws)

    control_ws = workbook.create_sheet("Control_Panel")
    _create_control_panel(control_ws, lists_ws)

    lists_ws.sheet_state = "hidden"
    workbook.active = workbook.index(deployment_ws)
    workbook.save(out)
    click.echo(f"Control workbook written to {out}")


def _create_deployment_steps(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 34
    worksheet.column_dimensions["C"].width = 88
    worksheet.column_dimensions["D"].width = 72

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "Deployment Steps - copy the tool, install dependencies, activate the license, and run"
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    worksheet["A1"].fill = HDR_FILL
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].border = BORDER
    worksheet.row_dimensions[1].height = 24

    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = (
        "Recommended portable setup: keep Capacity_Optimizer_Control.xlsx inside 'Tooling Control Panel', "
        "then use Project_Root_Folder = '..', Input_Load_Folder = 'Data_Input', "
        "Input_Master_Folder = 'Data_Input', Output_Folder = 'output', and place 'license.json' in the project root. "
        "中文说明见右侧列。"
    )
    worksheet["A2"].font = Font(color="7F6000", bold=True, size=10)
    worksheet["A2"].fill = MIGRATION_VAL_FILL
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet["A2"].border = BORDER
    worksheet.row_dimensions[2].height = 34

    _write_header(worksheet, 4, ["Step", "Action", "What To Do", "中文说明"])
    steps = [
        (
            "Step 1",
            "Copy the full folder",
            "Copy the entire 'capacity_optimizer' folder to the target computer. Do not copy only the Excel file.",
            "把整个 `capacity_optimizer` 文件夹完整复制到目标电脑，不要只复制 Excel 文件。",
        ),
        (
            "Step 2",
            "Install Python",
            "Install Python on the target computer and confirm 'python --version' works in Command Prompt or PowerShell.",
            "在目标电脑安装 Python，并确认在命令行里执行 `python --version` 可以正常返回版本号。",
        ),
        (
            "Step 3",
            "Run the setup batch file",
            "Double-click 'setup_requirements.bat' in the project root to install the required Python packages automatically.",
            "双击项目根目录里的 `setup_requirements.bat`，自动安装所需 Python 依赖。",
        ),
        (
            "Step 4",
            "Get the machine fingerprint",
            "If RSCP already gave you a trial/unbound license, skip to Step 6. Otherwise double-click 'get_machine_fingerprint.bat' in the project root. It creates 'machine_fingerprint.json' for this computer.",
            "如果 RSCP 已经直接给了试用版或 unbound 授权，可以跳到 Step 6。否则双击项目根目录里的 `get_machine_fingerprint.bat`，为当前电脑生成 `machine_fingerprint.json`。",
        ),
        (
            "Step 5",
            "Request the license file",
            "For a machine-locked license, send 'machine_fingerprint.json' to RSCP and request the signed 'license.json' for this computer. For a short-term trial, RSCP can issue an unbound license directly.",
            "如果要机绑授权，把 `machine_fingerprint.json` 发给 RSCP，并申请这台电脑专用的签名授权文件 `license.json`。如果只是短期试用，RSCP 也可以直接发 unbound 授权。",
        ),
        (
            "Step 6",
            "Place the license file",
            "Copy the signed 'license.json' into the project root, next to main.py and run_optimizer.bat.",
            "把签名后的 `license.json` 放到项目根目录，也就是和 `main.py`、`run_optimizer.bat` 同级的位置。",
        ),
        (
            "Step 7",
            "Open the control workbook",
            "Open 'Tooling Control Panel\\Capacity_Optimizer_Control.xlsx'. Save the workbook after any edits, then go to the 'Control_Panel' sheet after reading this page.",
            "打开 `Tooling Control Panel\\Capacity_Optimizer_Control.xlsx`。每次修改后先保存，再切换到 `Control_Panel`。",
        ),
        (
            "Step 8",
            "Check migration settings",
            "In Control_Panel, first check Project_Root_Folder, Input_Load_Folder, Input_Master_Folder, and Output_Folder. If the workbook stays in 'Tooling Control Panel', keep '..', 'Data_Input', 'Data_Input', and 'output'.",
            "先检查 `Project_Root_Folder`、`Input_Load_Folder`、`Input_Master_Folder`、`Output_Folder`。如果 workbook 仍然放在 `Tooling Control Panel` 目录中，默认保持 `.. / Data_Input / Data_Input / output` 即可。",
        ),
        (
            "Step 9",
            "Check run settings",
            "Then review Scenario_Name, Start_Year, Start_Month_Num, Horizon_Months, Run_Mode, Verbose, and Skip_Validation_Errors.",
            "再检查运行参数，例如 `Scenario_Name`、起始年月、`Horizon_Months`、`Run_Mode`、`Verbose`、`Skip_Validation_Errors`。",
        ),
        (
            "Step 10",
            "Run the tool",
            "Press Ctrl+S to save first. Then click 'Run Optimizer' in Control_Panel, or run: python main.py --input-template \"Tooling Control Panel/Capacity_Optimizer_Control.xlsx\". If required packages or license files are missing, the run batch file will stop and tell you what to do next.",
            "先按 `Ctrl+S` 保存，再在 `Control_Panel` 点击 `Run Optimizer`，或者在命令行运行：`python main.py --input-template \"Tooling Control Panel/Capacity_Optimizer_Control.xlsx\"`。如果依赖或授权文件缺失，运行批处理会先停止并明确告诉你下一步该做什么。",
        ),
        (
            "Step 11",
            "Review output",
            "Open the 'output' folder and check the generated Excel reports. If Run_Mode = Both, the comparison workbook is also generated.",
            "打开 `output` 文件夹查看结果。如果 `Run_Mode = Both`，还会额外生成一份 ModeA 和 ModeB 的对比报告。",
        ),
        (
            "Step 12",
            "If the run fails",
            "Check that the four migration settings point to valid folders, master_capacity.csv and planner files exist, Python packages were installed successfully, and a valid 'license.json' is present in the project root. If needed, run 'setup_requirements.bat' again or regenerate 'machine_fingerprint.json'.",
            "如果运行失败，优先检查四个迁移路径是否正确，`master_capacity.csv` 和 planner 文件是否存在，依赖包是否安装成功，以及项目根目录里是否放了有效的 `license.json`。必要时重新运行 `setup_requirements.bat` 或重新生成 `machine_fingerprint.json`。",
        ),
    ]

    for row_num, (step, action, detail, detail_cn) in enumerate(steps, start=5):
        worksheet[f"A{row_num}"] = step
        worksheet[f"B{row_num}"] = action
        worksheet[f"C{row_num}"] = detail
        worksheet[f"D{row_num}"] = detail_cn
        worksheet[f"A{row_num}"].fill = MIGRATION_FILL
        worksheet[f"A{row_num}"].font = Font(bold=True, color="7F6000")
        worksheet[f"B{row_num}"].font = Font(bold=True)
        worksheet[f"D{row_num}"].fill = SUBHDR_FILL
        worksheet[f"D{row_num}"].font = Font(color="1F1F1F", size=9)
        for col in ("A", "B", "C", "D"):
            worksheet[f"{col}{row_num}"].border = BORDER
            worksheet[f"{col}{row_num}"].alignment = Alignment(wrap_text=True, vertical="top")

    footer_row = 5 + len(steps) + 2
    worksheet.merge_cells(f"A{footer_row}:D{footer_row}")
    worksheet[f"A{footer_row}"] = "Where to continue: open the 'Control_Panel' sheet and start with the Migration Setup section."
    worksheet[f"A{footer_row}"].font = Font(color="1F4E79", bold=True, size=10)
    worksheet[f"A{footer_row}"].fill = SUBHDR_FILL
    worksheet[f"A{footer_row}"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet[f"A{footer_row}"].border = BORDER
    worksheet.freeze_panes = "A5"


def _create_instructions(worksheet) -> None:
    worksheet.column_dimensions["A"].width = 110
    lines = [
        ("CHEMICAL CAPACITY OPTIMIZER - EXCEL WORKFLOW", True),
        ("", False),
        ("1. Run setup_requirements.bat on a new computer.", False),
        ("2. For a short trial, ask RSCP for an unbound trial license and place license.json in the project root.", False),
        ("3. For a machine-locked license, run get_machine_fingerprint.bat and send machine_fingerprint.json to RSCP.", False),
        ("4. Place the signed license.json in the project root.", False),
        ("5. Fill in the Control_Panel sheet.", False),
        ("6. Planner and master data stay in CSV/Excel files on disk; the Python tool reads them directly.", False),
        ("7. Run: python main.py --input-template \"Tooling Control Panel/Capacity_Optimizer_Control.xlsx\"", False),
        ("   Or click the Run button inside Control_Panel.", False),
        ("8. The tool writes one Excel result workbook per mode, with dashboard/report sheets included.", False),
        ("", False),
        ("KEY SETTINGS", True),
        ("Project_Root_Folder: main project folder. Keep '..' when the workbook stays inside Tooling Control Panel.", False),
        ("Input_Load_Folder: planner folder name or path. Relative values are resolved from Project_Root_Folder.", False),
        ("Input_Master_Folder: master-data folder name or path. Relative values are resolved from Project_Root_Folder.", False),
        ("Output_Folder: output folder name or path. Relative values are resolved from Project_Root_Folder.", False),
        ("Output_FileName: base output file name; mode and timestamp are appended automatically.", False),
        ("Scenario_Name: planner scenario to load. If your input folder changes, type the scenario manually if needed.", False),
        ("Start_Year / Start_Month_Num: first planning bucket.", False),
        ("Horizon_Months: number of months to optimize.", False),
        ("Run_Mode: ModeA, ModeB, or Both.", False),
        ("Direct_Mode: keep 'Yes' for the folder-based CSV workflow.", False),
        ("Verbose: prints solver detail in the command window.", False),
        ("Skip_Validation_Errors: only use 'Yes' when you intentionally want a forced run.", False),
        ("", False),
        ("LICENSE", True),
        ("license.json: required in the project root before the optimizer can run.", False),
        ("Trial / unbound license: no machine fingerprint required; RSCP can issue it directly.", False),
        ("get_machine_fingerprint.bat: creates machine_fingerprint.json for RSCP to issue a machine-locked license.", False),
        ("Run_Info: each output workbook records license status, license ID, customer, expiry date, and binding mode.", False),
        ("", False),
        ("RESULT WORKBOOK CONTENT", True),
        ("Dashboard, Monthly_Trend, Bottleneck, WC_Heatmap, Product_Risk", False),
        ("Allocation_Detail, Allocation_Summary, Outsource_Summary, Unmet_Summary, WC_Load_Pct", False),
        ("Binary_Feasibility, Validation_Issues, Run_Info", False),
    ]

    for row_num, (text, bold) in enumerate(lines, start=1):
        cell = worksheet.cell(row_num, 1, text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(
            bold=bold,
            size=10 if bold else 9,
            color="1F4E79" if bold else "000000",
        )


def _create_control_panel(worksheet, lists_ws) -> None:
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 54
    worksheet.column_dimensions["C"].width = 64

    _write_header(worksheet, 1, ["Parameter", "Value", "Description / Example"])

    worksheet.merge_cells("A2:C2")
    worksheet["A2"] = "Migration Setup - update these cells first after copying the tool to another computer"
    worksheet["A2"].font = Font(bold=True, color="FFFFFF", size=11)
    worksheet["A2"].fill = BTN_FILL
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"].border = BORDER
    worksheet.row_dimensions[2].height = 22

    worksheet.merge_cells("A3:C3")
    worksheet["A3"] = (
        "Recommended portable setup: keep Project_Root_Folder = '..', "
        "Input folders = 'Data_Input', Output_Folder = 'output', and place 'license.json' in the project root. "
        "See 'Deployment_Steps' for the full setup checklist."
    )
    worksheet["A3"].font = Font(color="7F6000", bold=True, size=10)
    worksheet["A3"].fill = MIGRATION_VAL_FILL
    worksheet["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet["A3"].border = BORDER
    worksheet.row_dimensions[3].height = 32

    now = datetime.now()
    scenario_options = _scenario_options()
    default_scenario = scenario_options[0] if scenario_options else "Base Scenario"

    migration_rows = [
        ("Project_Root_Folder", DEFAULT_PROJECT_ROOT, "Project root folder; '..' means one level above this workbook"),
        ("Input_Load_Folder", "Data_Input", "Planner folder name or full path"),
        ("Input_Master_Folder", "Data_Input", "Master-data folder name or full path"),
        ("Output_Folder", "output", "Result folder name or full path"),
    ]
    run_rows = [
        ("Output_FileName", "capacity_result.xlsx", "Base name for result workbooks"),
        ("Scenario_Name", default_scenario, "Planner Scenario value; type manually if your data folder changes"),
        ("Start_Year", str(now.year), "First planning year"),
        ("Start_Month_Num", str(now.month), "First planning month number (1-12)"),
        ("Start_Month", f"{now.year}-{now.month:02d}", "Optional legacy field; Python derives this from year/month first"),
        ("Horizon_Months", "60", "Number of monthly buckets"),
        ("Run_Mode", "Both", "ModeA, ModeB, or Both"),
        ("Direct_Mode", "Yes", "Use direct folder-based CSV/Excel inputs"),
        ("Verbose", "No", "Print solver detail in the command window"),
        ("Skip_Validation_Errors", "No", "Run even when validation finds errors"),
        ("Run_Timestamp", "", "Filled by Python at runtime"),
        ("Notes", "", "Free text"),
    ]

    value_rows: dict[str, int] = {}
    current_row = 4
    for key, value, description in migration_rows:
        row_num = current_row
        value_rows[key] = row_num
        key_cell = worksheet.cell(row_num, 1, key)
        val_cell = worksheet.cell(row_num, 2, value)
        desc_cell = worksheet.cell(row_num, 3, description)

        key_cell.font = Font(bold=True, size=10)
        key_cell.fill = MIGRATION_FILL
        key_cell.border = BORDER

        val_cell.fill = MIGRATION_VAL_FILL
        val_cell.border = BORDER
        val_cell.alignment = Alignment(horizontal="left")
        val_cell.font = Font(bold=True, color="7F6000")

        desc_cell.font = Font(italic=True, color="7F6000", size=9)
        desc_cell.fill = MIGRATION_VAL_FILL
        desc_cell.border = BORDER
        desc_cell.alignment = Alignment(wrap_text=True)
        current_row += 1

    worksheet.merge_cells(f"A{current_row}:C{current_row}")
    worksheet[f"A{current_row}"] = "Run Settings"
    worksheet[f"A{current_row}"].font = Font(bold=True, color="FFFFFF", size=11)
    worksheet[f"A{current_row}"].fill = HDR_FILL
    worksheet[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet[f"A{current_row}"].border = BORDER
    worksheet.row_dimensions[current_row].height = 20
    current_row += 1

    for key, value, description in run_rows:
        row_num = current_row
        value_rows[key] = row_num
        key_cell = worksheet.cell(row_num, 1, key)
        val_cell = worksheet.cell(row_num, 2, value)
        desc_cell = worksheet.cell(row_num, 3, description)

        key_cell.font = Font(bold=True, size=10)
        key_cell.fill = CFG_FILL
        key_cell.border = BORDER

        val_cell.fill = VAL_FILL
        val_cell.border = BORDER
        val_cell.alignment = Alignment(horizontal="left")

        desc_cell.font = Font(italic=True, color="666666", size=9)
        desc_cell.border = BORDER
        desc_cell.alignment = Alignment(wrap_text=True)
        current_row += 1

    worksheet.freeze_panes = "B4"
    worksheet[f"A{current_row + 1}"] = "Run command"
    worksheet[f"A{current_row + 1}"].font = Font(bold=True, color="1F4E79", size=11)
    worksheet[f"B{current_row + 1}"] = 'python main.py --input-template "Tooling Control Panel/Capacity_Optimizer_Control.xlsx"'
    worksheet[f"B{current_row + 1}"].alignment = Alignment(wrap_text=True)

    worksheet[f"A{current_row + 3}"] = "Quick actions"
    worksheet[f"A{current_row + 3}"].font = Font(bold=True, color="1F4E79", size=11)
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 3}",
        bottom_right=f"C{current_row + 4}",
        label="Setup Dependencies",
        target=r"..\setup_requirements.bat",
    )
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 6}",
        bottom_right=f"C{current_row + 7}",
        label="Get Machine Fingerprint",
        target=r"..\get_machine_fingerprint.bat",
    )
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 9}",
        bottom_right=f"C{current_row + 10}",
        label="Run Optimizer",
        target=r"..\run_optimizer.bat",
    )
    _add_action_button(
        worksheet,
        top_left=f"B{current_row + 12}",
        bottom_right=f"C{current_row + 13}",
        label="Open Output Folder",
        target=r"..\output",
    )
    worksheet[f"A{current_row + 15}"] = "Button note"
    worksheet[f"A{current_row + 15}"].font = Font(bold=True, color="1F4E79", size=11)
    worksheet[f"B{current_row + 15}"] = (
        "On a new computer, click Setup Dependencies first. If RSCP gave you a trial/unbound license, "
        "place license.json in the project root and continue. Otherwise click Get Machine Fingerprint, "
        "send machine_fingerprint.json to RSCP, place the returned license.json in the project root, "
        "save the workbook (Ctrl+S), and click Run Optimizer."
    )
    worksheet[f"B{current_row + 15}"].alignment = Alignment(wrap_text=True)

    _add_list_validation(worksheet, f"B{value_rows['Run_Mode']}", lists_ws, "$A$2:$A$4")
    _add_list_validation(worksheet, f"B{value_rows['Direct_Mode']}", lists_ws, "$B$2:$B$3")
    _add_list_validation(worksheet, f"B{value_rows['Verbose']}", lists_ws, "$B$2:$B$3")
    _add_list_validation(worksheet, f"B{value_rows['Skip_Validation_Errors']}", lists_ws, "$B$2:$B$3")
    _add_list_validation(worksheet, f"B{value_rows['Start_Month_Num']}", lists_ws, "$C$2:$C$13")
    if lists_ws["D2"].value:
        last_scenario_row = max(2, lists_ws.max_row)
        _add_list_validation(
            worksheet,
            f"B{value_rows['Scenario_Name']}",
            lists_ws,
            f"$D$2:$D${last_scenario_row}",
        )


def _create_lists(worksheet) -> None:
    worksheet["A1"] = "Run_Mode"
    worksheet["A2"] = "ModeA"
    worksheet["A3"] = "ModeB"
    worksheet["A4"] = "Both"

    worksheet["B1"] = "Yes_No"
    worksheet["B2"] = "Yes"
    worksheet["B3"] = "No"

    worksheet["C1"] = "Month_Number"
    for month in range(1, 13):
        worksheet.cell(month + 1, 3, month)

    worksheet["D1"] = "Scenario_Name"
    for row_num, scenario in enumerate(_scenario_options(), start=2):
        worksheet.cell(row_num, 4, scenario)

    for column in range(1, 5):
        worksheet.column_dimensions[get_column_letter(column)].width = 22


def _scenario_options() -> list[str]:
    if os.path.isdir(DEFAULT_LOAD_DIR):
        try:
            options = discover_planner_scenarios(DEFAULT_LOAD_DIR)
            if options:
                return options
        except Exception:
            pass
    return ["Base Scenario", "Baseline", "Case90"]


def _write_header(worksheet, row: int, labels: list[str]) -> None:
    for col_num, label in enumerate(labels, start=1):
        cell = worksheet.cell(row, col_num)
        cell.value = label
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    worksheet.row_dimensions[row].height = 20


def _add_action_button(worksheet, top_left: str, bottom_right: str, label: str, target: str) -> None:
    worksheet.merge_cells(f"{top_left}:{bottom_right}")
    cell = worksheet[top_left]
    cell.value = label
    cell.hyperlink = target
    cell.fill = BTN_FILL
    cell.font = BTN_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

    start_col = worksheet[top_left].column
    end_col = worksheet[bottom_right].column
    start_row = worksheet[top_left].row
    end_row = worksheet[bottom_right].row
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = 22
        for col in range(start_col, end_col + 1):
            worksheet.cell(row, col).border = BORDER


def _add_list_validation(worksheet, target_ref: str, lists_ws, source_range: str) -> None:
    validation = DataValidation(
        type="list",
        formula1=f"=Lists!{source_range}",
        allow_blank=True,
    )
    worksheet.add_data_validation(validation)
    validation.add(worksheet[target_ref])


if __name__ == "__main__":
    main()
